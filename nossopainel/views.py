"""Views responsáveis por dashboards, cadastros e integrações do painel."""

import base64
from .services.wpp import _sanitize_response
from .email_utils import (
    send_profile_change_notification,
    send_password_change_notification,
    send_login_notification,
    send_2fa_enabled_notification
)
import codecs
import io
import json
import logging
import operator
import os
import random
import re
import requests
import threading
import time
import unicodedata
from pathlib import Path
from django.db.models import Sum, Q, Count, F, ExpressionWrapper, DurationField, Exists, OuterRef, Min, Prefetch
from django.db.models.functions import Upper, Coalesce, ExtractDay, Trim
from django.core.exceptions import ObjectDoesNotExist, ValidationError
from django.views.decorators.clickjacking import xframe_options_exempt
from django.views.decorators.cache import cache_page, never_cache
from django.shortcuts import render, redirect, get_object_or_404
from django.db.models.functions import ExtractMonth, ExtractYear
from django.views.decorators.http import require_http_methods, require_GET
from plotly.colors import sample_colorscale, make_colorscale
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views.decorators.vary import vary_on_cookie
from django.views.decorators.http import require_POST
from django.db.models.deletion import ProtectedError
from django.views.decorators.csrf import csrf_exempt
from django.core.validators import validate_email
from django.utils.timezone import localtime, now
from django.contrib.auth.views import LoginView
from django.views.generic.list import ListView
from datetime import timedelta, datetime, date
from django.utils.dateparse import parse_date
from django.forms.models import model_to_dict
from decimal import Decimal, InvalidOperation
from django.views.generic import DetailView
from django.contrib.auth.models import User
from babel.numbers import format_currency
from matplotlib.patches import Patch
from django.contrib import messages
from django.utils import timezone
from django.db import transaction
from django.conf import settings
import matplotlib.pyplot as plt
from plotly.offline import plot
from django.views import View
from .models import DDD_UF_MAP, extrair_pais_do_telefone
from .forms import LoginForm
from typing import Optional
import plotly.express as px
import geopandas as gpd
import pandas as pd
import calendar
import warnings
import inspect

from django.http import (
    HttpResponseBadRequest,
    HttpResponseNotFound,
    HttpResponse, JsonResponse,
    Http404
)
from .models import (
    Cliente, Servidor, Dispositivo,
    Aplicativo, Tipos_pgto, Plano,
    Mensalidade, ContaDoAplicativo,
    SessaoWpp, SecretTokenAPI,
    DadosBancarios, MensagemEnviadaWpp,
    DominiosDNS, PlanoIndicacao,
    HorarioEnvios, NotificationRead,
    UserActionLog, ClientePlanoHistorico,
    ContaReseller, TarefaMigracaoDNS,
    DispositivoMigracaoDNS, AssinaturaCliente,
    # Integração Bancária
    InstituicaoBancaria,
    ContaBancaria,
    ClienteContaBancaria,
    CobrancaPix,
    CredencialAPI,
    ConfiguracaoLimite,
    NotificacaoSistema,
    PlanoLinkPagamento,
)
from .utils import (
    envio_apos_novo_cadastro,
    validar_tel_whatsapp,
    criar_mensalidade,
    log_user_action,
    historico_iniciar,
    historico_encerrar_vigente,
    get_client_ip,
    extrair_dominio_de_url,
    calcular_valor_mensalidade,
    normalizar_dispositivo,
    normalizar_aplicativo,
    normalizar_servidor,
    get_or_create_dispositivo,
    get_or_create_aplicativo,
    get_or_create_servidor,
    enroll_client_in_campaign_if_eligible,
)
from .wpp_views import (
    cancelar_sessao_wpp,
    check_connection_wpp,
    conectar_wpp,
    desconectar_wpp,
    status_wpp,
    whatsapp,
)

# Constantes
PLANOS_MESES = {
    'mensal': 1,
    'bimestral': 2,
    'trimestral': 3,
    'semestral': 6,
    'anual': 12
}

LOG_DIR = os.path.join(settings.BASE_DIR, 'logs')
MESES_31_DIAS = [1, 3, 5, 7, 8, 10, 12]

# Opções de paginação para tabela do dashboard
PAGINATION_OPTIONS = [10, 25, 50, 100]
DEFAULT_PAGINATION = 10

# Opções de ordenação para tabela do dashboard (whitelist de segurança)
SORT_FIELDS = {
    'nome': 'nome',
    'data_adesao': 'data_adesao',
    'contas_count': 'contas_count',
    'ultimo_pagamento': 'ultimo_pagamento',
    'dt_vencimento': 'mensalidade__dt_vencimento',
}
DEFAULT_SORT = 'dt_vencimento'
DEFAULT_ORDER = 'asc'

warnings.filterwarnings(
    "ignore", message="errors='ignore' is deprecated", category=FutureWarning
)
logger = logging.getLogger(__name__)
url_api = os.getenv("URL_API_WPP")

class StaffRequiredMixin(UserPassesTestMixin):
    raise_exception = True

    def test_func(self):
        user = self.request.user
        return user.is_authenticated and (user.is_staff or user.is_superuser)

############################################ AUTH VIEW ############################################

# PÁGINA DE LOGIN
class Login(LoginView):
    template_name = 'login.html'
    form_class = LoginForm
    redirect_authenticated_user = True
    success_url = 'dashboard/'

    def form_valid(self, form):
        """Override to check for 2FA before logging in."""
        from .models import UserProfile
        from django.contrib.auth import authenticate

        username = form.cleaned_data.get('username')
        password = form.cleaned_data.get('password')

        logger.info(f'[LOGIN] Attempting login for user: {username}')

        # Authenticate user (but don't log in yet)
        user = authenticate(self.request, username=username, password=password)

        if user is None:
            logger.warning(f'[LOGIN] Authentication failed for user: {username}')
            return super().form_invalid(form)

        logger.info(f'[LOGIN] Authentication successful for user: {username} (ID: {user.id})')

        # Check if user has 2FA enabled
        try:
            profile = UserProfile.objects.get(user=user)
            logger.info(f'[LOGIN] UserProfile found. 2FA enabled: {profile.two_factor_enabled}, Secret exists: {bool(profile.two_factor_secret)}')

            if profile.two_factor_enabled and profile.two_factor_secret:
                # Store user_id in session for 2FA verification
                self.request.session['pending_2fa_user_id'] = user.id
                self.request.session['pending_2fa_username'] = user.username

                # ✅ SEGURANÇA: Armazenar dados de validação para prevenir session fixation
                from django.utils import timezone
                self.request.session['pending_2fa_timestamp'] = timezone.now().isoformat()
                self.request.session['pending_2fa_ip'] = self.request.META.get('REMOTE_ADDR', '')
                self.request.session['pending_2fa_user_agent'] = self.request.META.get('HTTP_USER_AGENT', '')[:200]

                # Force session save to ensure data persists
                self.request.session.modified = True
                self.request.session.save()

                logger.info(f'[LOGIN] 2FA required for user {user.id}. Session data saved. Redirecting to verify-2fa')
                logger.info(f'[LOGIN] Session key: {self.request.session.session_key}')
                logger.info(f'[LOGIN] Session data: pending_2fa_user_id={user.id}, pending_2fa_username={user.username}')

                return redirect('verify-2fa')
        except UserProfile.DoesNotExist:
            logger.warning(f'[LOGIN] UserProfile not found for user {user.id}. Proceeding without 2FA check.')
            pass

        # No 2FA or not enabled, proceed with normal login
        logger.info(f'[LOGIN] No 2FA required. Proceeding with normal login for user {user.id}')
        return super().form_valid(form)


from django_ratelimit.decorators import ratelimit

@ratelimit(key='ip', rate='5/m', method='POST', block=True)
@require_http_methods(["GET", "POST"])
def verify_2fa_code(request):
    """Verify 2FA code and complete login.

    Rate limiting: 5 tentativas por minuto por IP.
    Após 5 tentativas falhadas, o IP é bloqueado por 1 minuto.
    """
    from .models import UserProfile
    from django.contrib.auth import login, get_user_model
    from django_ratelimit.exceptions import Ratelimited

    # Verificar se foi bloqueado por rate limiting
    was_limited = getattr(request, 'limited', False)
    if was_limited:
        ip = request.META.get('REMOTE_ADDR', 'unknown')
        logger.warning(f'[2FA_RATELIMIT] Rate limit exceeded for IP {ip}')
        messages.error(request, 'Muitas tentativas de verificação. Aguarde 1 minuto e tente novamente.')
        return render(request, 'login_2fa.html', {'rate_limited': True})

    logger.info(f'[2FA_VERIFY] Accessed verify_2fa_code view. Method: {request.method}')
    logger.info(f'[2FA_VERIFY] Session key: {request.session.session_key}')
    logger.info(f'[2FA_VERIFY] Session data: {dict(request.session.items())}')

    # Check if there's a pending 2FA verification
    user_id = request.session.get('pending_2fa_user_id')
    username = request.session.get('pending_2fa_username')
    timestamp_str = request.session.get('pending_2fa_timestamp')
    stored_ip = request.session.get('pending_2fa_ip')
    stored_ua = request.session.get('pending_2fa_user_agent')

    logger.info(f'[2FA_VERIFY] Retrieved from session - user_id: {user_id}, username: {username}')

    # ✅ SEGURANÇA: Validação 1 - Dados existem
    if not all([user_id, username, timestamp_str]):
        logger.warning('[2FA_SECURITY] Missing session data (user_id, username, or timestamp)')
        messages.error(request, 'Sessão inválida. Faça login novamente.')
        request.session.flush()
        return redirect('login')

    # ✅ SEGURANÇA: Validação 2 - Timeout de 5 minutos
    from django.utils import timezone
    from datetime import timedelta, datetime
    try:
        timestamp = datetime.fromisoformat(timestamp_str)
        if timezone.now() - timestamp > timedelta(minutes=5):
            logger.warning(f'[2FA_SECURITY] Session expired for user {user_id} (timeout > 5 min)')
            # ✅ SEGURANÇA: Mensagem genérica para não revelar motivo específico
            messages.error(request, 'Sessão inválida. Faça login novamente.')
            request.session.flush()
            return redirect('login')
    except (ValueError, TypeError) as e:
        logger.error(f'[2FA_SECURITY] Invalid timestamp format: {timestamp_str} - {e}')
        messages.error(request, 'Sessão inválida. Faça login novamente.')
        request.session.flush()
        return redirect('login')

    # ✅ SEGURANÇA: Validação 3 - IP não mudou
    current_ip = request.META.get('REMOTE_ADDR', '')
    if stored_ip and current_ip != stored_ip:
        logger.warning(f'[2FA_SECURITY] IP mismatch for user {user_id}: {stored_ip} != {current_ip}')
        # ✅ SEGURANÇA: Mensagem genérica para não revelar motivo específico
        messages.error(request, 'Sessão inválida. Faça login novamente.')
        request.session.flush()
        return redirect('login')

    # ✅ SEGURANÇA: Validação 4 - User-Agent não mudou (warning only)
    current_ua = request.META.get('HTTP_USER_AGENT', '')[:200]
    if stored_ua and current_ua != stored_ua:
        logger.warning(f'[2FA_SECURITY] User-Agent mismatch for user {user_id}')
        # Não bloqueia, apenas registra (pode ser legítimo)

    User = get_user_model()

    try:
        user = User.objects.get(id=user_id)
        profile = UserProfile.objects.get(user=user)
        logger.info(f'[2FA_VERIFY] User found: {user.username} (ID: {user.id})')
    except (User.DoesNotExist, UserProfile.DoesNotExist) as e:
        logger.error(f'[2FA_VERIFY] User or profile not found for user_id {user_id}: {str(e)}')
        # ✅ SEGURANÇA: Mensagem genérica para não revelar existência do usuário
        messages.error(request, 'Sessão inválida. Faça login novamente.')
        request.session.flush()
        return redirect('login')

    if request.method == 'POST':
        logger.info(f'[2FA_VERIFY] POST request - processing 2FA code verification')
        code = request.POST.get('code', '').strip()

        if not code:
            messages.error(request, 'Digite o código de verificação.')
            return render(request, 'login_2fa.html')

        # Verify code
        if profile.verify_2fa_code(code):
            # Code is valid, complete login
            login(request, user, backend='django.contrib.auth.backends.ModelBackend')

            # ✅ SEGURANÇA: Regenerar session ID para prevenir session fixation
            request.session.cycle_key()

            # ✅ SEGURANÇA: Limpar dados temporários de 2FA
            request.session.pop('pending_2fa_user_id', None)
            request.session.pop('pending_2fa_username', None)
            request.session.pop('pending_2fa_timestamp', None)
            request.session.pop('pending_2fa_ip', None)
            request.session.pop('pending_2fa_user_agent', None)

            # Log de segurança para login bem-sucedido
            ip = request.META.get('REMOTE_ADDR', 'unknown')
            security_logger = logging.getLogger('security')
            security_logger.info(
                f'[2FA_SUCCESS] Successful 2FA login | '
                f'User: {user.username} (ID: {user.id}) | '
                f'IP: {ip}'
            )

            # Send login notification if enabled
            try:
                if profile.email_on_login:
                    ip_address = get_client_ip(request) or 'Desconhecido'
                    send_login_notification(user, ip_address=ip_address, user_agent=request.META.get('HTTP_USER_AGENT', 'Desconhecido'))
            except Exception as e:
                logger.warning(f'[2FA_LOGIN] Email notification failed for user {user.id}: {str(e)}')

            messages.success(request, f'Bem-vindo, {user.first_name or user.username}!')
            return redirect('dashboard')
        else:
            # Check if it's a backup code
            if profile.use_backup_code(code):
                # Backup code is valid, complete login
                login(request, user, backend='django.contrib.auth.backends.ModelBackend')

                # ✅ SEGURANÇA: Regenerar session ID para prevenir session fixation
                request.session.cycle_key()

                # Set flag for LoginLog signal to detect backup code usage
                request.session['backup_code_used'] = True

                # ✅ SEGURANÇA: Limpar dados temporários de 2FA
                request.session.pop('pending_2fa_user_id', None)
                request.session.pop('pending_2fa_username', None)
                request.session.pop('pending_2fa_timestamp', None)
                request.session.pop('pending_2fa_ip', None)
                request.session.pop('pending_2fa_user_agent', None)

                # Log de segurança para uso de backup code (importante para auditoria)
                ip = request.META.get('REMOTE_ADDR', 'unknown')
                security_logger = logging.getLogger('security')
                security_logger.warning(
                    f'[2FA_BACKUP_CODE] Backup code used for login | '
                    f'User: {user.username} (ID: {user.id}) | '
                    f'IP: {ip} | '
                    f'Remaining backup codes: {len(profile.two_factor_backup_codes)}'
                )

                # Send login notification if enabled
                try:
                    if profile.email_on_login:
                        ip_address = get_client_ip(request) or 'Desconhecido'
                        send_login_notification(user, ip_address=ip_address, user_agent=request.META.get('HTTP_USER_AGENT', 'Desconhecido'))
                except Exception as e:
                    logger.warning(f'[2FA_LOGIN] Email notification failed for user {user.id}: {str(e)}')

                messages.warning(request, 'Login realizado com código de backup. Considere regenerar seus códigos.')
                return redirect('dashboard')
            else:
                # Log de segurança para tentativa falhada
                ip = request.META.get('REMOTE_ADDR', 'unknown')
                user_agent = request.META.get('HTTP_USER_AGENT', 'unknown')
                security_logger = logging.getLogger('security')
                security_logger.warning(
                    f'[2FA_FAILED] Invalid 2FA code attempt | '
                    f'User: {user.username} (ID: {user.id}) | '
                    f'IP: {ip} | '
                    f'User-Agent: {user_agent}'
                )
                messages.error(request, 'Código inválido. Tente novamente.')
                return render(request, 'login_2fa.html')

    # GET request, show 2FA form
    logger.info(f'[2FA_VERIFY] GET request - rendering 2FA form for user: {username}')
    return render(request, 'login_2fa.html')

# PÁGINA DE ERRO 404
def not_found(request, exception):
    """Renderiza a página 404 personalizada utilizada em diversos entrypoints."""
    return render(request, 'pages/404-error.html')

############################################ LIST VIEW ############################################

class CarregarContasDoAplicativo(LoginRequiredMixin, View):
    """Retorna as contas de aplicativo associadas a um cliente para exibição no modal."""

    def get(self, request):
        """Lista contas do cliente autenticado já serializadas para o modal."""
        cliente_id = self.request.GET.get("cliente_id")
        cliente = get_object_or_404(Cliente, id=cliente_id, usuario=self.request.user)
        conta_app = (
            ContaDoAplicativo.objects.filter(cliente=cliente, usuario=self.request.user)
            .select_related('app', 'dispositivo', 'cliente__dispositivo')
        )

        conta_app_json = []

        for conta in conta_app:
            nome_aplicativo = conta.app.nome
            nome_dispositivo = (
                conta.dispositivo.nome if conta.dispositivo
                else (conta.cliente.dispositivo.nome if conta.cliente.dispositivo
                else 'Não especificado')
            )
            conta_json = model_to_dict(conta)
            conta_json['nome_aplicativo'] = nome_aplicativo
            conta_json['nome_dispositivo'] = nome_dispositivo
            conta_json['logo_url'] = conta.app.get_logo_url()
            conta_app_json.append(conta_json)

        conta_app_json = sorted(conta_app_json, key=operator.itemgetter('nome_aplicativo'))

        return JsonResponse({"conta_app": conta_app_json}, safe=False)


class CarregarQuantidadesMensalidades(LoginRequiredMixin, View):
    """Expõe um resumo de mensalidades (pagas, pendentes e canceladas) no modal do cliente."""

    def get(self, request):
        """Calcula as contagens por status e devolve um payload JSON para o modal."""
        cliente_id = self.request.GET.get("cliente_id")
        cliente = get_object_or_404(Cliente, id=cliente_id, usuario=self.request.user)
        hoje = timezone.localtime().date()
        mensalidades_totais = Mensalidade.objects.filter(
            usuario=self.request.user,
            cliente=cliente
        ).select_related('cliente__assinatura').order_by('-id').values(
            'id', 'dt_vencimento', 'dt_pagamento', 'valor', 'pgto', 'cancelado',
            'cliente__assinatura__em_campanha',
            # ⭐ FASE 2.5: Novos campos de rastreamento de campanhas
            'gerada_em_campanha',
            'dados_historicos_verificados',
            'valor_base_plano',
            'desconto_campanha',
            'desconto_progressivo',
            'tipo_campanha',
            'numero_mes_campanha'
        )
        mensalidades_pagas = Mensalidade.objects.filter(usuario=self.request.user, pgto=True, cliente=cliente)
        mensalidades_pendentes = Mensalidade.objects.filter(usuario=self.request.user, dt_pagamento=None, pgto=False, cancelado=False, dt_cancelamento=None, dt_vencimento__lt=hoje, cliente=cliente)
        mensalidades_canceladas = Mensalidade.objects.filter(usuario=self.request.user, cancelado=True, cliente=cliente)

        qtd_mensalidades_pagas = 0
        for mensalidade in mensalidades_pagas:
            if mensalidade.cliente.id == cliente.id:
                qtd_mensalidades_pagas += 1

        qtd_mensalidades_pendentes = 0
        for mensalidade in mensalidades_pendentes:
            if mensalidade.cliente.id == cliente.id:
                qtd_mensalidades_pendentes += 1

        qtd_mensalidades_canceladas = 0
        for mensalidade in mensalidades_canceladas:
            if mensalidade.cliente.id == cliente.id:
                qtd_mensalidades_canceladas += 1

        data = {
            'mensalidades_totais': list(mensalidades_totais),
            'qtd_mensalidades_pagas': qtd_mensalidades_pagas,
            'qtd_mensalidades_pendentes': qtd_mensalidades_pendentes,
            'qtd_mensalidades_canceladas': qtd_mensalidades_canceladas
        }

        return JsonResponse(data)


class CarregarInidicacoes(LoginRequiredMixin, View):

    def get(self, request):
        from nossopainel.utils import calcular_desconto_progressivo_total
        from .models import DescontoProgressivoIndicacao

        cliente_id = self.request.GET.get("cliente_id")
        cliente = get_object_or_404(Cliente, id=cliente_id, usuario=self.request.user)
        indicados = Cliente.objects.filter(indicado_por=cliente, usuario=self.request.user).order_by('-id')

        # Calcular informações sobre descontos progressivos
        desconto_info = calcular_desconto_progressivo_total(cliente)

        # Buscar quais indicados geram desconto ativo
        descontos_ativos_ids = set()
        if desconto_info["qtd_descontos_ativos"] > 0:
            descontos_ativos = DescontoProgressivoIndicacao.objects.filter(
                cliente_indicador=cliente,
                ativo=True
            ).values_list('cliente_indicado_id', flat=True)
            descontos_ativos_ids = set(descontos_ativos)

        # Converter indicados para dicionários e adicionar flag de desconto
        indicados_list = []
        for indicado in indicados:
            indicado_dict = {
                'id': indicado.id,
                'nome': indicado.nome,
                'data_adesao': indicado.data_adesao.strftime('%Y-%m-%d') if indicado.data_adesao else None,
                'cancelado': indicado.cancelado,
                'tem_desconto_ativo': indicado.id in descontos_ativos_ids,
            }
            indicados_list.append(indicado_dict)

        data = {
            'indicacoes': indicados_list,
            'desconto_progressivo': {
                'ativo': desconto_info["plano"] is not None and desconto_info["plano"].status,
                'valor_total': float(desconto_info["valor_total"]),
                'qtd_descontos_ativos': desconto_info["qtd_descontos_ativos"],
                'qtd_descontos_aplicados': desconto_info["qtd_descontos_aplicados"],
                'limite_indicacoes': desconto_info["limite_indicacoes"],
            }
        }

        return JsonResponse(data)


class ClientesCancelados(LoginRequiredMixin, ListView):
    """
    View para listar clientes, considerando clientes cancelados e ativos.
    """
    model = Cliente
    template_name = "pages/clientes-cancelados.html"
    paginate_by = 15

    def get_queryset(self):
        """
        Retorna a queryset de clientes para a exibição na página.

        Filtra os clientes do usuário atual e os ordena pela data de adesão.
        Se houver uma consulta (q) na URL, filtra os clientes cujo nome contenha o valor da consulta.
        """
        query = self.request.GET.get("q")
        queryset = (
            Cliente.objects.filter(usuario=self.request.user, cancelado=True)
            .order_by("-data_cancelamento")
        )
        
        if query:
            queryset = queryset.filter(nome__icontains=query)
        return queryset
    
    def get_context_data(self, **kwargs):
        """
        Retorna o contexto dos dados para serem exibidos no template.

        Adiciona informações adicionais ao contexto, como objetos relacionados e variáveis de controle de página.
        """
        context = super().get_context_data(**kwargs)
        page_group = 'clientes'
        page = 'lista-clientes'
        range_num = range(1,32)

        # Adiciona dispositivos e aplicativos para o modal de edição
        dispositivos = Dispositivo.objects.all()
        aplicativos = Aplicativo.objects.all()

        # Serializa para JSON (para uso no JavaScript)
        dispositivos_json = [
            {'id': d.id, 'nome': d.nome}
            for d in dispositivos
        ]
        aplicativos_json = [
            {'id': a.id, 'nome': a.nome, 'device_has_mac': a.device_has_mac}
            for a in aplicativos
        ]

        context.update(
            {
                "range": range_num,
                "page_group": page_group,
                "page": page,
                "dispositivos": dispositivos,
                "aplicativos": aplicativos,
                "dispositivos_json": json.dumps(dispositivos_json),
                "aplicativos_json": json.dumps(aplicativos_json),
            }
        )
        return context
    

class TabelaDashboardAjax(LoginRequiredMixin, ListView):
    model = Cliente
    template_name = "partials/table-clients.html"

    def get_paginate_by(self, queryset):
        """Retorna quantidade de itens por página baseado no parâmetro da URL ou session."""
        per_page = self.request.GET.get('per_page')
        if per_page:
            try:
                per_page = int(per_page)
                if per_page in PAGINATION_OPTIONS:
                    self.request.session['dashboard_per_page'] = per_page
                    return per_page
            except (ValueError, TypeError):
                pass
        return self.request.session.get('dashboard_per_page', DEFAULT_PAGINATION)

    def get_sort_params(self):
        """Extrai e valida parâmetros de ordenação da URL."""
        sort_field = self.request.GET.get('sort', DEFAULT_SORT)
        sort_order = self.request.GET.get('order', DEFAULT_ORDER)

        # Validação: só aceita campos do whitelist
        if sort_field not in SORT_FIELDS:
            sort_field = DEFAULT_SORT
        # Validação: só aceita 'asc' ou 'desc'
        if sort_order not in ('asc', 'desc'):
            sort_order = DEFAULT_ORDER

        return sort_field, sort_order

    def get_queryset(self):
        """
        Retorna a queryset para a listagem de clientes no dashboard.

        Filtra os clientes que não foram cancelados e possuem mensalidades não canceladas, não pagas e sem data de pagamento definida.
        Ordena a queryset pelo campo especificado via parâmetros sort/order.
        Realiza a operação distinct() para evitar duplicatas na listagem.
        Caso haja um valor de busca na URL (parâmetro 'q'), filtra a queryset pelos clientes cujo nome contém o valor de busca.
        """
        query = self.request.GET.get("q")
        queryset = (
            Cliente.objects
            .select_related('dispositivo', 'sistema', 'servidor', 'forma_pgto', 'plano')
            .annotate(
                contas_count=Count('conta_aplicativo', distinct=True)
            )
            .prefetch_related(
                Prefetch(
                    'conta_aplicativo',
                    queryset=ContaDoAplicativo.objects.select_related('dispositivo', 'app'),
                    to_attr='contas_list'
                ),
                Prefetch(
                    'mensalidade_set',
                    queryset=Mensalidade.objects.filter(
                        cancelado=False,
                        dt_cancelamento=None,
                        dt_pagamento=None,
                        pgto=False
                    )
                )
            )
            .filter(cancelado=False).filter(
                mensalidade__cancelado=False,
                mensalidade__dt_cancelamento=None,
                mensalidade__dt_pagamento=None,
                mensalidade__pgto=False,
                usuario=self.request.user,
            ).distinct()
        )
        if query:
            queryset = queryset.filter(
                Q(nome__icontains=query) | Q(telefone__icontains=query)
            )

        # Aplicar ordenação server-side
        sort_field, sort_order = self.get_sort_params()
        db_field = SORT_FIELDS[sort_field]
        if sort_order == 'desc':
            db_field = f'-{db_field}'
        queryset = queryset.order_by(db_field)

        time.sleep(0.5)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["hoje"] = timezone.localtime().date()
        context['pagination_options'] = PAGINATION_OPTIONS
        context['current_per_page'] = self.get_paginate_by(self.get_queryset())
        # Adicionar estado de ordenação ao contexto
        sort_field, sort_order = self.get_sort_params()
        context['current_sort'] = sort_field
        context['current_order'] = sort_order
        return context


class ModalDNSJsonView(LoginRequiredMixin, View):
    def get(self, request):
        dns = DominiosDNS.objects.filter(usuario=request.user).order_by(
            "-status", "-data_online", "-servidor", "dominio"
        )
        data = []
        for d in dns:
            data.append({
                "status": True if d.status == "online" else False,
                "dominio": d.dominio,
                "monitorado": d.monitorado,
                "servidor": str(d.servidor.nome),
                "acesso_canais": d.acesso_canais if d.acesso_canais else "",
                "data_online": d.data_online.strftime('%d/%m/%Y %H:%M') if d.data_online else "",
                "data_offline": d.data_offline.strftime('%d/%m/%Y %H:%M') if d.data_offline else "",
                "data_ultima_verificacao": d.data_ultima_verificacao.strftime('%d/%m/%Y %H:%M') if d.data_ultima_verificacao else "",
            })
        return JsonResponse({"dns": data})


class LogFilesListView(LoginRequiredMixin, StaffRequiredMixin, View):
    def get(self, request):
        base_path = Path(LOG_DIR)
        if not base_path.exists():
            return JsonResponse({"files": []})

        resolved_base = base_path.resolve()
        log_files = []
        for path in resolved_base.rglob("*.log"):
            if not path.is_file():
                continue
            try:
                relative_path = path.relative_to(resolved_base)
            except ValueError:
                continue
            log_files.append(str(relative_path).replace("\\", "/"))

        return JsonResponse({"files": sorted(log_files)})


class LogFileContentView(LoginRequiredMixin, StaffRequiredMixin, View):
    def get(self, request):
        filename = request.GET.get("file", "").strip()
        if not filename:
            return JsonResponse({'error': 'Arquivo não informado.'}, status=400)

        base_path = Path(LOG_DIR)
        if not base_path.exists():
            raise Http404('Arquivo não encontrado.')

        resolved_base = base_path.resolve()
        candidate = (resolved_base / filename).resolve()

        if candidate == resolved_base or resolved_base not in candidate.parents:
            raise Http404('Arquivo não permitido.')

        if candidate.suffix != '.log' or not candidate.is_file():
            raise Http404('Arquivo não encontrado.')

        with candidate.open(encoding='utf-8', errors='replace') as handler:
            lines = handler.readlines()[-2000:]

        return JsonResponse({'content': ''.join(lines)})

class UserActionLogListView(LoginRequiredMixin, View):
    def get(self, request):
        try:
            limit = int(request.GET.get('limit', 50))
        except (TypeError, ValueError):
            limit = 50
        limit = max(1, min(limit, 200))

        user_is_staff = request.user.is_staff or request.user.is_superuser
        queryset = UserActionLog.objects.all() if user_is_staff else UserActionLog.objects.filter(usuario=request.user)

        target_user = None
        user_param = request.GET.get('user')
        if user_is_staff and user_param:
            try:
                target_user = User.objects.get(pk=int(user_param))
            except (ValueError, User.DoesNotExist):
                try:
                    target_user = User.objects.get(username=user_param)
                except User.DoesNotExist:
                    target_user = None
        if target_user:
            queryset = queryset.filter(usuario=target_user)

        action_filter = request.GET.get('action')
        valid_actions = {choice for choice, _ in UserActionLog.ACTION_CHOICES}
        if action_filter in valid_actions:
            queryset = queryset.filter(acao=action_filter)

        search_term = request.GET.get('search')
        if search_term:
            queryset = queryset.filter(
                Q(mensagem__icontains=search_term)
                | Q(entidade__icontains=search_term)
                | Q(objeto_repr__icontains=search_term)
            )

        start_date = request.GET.get('since')
        end_date = request.GET.get('until')
        if start_date:
            parsed = parse_date(start_date)
            if parsed:
                queryset = queryset.filter(criado_em__date__gte=parsed)
        if end_date:
            parsed = parse_date(end_date)
            if parsed:
                queryset = queryset.filter(criado_em__date__lte=parsed)

        queryset = queryset.select_related('usuario')
        total_registros = queryset.count()
        queryset = queryset.order_by('-criado_em')[:limit]

        dados = []
        for log in queryset:
            dados.append({
                'id': log.id,
                'timestamp': timezone.localtime(log.criado_em).strftime('%d/%m/%Y %H:%M:%S'),
                'acao': log.acao,
                'acao_label': log.get_acao_display(),
                'entidade': log.entidade,
                'objeto_id': log.objeto_id,
                'objeto_repr': log.objeto_repr,
                'mensagem': log.mensagem,
                'extras': log.extras or {},
                'ip': log.ip,
                'request_path': log.request_path,
                'usuario': {
                    'id': log.usuario_id,
                    'username': log.usuario.get_username(),
                    'nome': log.usuario.get_full_name() or log.usuario.get_username(),
                },
            })

        return JsonResponse({
            'results': dados,
            'actions': [{'value': value, 'label': label} for value, label in UserActionLog.ACTION_CHOICES],
            'total': total_registros,
            'limit': limit,
            'can_view_all': user_is_staff,
            'selected_user': target_user.id if target_user else request.user.id,
            'users': [
                {
                    'id': user.id,
                    'username': user.get_username(),
                    'nome': user.get_full_name() or user.get_username(),
                }
                for user in User.objects.filter(is_active=True).order_by('username')
            ] if user_is_staff else [],
        })

class TabelaDashboard(LoginRequiredMixin, ListView):
    """
    View para listagem de clientes e outras informações exibidas no dashboard.
    """
    login_url = "login"
    model = Cliente
    template_name = "dashboard.html"

    def get_paginate_by(self, queryset):
        """Retorna quantidade de itens por página baseado no parâmetro da URL ou session."""
        per_page = self.request.GET.get('per_page')
        if per_page:
            try:
                per_page = int(per_page)
                if per_page in PAGINATION_OPTIONS:
                    self.request.session['dashboard_per_page'] = per_page
                    return per_page
            except (ValueError, TypeError):
                pass
        return self.request.session.get('dashboard_per_page', DEFAULT_PAGINATION)

    def get_sort_params(self):
        """Extrai e valida parâmetros de ordenação da URL."""
        sort_field = self.request.GET.get('sort', DEFAULT_SORT)
        sort_order = self.request.GET.get('order', DEFAULT_ORDER)

        # Validação: só aceita campos do whitelist
        if sort_field not in SORT_FIELDS:
            sort_field = DEFAULT_SORT
        # Validação: só aceita 'asc' ou 'desc'
        if sort_order not in ('asc', 'desc'):
            sort_order = DEFAULT_ORDER

        return sort_field, sort_order

    def get_queryset(self):

        query = self.request.GET.get("q")
        queryset = (
            Cliente.objects
            .select_related('dispositivo', 'sistema', 'servidor', 'forma_pgto', 'plano')
            .annotate(
                contas_count=Count('conta_aplicativo', distinct=True)
            )
            .prefetch_related(
                Prefetch(
                    'conta_aplicativo',
                    queryset=ContaDoAplicativo.objects.select_related('dispositivo', 'app'),
                    to_attr='contas_list'
                ),
                Prefetch(
                    'mensalidade_set',
                    queryset=Mensalidade.objects.filter(
                        cancelado=False,
                        dt_cancelamento=None,
                        dt_pagamento=None,
                        pgto=False
                    )
                )
            )
            .filter(cancelado=False).filter(
                mensalidade__cancelado=False,
                mensalidade__dt_cancelamento=None,
                mensalidade__dt_pagamento=None,
                mensalidade__pgto=False,
                usuario=self.request.user,
            ).distinct()
        )
        if query:
            queryset = queryset.filter(nome__icontains=query)

        # Aplicar ordenação server-side
        sort_field, sort_order = self.get_sort_params()
        db_field = SORT_FIELDS[sort_field]
        if sort_order == 'desc':
            db_field = f'-{db_field}'
        queryset = queryset.order_by(db_field)

        return queryset
    
    def get_context_data(self, **kwargs):
        """
        Retorna o contexto de dados para serem exibidos no dashboard.

        Inicializa as variáveis necessárias, como a moeda utilizada, a data de hoje e o ano atual.
        Calcula o total de clientes baseado na queryset.
        Obtém o mês atual.
        Define a página atual como 'dashboard'.
        Filtra os clientes em atraso.
        Calcula o valor total pago no mês atual.
        Calcula a quantidade de mensalidades pagas no mês atual.
        Calcula o valor total a receber no mês atual.
        Calcula a quantidade de mensalidades a receber na próxima semana.
        Calcula a quantidade de novos clientes no mês atual.
        Calcula a quantidade de clientes cancelados no mês atual.
        Obtém a lista de aplicativos do usuário ordenados por nome.
        Atualiza o contexto com as informações calculadas.
        Retorna o contexto atualizado.
        """
        moeda = "BRL"
        page = 'dashboard'
        hoje = timezone.localtime().date()
        f_name = self.request.user.first_name
        ano_atual = timezone.localtime().year
        proxima_semana = hoje + timedelta(days=7)
        dt_inicio = str(self.request.user.date_joined)
        context = super().get_context_data(**kwargs)
        total_clientes = self.get_queryset().count()
        mes_atual = timezone.localtime().date().month

        # Variáveis para context do modal de edição do cadastro do cliente
        indicadores = Cliente.objects.filter(usuario=self.request.user).order_by('nome')
        servidores = Servidor.objects.filter(usuario=self.request.user).order_by('nome')
        formas_pgtos = Tipos_pgto.objects.filter(usuario=self.request.user).select_related('conta_bancaria__instituicao', 'dados_bancarios').order_by('nome')
        planos = Plano.objects.filter(usuario=self.request.user).order_by('nome', 'telas', 'valor')
        dispositivos = Dispositivo.objects.filter(usuario=self.request.user).order_by('nome')
        aplicativos = Aplicativo.objects.filter(usuario=self.request.user).order_by('nome')

        # Serializa dispositivos e aplicativos para JSON (para uso no JavaScript do modal de edição)
        dispositivos_json = [
            {'id': d.id, 'nome': d.nome}
            for d in dispositivos
        ]
        aplicativos_json = [
            {'id': a.id, 'nome': a.nome, 'device_has_mac': a.device_has_mac}
            for a in aplicativos
        ]

        clientes_em_atraso = Cliente.objects.filter(
            cancelado=False,
            mensalidade__cancelado=False,
            mensalidade__dt_pagamento=None,
            mensalidade__pgto=False,
            mensalidade__dt_vencimento__lt=hoje,
            usuario=self.request.user,
        ).count()

        valor_total_pago = (
            Mensalidade.objects.filter(
                cancelado=False,
                dt_pagamento__year=ano_atual,
                dt_pagamento__month=mes_atual,
                usuario=self.request.user,
                pgto=True,
            ).aggregate(valor_total=Sum("valor"))["valor_total"]
            or 0
        )
        valor_total_pago = format_currency(valor_total_pago, moeda)

        total_pago_qtd = Mensalidade.objects.filter(
            cancelado=False,
            dt_pagamento__year=ano_atual,
            dt_pagamento__month=mes_atual,
            usuario=self.request.user,
            pgto=True,
        ).count()

        valor_total_receber = (
            Mensalidade.objects.filter(
                cancelado=False,
                dt_vencimento__year=ano_atual,
                dt_vencimento__month=mes_atual,
                usuario=self.request.user,
                pgto=False,
            ).aggregate(valor_total=Sum("valor"))["valor_total"]
            or 0
        )
        valor_total_receber = format_currency(valor_total_receber, moeda)

        total_receber_qtd = Mensalidade.objects.filter(
            cancelado=False,
            dt_vencimento__year=ano_atual,
            dt_vencimento__month=mes_atual,
            dt_vencimento__day__gte=hoje.day,
            usuario=self.request.user,
            pgto=False,
        ).count()

        novos_clientes_qtd = (
            Cliente.objects.filter(
                cancelado=False,
                data_adesao__year=ano_atual,
                data_adesao__month=mes_atual,
                usuario=self.request.user,
            )
        ).count()

        clientes_cancelados_qtd = Cliente.objects.filter(
            cancelado=True,
            data_cancelamento__year=ano_atual,
            data_cancelamento__month=mes_atual,
            usuario=self.request.user,
        ).count()

        anos_adesao = (
            Cliente.objects.filter(usuario=self.request.user)
            .annotate(ano=ExtractYear('data_adesao'))
            .values_list('ano', flat=True)
            .distinct()
            .order_by('-ano')
        )

        # Resumo dos planos de adesão
        planos_adesao = (
            Cliente.objects
            .filter(usuario=self.request.user, cancelado=False)
            .annotate(nome_norm=Upper(Trim(F('plano__nome'))))
            .values('nome_norm')
            .annotate(qtd_adesoes=Count('id'))
            .order_by('nome_norm')
        )
        planos_cadastrados_norm = list(
            Plano.objects.filter(usuario=self.request.user)
            .annotate(nome_norm=Upper(Trim(F('nome'))))
            .values_list('nome_norm', flat=True)
            .distinct()
        )
        contagens = {row['nome_norm']: int(row['qtd_adesoes']) for row in planos_adesao}
        ordem_norm = ['MENSAL', 'TRIMESTRAL', 'SEMESTRAL', 'ANUAL']
        planos_ordenados_norm = sorted(
            set(planos_cadastrados_norm),
            key=lambda n: (ordem_norm.index(n) if n in ordem_norm else len(ordem_norm), n)
        )
        def label(n):
            mapa = {'MENSAL': 'Mensal', 'TRIMESTRAL': 'Trimestral', 'SEMESTRAL': 'Semestral', 'ANUAL': 'Anual'}
            return mapa.get(n, n.title())
        total = int(total_clientes) if total_clientes else 0
        planos_resumo = []
        for nome_norm in planos_ordenados_norm:
            qtd = contagens.get(nome_norm, 0)
            pct = (qtd / total * 100) if total else 0
            planos_resumo.append({'nome': label(nome_norm), 'qtd': qtd, 'pct': pct})
        # Fim resumo dos planos de adesão

        lista_meses = [
            (1, 'Jan'), (2, 'Fev'), (3, 'Mar'), (4, 'Abr'), (5, 'Mai'), (6, 'Jun'),
            (7, 'Jul'), (8, 'Ago'), (9, 'Set'), (10, 'Out'), (11, 'Nov'), (12, 'Dez')
        ]

        data_atual = timezone.localtime().date()

        query_telas = (
            Cliente.objects
            .filter(usuario=self.request.user, cancelado=False, plano__usuario=self.request.user)
            .select_related("plano")
            .only("id", "plano__telas", "plano__nome")
        )
        total_telas = query_telas.aggregate(
            total=Coalesce(Sum("plano__telas"), 0)
        )["total"]

        range_num = range(1,32)

        # Contar clientes ativos sem forma de pagamento
        clientes_sem_forma_pgto = Cliente.objects.filter(
            usuario=self.request.user,
            cancelado=False,
            forma_pgto__isnull=True
        ).count()

        # ============================================================
        # ALERTAS FASTDEPIX - Planos sem link ou com valor divergente
        # ============================================================
        alertas_fastdepix = {
            'planos_sem_link': [],
            'planos_valor_divergente': [],
            'tem_alertas': False
        }

        # Verificar se usuário tem conta FastDePix com link_fastdepix
        contas_fastdepix = ContaBancaria.objects.filter(
            usuario=self.request.user,
            instituicao__tipo_integracao='fastdepix',
            tipo_cobranca_fastdepix='link_fastdepix',
            ativo=True
        ).select_related('instituicao')

        if contas_fastdepix.exists():
            planos_usuario = Plano.objects.filter(usuario=self.request.user)

            for conta in contas_fastdepix:
                # Buscar links existentes para esta conta
                links_conta = PlanoLinkPagamento.objects.filter(
                    conta_bancaria=conta
                ).select_related('plano')
                planos_com_link = {link.plano_id for link in links_conta}

                # Planos sem link
                for plano in planos_usuario:
                    if plano.id not in planos_com_link:
                        alertas_fastdepix['planos_sem_link'].append({
                            'plano_id': plano.id,
                            'plano_nome': f"{plano.nome} ({plano.telas} tela(s))",
                            'plano_valor': float(plano.valor),
                            'conta_id': conta.id,
                            'conta_nome': conta.nome_identificacao
                        })

                # Planos com valor divergente
                for link in links_conta:
                    if link.valor_divergente:
                        alertas_fastdepix['planos_valor_divergente'].append({
                            'plano_id': link.plano.id,
                            'plano_nome': f"{link.plano.nome} ({link.plano.telas} tela(s))",
                            'valor_atual': float(link.plano.valor),
                            'valor_configurado': float(link.valor_configurado),
                            'conta_id': conta.id,
                            'conta_nome': conta.nome_identificacao
                        })

            alertas_fastdepix['tem_alertas'] = (
                len(alertas_fastdepix['planos_sem_link']) > 0 or
                len(alertas_fastdepix['planos_valor_divergente']) > 0
            )

        context.update(
            {
                "hoje": hoje,
                "page": page,
                "range": range_num,
                "nome_user": f_name,
                "aplicativos": aplicativos,
                "data_criacao_user": dt_inicio,
                "total_clientes": total_clientes,
                "valor_total_pago": valor_total_pago,
                "novos_clientes_qtd": novos_clientes_qtd,
                "clientes_em_atraso": clientes_em_atraso,
                "valor_total_receber": valor_total_receber,
                "valor_total_pago_qtd": total_pago_qtd,
                "valor_total_receber_qtd": total_receber_qtd,
                "clientes_cancelados_qtd": clientes_cancelados_qtd,
                "clientes_sem_forma_pgto": clientes_sem_forma_pgto,
                "total_telas_ativas": int(total_telas),
                'planos_resumo': planos_resumo,
                ## alertas FastDePix
                "alertas_fastdepix": alertas_fastdepix,
                ## context para modal de edição
                "planos": planos,
                "servidores": servidores,
                "indicadores": indicadores,
                "dispositivos": dispositivos,
                "formas_pgtos": formas_pgtos,
                "dispositivos_json": json.dumps(dispositivos_json),
                "aplicativos_json": json.dumps(aplicativos_json),
                ## context para o gráfico de adesões e cancelamentos
                "anos_adesao": anos_adesao,
                "lista_meses": lista_meses,
                "anuo_atual": ano_atual,
                "data_atual": data_atual,
                ## context para paginação configurável
                "pagination_options": PAGINATION_OPTIONS,
                "current_per_page": self.get_paginate_by(self.get_queryset()),
                ## context para ordenação server-side
                "current_sort": self.get_sort_params()[0],
                "current_order": self.get_sort_params()[1],
            }
        )
        return context


@login_required
def send_message_wpp(request):
    """
    Envia mensagens WhatsApp em massa via WPPConnect API.

    Funcionalidades:
    - Validação robusta via Django Form
    - Sistema de controle via session (pausar/retomar)
    - Mascaramento de telefones nos logs de arquivo
    - Delay incremental (> 200 envios/dia = +10s)
    - Timeout e retry em requests HTTP
    - Registro em UserActionLog
    - Sleep interruptível (parada em ~2s)
    - Constraint UNIQUE previne duplicatas
    - PROCESSAMENTO EM BACKGROUND (threading) para logs em tempo real

    SEGURANÇA:
    - CSRF protected
    - Validação de inputs (XSS, injection, DoS)
    - Path traversal prevention nos logs
    - Rate limiting via constraint UNIQUE
    - Isolamento por sessão (multi-user safe)
    """
    from django.db import IntegrityError
    from .forms import EnvioWhatsAppForm
    from .utils import mask_phone_number, get_envios_hoje
    import threading

    if request.method != 'POST':
        return JsonResponse({'error': 'Método inválido'}, status=400)

    # Previne múltiplas requisições simultâneas do mesmo usuário
    if request.session.get('envio_em_progresso', False):
        return JsonResponse({
            'error': 'Já existe um envio em progresso. Aguarde a conclusão ou pause o envio atual.'
        }, status=409)

    # Validação com Django Form
    form = EnvioWhatsAppForm(request.POST, request.FILES)
    if not form.is_valid():
        return JsonResponse({'error': form.errors}, status=400)

    # Dados validados
    tipo_envio = form.cleaned_data['options']
    mensagem = form.cleaned_data['mensagem']
    imagem = form.cleaned_data.get('imagem')
    telefones_file = form.cleaned_data.get('telefones')

    # Setup
    usuario = request.user
    sessao = get_object_or_404(SessaoWpp, usuario=usuario)
    token = sessao.token
    BASE_URL = url_api + '/{}/send-{}'

    # Logs
    import re
    safe_username = re.sub(r'[^\w\-]', '', usuario.username)
    log_directory = Path('./logs/Envios manuais/')
    log_directory.mkdir(parents=True, exist_ok=True)
    log_filename = log_directory / f'{safe_username}.log'
    log_result_filename = log_directory / f'{safe_username}_send_result.log'

    # Imagem base64
    imagem_base64 = None
    if imagem:
        imagem_base64 = base64.b64encode(imagem.read()).decode('utf-8')

    # Coleta de telefones (validação inicial)
    telefones = []
    if tipo_envio == 'ativos':
        clientes = Cliente.objects.filter(usuario=usuario, cancelado=False, nao_enviar_msgs=False)
        telefones = [re.sub(r'\s+|\W', '', c.telefone) for c in clientes]
    elif tipo_envio == 'cancelados':
        clientes = Cliente.objects.filter(
            usuario=usuario,
            cancelado=True,
            data_cancelamento__lte=timezone.now() - timedelta(days=7),
            nao_enviar_msgs=False
        )
        telefones = [re.sub(r'\s+|\W', '', c.telefone) for c in clientes]
    elif tipo_envio == 'avulso':
        if telefones_file:
            lines = telefones_file.read().decode('utf-8').splitlines()
            telefones = [re.sub(r'\s+|\W', '', tel) for tel in lines if tel.strip()]

    if not telefones:
        return JsonResponse({'error': 'Nenhum telefone válido encontrado'}, status=400)

    # Marca envio como em progresso
    request.session['envio_em_progresso'] = True
    request.session['stop_envio'] = False
    request.session.save()

    # Prepara dados para a thread
    thread_data = {
        'session_key': request.session.session_key,
        'user_id': usuario.id,
        'tipo_envio': tipo_envio,
        'mensagem': mensagem,
        'imagem_nome': str(imagem) if imagem else None,
        'imagem_base64': imagem_base64,
        'telefones': telefones,
        'token': token,
        'BASE_URL': BASE_URL,
        'log_filename': str(log_filename),
        'log_result_filename': str(log_result_filename),
        'remote_addr': request.META.get('REMOTE_ADDR'),
        'request_path': request.path
    }

    # Função executada em background (thread separada)
    def processar_envios_background(data):
        """
        Processa envios WhatsApp em thread separada para permitir
        que o frontend exiba logs em tempo real via polling.
        """
        from django.contrib.sessions.backends.db import SessionStore
        from django.contrib.auth import get_user_model
        from django.db import IntegrityError
        from .utils import mask_phone_number, get_envios_hoje
        from wpp.api_connection import check_number_status

        User = get_user_model()

        # Reconstrói objetos necessários
        usuario = User.objects.get(id=data['user_id'])
        session = SessionStore(session_key=data['session_key'])

        tipo_envio = data['tipo_envio']
        mensagem = data['mensagem']
        imagem_nome = data['imagem_nome']
        imagem_base64 = data['imagem_base64']
        telefones = data['telefones']
        token = data['token']
        BASE_URL = data['BASE_URL']
        log_filename = data['log_filename']
        log_result_filename = data['log_result_filename']
        remote_addr = data['remote_addr']
        request_path = data['request_path']

        def log_result(file_path, content):
            """Escreve log em arquivo com timestamp."""
            with codecs.open(file_path, 'a', encoding='utf-8') as f:
                f.write(f"[{datetime.now().strftime('%d-%m-%Y %H:%M:%S')}] {content}\n")

        def enviar_mensagem_com_retry(url, telefone, max_retries=3):
            """
            Envia mensagem com retry automático e timeout.
            Cria o registro no BD ANTES de enviar (previne duplicatas).
            """
            hoje = timezone.now().date()

            # Check: já enviou hoje?
            if MensagemEnviadaWpp.objects.filter(
                usuario=usuario,
                telefone=telefone,
                data_envio=hoje
            ).exists():
                log_result(log_result_filename, f"{telefone} - ⚠️ Já enviado hoje (ignorado)")
                return False

            # Validação do número via WhatsApp API
            numero_existe = check_number_status(telefone, token, usuario.username)
            if not numero_existe or not numero_existe.get('status'):
                log_result(log_result_filename, f"{telefone} - ❌ Número inválido no WhatsApp")
                if tipo_envio == 'avulso':
                    from .models import TelefoneLeads
                    TelefoneLeads.objects.filter(telefone=telefone, usuario=usuario).delete()
                return False

            # PASSO 1: Cria registro PRIMEIRO (previne duplicata)
            try:
                registro = MensagemEnviadaWpp.objects.create(usuario=usuario, telefone=telefone)
            except IntegrityError:
                log_result(log_result_filename, f"{telefone} - ⚠️ Já enviado hoje (ignorado)")
                return False

            headers = {
                'Content-Type': 'application/json',
                'Accept': 'application/json',
                'Authorization': f'Bearer {token}'
            }
            body = {
                'phone': telefone,
                'isGroup': False,
                'message': mensagem
            }
            if imagem_nome:
                body.update({
                    'filename': imagem_nome,
                    'caption': mensagem,
                    'base64': f'data:image/png;base64,{imagem_base64}'
                })

            # PASSO 2: Tenta enviar (com rollback se falhar)
            try:
                for attempt in range(max_retries):
                    try:
                        response = requests.post(url, headers=headers, json=body, timeout=30)

                        if response.status_code in [200, 201]:
                            masked_phone = mask_phone_number(telefone)
                            log_result(log_filename, f"[TIPO][Manual] [USUÁRIO][{usuario}] [TELEFONE][{masked_phone}] Mensagem enviada!")
                            log_result(log_result_filename, f"{telefone} - ✅ Mensagem enviada")
                            return True
                        else:
                            try:
                                response_data = response.json()
                                error_message = response_data.get('message', 'Erro desconhecido')
                            except json.decoder.JSONDecodeError:
                                # Sanitiza para evitar registrar HTML de páginas de erro
                                sanitized = _sanitize_response(response.text)
                                if isinstance(sanitized, dict):
                                    error_message = sanitized.get('mensagem', 'Erro desconhecido')
                                else:
                                    error_message = str(sanitized)

                            masked_phone = mask_phone_number(telefone)
                            log_result(log_filename, f"[TIPO][Manual] [USUÁRIO][{usuario}] [TELEFONE][{masked_phone}] [CODE][{response.status_code}] - {error_message}")

                            if response.status_code >= 500 and attempt < max_retries - 1:
                                wait_time = 2 ** attempt
                                time.sleep(wait_time)
                                continue
                            else:
                                registro.delete()
                                log_result(log_result_filename, f"{telefone} - ❌ Não enviada (consultar log)")
                                return False

                    except requests.exceptions.Timeout:
                        masked_phone = mask_phone_number(telefone)
                        log_result(log_filename, f"[TIPO][Manual] [USUÁRIO][{usuario}] [TELEFONE][{masked_phone}] [TIMEOUT] Timeout após 30s")
                        if attempt < max_retries - 1:
                            time.sleep(2 ** attempt)
                            continue
                        else:
                            registro.delete()
                            log_result(log_result_filename, f"{telefone} - ❌ Timeout (consultar log)")
                            return False

                    except requests.exceptions.RequestException as e:
                        registro.delete()
                        masked_phone = mask_phone_number(telefone)
                        log_result(log_filename, f"[TIPO][Manual] [USUÁRIO][{usuario}] [TELEFONE][{masked_phone}] [ERRO] {str(e)}")
                        log_result(log_result_filename, f"{telefone} - ❌ Erro de conexão")
                        return False

            except Exception as e:
                registro.delete()
                masked_phone = mask_phone_number(telefone)
                log_result(log_filename, f"[TIPO][Manual] [USUÁRIO][{usuario}] [TELEFONE][{masked_phone}] [ERRO INESPERADO] {str(e)}")
                log_result(log_result_filename, f"{telefone} - ❌ Erro inesperado")
                return False

            registro.delete()
            return False

        try:
            # UserActionLog: início
            UserActionLog.objects.create(
                usuario=usuario,
                acao='other',
                entidade='envio_whatsapp',
                mensagem='Envio WhatsApp iniciado',
                ip=remote_addr,
                extras={
                    'tipo_envio': tipo_envio,
                    'total_telefones': len(telefones),
                    'com_imagem': bool(imagem_nome),
                    'path': request_path
                }
            )

            # Processamento
            url_envio = BASE_URL.format(usuario, 'image' if imagem_nome else 'message')
            total = len(telefones)
            enviadas = 0
            envios_hoje_inicial = get_envios_hoje(usuario)

            log_result(log_result_filename, f"\n=== INICIANDO ENVIO ===\nTotal: {total} telefones\n")

            for idx, telefone in enumerate(telefones, 1):
                # CHECK: Flag de parada
                if session.get('stop_envio', False):
                    log_result(log_result_filename, f"\n=== ENVIO PARADO PELO USUÁRIO ===\nEnviadas: {enviadas}/{total}\n")
                    UserActionLog.objects.create(
                        usuario=usuario,
                        acao='other',
                        entidade='envio_whatsapp',
                        mensagem='Envio WhatsApp pausado pelo usuário',
                        ip=remote_addr,
                        extras={
                            'tipo_envio': tipo_envio,
                            'enviadas': enviadas,
                            'total': total,
                            'percentual': round((enviadas / total) * 100, 2),
                            'path': request_path
                        }
                    )
                    return

                # Envia mensagem
                if enviar_mensagem_com_retry(url_envio, telefone):
                    enviadas += 1

                # Delay inteligente
                envios_hoje_atual = envios_hoje_inicial + enviadas
                base_delay = random.uniform(5, 12)
                if envios_hoje_atual > 200:
                    extra_delay = 10
                    total_delay = base_delay + extra_delay
                    log_result(log_result_filename, f"⚠️ Limite de 200 envios/dia excedido. Delay aumentado para {total_delay:.1f}s")
                else:
                    total_delay = base_delay

                # Sleep interruptível
                chunks = int(total_delay)
                remainder = total_delay - chunks
                for _ in range(chunks):
                    if session.get('stop_envio', False):
                        log_result(log_result_filename, f"\n=== ENVIO PARADO DURANTE SLEEP ===\n")
                        return
                    time.sleep(1)
                if remainder > 0:
                    time.sleep(remainder)

            # UserActionLog: conclusão
            UserActionLog.objects.create(
                usuario=usuario,
                acao='other',
                entidade='envio_whatsapp',
                mensagem='Envio WhatsApp concluído com sucesso',
                ip=remote_addr,
                extras={
                    'tipo_envio': tipo_envio,
                    'enviadas': enviadas,
                    'total': total,
                    'percentual': round((enviadas / total) * 100, 2) if total > 0 else 0,
                    'path': request_path
                }
            )

            log_result(log_result_filename, f"\n=== ENVIO CONCLUÍDO ===\nEnviadas: {enviadas}/{total}\n")

        except Exception as e:
            log_result(log_result_filename, f"\n=== ERRO INESPERADO ===\n{str(e)}\n")
        finally:
            # Cleanup da sessão
            session['envio_em_progresso'] = False
            session['stop_envio'] = False
            session.save()

    # Inicia thread de processamento
    thread = threading.Thread(target=processar_envios_background, args=(thread_data,), daemon=True)
    thread.start()

    # UserActionLog: início (registro imediato)
    UserActionLog.objects.create(
        usuario=usuario,
        acao='other',
        entidade='envio_whatsapp',
        mensagem='Envio WhatsApp aceito para processamento',
        ip=request.META.get('REMOTE_ADDR'),
        extras={
            'tipo_envio': tipo_envio,
            'total_telefones': len(telefones),
            'com_imagem': bool(imagem),
            'path': request.path
        }
    )

    # Retorna imediatamente (status 202 = Accepted)
    return JsonResponse({
        'success': 'Envio iniciado com sucesso',
        'total': len(telefones),
        'message': 'Os logs serão exibidos em tempo real no console'
    }, status=202)



@require_http_methods(["GET"])
@login_required
def secret_token_api(request):
    """
        Função de view para consultar o Secret Token da API WPP Connect
    """
    if not request.user.is_superuser:
        return JsonResponse({'error_message': 'Permissão negada.'}, status=403)

    token = (
        SecretTokenAPI.objects.filter(usuario=request.user)
        .order_by('-dt_criacao')
        .values_list('token', flat=True)
        .first()
    )

    if not token:
        return JsonResponse({'error_message': 'Token não encontrado.'}, status=404)

    return JsonResponse({'stkn': token}, status=200)

@login_required
def get_session_wpp(request):
    """
        Função de view para consultar o Token da sessão WhatsApp do usuário da requisição.
    """
    if request.method == 'GET':
        if request.user:
            sessao = get_object_or_404(SessaoWpp, usuario=request.user)
            token = sessao.token
        else:
            return JsonResponse({"error_message": "Usuário da requisição não identificado."}, status=500)
    else:
        return JsonResponse({"error_message": "Método da requisição não permitido."}, status=500)

    return JsonResponse({"token": token}, status=200)


@require_http_methods(["GET"])
@login_required
def get_logs_wpp(request):
    """
    Retorna logs de envios do usuário atual.

    SEGURANÇA:
    - Valida username contra path traversal (../)
    - Normaliza path para prevenir acesso fora de LOG_DIR
    - Verifica que path final está dentro do diretório esperado
    """
    import re

    # Sanitiza username (previne path traversal)
    username = request.user.username
    # Remove caracteres perigosos: apenas alfanuméricos, underscore, hífen
    safe_username = re.sub(r'[^\w\-]', '', username)

    # Constrói path
    log_dir = Path(LOG_DIR) / "Envios manuais"
    log_path = log_dir / f"{safe_username}_send_result.log"

    # Normaliza e resolve path (previne ../ e symlinks)
    log_path = log_path.resolve()
    log_dir = log_dir.resolve()

    # Verifica se log_path está dentro de log_dir (previne traversal)
    try:
        log_path.relative_to(log_dir)
    except ValueError:
        # Path está fora do diretório permitido
        return JsonResponse({'error': 'Acesso negado'}, status=403)

    if not log_path.exists():
        return JsonResponse({'logs': ''})

    logs = log_path.read_text(encoding='utf-8', errors='ignore')
    return JsonResponse({'logs': logs})


@require_http_methods(["POST"])
@login_required
def parar_envio(request):
    """
    Endpoint para usuário pausar envios WhatsApp em progresso.

    Sistema de controle via session:
    - session['envio_em_progresso']: Boolean indicando se há envio ativo
    - session['stop_envio']: Flag para interromper loop de envios

    Response:
    - 200: Flag setada com sucesso, envio será pausado
    - 409: Nenhum envio em progresso para pausar

    SEGURANÇA:
    - Isolamento por sessão (cada usuário controla apenas seus próprios envios)
    - CSRF protected
    """
    if not request.session.get('envio_em_progresso', False):
        return JsonResponse({
            'error': 'Nenhum envio em progresso'
        }, status=409)

    request.session['stop_envio'] = True
    request.session.save()

    return JsonResponse({
        'success': 'Envio será pausado em breve'
    }, status=200)


@require_http_methods(["GET"])
@login_required
def status_envio(request):
    """
    Retorna status atual do envio para sincronizar UI.

    Used by JavaScript para:
    - Exibir botão correto (Enviar/Parar/Retomar)
    - Polling a cada 5s durante envio ativo
    - Prevenir múltiplas requisições simultâneas

    Response JSON:
        {
            "em_progresso": bool,  # True se há envio rodando
            "pausado": bool,       # True se foi pausado pelo usuário
            "total_hoje": int      # Quantidade de envios realizados hoje
        }
    """
    from .utils import get_envios_hoje

    return JsonResponse({
        'em_progresso': request.session.get('envio_em_progresso', False),
        'pausado': request.session.get('stop_envio', False),
        'total_hoje': get_envios_hoje(request.user)
    })


@require_http_methods(["POST"])
@login_required
def limpar_log_wpp(request):
    """
    Deleta fisicamente o arquivo de log de envios WhatsApp do usuário.

    REGRAS:
    - Só permite limpar se NÃO houver envio em progresso
    - Deleta ambos os arquivos: log principal e log de resultados

    Response:
    - 200: Log deletado com sucesso
    - 409: Não pode limpar durante envio em progresso

    SEGURANÇA:
    - Path traversal prevention (sanitiza username)
    - Isolamento por usuário (cada um limpa apenas seus logs)
    - CSRF protected
    """
    # Verifica se há envio em progresso
    if request.session.get('envio_em_progresso', False):
        return JsonResponse({
            'error': 'Não é possível limpar logs durante um envio em progresso'
        }, status=409)

    # Path dos arquivos de log
    import re
    safe_username = re.sub(r'[^\w\-]', '', request.user.username)
    log_directory = Path('./logs/Envios manuais/')
    log_filename = log_directory / f'{safe_username}.log'
    log_result_filename = log_directory / f'{safe_username}_send_result.log'

    try:
        # Deleta arquivo de log principal (se existir)
        if log_filename.exists():
            log_filename.unlink()

        # Deleta arquivo de log de resultados (se existir)
        if log_result_filename.exists():
            log_result_filename.unlink()

        # Log da ação
        UserActionLog.objects.create(
            usuario=request.user,
            acao='other',
            entidade='log_whatsapp',
            mensagem='Arquivo de log de envios WhatsApp foi limpo',
            ip=request.META.get('REMOTE_ADDR'),
            extras={
                'path': request.path,
                'arquivos_deletados': [str(log_filename), str(log_result_filename)]
            }
        )

        return JsonResponse({
            'success': 'Logs limpos com sucesso'
        }, status=200)

    except Exception as e:
        return JsonResponse({
            'error': f'Erro ao limpar logs: {str(e)}'
        }, status=500)


@login_required
def profile_page(request):
    from .models import UserProfile, UserActionLog

    user = request.user
    dados_bancarios = DadosBancarios.objects.filter(usuario=user).first()

    # Obter ou criar perfil do usuário
    profile, created = UserProfile.objects.get_or_create(user=user)

    dt_inicio = user.date_joined.strftime('%d/%m/%Y') if user.date_joined else '--'
    f_name = user.first_name or '--'
    l_name = user.last_name or '--'
    email = user.email or '--'

    beneficiario = dados_bancarios.beneficiario if dados_bancarios else '--'
    instituicao = dados_bancarios.instituicao if dados_bancarios else '--'
    tipo_chave = dados_bancarios.tipo_chave if dados_bancarios else '--'
    chave = dados_bancarios.chave if dados_bancarios else '--'
    wpp = dados_bancarios.wpp if dados_bancarios else '--'

    ip = get_client_ip(request)
    localizacao = get_location_from_ip(ip)

    # ========== ESTATÍSTICAS DO USUÁRIO ==========
    # Total de clientes ativos
    total_clientes = Cliente.objects.filter(usuario=user, cancelado=False).count()

    # Valor de negócio (soma dos valores dos planos dos clientes ativos)
    receita_mensal = Cliente.objects.filter(
        usuario=user,
        cancelado=False
    ).aggregate(
        total=Sum('plano__valor')
    )['total'] or 0

    # Dias no sistema
    dias_sistema = (timezone.now().date() - user.date_joined.date()).days if user.date_joined else 0

    # Última atividade registrada
    ultima_acao = UserActionLog.objects.filter(usuario=user).order_by('-criado_em').first()
    if ultima_acao:
        diff = timezone.now() - ultima_acao.criado_em
        if diff.days > 0:
            ultima_atividade = f"Há {diff.days} dia{'s' if diff.days > 1 else ''}"
        elif diff.seconds // 3600 > 0:
            horas = diff.seconds // 3600
            ultima_atividade = f"Há {horas} hora{'s' if horas > 1 else ''}"
        elif diff.seconds // 60 > 0:
            minutos = diff.seconds // 60
            ultima_atividade = f"Há {minutos} minuto{'s' if minutos > 1 else ''}"
        else:
            ultima_atividade = "Agora mesmo"
    else:
        ultima_atividade = "Nunca"

    # Total de clientes cancelados
    total_cancelados = Cliente.objects.filter(usuario=user, cancelado=True).count()

    # Total de mensalidades recebidas este mês
    mes_atual = timezone.now().month
    ano_atual = timezone.now().year
    mensalidades_mes = Mensalidade.objects.filter(
        usuario=user,
        pgto=True,
        dt_pagamento__month=mes_atual,
        dt_pagamento__year=ano_atual
    ).count()

    return render(
        request,
        'pages/perfil.html',
        {
            'beneficiario': beneficiario,
            'localizacao': localizacao,
            'instituicao': instituicao,
            'tipo_chave': tipo_chave,
            'sobrenome_user': l_name,
            'dt_inicio': dt_inicio,
            'nome_user': f_name,
            'username': user.username,
            'email': email,
            'chave': chave,
            'wpp': wpp,
            'avatar_url': profile.get_avatar_url(),
            # Estatísticas
            'total_clientes': total_clientes,
            'receita_mensal': receita_mensal,
            'dias_sistema': dias_sistema,
            'ultima_atividade': ultima_atividade,
            'total_cancelados': total_cancelados,
            'mensalidades_mes': mensalidades_mes,
            # Preferências
            'theme_preference': profile.theme_preference,
            'email_on_profile_change': profile.email_on_profile_change,
            'email_on_password_change': profile.email_on_password_change,
            'email_on_login': profile.email_on_login,
            'profile_public': profile.profile_public,
            'show_email': profile.show_email,
            'show_phone': profile.show_phone,
            'show_statistics': profile.show_statistics,
            # 2FA
            'two_factor_enabled': profile.two_factor_enabled,
            'has_backup_codes': bool(profile.two_factor_backup_codes),
        },
    )


@login_required
def exportar_clientes_excel(request):
    """Exporta clientes do usuário para Excel (.xlsx)"""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from django.db.models import Prefetch

    usuario = request.user
    clientes = Cliente.objects.filter(usuario=usuario).select_related(
        'servidor', 'dispositivo', 'sistema', 'plano', 'forma_pgto', 'indicado_por'
    ).prefetch_related(
        Prefetch('conta_aplicativo', queryset=ContaDoAplicativo.objects.select_related('app', 'dispositivo'))
    )

    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"

    # Estilos
    header_font = Font(bold=True, size=14)
    subheader_font = Font(bold=True, size=11)
    table_header_font = Font(bold=True, color="FFFFFF")
    table_header_fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
    font_verde = Font(color="008000")  # Verde para Ativo
    font_vermelha = Font(color="FF0000")  # Vermelho para Cancelado
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # === CABEÇALHO DO USUÁRIO ===
    dados_bancarios = DadosBancarios.objects.filter(usuario=usuario).first()
    telefone_user = dados_bancarios.wpp if dados_bancarios else ""

    ws['A1'] = "RELATÓRIO DE CLIENTES"
    ws['A1'].font = header_font

    ws['A2'] = f"Nome: {usuario.first_name} {usuario.last_name}"
    ws['A3'] = f"Usuário: {usuario.username}"
    ws['A4'] = f"Telefone: {telefone_user}"
    ws['A5'] = f"Total de Clientes: {clientes.count()}"

    # === CABEÇALHO DA TABELA (linha 7) ===
    colunas = [
        'Servidor', 'Dispositivo', 'Aplicativo', 'Device ID', 'Email', 'Device Key',
        'Nome', 'Telefone', 'Indicado Por', 'Data Vencimento', 'Forma Pgto',
        'Tipo Plano', 'Plano Valor', 'Qtd. Telas', 'Data Adesão', 'Status'
    ]

    for col, titulo in enumerate(colunas, 1):
        cell = ws.cell(row=7, column=col, value=titulo)
        cell.font = table_header_font
        cell.fill = table_header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # === DADOS DOS CLIENTES ===
    linha = 8
    for cliente in clientes:
        contas = cliente.conta_aplicativo.all()

        if contas.exists():
            # Concatenar dados de múltiplas contas (incluindo o nome do app)
            aplicativos = " - ".join([c.app.nome for c in contas if c.app])
            device_ids = " - ".join([c.device_id for c in contas if c.device_id])
            emails = " - ".join([c.email for c in contas if c.email])
            device_keys = " - ".join([c.device_key for c in contas if c.device_key])
        else:
            aplicativos = cliente.sistema.nome if cliente.sistema else ""
            device_ids = ""
            emails = ""
            device_keys = ""

        dados = [
            cliente.servidor.nome if cliente.servidor else "",
            cliente.dispositivo.nome if cliente.dispositivo else "",
            aplicativos,
            device_ids,
            emails,
            device_keys,
            cliente.nome,
            cliente.telefone,
            cliente.indicado_por.nome if cliente.indicado_por else "",
            cliente.data_vencimento.strftime('%d/%m/%Y') if cliente.data_vencimento else "",
            cliente.forma_pgto.nome if cliente.forma_pgto else "",
            cliente.plano.nome if cliente.plano else "",
            float(cliente.plano.valor) if cliente.plano else 0,
            cliente.plano.telas if cliente.plano else 0,
            cliente.data_adesao.strftime('%d/%m/%Y') if cliente.data_adesao else "",
            "Cancelado" if cliente.cancelado else "Ativo",
        ]

        for col, valor in enumerate(dados, 1):
            cell = ws.cell(row=linha, column=col, value=valor)
            cell.border = thin_border

        # Aplicar cor na coluna Status
        status_col = len(colunas)
        status_cell = ws.cell(row=linha, column=status_col)
        if cliente.cancelado:
            status_cell.font = font_vermelha
        else:
            status_cell.font = font_verde

        linha += 1

    # Ajustar largura das colunas
    for col in range(1, len(colunas) + 1):
        ws.column_dimensions[ws.cell(row=7, column=col).column_letter].width = 15

    # Salvar em bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Retornar resposta
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="clientes_{usuario.username}.xlsx"'
    return response


@login_required
def generate_graphic_columns_per_month(request):
    ano_atual = now().year
    mes = int(request.GET.get("mes", now().month))
    usuario = request.user

    if not (1 <= mes <= 12):
        return HttpResponse("Mês inválido", status=400)

    dados_adesoes = (
        Cliente.objects.filter(
            data_adesao__year=ano_atual,
            data_adesao__month=mes,
            usuario=usuario,
        )
        .annotate(dia=ExtractDay("data_adesao"))
        .values("dia")
        .annotate(total=Count("id"))
        .order_by("dia")
    )

    dados_cancelamentos = (
        Cliente.objects.filter(
            data_cancelamento__year=ano_atual,
            data_cancelamento__month=mes,
            usuario=usuario,
        )
        .annotate(dia=ExtractDay("data_cancelamento"))
        .values("dia")
        .annotate(total=Count("id"))
        .order_by("dia")
    )

    adesoes_dict = {dado["dia"]: dado["total"] for dado in dados_adesoes}
    cancelamentos_dict = {dado["dia"]: dado["total"] for dado in dados_cancelamentos}

    dias = []
    adesoes = []
    cancelamentos = []

    total_dias_mes = calendar.monthrange(ano_atual, mes)[1]

    for dia in range(1, total_dias_mes + 1):
        if dia in adesoes_dict or dia in cancelamentos_dict:
            dias.append(str(dia))
            adesoes.append(adesoes_dict.get(dia, 0))
            cancelamentos.append(cancelamentos_dict.get(dia, 0))

    total_adesoes = sum(adesoes)
    total_cancelamentos = sum(cancelamentos)
    saldo_final = total_adesoes - total_cancelamentos

    plt.figure(figsize=(7, 3))
    plt.bar(dias, adesoes, color="#4CAF50", width=0.4, label="Adesões")
    plt.bar(dias, cancelamentos, color="#F44336", width=0.4, bottom=adesoes, label="Cancelamentos")

    for i, valor in enumerate(adesoes):
        if valor > 0:
            plt.text(i, valor / 2, str(valor), ha='center', va='center', fontsize=10, color='white', fontweight='bold')

    for i, valor in enumerate(cancelamentos):
        if valor > 0:
            plt.text(
                i,
                adesoes[i] + valor / 2,
                str(valor),
                ha='center',
                va='center',
                fontsize=10,
                color='white',
                fontweight='bold',
            )

    nomes_pt = [
        "",
        "Janeiro",
        "Fevereiro",
        "Março",
        "Abril",
        "Maio",
        "Junho",
        "Julho",
        "Agosto",
        "Setembro",
        "Outubro",
        "Novembro",
        "Dezembro",
    ]
    nome_mes = nomes_pt[mes]
    plt.title(f"Adesões e Cancelamentos por mês - {nome_mes} {ano_atual}", fontsize=14)
    plt.xlabel("Dia", fontsize=12)
    plt.ylabel("Quantidade", fontsize=12)
    plt.xticks(fontsize=10, fontweight='bold')
    plt.yticks(fontsize=10)

    cor_saldo = "#624BFF" if saldo_final >= 0 else "#F44336"
    texto_saldo = f"Saldo {nome_mes}: {'+' if saldo_final > 0 else ''}{saldo_final}"

    saldo_patch = Patch(color=cor_saldo, label=texto_saldo)
    plt.legend(
        handles=[
            Patch(color="#4CAF50", label=f"Adesões: {total_adesoes}"),
            Patch(color="#F44336", label=f"Cancelamentos: {total_cancelamentos}"),
            saldo_patch,
        ]
    )

    plt.gca().spines['top'].set_visible(False)
    plt.gca().spines['right'].set_visible(False)

    buffer = io.BytesIO()
    plt.tight_layout()
    plt.savefig(buffer, format='png', bbox_inches="tight", dpi=100)
    buffer.seek(0)
    plt.close()

    return HttpResponse(buffer.getvalue(), content_type="image/png")


def _month_name_pt(month: int) -> str:
    nomes = [
        "",
        "Janeiro",
        "Fevereiro",
        "Mar\u00e7o",
        "Abril",
        "Maio",
        "Junho",
        "Julho",
        "Agosto",
        "Setembro",
        "Outubro",
        "Novembro",
        "Dezembro",
    ]
    if 1 <= month < len(nomes):
        return nomes[month]
    return str(month)


def _month_abbr_pt(month: int) -> str:
    abreviacoes = [
        "Jan",
        "Fev",
        "Mar",
        "Abr",
        "Mai",
        "Jun",
        "Jul",
        "Ago",
        "Set",
        "Out",
        "Nov",
        "Dez",
    ]
    if 1 <= month <= len(abreviacoes):
        return abreviacoes[month - 1]
    return str(month)


def _dataset_adesao_cancelamentos_mensal(usuario, year: int, month: int):
    dados_adesoes = (
        Cliente.objects.filter(
            data_adesao__year=year,
            data_adesao__month=month,
            usuario=usuario,
        )
        .annotate(dia=ExtractDay("data_adesao"))
        .values("dia")
        .annotate(total=Count("id"))
        .order_by("dia")
    )

    dados_cancelamentos = (
        Cliente.objects.filter(
            data_cancelamento__year=year,
            data_cancelamento__month=month,
            usuario=usuario,
        )
        .annotate(dia=ExtractDay("data_cancelamento"))
        .values("dia")
        .annotate(total=Count("id"))
        .order_by("dia")
    )

    adesoes_dict = {item["dia"]: item["total"] for item in dados_adesoes}
    cancelamentos_dict = {item["dia"]: item["total"] for item in dados_cancelamentos}

    categorias = []
    adesoes = []
    cancelamentos = []
    saldo = []

    total_dias = calendar.monthrange(year, month)[1]
    for dia in range(1, total_dias + 1):
        if dia in adesoes_dict or dia in cancelamentos_dict:
            valor_adesao = adesoes_dict.get(dia, 0)
            valor_cancelamento = cancelamentos_dict.get(dia, 0)
            categorias.append(str(dia))
            adesoes.append(valor_adesao)
            cancelamentos.append(valor_cancelamento)
            saldo.append(valor_adesao - valor_cancelamento)

    total_adesoes = sum(adesoes)
    total_cancelamentos = sum(cancelamentos)
    saldo_final = total_adesoes - total_cancelamentos

    return {
        "mode": "monthly",
        "categories": categorias,
        "series": [
            {"key": "adesoes", "name": "Ades\u00f5es", "type": "bar", "data": adesoes},
            {"key": "cancelamentos", "name": "Cancelamentos", "type": "bar", "data": cancelamentos},
            {"key": "saldo", "name": "Saldo", "type": "line", "data": saldo},
        ],
        "summary": {
            "total_adesoes": total_adesoes,
            "total_cancelamentos": total_cancelamentos,
            "saldo": saldo_final,
            "saldo_label": f"{'+' if saldo_final > 0 else ''}{saldo_final}",
        },
        "meta": {
            "mode": "monthly",
            "month": month,
            "month_name": _month_name_pt(month),
            "year": year,
            "range_label": f"{_month_name_pt(month)} {year}",
        },
    }


def _dataset_adesao_cancelamentos_lifetime(usuario):
    dados_adesoes = (
        Cliente.objects.filter(usuario=usuario, data_adesao__isnull=False)
        .annotate(ano=ExtractYear("data_adesao"), mes=ExtractMonth("data_adesao"))
        .values("ano", "mes")
        .annotate(total=Count("id"))
        .order_by("ano", "mes")
    )

    dados_cancelamentos = (
        Cliente.objects.filter(usuario=usuario, data_cancelamento__isnull=False)
        .annotate(ano=ExtractYear("data_cancelamento"), mes=ExtractMonth("data_cancelamento"))
        .values("ano", "mes")
        .annotate(total=Count("id"))
        .order_by("ano", "mes")
    )

    adesoes_dict = {(item["ano"], item["mes"]): item["total"] for item in dados_adesoes}
    cancelamentos_dict = {(item["ano"], item["mes"]): item["total"] for item in dados_cancelamentos}

    todos_periodos = sorted(set(adesoes_dict.keys()) | set(cancelamentos_dict.keys()))

    categorias = []
    adesoes = []
    cancelamentos = []
    saldo = []

    for ano, mes in todos_periodos:
        valor_adesao = adesoes_dict.get((ano, mes), 0)
        valor_cancelamento = cancelamentos_dict.get((ano, mes), 0)
        categorias.append(f"{_month_abbr_pt(mes)} {ano}")
        adesoes.append(valor_adesao)
        cancelamentos.append(valor_cancelamento)
        saldo.append(valor_adesao - valor_cancelamento)

    total_adesoes = sum(adesoes)
    total_cancelamentos = sum(cancelamentos)
    saldo_final = total_adesoes - total_cancelamentos

    meta = {
        "mode": "lifetime",
        "title": "Ades\u00e3o e Cancelamentos - Hist\u00f3rico completo",
    }
    if todos_periodos:
        ano_inicio, mes_inicio = todos_periodos[0]
        ano_fim, mes_fim = todos_periodos[-1]
        meta.update(
            {
                "start_year": ano_inicio,
                "start_month": mes_inicio,
                "end_year": ano_fim,
                "end_month": mes_fim,
                "range_label": f"De {_month_abbr_pt(mes_inicio)} {ano_inicio} a {_month_abbr_pt(mes_fim)} {ano_fim}",
            }
        )

    return {
        "mode": "lifetime",
        "categories": categorias,
        "series": [
            {"key": "adesoes", "name": "Ades\u00f5es", "type": "bar", "data": adesoes},
            {"key": "cancelamentos", "name": "Cancelamentos", "type": "bar", "data": cancelamentos},
            {"key": "saldo", "name": "Saldo", "type": "line", "data": saldo},
        ],
        "summary": {
            "total_adesoes": total_adesoes,
            "total_cancelamentos": total_cancelamentos,
            "saldo": saldo_final,
            "saldo_label": f"{'+' if saldo_final > 0 else ''}{saldo_final}",
        },
        "meta": meta,
    }


def _dataset_adesao_cancelamentos_anual(usuario, year: int):
    dados_adesoes = (
        Cliente.objects.filter(data_adesao__year=year, usuario=usuario)
        .annotate(mes=ExtractMonth("data_adesao"))
        .values("mes")
        .annotate(total=Count("id"))
        .order_by("mes")
    )

    dados_cancelamentos = (
        Cliente.objects.filter(data_cancelamento__year=year, usuario=usuario)
        .annotate(mes=ExtractMonth("data_cancelamento"))
        .values("mes")
        .annotate(total=Count("id"))
        .order_by("mes")
    )

    adesoes_dict = {item["mes"]: item["total"] for item in dados_adesoes}
    cancelamentos_dict = {item["mes"]: item["total"] for item in dados_cancelamentos}

    categorias = []
    adesoes = []
    cancelamentos = []
    saldo = []

    for mes in range(1, 13):
        valor_adesao = adesoes_dict.get(mes, 0)
        valor_cancelamento = cancelamentos_dict.get(mes, 0)
        if valor_adesao or valor_cancelamento:
            categorias.append(_month_abbr_pt(mes))
            adesoes.append(valor_adesao)
            cancelamentos.append(valor_cancelamento)
            saldo.append(valor_adesao - valor_cancelamento)

    total_adesoes = sum(adesoes)
    total_cancelamentos = sum(cancelamentos)
    saldo_final = total_adesoes - total_cancelamentos

    return {
        "mode": "annual",
        "categories": categorias,
        "series": [
            {"key": "adesoes", "name": "Ades\u00f5es", "type": "bar", "data": adesoes},
            {"key": "cancelamentos", "name": "Cancelamentos", "type": "bar", "data": cancelamentos},
            {"key": "saldo", "name": "Saldo", "type": "line", "data": saldo},
        ],
        "summary": {
            "total_adesoes": total_adesoes,
            "total_cancelamentos": total_cancelamentos,
            "saldo": saldo_final,
            "saldo_label": f"{'+' if saldo_final > 0 else ''}{saldo_final}",
        },
        "meta": {
            "mode": "annual",
            "year": year,
            "range_label": str(year),
        },
    }


@login_required
@require_GET
def adesoes_cancelamentos_api(request):
    modo = (request.GET.get("mode", "monthly") or "monthly").lower()
    hoje = timezone.localdate()
    usuario = request.user

    try:
        ano = int(request.GET.get("year", hoje.year))
    except (TypeError, ValueError):
        ano = hoje.year

    if modo not in {"monthly", "annual", "lifetime"}:
        modo = "monthly"

    if modo == "annual":
        resultado = _dataset_adesao_cancelamentos_anual(usuario, ano)
        resultado["meta"]["title"] = f"Ades\u00e3o e Cancelamentos por ano - {ano}"
    elif modo == "lifetime":
        resultado = _dataset_adesao_cancelamentos_lifetime(usuario)
    else:
        try:
            mes = int(request.GET.get("month", hoje.month))
        except (TypeError, ValueError):
            mes = hoje.month
        mes = max(1, min(12, mes))
        resultado = _dataset_adesao_cancelamentos_mensal(usuario, ano, mes)
        resultado["meta"]["title"] = f"Ades\u00e3o e Cancelamentos por m\u00eas - {resultado['meta']['month_name']} {ano}"

    return JsonResponse(resultado)


@login_required
def api_listar_todos_clientes(request):
    """
    API para listar todos os clientes do usuário ordenados por data de adesão (mais recente primeiro).
    Retorna: nome, telefone, tem_assinatura, cancelado, data_adesao, logo_servidor
    """
    usuario = request.user
    clientes = Cliente.objects.filter(usuario=usuario).select_related('servidor').order_by('-data_adesao')

    resultado = []
    for cliente in clientes:
        # Obter logo do servidor ou default
        if cliente.servidor:
            logo_url = cliente.servidor.get_imagem_url(usuario)
        else:
            logo_url = '/static/assets/images/logo-apps/default.png'

        resultado.append({
            'id': cliente.id,
            'nome': cliente.nome,
            'telefone': cliente.telefone,
            'tem_assinatura': cliente.tem_assinatura,
            'cancelado': cliente.cancelado,
            'data_adesao': cliente.data_adesao.strftime('%d/%m/%Y') if cliente.data_adesao else None,
            'logo_servidor': logo_url,
            'servidor_nome': cliente.servidor.nome if cliente.servidor else None,
        })

    return JsonResponse({
        'clientes': resultado,
        'total': len(resultado)
    })


@login_required
def api_remover_cliente_sem_assinatura(request, cliente_id):
    """
    Remove um cliente que NÃO possui assinatura/mensalidade.
    Apenas clientes sem assinatura podem ser removidos por esta API.
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método não permitido'}, status=405)

    usuario = request.user

    try:
        cliente = Cliente.objects.get(id=cliente_id, usuario=usuario)

        # Verificar se o cliente tem assinatura
        if cliente.tem_assinatura:
            return JsonResponse({
                'success': False,
                'error': 'Este cliente possui assinatura e não pode ser removido por aqui.'
            }, status=400)

        # Verificar se tem mensalidades
        if Mensalidade.objects.filter(cliente=cliente).exists():
            return JsonResponse({
                'success': False,
                'error': 'Este cliente possui mensalidades e não pode ser removido.'
            }, status=400)

        nome_cliente = cliente.nome
        cliente.delete()

        return JsonResponse({
            'success': True,
            'message': f'Cliente "{nome_cliente}" removido com sucesso.'
        })

    except Cliente.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Cliente não encontrado.'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Erro ao remover cliente: {str(e)}'
        }, status=500)


@login_required
def generate_graphic_columns_per_year(request):
    ano = request.GET.get("ano", timezone.now().year)
    usuario = request.user

    dados_adesoes = (
        Cliente.objects.filter(data_adesao__year=ano, usuario=usuario)
        .annotate(mes=ExtractMonth("data_adesao"))
        .values("mes")
        .annotate(total=Count("id"))
        .order_by("mes")
    )

    dados_cancelamentos = (
        Cliente.objects.filter(data_cancelamento__year=ano, usuario=usuario)
        .annotate(mes=ExtractMonth("data_cancelamento"))
        .values("mes")
        .annotate(total=Count("id"))
        .order_by("mes")
    )

    meses = []
    adesoes = []
    cancelamentos = []

    adesoes_dict = {dado["mes"]: dado["total"] for dado in dados_adesoes}
    cancelamentos_dict = {dado["mes"]: dado["total"] for dado in dados_cancelamentos}

    for mes in range(1, 13):
        if mes in adesoes_dict or mes in cancelamentos_dict:
            meses.append(calendar.month_abbr[mes])
            adesoes.append(adesoes_dict.get(mes, 0))
            cancelamentos.append(cancelamentos_dict.get(mes, 0))

    total_adesoes = sum(adesoes)
    total_cancelamentos = sum(cancelamentos)
    saldo_final = total_adesoes - total_cancelamentos

    plt.figure(figsize=(7, 3))
    plt.bar(meses, adesoes, color="#4CAF50", width=0.4, label="Adesões")
    plt.bar(meses, cancelamentos, color="#F44336", width=0.4, bottom=adesoes, label="Cancelamentos")

    for i, valor in enumerate(adesoes):
        if valor > 0:
            plt.text(i, valor / 2, str(valor), ha='center', va='center', fontsize=10, color='white', fontweight='bold')

    for i, valor in enumerate(cancelamentos):
        if valor > 0:
            plt.text(
                i,
                adesoes[i] + valor / 2,
                str(valor),
                ha='center',
                va='center',
                fontsize=10,
                color='white',
                fontweight='bold',
            )

    plt.title(f'Adesão e Cancelamentos por ano - {ano}', fontsize=14)
    plt.xlabel('Mês', fontsize=12)
    plt.ylabel('Quantidade', fontsize=12)
    plt.xticks(fontsize=10, fontweight='bold')
    plt.yticks(fontsize=10)

    cor_saldo = "#624BFF" if saldo_final >= 0 else "#F44336"
    texto_saldo = f"Saldo {ano}: {'+' if saldo_final > 0 else ''}{saldo_final}"

    saldo_patch = Patch(color=cor_saldo, label=texto_saldo)
    plt.legend(
        handles=[
            Patch(color="#4CAF50", label=f"Adesões: {total_adesoes}"),
            Patch(color="#F44336", label=f"Cancelamentos: {total_cancelamentos}"),
            saldo_patch,
        ]
    )

    plt.gca().spines['top'].set_visible(False)
    plt.gca().spines['right'].set_visible(False)

    buffer = io.BytesIO()
    plt.savefig(buffer, format='png', bbox_inches="tight", dpi=100)
    buffer.seek(0)
    plt.close()

    return HttpResponse(buffer.getvalue(), content_type="image/png")


def user_cache_key(request):
    return f"user-{request.user.id}" if request.user.is_authenticated else "anonymous"


def cache_page_by_user(timeout):
    def decorator(view_func):
        def _wrapped_view(request, *args, **kwargs):
            return cache_page(timeout, key_prefix=user_cache_key(request))(view_func)(request, *args, **kwargs)

        return _wrapped_view

    return decorator


@login_required
@cache_page_by_user(60 * 5)
def mapa_clientes_data(request):
    usuario = request.user

    dados = dict(
        Cliente.objects.filter(cancelado=False, usuario=usuario)
        .values("uf")
        .annotate(total=Count("id"))
        .values_list("uf", "total")
    )

    total_geral = sum(dados.values())
    clientes_internacionais = Cliente.objects.filter(cancelado=False, usuario=usuario, uf__isnull=True).count()

    mapa = gpd.read_file("archives/brasil_estados.geojson")

    def _normalizar_estado(nome):
        if not isinstance(nome, str):
            return ""
        return unicodedata.normalize("NFKD", nome).encode("ascii", "ignore").decode("ascii").lower()

    siglas = {
        "acre": "AC",
        "alagoas": "AL",
        "amapa": "AP",
        "amazonas": "AM",
        "bahia": "BA",
        "ceara": "CE",
        "distrito federal": "DF",
        "espirito santo": "ES",
        "goias": "GO",
        "maranhao": "MA",
        "mato grosso": "MT",
        "mato grosso do sul": "MS",
        "minas gerais": "MG",
        "para": "PA",
        "paraiba": "PB",
        "parana": "PR",
        "pernambuco": "PE",
        "piaui": "PI",
        "rio de janeiro": "RJ",
        "rio grande do norte": "RN",
        "rio grande do sul": "RS",
        "rondonia": "RO",
        "roraima": "RR",
        "santa catarina": "SC",
        "sao paulo": "SP",
        "sergipe": "SE",
        "tocantins": "TO",
    }

    mapa["sigla"] = mapa["name"].apply(lambda nome: siglas.get(_normalizar_estado(nome)))
    mapa = mapa.dropna(subset=["sigla"])
    mapa["clientes"] = mapa["sigla"].apply(lambda uf: dados.get(uf, 0))
    mapa["porcentagem"] = mapa["clientes"].apply(
        lambda x: round((x / total_geral) * 100, 1) if total_geral > 0 else 0
    )

    mapa = mapa.drop(columns=["created_at", "updated_at"], errors="ignore")
    geojson_data = json.loads(mapa.to_json())

    for feature in geojson_data.get("features", []):
        props = feature.get("properties", {})
        props["clientes"] = int(props.get("clientes", 0) or 0)
        props["porcentagem"] = float(props.get("porcentagem", 0) or 0)
        feature["properties"] = props

    max_clientes = int(max(mapa["clientes"]) if mapa["clientes"].any() else 0)

    # Conta clientes por país (excluindo Brasil)
    clientes_por_pais = list(
        Cliente.objects.filter(cancelado=False, usuario=usuario)
        .exclude(pais='BR')
        .exclude(pais__isnull=True)
        .values('pais')
        .annotate(total=Count('id'))
        .order_by('-total')
    )

    # Mapa de códigos de país para nomes legíveis
    PAIS_NOMES = {
        'US': 'Estados Unidos', 'PT': 'Portugal', 'ES': 'Espanha',
        'IT': 'Itália', 'FR': 'França', 'DE': 'Alemanha',
        'GB': 'Reino Unido', 'AR': 'Argentina', 'CL': 'Chile',
        'CH': 'Suíça', 'NL': 'Países Baixos', 'IE': 'Irlanda',
        'AU': 'Austrália', 'JP': 'Japão', 'MX': 'México',
        'CO': 'Colômbia', 'PE': 'Peru', 'VE': 'Venezuela',
        'UY': 'Uruguai', 'PY': 'Paraguai', 'BO': 'Bolívia',
        'EC': 'Equador', 'PA': 'Panamá', 'CR': 'Costa Rica',
        'CA': 'Canadá', 'BE': 'Bélgica', 'AT': 'Áustria',
        'SE': 'Suécia', 'NO': 'Noruega', 'DK': 'Dinamarca',
        'FI': 'Finlândia', 'PL': 'Polônia', 'CZ': 'República Tcheca',
        'RU': 'Rússia', 'CN': 'China', 'KR': 'Coreia do Sul',
        'IN': 'Índia', 'ZA': 'África do Sul', 'AE': 'Emirados Árabes',
        'IL': 'Israel', 'TR': 'Turquia', 'GR': 'Grécia',
        'NZ': 'Nova Zelândia', 'SG': 'Singapura', 'HK': 'Hong Kong',
        'TW': 'Taiwan', 'TH': 'Tailândia', 'MY': 'Malásia',
        'PH': 'Filipinas', 'ID': 'Indonésia', 'VN': 'Vietnã',
        'EG': 'Egito', 'MA': 'Marrocos', 'NG': 'Nigéria',
        'AO': 'Angola', 'MZ': 'Moçambique', 'CV': 'Cabo Verde',
    }

    # Adiciona nome legível a cada país
    for item in clientes_por_pais:
        item['nome'] = PAIS_NOMES.get(item['pais'], item['pais'])

    # Lista de estados com clientes (ordenada por quantidade)
    UF_NOMES = {
        'AC': 'Acre', 'AL': 'Alagoas', 'AP': 'Amapá', 'AM': 'Amazonas',
        'BA': 'Bahia', 'CE': 'Ceará', 'DF': 'Distrito Federal', 'ES': 'Espírito Santo',
        'GO': 'Goiás', 'MA': 'Maranhão', 'MT': 'Mato Grosso', 'MS': 'Mato Grosso do Sul',
        'MG': 'Minas Gerais', 'PA': 'Pará', 'PB': 'Paraíba', 'PR': 'Paraná',
        'PE': 'Pernambuco', 'PI': 'Piauí', 'RJ': 'Rio de Janeiro', 'RN': 'Rio Grande do Norte',
        'RS': 'Rio Grande do Sul', 'RO': 'Rondônia', 'RR': 'Roraima', 'SC': 'Santa Catarina',
        'SP': 'São Paulo', 'SE': 'Sergipe', 'TO': 'Tocantins',
    }

    clientes_por_estado = list(
        Cliente.objects.filter(cancelado=False, usuario=usuario, uf__isnull=False)
        .values('uf')
        .annotate(total=Count('id'))
        .order_by('-total')
    )

    # Adiciona nome legível e percentual a cada estado
    for item in clientes_por_estado:
        item['nome'] = UF_NOMES.get(item['uf'], item['uf'])
        item['percentual'] = round((item['total'] / total_geral) * 100, 1) if total_geral > 0 else 0

    return JsonResponse(
        {
            "features": geojson_data.get("features", []),
            "summary": {
                "total_geral": int(total_geral),
                "fora_pais": int(clientes_internacionais),
                "max_clientes": max_clientes,
                "por_pais": clientes_por_pais,
                "por_estado": clientes_por_estado,
            },
        }
    )


@login_required
@cache_page_by_user(60 * 5)
def clientes_servidor_data(request):
    usuario = request.user

    servidores = list(
        Servidor.objects.filter(usuario=usuario)
        .order_by("nome")
        .values_list("nome", flat=True)
    )

    filtro = (request.GET.get("servidor") or "todos").strip()
    selected = filtro if filtro in servidores else "todos"

    base_queryset = Cliente.objects.filter(usuario=usuario, cancelado=False)

    if selected != "todos":
        queryset = base_queryset.filter(servidor__nome=selected)
        agregados = (
            queryset.values("sistema__nome")
            .annotate(total=Count("id"))
            .order_by("-total", "sistema__nome")
        )
        mode = "aplicativo"
    else:
        queryset = base_queryset
        agregados = (
            queryset.values("servidor__nome")
            .annotate(total=Count("id"))
            .order_by("-total", "servidor__nome")
        )
        mode = "servidor"

    total = sum(item["total"] for item in agregados)

    segments = []
    for item in agregados:
        if mode == "servidor":
            label = item["servidor__nome"] or "Sem servidor"
        else:
            label = item["sistema__nome"] or "Sem aplicativo"
        valor = int(item["total"])
        percent = round((valor / total) * 100, 2) if total > 0 else 0.0
        segments.append(
            {
                "label": label,
                "value": valor,
                "percent": percent,
            }
        )

    segments.sort(key=lambda x: (-x["value"], x["label"]))

    options = ["Todos Servidores"] + servidores
    selected_label = selected if selected != "todos" else "Todos os Servidores"

    return JsonResponse(
        {
            "options": options,
            "selected": selected_label,
            "mode": mode,
            "total": int(total),
            "segments": segments,
        }
    )


@login_required
@never_cache
def notifications_dropdown(request):
    hoje = timezone.localdate()
    tipos = [Tipos_pgto.CARTAO, Tipos_pgto.BOLETO]

    # Mensalidades vencidas
    mensalidades_vencidas = (
        Mensalidade.objects.select_related("cliente", "cliente__forma_pgto", "cliente__plano")
        .filter(
            usuario=request.user,
            pgto=False,
            cancelado=False,
            cliente__cancelado=False,
            cliente__forma_pgto__nome__in=tipos,
            dt_vencimento__lt=hoje,
        )
        .exclude(notifications_read__usuario=request.user)
        .annotate(
            dias_atraso=ExpressionWrapper(
                hoje - F("dt_vencimento"), output_field=DurationField()
            )
        )
        .order_by("dt_vencimento")
    )

    # Notificações do sistema (limite MEI, mudança de plano, etc.)
    notificacoes_sistema = NotificacaoSistema.objects.filter(
        usuario=request.user,
        lida=False
    ).order_by('-criada_em')[:10]

    itens = list(mensalidades_vencidas[:15])
    context = {
        "notif_items": itens,
        "notif_count": mensalidades_vencidas.count() + notificacoes_sistema.count(),
        "notificacoes_sistema": notificacoes_sistema,
    }
    return render(request, "notificacoes/dropdown.html", context)

@login_required
@require_POST
def notifications_mark_all_read(request):
    hoje = timezone.localdate()
    tipos = [Tipos_pgto.CARTAO, Tipos_pgto.BOLETO]
    ids = list(
        Mensalidade.objects.filter(
            usuario=request.user,
            pgto=False, cancelado=False,
            cliente__cancelado=False,
            cliente__forma_pgto__nome__in=tipos,
            dt_vencimento__lt=hoje,
        ).values_list("id", flat=True)
    )

    objetos = [
        NotificationRead(usuario=request.user, mensalidade_id=mensalidade_id)
        for mensalidade_id in ids
    ]
    NotificationRead.objects.bulk_create(objetos, ignore_conflicts=True)

    # Marcar também notificações do sistema como lidas
    notif_sistema_marcadas = NotificacaoSistema.objects.filter(
        usuario=request.user,
        lida=False
    ).update(lida=True, data_leitura=timezone.now())

    request.session.pop("notif_read_ids", None)

    return JsonResponse({"ok": True, "cleared": len(ids), "sistema_cleared": notif_sistema_marcadas})

class NotificationsModalView(LoginRequiredMixin, ListView):
    model = Mensalidade
    template_name = "notificacoes/_lista_modal.html"
    context_object_name = "mensalidades"
    paginate_by = 15

    def get_queryset(self):
        hoje = timezone.localdate()
        return (
            Mensalidade.objects
            .select_related("cliente", "cliente__forma_pgto", "cliente__plano")
            .filter(
                usuario=self.request.user,
                pgto=False,
                cancelado=False,
                cliente__cancelado=False,
                cliente__forma_pgto__nome__in=[Tipos_pgto.CARTAO, Tipos_pgto.BOLETO],
                dt_vencimento__lt=hoje,
            )
            .annotate(
                ja_lida=Exists(
                    NotificationRead.objects.filter(
                        usuario=self.request.user,
                        mensalidade=OuterRef("pk"),
                    )
                ),
                dias_atraso=ExpressionWrapper(
                    hoje - F("dt_vencimento"), output_field=DurationField()
                ),
            )
            .order_by("ja_lida", "dt_vencimento")
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Adicionar notificações do sistema
        context['notificacoes_sistema'] = NotificacaoSistema.objects.filter(
            usuario=self.request.user,
            lida=False
        ).order_by('-criada_em')[:50]
        return context

@login_required
def notifications_count(request):
    hoje = timezone.localdate()
    tipos = [Tipos_pgto.CARTAO, Tipos_pgto.BOLETO]

    # Contagem de mensalidades vencidas
    count_mensalidades = (
        Mensalidade.objects
        .filter(
            usuario=request.user,
            pgto=False, cancelado=False,
            cliente__cancelado=False,
            cliente__forma_pgto__nome__in=tipos,
            dt_vencimento__lt=hoje,
        )
        .exclude(notifications_read__usuario=request.user)
        .count()
    )

    # Contagem de notificações do sistema
    count_sistema = NotificacaoSistema.objects.filter(
        usuario=request.user,
        lida=False
    ).count()

    return JsonResponse({"count": count_mensalidades + count_sistema})

class MensalidadeDetailView(LoginRequiredMixin, DetailView):
    model = Mensalidade
    template_name = "mensalidades/detalhe.html"

    def get_queryset(self):
        # Restringe ao usuário logado
        return Mensalidade.objects.select_related("cliente", "cliente__plano", "cliente__forma_pgto").filter(
            usuario=self.request.user
        )

############################################ UPDATE VIEW ############################################

@require_POST
@login_required
def HorarioEnvio(request):
    usuario = request.user
    horario = request.POST.get('horario')

    try:
        obj_existente = HorarioEnvio.objects.get(usuario=usuario)
    except HorarioEnvio.DoesNotExist:
        obj_existente = None

    if request.POST.get('ativar'):
        if not obj_existente:
            obj = HorarioEnvio.objects.create(usuario=usuario, horario=horario, ativo=True)
            obj.save()
        else:
            obj_existente.horario = horario
            obj_existente.ativo = True
            obj_existente.save()

    elif request.POST.get('desativar'):
        if not obj_existente:
            obj = HorarioEnvio.objects.create(usuario=usuario, horario=None, ativo=False)
            obj.save()
        else:
            obj_existente.horario = None
            obj_existente.ativo = False
            obj_existente.save()

    return JsonResponse({"success_message": "Horário de envio atualizado com sucesso."}, status=200)


@login_required
def session_wpp(request):
    """
    Função de view para criar ou deletar uma sessão do WhatsApp
    """
    if request.method == 'DELETE':
        try:
            sessao = SessaoWpp.objects.filter(usuario=request.user)
            sessao.delete()
            # Realizar outras ações necessárias após a exclusão, se houver
            return JsonResponse({"success_message_session": "Sessão deletada com sucesso."}, status=200)
        except ObjectDoesNotExist:
            return JsonResponse({"error_message": "A sessão não existe."}, status=404)
        except Exception as e:
            return JsonResponse({"error_message": str(e)}, status=500)

    elif request.method in ['PUT', 'PATCH']:
        body = json.loads(request.body)
        token = body.get('token')
        try:
            sessao, created = SessaoWpp.objects.update_or_create(
                usuario=request.user.username,
                defaults={
                    "user": request.user,
                    "token": token,
                    "dt_inicio": timezone.localtime(),
                }
            )
            # Realizar outras ações necessárias após salvar/atualizar, se houver
            return JsonResponse({"success_message_session": "Sessão salva/atualizada com sucesso."}, status=200)
        except Exception as e:
            return JsonResponse({"error_message": str(e)}, status=500)

    else:
        return JsonResponse({"error_message": "Método da requisição não permitido."}, status=405)


@login_required
@transaction.atomic
def reactivate_customer(request, cliente_id):
    """
    View para reativar um cliente anteriormente cancelado.
    Avalia a última mensalidade e cria nova se necessário.

    Regras:
    - Se cancelado há menos de 7 dias: reativa mensalidade existente
    - Se cancelado há mais de 7 dias: cria nova mensalidade com vencimento atual
    - Sempre verifica se já existe mensalidade em aberto para evitar duplicação
    """
    try:
        cliente = Cliente.objects.get(pk=cliente_id, usuario=request.user)
    except Cliente.DoesNotExist:
        return JsonResponse({"error_message": "Cliente não encontrado."}, status=404)

    data_hoje = timezone.localdate()
    sete_dias_atras = data_hoje - timedelta(days=7)
    nova_mensalidade_criada = False
    mensalidade_reativada = False

    logger.info('[%s] [USER][%s] Iniciando reativação do cliente ID: %s',
                timezone.localtime(), request.user, cliente_id)

    # Atualiza os campos de reativação
    cliente.cancelado = False
    cliente.data_cancelamento = None
    cliente.data_vencimento = data_hoje

    # ⭐ ASSOCIAR FORMA DE PAGAMENTO (se fornecida no POST)
    forma_pgto_id = request.POST.get('forma_pagamento_id')
    if forma_pgto_id:
        try:
            forma_pgto = Tipos_pgto.objects.get(id=forma_pgto_id, usuario=request.user)
            cliente.forma_pgto = forma_pgto

            logger.info('[%s] [USER][%s] Cliente ID %s associado à forma de pagamento ID %s',
                        timezone.localtime(), request.user, cliente_id, forma_pgto_id)

            # Se tem conta bancária, criar/atualizar associação ClienteContaBancaria
            if forma_pgto.conta_bancaria:
                # Desativar associações anteriores
                ClienteContaBancaria.objects.filter(
                    cliente=cliente,
                    ativo=True
                ).update(ativo=False)

                # Criar nova associação
                ClienteContaBancaria.objects.update_or_create(
                    cliente=cliente,
                    conta_bancaria=forma_pgto.conta_bancaria,
                    defaults={'ativo': True}
                )

                logger.info('[%s] [USER][%s] Cliente ID %s associado à conta bancária ID %s',
                            timezone.localtime(), request.user, cliente_id, forma_pgto.conta_bancaria.id)

        except Tipos_pgto.DoesNotExist:
            logger.warning('[%s] [USER][%s] Forma de pagamento ID %s não encontrada',
                           timezone.localtime(), request.user, forma_pgto_id)

    cliente.save()

    logger.info('[%s] [USER][%s] Cliente ID %s marcado como ativo',
                timezone.localtime(), request.user, cliente_id)

    # histórico: inicia novo período vigente
    try:
        historico_iniciar(cliente, inicio=data_hoje, motivo='reactivate')
    except Exception:
        pass

    # ⭐ FASE 2: Gerenciamento de campanha na reativação
    # REGRA:
    #   - Se cancelado há ≤ 7 dias: PRESERVA campanha anterior (não chama enroll que zera)
    #   - Se cancelado há > 7 dias: REMOVE campanha (cliente deve trocar plano para nova promoção)
    try:
        from nossopainel.models import AssinaturaCliente

        # Calcula há quantos dias o cliente estava cancelado
        dias_cancelado = None
        if cliente.data_cancelamento:
            dias_cancelado = (data_hoje - cliente.data_cancelamento).days

        if dias_cancelado is not None and dias_cancelado <= 7:
            # ≤ 7 dias: Preserva campanha anterior (não faz nada, dados já estão na AssinaturaCliente)
            logger.info(
                f"[CAMPANHA] Cliente {cliente.id} reativado em {dias_cancelado} dias. "
                f"Campanha anterior preservada."
            )
        else:
            # > 7 dias: Remove campanha - cliente precisa trocar plano para nova promoção
            try:
                assinatura = AssinaturaCliente.objects.filter(cliente=cliente, ativo=True).first()
                if assinatura and assinatura.em_campanha:
                    assinatura.em_campanha = False
                    assinatura.campanha_data_adesao = None
                    assinatura.campanha_mensalidades_pagas = 0
                    assinatura.campanha_duracao_total = None
                    assinatura.save()
                    logger.info(
                        f"[CAMPANHA] Cliente {cliente.id} reativado após {dias_cancelado} dias. "
                        f"Campanha removida - deve trocar plano para nova promoção."
                    )
            except Exception as e:
                logger.error(f"Erro ao remover campanha do cliente {cliente.id}: {e}", exc_info=True)

    except Exception as e:
        logger.error(f"Erro ao gerenciar campanha durante reativação: {e}", exc_info=True)

    try:
        # ⭐ FASE 2: Calcula valor considerando campanhas promocionais + descontos progressivos
        valor_mensalidade = calcular_valor_mensalidade(cliente)

        logger.info('[%s] [USER][%s] Valor mensalidade calculado para cliente ID %s: R$ %s',
                    timezone.localtime(), request.user, cliente_id, valor_mensalidade)

        # PROTEÇÃO CONTRA DUPLICAÇÃO: Verifica se já existe mensalidade em aberto
        mensalidade_em_aberto = Mensalidade.objects.filter(
            cliente=cliente,
            pgto=False,
            cancelado=False,
            dt_vencimento__gte=data_hoje
        ).first()

        if mensalidade_em_aberto:
            # Já existe mensalidade em aberto, apenas atualiza o valor
            logger.info('[%s] [USER][%s] Cliente ID %s já possui mensalidade em aberto (ID: %s). Atualizando valor.',
                        timezone.localtime(), request.user, cliente_id, mensalidade_em_aberto.id)
            mensalidade_em_aberto.valor = valor_mensalidade
            mensalidade_em_aberto.save()
            mensalidade_reativada = True
        else:
            # Obtém a última mensalidade do cliente
            ultima_mensalidade = Mensalidade.objects.filter(cliente=cliente).order_by('-dt_vencimento').first()

            if ultima_mensalidade:
                # Verifica se a mensalidade foi cancelada há menos de 7 dias
                if sete_dias_atras <= ultima_mensalidade.dt_vencimento <= data_hoje:
                    # CASO 1: Cancelado há menos de 7 dias - Reativa a mensalidade existente
                    logger.info('[%s] [USER][%s] Cliente ID %s cancelado há menos de 7 dias. Reativando mensalidade ID: %s',
                                timezone.localtime(), request.user, cliente_id, ultima_mensalidade.id)

                    ultima_mensalidade.cancelado = False
                    ultima_mensalidade.dt_cancelamento = None
                    ultima_mensalidade.valor = valor_mensalidade
                    ultima_mensalidade.save()
                    mensalidade_reativada = True
                else:
                    # CASO 2: Cancelado há mais de 7 dias - Cria nova mensalidade
                    logger.info('[%s] [USER][%s] Cliente ID %s cancelado há mais de 7 dias. Criando nova mensalidade.',
                                timezone.localtime(), request.user, cliente_id)

                    # Mantém a mensalidade anterior como cancelada
                    ultima_mensalidade.cancelado = True
                    ultima_mensalidade.dt_cancelamento = data_hoje
                    ultima_mensalidade.save()

                    # Cria nova mensalidade com desconto progressivo aplicado
                    nova_mensalidade = Mensalidade.objects.create(
                        cliente=cliente,
                        valor=valor_mensalidade,
                        dt_vencimento=data_hoje,
                        usuario=cliente.usuario
                    )
                    nova_mensalidade_criada = True

                    logger.info('[%s] [USER][%s] Nova mensalidade ID %s criada para cliente ID %s',
                                timezone.localtime(), request.user, nova_mensalidade.id, cliente_id)
            else:
                # CASO 3: Não existe mensalidade anterior - Cria primeira mensalidade
                logger.info('[%s] [USER][%s] Cliente ID %s sem mensalidade anterior. Criando primeira mensalidade.',
                            timezone.localtime(), request.user, cliente_id)

                nova_mensalidade = Mensalidade.objects.create(
                    cliente=cliente,
                    valor=valor_mensalidade,
                    dt_vencimento=data_hoje,
                    usuario=cliente.usuario
                )
                nova_mensalidade_criada = True

                logger.info('[%s] [USER][%s] Primeira mensalidade ID %s criada para cliente ID %s',
                            timezone.localtime(), request.user, nova_mensalidade.id, cliente_id)

    except Exception as erro:
        logger.error('[%s] [USER][%s] [IP][%s] [ERRO][%s]',
                     timezone.localtime(), request.user,
                     get_client_ip(request) or 'N/A', erro,
                     exc_info=True)
        return JsonResponse({"error_message": "Erro ao processar mensalidade na reativação."}, status=500)

    logger.info('[%s] [USER][%s] Reativação do cliente ID %s concluída. Nova mensalidade: %s | Mensalidade reativada: %s',
                timezone.localtime(), request.user, cliente_id, nova_mensalidade_criada, mensalidade_reativada)

    log_user_action(
        request=request,
        action=UserActionLog.ACTION_REACTIVATE,
        instance=cliente,
        message="Cliente reativado.",
        extra={
            "nova_mensalidade_criada": nova_mensalidade_criada,
            "mensalidade_reativada": mensalidade_reativada,
        },
    )
    return JsonResponse({"success_message_activate": "Reativação feita com sucesso!"})


# AÇÃO DE PAGAR MENSALIDADE
@login_required
def pay_monthly_fee(request, mensalidade_id):
    """
    Função de view para pagar uma mensalidade.

    Proteção contra duplicação:
    - Usa select_for_update() para bloquear a mensalidade durante processamento
    - Verifica se já está paga antes de processar
    - Verifica se foi criada mensalidade futura nos últimos 60 segundos
    """
    from django.db import transaction

    hoje = timezone.localtime().date()

    try:
        with transaction.atomic():
            # Bloqueia a mensalidade para evitar processamento duplicado
            mensalidade = Mensalidade.objects.select_for_update(nowait=True).get(
                pk=mensalidade_id,
                usuario=request.user
            )

            # PROTEÇÃO: Verifica se já está paga
            if mensalidade.pgto:
                logger.info(f"[ANTI-DUP] Mensalidade {mensalidade_id} já está paga. Ignorando requisição duplicada.")
                return JsonResponse({"success_message_invoice": "Mensalidade já estava paga!"}, status=200)

            # Verifica se a mensalidade está atrasada por mais de 7 dias
            if mensalidade.dt_vencimento < hoje - timedelta(days=7):
                return JsonResponse({"error_message": "erro"})

            # PROTEÇÃO: Verifica se já existe mensalidade futura (criada por outra requisição)
            if Mensalidade.objects.filter(
                cliente=mensalidade.cliente,
                dt_vencimento__gt=mensalidade.dt_vencimento,
                pgto=False,
                cancelado=False
            ).exists():
                logger.warning(f"[ANTI-DUP] Mensalidade futura já existe para cliente {mensalidade.cliente.nome}. Pagamento já foi processado.")
                return JsonResponse({"success_message_invoice": "Mensalidade paga!"}, status=200)

            # Realiza as modificações na mensalidade paga
            mensalidade.dt_pagamento = timezone.localtime().date()
            mensalidade.pgto = True
            # Registra a forma de pagamento usada (histórico imutável)
            if not mensalidade.forma_pgto:
                mensalidade.forma_pgto = mensalidade.cliente.forma_pgto
            mensalidade.save()

    except Mensalidade.DoesNotExist:
        return JsonResponse({"error_message": "Mensalidade não encontrada."}, status=404)
    except Exception as e:
        # Se a linha está bloqueada (nowait=True), outra requisição está processando
        if 'could not obtain lock' in str(e).lower() or 'lock' in str(e).lower():
            logger.info(f"[ANTI-DUP] Mensalidade {mensalidade_id} está sendo processada por outra requisição.")
            return JsonResponse({"success_message_invoice": "Pagamento em processamento..."}, status=200)
        raise

    try:
        log_user_action(
            request=request,
            action=UserActionLog.ACTION_PAYMENT,
            instance=mensalidade,
            message="Mensalidade marcada como paga.",
            extra={
                "cliente": mensalidade.cliente_id,
                "valor": str(mensalidade.valor),
                "dt_pagamento": mensalidade.dt_pagamento.strftime('%Y-%m-%d') if mensalidade.dt_pagamento else '',
            },
        )

    except Exception as erro:
        logger.error('[%s] [USER][%s] [IP][%s] [ERRO][%s]', timezone.localtime(), request.user, get_client_ip(request) or 'N/A', erro, exc_info=True)
        return JsonResponse({"error_message": "Ocorreu um erro ao tentar pagar essa mensalidade."}, status=500)

    # Retorna uma resposta JSON indicando que a mensalidade foi paga com sucesso
    return JsonResponse({"success_message_invoice": "Mensalidade paga!"}, status=200)


# AÇÃO PARA CANCELAMENTO DE CLIENTE
@login_required
def cancel_customer(request, cliente_id):
    if request.user.is_authenticated:
        cliente = Cliente.objects.get(pk=cliente_id, usuario=request.user)

        # Realiza as modificações no cliente
        cliente.cancelado = True
        cliente.data_cancelamento = timezone.localtime().date()
        try:
            cliente.save()
        except Exception as erro:
            logger.error('[%s] [USER][%s] [IP][%s] [ERRO][%s]', timezone.localtime(), request.user, get_client_ip(request) or 'N/A', erro, exc_info=True)
            return JsonResponse({"error_message": "Ocorreu um erro ao tentar cancelar esse cliente."}, status=500)

        # Cancelar todas as mensalidades relacionadas ao cliente
        mensalidades = cliente.mensalidade_set.filter(dt_vencimento__gte=timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0), pgto=False, cancelado=False)
        mensalidades_count = mensalidades.count()
        for mensalidade in mensalidades:
            mensalidade.cancelado = True
            mensalidade.dt_cancelamento = timezone.localtime().date()
            mensalidade.save()

        # Encerra histórico vigente na data do cancelamento
        try:
            historico_encerrar_vigente(cliente, timezone.localdate())
        except Exception:
            pass

        # Retorna uma resposta JSON indicando que o cliente foi cancelado com sucesso
        log_user_action(
            request=request,
            action=UserActionLog.ACTION_CANCEL,
            instance=cliente,
            message="Cliente cancelado.",
            extra={
                "mensalidades_canceladas": mensalidades_count,
            },
        )
        return JsonResponse({"success_message_cancel": "Eita! mais um cliente cancelado?! "})
    else:
        return redirect("login")


# Função para extrair UF pelo DDD
def extrair_uf_do_telefone(telefone):
    telefone_digits = re.sub(r'\D+', '', telefone)
    if telefone_digits.startswith('55'):
        telefone_digits = telefone_digits[2:]
    ddd = telefone_digits[:2]
    return DDD_UF_MAP.get(ddd, '')

# Função ajustada para calcular a nova data de vencimento com base no plano e data base
def calcular_nova_data_vencimento(plano_nome, data_base):
    meses = PLANOS_MESES.get(plano_nome.lower(), 0)
    mes = data_base.month + meses
    ano = data_base.year

    if mes > 12:
        mes -= 12
        ano += 1

    dia = data_base.day
    if dia == 31 and mes not in MESES_31_DIAS:
        dia = 1
        mes += 1
        if mes > 12:
            mes = 1
            ano += 1

    try:
        return date(year=ano, month=mes, day=dia)
    except ValueError:
        return date(year=ano, month=mes, day=1)


@login_required
def api_cliente_contas(request, cliente_id):
    """
    API para carregar as contas de aplicativos de um cliente.
    Retorna JSON com todas as contas incluindo dispositivo, app, credenciais e status principal.
    """
    try:
        cliente = Cliente.objects.get(id=cliente_id, usuario=request.user)

        # Busca todas as contas do cliente com relacionamentos
        contas = ContaDoAplicativo.objects.filter(cliente=cliente).select_related(
            'dispositivo', 'app'
        ).order_by('-is_principal', 'id')  # Principal primeiro, depois por ordem de criação

        # Serializa para JSON
        contas_data = []
        for conta in contas:
            conta_dict = {
                'id': conta.id,
                'dispositivo_id': conta.dispositivo.id if conta.dispositivo else None,
                'dispositivo_nome': conta.dispositivo.nome if conta.dispositivo else None,
                'app_id': conta.app.id,
                'app_nome': conta.app.nome,
                'app_device_has_mac': conta.app.device_has_mac,
                'device_id': conta.device_id or '',
                'email': conta.email or '',
                'device_key': conta.device_key or '',
                'is_principal': conta.is_principal,
                'verificado': conta.verificado,
            }
            contas_data.append(conta_dict)

        return JsonResponse({
            'success': True,
            'contas': contas_data,
            'total': len(contas_data)
        })

    except Cliente.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Cliente não encontrado'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@transaction.atomic
def edit_customer(request, cliente_id):
    if request.method != "POST":
        return redirect("dashboard")

    try:
        post = request.POST
        user = request.user
        token = SessaoWpp.objects.filter(usuario=user, is_active=True).first()
        cliente = get_object_or_404(Cliente, pk=cliente_id, usuario=user)
        mensalidade = get_object_or_404(Mensalidade, cliente=cliente, pgto=False, cancelado=False, usuario=user)

        original_cliente = {
            "nome": cliente.nome,
            "telefone": cliente.telefone,
            "uf": cliente.uf,
            "indicado_por": cliente.indicado_por,
            "servidor": cliente.servidor,
            "forma_pgto": cliente.forma_pgto,
            "plano": cliente.plano,
            "data_vencimento": cliente.data_vencimento,
            "nao_enviar_msgs": cliente.nao_enviar_msgs,
            "notas": cliente.notas,
        }
        original_mensalidade = {
            "dt_vencimento": mensalidade.dt_vencimento,
            "valor": mensalidade.valor,
        }

        # Nome
        nome = post.get("nome", "").strip()
        if cliente.nome != nome:
            cliente.nome = nome

        # Telefone
        telefone_novo = post.get("telefone", "").strip()
        if cliente.telefone != telefone_novo:

            # Valida o novo telefone
            resultado_wpp = validar_tel_whatsapp(telefone_novo, token.token, user)
            cliente_existe_telefone = resultado_wpp.get("cliente_existe_telefone")
            telefone_novo = resultado_wpp.get("telefone_validado_wpp")

            if not resultado_wpp.get("wpp"):
                logger.warning(
                    f"[WARN][EDITAR CLIENTE] O número {telefone_novo} não possui um WhatsApp."
                )
                return JsonResponse({
                    "error": True,
                    "error_message_edit": (
                        f"O número {telefone_novo} não possui um WhatsApp.<br>"
                        "O cadastro do cliente precisa ter um número ativo no WhatsApp."
                    ),
                }, status=500)

            # Se o novo número existe para outro cliente cadastrado
            if cliente_existe_telefone:
                # Garante que não é o próprio cliente que está editando
                cliente_existente = Cliente.objects.filter(
                    telefone=cliente_existe_telefone,
                    usuario=user
                ).exclude(pk=cliente_id).first()
                if cliente_existente:
                    logger.warning(
                        f"[WARN][EDITAR CLIENTE] Já existe cliente para telefone {cliente_existente.telefone} (ID: {cliente_existente.id})"
                    )
                    return JsonResponse({
                        "error": True,
                        "error_message_edit": (
                            "Há um cliente cadastrado com o telefone informado! <br><br>"
                            f"<strong>Nome:</strong> {cliente_existente.nome} <br>"
                            f"<strong>Telefone:</strong> {cliente_existente.telefone}"
                        ),
                    }, status=500)
                # Se chegou aqui, está tudo certo (ou é ele mesmo)
            # Atualiza telefone, UF e País
            cliente.telefone = telefone_novo
            cliente.uf = extrair_uf_do_telefone(telefone_novo)
            cliente.pais = extrair_pais_do_telefone(telefone_novo)

        # Indicação
        indicado_nome = post.get("indicado_por")
        if indicado_nome:
            indicado = Cliente.objects.filter(nome=indicado_nome, usuario=user).first()
            if indicado and cliente.indicado_por != indicado:
                cliente.indicado_por = indicado

        # Servidor
        servidor = get_object_or_404(Servidor, nome=post.get("servidor"), usuario=user)
        if cliente.servidor != servidor:
            cliente.servidor = servidor

        # Forma de pagamento (agora usando ID ao invés de nome)
        forma_pgto_id = post.get("forma_pgto", "")
        if forma_pgto_id and forma_pgto_id.isdigit():
            forma_pgto = Tipos_pgto.objects.filter(pk=int(forma_pgto_id), usuario=user).first()
            if forma_pgto and cliente.forma_pgto != forma_pgto:
                # Valida se a nova forma de pagamento não está bloqueada por limite
                if forma_pgto.esta_bloqueada:
                    return JsonResponse({
                        "error": True,
                        "error_message_edit": (
                            "A forma de pagamento selecionada atingiu o limite de faturamento "
                            "e não pode receber novos clientes.<br>"
                            "Selecione outra forma de pagamento disponível."
                        ),
                    }, status=400)
                cliente.forma_pgto = forma_pgto

        # Plano (agora usando ID ao invés de string formatada)
        plano_id = post.get("plano", "")
        if plano_id and plano_id.isdigit():
            plano = Plano.objects.filter(pk=int(plano_id), usuario=user).first()
            if plano and cliente.plano != plano:
                cliente.plano = plano
                mensalidade.valor = calcular_valor_mensalidade(cliente)
                # Atualiza histórico de planos: encerra vigente e inicia novo
                hoje = timezone.localdate()
                try:
                    historico_encerrar_vigente(cliente, fim=hoje - timedelta(days=1))
                    historico_iniciar(cliente, plano=plano, inicio=hoje, motivo='plan_change')
                except Exception:
                    pass

                # ⭐ FASE 2.5: Sync AssinaturaCliente.plano and reset campaign tracking when plan changes
                try:
                    from .models import AssinaturaCliente
                    assinatura = AssinaturaCliente.objects.get(cliente=cliente, ativo=True)

                    # Sincroniza o plano da assinatura com o novo plano do cliente
                    assinatura.plano = plano

                    # Se estava em campanha, reseta os campos de rastreamento
                    if assinatura.em_campanha:
                        assinatura.em_campanha = False
                        assinatura.campanha_data_adesao = None
                        assinatura.campanha_mensalidades_pagas = 0
                        assinatura.campanha_duracao_total = None
                        logger.info(
                            f"[CAMPANHA] Rastreamento de campanha resetado para {cliente.nome} devido a mudança de plano"
                        )

                    # Salva a assinatura (sempre, não só quando em campanha)
                    assinatura.save()

                except AssinaturaCliente.DoesNotExist:
                    pass  # No subscription record

                # ⭐ FASE 2.5: Re-enroll client in new plan's campaign if applicable
                if plano.campanha_ativa:
                    from .utils import enroll_client_in_campaign_if_eligible
                    enroll_result = enroll_client_in_campaign_if_eligible(cliente)
                    if enroll_result:
                        # Recalculate mensalidade value with new campaign
                        valor_base = plano.valor
                        valor_campanha = calcular_valor_mensalidade(cliente)

                        # Atualizar campos de rastreamento da mensalidade atual
                        mensalidade.valor = valor_campanha
                        mensalidade.gerada_em_campanha = True
                        mensalidade.valor_base_plano = valor_base
                        mensalidade.desconto_campanha = valor_base - valor_campanha
                        mensalidade.tipo_campanha = plano.campanha_tipo
                        mensalidade.numero_mes_campanha = 1  # Primeira mensalidade da campanha

                        logger.info(
                            f"[CAMPANHA] Cliente {cliente.nome} inscrito na campanha do novo plano '{plano.nome}'"
                        )
                else:
                    # Plano Regular - resetar campos de campanha da mensalidade
                    mensalidade.gerada_em_campanha = False
                    mensalidade.desconto_campanha = Decimal("0.00")
                    mensalidade.tipo_campanha = None
                    mensalidade.numero_mes_campanha = None

        # Data de vencimento
        data_vencimento_str = post.get("dt_pgto", "").strip()
        data_base = None

        for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
            try:
                data_base = datetime.strptime(data_vencimento_str, fmt).date()
                break
            except ValueError:
                continue

        if not data_base:
            return JsonResponse({
                "error": True,
                "error_message_edit": "Data de vencimento inválida. Use o formato DD/MM/AAAA ou AAAA-MM-DD."
            }, status=400)

        nova_data_vencimento = data_base

        if cliente.data_vencimento != nova_data_vencimento:
            cliente.data_vencimento = nova_data_vencimento
            mensalidade.dt_vencimento = nova_data_vencimento
            if plano:
                mensalidade.valor = calcular_valor_mensalidade(cliente)

        # Não enviar mensagens automáticas
        nao_enviar_msgs = post.get("nao_enviar_msgs") == "on"
        if cliente.nao_enviar_msgs != nao_enviar_msgs:
            cliente.nao_enviar_msgs = nao_enviar_msgs

        # Notas
        notas = post.get("notas", "").strip()
        if cliente.notas != notas:
            cliente.notas = notas

        # ========== PROCESSAMENTO DE CONTAS DE APLICATIVOS ==========
        # Processa dados de contas do aplicativo enviadas pelo formulário
        conta_principal_id = None
        contas_atualizadas = []

        # Verifica se há contas sendo editadas no formulário
        has_conta_fields = any(key.startswith('conta_') and '_' in key[6:] for key in post.keys())

        # Só desmarca todas as contas se houver contas sendo editadas
        # (evita desmarcar a principal quando apenas outros campos do cliente são alterados)
        if has_conta_fields:
            ContaDoAplicativo.objects.filter(cliente=cliente).update(is_principal=False)

        for key in post.keys():
            # Identifica campos de contas: conta_{id}_campo
            if key.startswith('conta_') and '_' in key[6:]:
                try:
                    # Extrai ID da conta
                    parts = key.split('_')
                    if len(parts) >= 3:
                        conta_id = int(parts[1])
                        campo = '_'.join(parts[2:])  # Pode ser 'is_principal', 'dispositivo_id', 'app_id', etc.

                        # Verifica se esta conta já foi processada neste loop
                        if conta_id not in contas_atualizadas:
                            contas_atualizadas.append(conta_id)

                            # Busca a conta
                            conta = ContaDoAplicativo.objects.filter(
                                id=conta_id,
                                cliente=cliente
                            ).first()

                            if conta:
                                # Atualiza campos da conta
                                dispositivo_id = post.get(f'conta_{conta_id}_dispositivo_id')
                                app_id = post.get(f'conta_{conta_id}_app_id')
                                device_id = post.get(f'conta_{conta_id}_device_id', '').strip()
                                email = post.get(f'conta_{conta_id}_email', '').strip()
                                device_key = post.get(f'conta_{conta_id}_device_key', '').strip()
                                is_principal = post.get(f'conta_{conta_id}_is_principal') == 'on'

                                # Atualiza dispositivo
                                if dispositivo_id:
                                    dispositivo = Dispositivo.objects.filter(id=dispositivo_id, usuario=user).first()
                                    if dispositivo:
                                        conta.dispositivo = dispositivo

                                # Atualiza aplicativo
                                if app_id:
                                    app = Aplicativo.objects.filter(id=app_id, usuario=user).first()
                                    if app:
                                        conta.app = app

                                # Atualiza credenciais
                                conta.device_id = device_id or None
                                conta.email = email or None
                                conta.device_key = device_key or None

                                # Atualiza status de conta principal
                                conta.is_principal = is_principal

                                conta.save()

                                # Log para debug
                                if is_principal:
                                    logger.info(
                                        f"[EDIT_CUSTOMER] Conta {conta_id} marcada como principal para cliente {cliente.nome}"
                                    )

                except (ValueError, IndexError):
                    continue  # Ignora campos mal formatados

        # ========== VALIDAÇÃO: GARANTIR QUE SEMPRE HAJA UMA CONTA PRINCIPAL ==========
        # Se contas foram editadas, garante que ao menos uma seja principal
        if has_conta_fields:
            # Verifica se alguma conta foi marcada como principal
            tem_principal = ContaDoAplicativo.objects.filter(cliente=cliente, is_principal=True).exists()

            if not tem_principal:
                # Se nenhuma conta foi marcada como principal, marca a primeira
                primeira_conta = ContaDoAplicativo.objects.filter(cliente=cliente).order_by('id').first()
                if primeira_conta:
                    primeira_conta.is_principal = True
                    primeira_conta.save()
                    logger.warning(
                        f"[EDIT_CUSTOMER] Nenhuma conta foi marcada como principal para cliente {cliente.nome}. "
                        f"Marcando automaticamente a conta ID {primeira_conta.id} como principal."
                    )

        # A sincronização dos campos Cliente.dispositivo e Cliente.sistema com a conta principal
        # é feita automaticamente pelo signal sincronizar_conta_principal quando conta.save() é chamado
        # Por isso, usamos update_fields para salvar APENAS os campos do formulário,
        # evitando sobrescrever dispositivo/sistema que o signal gerencia
        cliente.save(update_fields=[
            'nome', 'telefone', 'uf', 'indicado_por', 'servidor',
            'forma_pgto', 'plano', 'data_vencimento', 'nao_enviar_msgs', 'notas'
        ])
        mensalidade.save()

        changes = {}

        def _add_change(key, old, new):
            if old != new:
                changes[key] = (old, new)

        _add_change("nome", original_cliente["nome"], cliente.nome)
        _add_change("telefone", original_cliente["telefone"], cliente.telefone)
        _add_change("uf", original_cliente["uf"], cliente.uf)
        _add_change("indicado_por", original_cliente["indicado_por"], cliente.indicado_por)
        _add_change("servidor", original_cliente["servidor"], cliente.servidor)
        _add_change("forma_pgto", original_cliente["forma_pgto"], cliente.forma_pgto)
        _add_change("plano", original_cliente["plano"], cliente.plano)
        _add_change("data_vencimento", original_cliente["data_vencimento"], cliente.data_vencimento)
        _add_change("nao_enviar_msgs", original_cliente["nao_enviar_msgs"], cliente.nao_enviar_msgs)
        _add_change("notas", original_cliente["notas"], cliente.notas)
        _add_change("mensalidade.dt_vencimento", original_mensalidade["dt_vencimento"], mensalidade.dt_vencimento)
        _add_change("mensalidade.valor", original_mensalidade["valor"], mensalidade.valor)

        mensagem_log = "Cliente atualizado." if changes else "Cliente salvo sem alteracoes."
        log_user_action(
            request=request,
            action=UserActionLog.ACTION_UPDATE,
            instance=cliente,
            message=mensagem_log,
            extra=changes if changes else None,
        )

        return JsonResponse({
            "success": True,
            "success_message": f"<strong>{cliente.nome}</strong> foi atualizado com sucesso."
        }, status=200)

    except Exception as e:
        logger.error('[%s] [ERROR][EDITAR CLIENTE] [USER][%s] [IP][%s] [%s]',
                     timezone.localtime(), request.user,
                     get_client_ip(request) or 'N/A', e, exc_info=True)
        return JsonResponse({
            "error": True,
            "error_message": "Ocorreu um erro ao tentar atualizar esse cliente."
        }, status=500)



# AÇÃO PARA EDITAR O OBJETO PLANO MENSAL
@login_required
def edit_payment_plan(request, plano_id):
    """
    Função de view para editar um plano de adesão mensal.
    """
    plano_mensal = get_object_or_404(Plano, pk=plano_id, usuario=request.user)

    original_plano = {
        "nome": plano_mensal.nome,
        "telas": plano_mensal.telas,
        "valor": plano_mensal.valor,
        "campanha_ativa": plano_mensal.campanha_ativa,  # ⭐ FASE 2
    }

    planos_mensalidades = Plano.objects.all().order_by('nome')

    if request.method == "POST":
        nome = request.POST.get("nome")
        telas = request.POST.get("telas")
        valor = request.POST.get("valor")

        if nome and valor:
            plano_mensal.nome = nome
            plano_mensal.telas = telas
            plano_mensal.valor = Decimal(valor.replace(',', '.'))

            # ⭐ FASE 2: Update campaign data
            campanha_ativa_nova = request.POST.get('campanha_ativa') == 'on'
            campanha_ativa_atual = plano_mensal.campanha_ativa

            # Validar duplicidade ao mudar status da campanha
            if campanha_ativa_nova != campanha_ativa_atual:
                # Verificar se já existe plano com a nova configuração
                planos_conflitantes = Plano.objects.filter(
                    nome=nome,
                    valor=Decimal(valor.replace(',', '.')),
                    telas=telas,
                    usuario=request.user,
                    campanha_ativa=campanha_ativa_nova
                ).exclude(pk=plano_mensal.pk)

                if planos_conflitantes.exists():
                    tipo = "com campanha" if campanha_ativa_nova else "regular"
                    return render(
                        request,
                        "pages/cadastro-plano-adesao.html",
                        {
                            'planos_mensalidades': planos_mensalidades,
                            "error_message": f"Já existe um Plano {tipo} para esta configuração!",
                        },
                    )

            plano_mensal.campanha_ativa = campanha_ativa_nova

            if campanha_ativa_nova:
                plano_mensal.campanha_tipo = request.POST.get('campanha_tipo', 'FIXO')
                plano_mensal.campanha_data_inicio = request.POST.get('campanha_data_inicio') or None
                plano_mensal.campanha_data_fim = request.POST.get('campanha_data_fim') or None
                plano_mensal.campanha_duracao_meses = request.POST.get('campanha_duracao_meses') or None

                if plano_mensal.campanha_tipo == 'FIXO':
                    campanha_valor_fixo = request.POST.get('campanha_valor_fixo')
                    plano_mensal.campanha_valor_fixo = Decimal(campanha_valor_fixo.replace(',', '.')) if campanha_valor_fixo else None
                else:  # PERSONALIZADO
                    for i in range(1, 13):
                        campo = f'campanha_valor_mes_{i}'
                        valor_mes = request.POST.get(campo)
                        setattr(plano_mensal, campo, Decimal(valor_mes.replace(',', '.')) if valor_mes else None)
            else:
                # Clear campaign data if not active
                plano_mensal.campanha_tipo = None
                plano_mensal.campanha_data_inicio = None
                plano_mensal.campanha_data_fim = None
                plano_mensal.campanha_duracao_meses = None
                plano_mensal.campanha_valor_fixo = None
                plano_mensal.campanha_valor_mes_1 = None
                plano_mensal.campanha_valor_mes_2 = None
                plano_mensal.campanha_valor_mes_3 = None
                plano_mensal.campanha_valor_mes_4 = None
                plano_mensal.campanha_valor_mes_5 = None
                plano_mensal.campanha_valor_mes_6 = None
                plano_mensal.campanha_valor_mes_7 = None
                plano_mensal.campanha_valor_mes_8 = None
                plano_mensal.campanha_valor_mes_9 = None
                plano_mensal.campanha_valor_mes_10 = None
                plano_mensal.campanha_valor_mes_11 = None
                plano_mensal.campanha_valor_mes_12 = None

            try:
                plano_mensal.save()
                changes = {}
                if original_plano["nome"] != plano_mensal.nome:
                    changes["nome"] = (original_plano["nome"], plano_mensal.nome)
                if original_plano["telas"] != plano_mensal.telas:
                    changes["telas"] = (original_plano["telas"], plano_mensal.telas)
                if original_plano["valor"] != plano_mensal.valor:
                    changes["valor"] = (original_plano["valor"], plano_mensal.valor)
                # ⭐ FASE 2: Tracking de mudanças da campanha
                if original_plano["campanha_ativa"] != plano_mensal.campanha_ativa:
                    changes["campanha_ativa"] = (
                        original_plano["campanha_ativa"],
                        plano_mensal.campanha_ativa
                    )
                mensagem_plano = "Plano atualizado." if changes else "Plano salvo sem alteracoes."
                log_user_action(
                    request=request,
                    action=UserActionLog.ACTION_UPDATE,
                    instance=plano_mensal,
                    message=mensagem_plano,
                    extra=changes if changes else None,
                )


            except ValidationError as erro1:
                logger.error('[%s][USER][%s] [IP][%s] [ERRO][%s]', timezone.localtime(), request.user, get_client_ip(request) or 'N/A', erro1, exc_info=True)
                # Capturando a exceção ValidationError e renderizando a página novamente com a mensagem de erro
                return render(
                    request,
                    "pages/cadastro-plano-adesao.html",
                    {
                        'planos_mensalidades': planos_mensalidades,
                        "error_message": "Já existe um plano com este nome!",
                    },
                )

            except Exception as erro2:
                logger.error('[%s] [USER][%s] [IP][%s] [ERRO][%s]', timezone.localtime(), request.user, get_client_ip(request) or 'N/A', erro2, exc_info=True)
                # Capturando outros possíveis erros ao tentar salvar o plano e renderizando a página novamente com a mensagem de erro
                return render(
                    request,
                    "pages/cadastro-plano-adesao.html",
                    {
                        'planos_mensalidades': planos_mensalidades,
                        "error_message": "Ocorreu um erro ao salvar o plano. Verifique o log!",
                    },
                )
            
            # Em caso de sucesso, renderiza a página novamente com a mensagem de sucesso
            return render(
                request,
                "pages/cadastro-plano-adesao.html",
                {"planos_mensalidades": planos_mensalidades, "success_update": True},
            )

        else:
            return render(
                request,
                "pages/cadastro-plano-adesao.html",
                {
                    "planos_mensalidades": planos_mensalidades,
                    "error_message": "O campo do valor não pode estar em branco.",
                },
            )

    # Redireciona para a página de cadastro de plano de adesão se o método HTTP não for POST
    return redirect("cadastro-plano-adesao")


# AÇÃO PARA EDITAR O OBJETO SERVIDOR
@login_required
def edit_server(request, servidor_id):
    servidor = get_object_or_404(Servidor, pk=servidor_id, usuario=request.user)
    original_nome = servidor.nome

    servidores = Servidor.objects.filter(usuario=request.user).order_by('nome')

    if request.method == "POST":
        nome = request.POST.get("nome")
        imagem = request.FILES.get("imagem")

        if nome:
            servidor.nome = nome
            try:
                servidor.save()
                changes = {}
                if original_nome != servidor.nome:
                    changes["nome"] = (original_nome, servidor.nome)

                # Atualizar ou criar imagem do servidor
                if imagem:
                    from .models import ServidorImagem
                    servidor_imagem, created = ServidorImagem.objects.get_or_create(
                        servidor=servidor,
                        usuario=request.user
                    )
                    servidor_imagem.imagem = imagem
                    servidor_imagem.save()
                    changes["imagem"] = "atualizada" if not created else "adicionada"

                mensagem = "Servidor atualizado." if changes else "Servidor salvo sem alteracoes."
                log_user_action(
                    request=request,
                    action=UserActionLog.ACTION_UPDATE,
                    instance=servidor,
                    message=mensagem,
                    extra=changes if changes else None,
                )


            except ValidationError as erro1:
                logger.error('[%s] [USER][%s] [IP][%s] [ERRO][%s]', timezone.localtime(), request.user, get_client_ip(request) or 'N/A', erro1, exc_info=True)
                # Capturando outras exceções e renderizando a página novamente com a mensagem de erro
                return render(
                    request,
                    "pages/cadastro-servidor.html",
                    {
                        'servidores': servidores,
                        "error_message": "Já existe um servidor com este nome!",
                    },
                )

            except Exception as erro2:
                logger.error('[%s] [USER][%s] [IP][%s] [ERRO][%s]', timezone.localtime(), request.user, get_client_ip(request) or 'N/A', erro2, exc_info=True)
                # Capturando outros possíveis erros ao tentar salvar o servidor
                return render(
                    request,
                    "pages/cadastro-servidor.html",
                    {
                        'servidores': servidores,
                        "error_message": "Ocorreu um erro ao salvar o servidor. Verifique o log!",
                    },
                )
            
            return render(
                request,
                "pages/cadastro-servidor.html",
                {"servidores": servidores, "success_update": True},
            )

    return redirect("cadastro-servidor")


# AÇÃO PARA EDITAR O OBJETO DISPOSITIVO
@login_required
def edit_device(request, dispositivo_id):
    dispositivo = get_object_or_404(Dispositivo, pk=dispositivo_id, usuario=request.user)
    original_nome = dispositivo.nome

    dispositivos = Dispositivo.objects.filter(usuario=request.user).order_by('nome')

    if request.method == "POST":
        nome = request.POST.get("nome")

        if nome:         
            dispositivo.nome = nome
            try:
                dispositivo.save()
                changes = {}
                if original_nome != dispositivo.nome:
                    changes["nome"] = (original_nome, dispositivo.nome)
                mensagem = "Dispositivo atualizado." if changes else "Dispositivo salvo sem alteracoes."
                log_user_action(
                    request=request,
                    action=UserActionLog.ACTION_UPDATE,
                    instance=dispositivo,
                    message=mensagem,
                    extra=changes if changes else None,
                )


            except ValidationError as erro1:
                logger.error('[%s] [USER][%s] [IP][%s] [ERRO][%s]', timezone.localtime(), request.user, get_client_ip(request) or 'N/A', erro1, exc_info=True)
                # Capturando outras exceções e renderizando a página novamente com a mensagem de erro
                return render(
                    request,
                    "pages/cadastro-dispositivo.html",
                    {
                        'dispositivos': dispositivos,
                        "error_message": "Já existe um dispositivo com este nome!",
                    },
                )

            except Exception as erro2:
                logger.error('[%s] [USER][%s] [IP][%s] [ERRO][%s]', timezone.localtime(), request.user, get_client_ip(request) or 'N/A', erro2, exc_info=True)
                # Capturando outros possíveis erros ao tentar salvar o dispositivo
                return render(
                    request,
                    "pages/cadastro-dispositivo.html",
                    {
                        'dispositivos': dispositivos,
                        "error_message": "Ocorreu um erro ao salvar o dispositivo. Verifique o log!",
                    },
                )
            
            return render(
                request,
                "pages/cadastro-dispositivo.html",
                {"dispositivos": dispositivos, "success_update": True},
            )

    return redirect("cadastro-dispositivo")


# AÇÃO PARA EDITAR O OBJETO APLICATIVO
@login_required
def editar_app(request, aplicativo_id):
    aplicativo = get_object_or_404(Aplicativo, pk=aplicativo_id, usuario=request.user)

    aplicativos = Aplicativo.objects.filter(usuario=request.user).order_by('nome')

    if request.method == "POST":
        nome = request.POST.get("nome")
        mac = request.POST.get("mac")

        if nome and mac:
            aplicativo.nome = nome
            aplicativo.device_has_mac = True if mac == "true" else False

            try:
                aplicativo.save()

            except ValidationError as erro1:
                logger.error('[%s] [USER][%s] [IP][%s] [ERRO][%s]', timezone.localtime(), request.user, get_client_ip(request) or 'N/A', erro1, exc_info=True)
                # Capturando outras exceções e renderizando a página novamente com a mensagem de erro
                return render(
                    request,
                    "pages/cadastro-aplicativo.html",
                    {
                        'aplicativos': aplicativos,
                        "error_message": "Já existe um aplicativo com este nome!",
                    },
                )

            except Exception as erro2:
                logger.error('[%s] [USER][%s] [IP][%s] [ERRO][%s]', timezone.localtime(), request.user, get_client_ip(request) or 'N/A', erro2, exc_info=True)
                # Capturando outros possíveis erros ao tentar salvar o aplicativo
                return render(
                    request,
                    "pages/cadastro-aplicativo.html",
                    {
                        'aplicativos': aplicativos,
                        "error_message": "Ocorreu um erro ao salvar o aplicativo. Verifique o log!",
                    },
                )
            
            return render(
                request,
                "pages/cadastro-aplicativo.html",
                {"aplicativos": aplicativos, "success_update": True},
            )

    return redirect("cadastro-aplicativo")


@login_required
def edit_profile(request):
    if request.method == "POST":
        if request.user.is_authenticated:
            try:
                user = request.user
                original_usuario = {
                    "first_name": user.first_name or "",
                    "last_name": user.last_name or "",
                    "email": user.email or "",
                }
                dados_usuario = user
                dados_usuario.last_name = request.POST.get('sobrenome', '').strip()
                dados_usuario.first_name = request.POST.get('nome', '').strip()
                dados_usuario.email = request.POST.get('email', '').strip()
                dados_usuario.save()

                dados_bancarios = DadosBancarios.objects.filter(usuario=user).first()
                dados_bancarios_criado = False
                original_bancarios = None
                if dados_bancarios:
                    original_bancarios = {
                        "beneficiario": dados_bancarios.beneficiario or "",
                        "instituicao": dados_bancarios.instituicao or "",
                        "tipo_chave": dados_bancarios.tipo_chave or "",
                        "chave": dados_bancarios.chave or "",
                        "wpp": dados_bancarios.wpp or "",
                    }
                else:
                    dados_bancarios = DadosBancarios(usuario=user)
                    dados_bancarios_criado = True

                beneficiario = request.POST.get('beneficiario', '').strip()
                instituicao = request.POST.get('instituicao', '').strip()
                tipo_chave = request.POST.get('tipo_chave', '').strip()
                chave = request.POST.get('chave', '').strip()
                wpp = request.POST.get('wpp', '').strip()

                dados_bancarios.beneficiario = beneficiario
                dados_bancarios.instituicao = instituicao
                dados_bancarios.tipo_chave = tipo_chave
                dados_bancarios.chave = chave
                dados_bancarios.wpp = wpp
                dados_bancarios.save()

                changes = {}
                if original_usuario["first_name"] != dados_usuario.first_name:
                    changes["usuario.first_name"] = (original_usuario["first_name"], dados_usuario.first_name)
                if original_usuario["last_name"] != dados_usuario.last_name:
                    changes["usuario.last_name"] = (original_usuario["last_name"], dados_usuario.last_name)
                if original_usuario["email"] != dados_usuario.email:
                    changes["usuario.email"] = (original_usuario["email"], dados_usuario.email)

                if dados_bancarios_criado:
                    changes["dados_bancarios"] = {
                        "created": True,
                        "beneficiario": beneficiario,
                        "instituicao": instituicao,
                        "tipo_chave": tipo_chave,
                        "chave": chave,
                        "wpp": wpp,
                    }
                else:
                    updated_bancarios = {
                        "beneficiario": beneficiario,
                        "instituicao": instituicao,
                        "tipo_chave": tipo_chave,
                        "chave": chave,
                        "wpp": wpp,
                    }
                    for campo, antigo_valor in original_bancarios.items():
                        novo_valor = updated_bancarios.get(campo, "")
                        if antigo_valor != novo_valor:
                            changes[f"dados_bancarios.{campo}"] = (antigo_valor, novo_valor)

                mensagem = "Perfil atualizado." if changes else "Perfil salvo sem alterações."
                log_user_action(
                    request=request,
                    action=UserActionLog.ACTION_UPDATE,
                    instance=request.user,
                    message=mensagem,
                    extra=changes if changes else None,
                )

                # Enviar email de notificação (se habilitado e houve alterações)
                if changes:
                    try:
                        ip_address = get_client_ip(request) or 'Desconhecido'
                        # Formatar mudanças para o email
                        formatted_changes = {}
                        for key, value in changes.items():
                            if isinstance(value, tuple) and len(value) == 2:
                                formatted_changes[key] = {'old': value[0], 'new': value[1]}
                            else:
                                formatted_changes[key] = {'old': '', 'new': str(value)}

                        send_profile_change_notification(
                            request.user,
                            change_type='profile',
                            changes_detail=formatted_changes,
                            ip_address=ip_address
                        )
                    except Exception as e:
                        logger.warning(f'[EDIT_PROFILE] Email notification failed for user {request.user.id}: {str(e)}')

                messages.success(request, 'Perfil editado com sucesso!')
            except Exception as e:
                messages.error(request, 'Ocorreu um erro ao editar o perfil. Verifique o log!')
                logger.error('[%s] [USER][%s] [IP][%s] [ERRO][%s]', timezone.localtime(), request.user, get_client_ip(request) or 'N/A', e, exc_info=True)
        else:
            messages.error(request, 'Usuário da requisição não identificado!')
    else:
        messages.error(request, 'Método da requisição não permitido!')

    return redirect('perfil')


@login_required
@require_http_methods(["POST"])
def upload_avatar(request):
    """Processa upload de avatar do usuário com validação e otimização."""
    from .models import UserProfile
    from django.core.files.base import ContentFile
    from django.http import JsonResponse

    try:
        if 'avatar' not in request.FILES:
            return JsonResponse({'error': 'Nenhum arquivo enviado'}, status=400)

        avatar_file = request.FILES['avatar']

        # Validar tamanho (5MB)
        if avatar_file.size > 5 * 1024 * 1024:
            return JsonResponse({'error': 'Arquivo muito grande. Máximo: 5MB'}, status=400)

        # Validar tipo
        allowed_types = ['image/jpeg', 'image/png', 'image/gif', 'image/webp']
        if avatar_file.content_type not in allowed_types:
            return JsonResponse({'error': 'Formato inválido. Use JPG, PNG, GIF ou WEBP'}, status=400)

        try:
            from PIL import Image

            # Processar imagem
            img = Image.open(avatar_file)

            # Converter RGBA para RGB se necessário
            if img.mode in ('RGBA', 'LA', 'P'):
                background = Image.new('RGB', img.size, (255, 255, 255))
                if img.mode == 'P':
                    img = img.convert('RGBA')
                background.paste(img, mask=img.split()[-1] if img.mode == 'RGBA' else None)
                img = background

            # Crop para quadrado (se necessário)
            width, height = img.size
            if width != height:
                size = min(width, height)
                left = (width - size) / 2
                top = (height - size) / 2
                right = (width + size) / 2
                bottom = (height + size) / 2
                img = img.crop((left, top, right, bottom))

            # Redimensionar para 500x500
            img = img.resize((500, 500), Image.Resampling.LANCZOS)

            # Salvar em buffer
            buffer = io.BytesIO()
            img.save(buffer, format='JPEG', quality=85, optimize=True)
            buffer.seek(0)

        except ImportError:
            return JsonResponse({'error': 'Biblioteca Pillow não instalada'}, status=500)
        except Exception as e:
            logger.error(f'[AVATAR_UPLOAD] Erro ao processar imagem: {str(e)}', exc_info=True)
            return JsonResponse({'error': 'Erro ao processar imagem'}, status=500)

        # Obter ou criar perfil
        profile, created = UserProfile.objects.get_or_create(user=request.user)

        # Deletar avatar antigo
        if profile.avatar:
            profile.delete_old_avatar()

        # Salvar novo avatar
        filename = f'avatar_{request.user.id}_{timezone.now().timestamp()}.jpg'
        profile.avatar.save(filename, ContentFile(buffer.read()), save=True)

        # Log da ação
        log_user_action(
            request=request,
            action=UserActionLog.ACTION_UPDATE,
            instance=request.user,
            message='Avatar atualizado',
            extra={'avatar_url': profile.avatar.url}
        )

        # Enviar email de notificação (se habilitado)
        try:
            ip_address = get_client_ip(request) or 'Desconhecido'
            send_profile_change_notification(
                request.user,
                change_type='avatar',
                changes_detail={'avatar': 'Imagem de perfil atualizada'},
                ip_address=ip_address
            )
        except Exception as e:
            logger.warning(f'[UPLOAD_AVATAR] Email notification failed for user {request.user.id}: {str(e)}')

        return JsonResponse({
            'success': True,
            'avatar_url': profile.avatar.url,
            'message': 'Avatar atualizado com sucesso!'
        })

    except Exception as e:
        logger.error(f'[AVATAR_UPLOAD] User: {request.user.id}, Error: {str(e)}', exc_info=True)
        return JsonResponse({'error': 'Erro ao processar avatar'}, status=500)


@login_required
@require_http_methods(["POST"])
def remove_avatar(request):
    """Remove avatar do usuário e restaura para padrão."""
    from .models import UserProfile
    from django.http import JsonResponse

    try:
        profile = UserProfile.objects.get(user=request.user)

        if profile.avatar:
            profile.delete_old_avatar()
            profile.avatar = None
            profile.save()

            log_user_action(
                request=request,
                action=UserActionLog.ACTION_UPDATE,
                instance=request.user,
                message='Avatar removido'
            )

            return JsonResponse({
                'success': True,
                'avatar_url': '/static/assets/images/avatar/default-avatar.svg',
                'message': 'Avatar removido com sucesso!'
            })

        return JsonResponse({'error': 'Nenhum avatar para remover'}, status=400)

    except UserProfile.DoesNotExist:
        return JsonResponse({'error': 'Perfil não encontrado'}, status=404)
    except Exception as e:
        logger.error(f'[AVATAR_REMOVE] User: {request.user.id}, Error: {str(e)}', exc_info=True)
        return JsonResponse({'error': 'Erro ao remover avatar'}, status=500)


@login_required
def profile_activity_history(request):
    """Retorna histórico de alterações do perfil do usuário."""
    from .models import UserActionLog
    from django.http import JsonResponse
    from django.db.models import Q

    # Buscar logs onde:
    # 1. O usuário fez a ação (usuario=request.user)
    # 2. E a ação foi sobre o próprio perfil (entidade='User' e objeto_id=user.id)
    logs = UserActionLog.objects.filter(
        Q(usuario=request.user) &
        Q(entidade='User') &
        Q(objeto_id=str(request.user.id))
    ).order_by('-criado_em')[:10]

    history = []
    for log in logs:
        history.append({
            'timestamp': log.criado_em.strftime('%d/%m/%Y %H:%M'),
            'action': log.get_acao_display(),
            'message': log.mensagem,
            'ip': log.ip or '--',
        })

    return JsonResponse({'history': history})


@login_required
@require_http_methods(["POST"])
def change_password(request):
    """Processa alteração de senha do usuário com validações."""
    from django.contrib.auth.password_validation import validate_password
    from django.core.exceptions import ValidationError as DjangoValidationError
    from django.contrib.auth import update_session_auth_hash
    from django.http import JsonResponse

    try:
        current_password = request.POST.get('current_password', '').strip()
        new_password = request.POST.get('new_password', '').strip()
        confirm_password = request.POST.get('confirm_password', '').strip()

        # Validar senha atual
        if not request.user.check_password(current_password):
            return JsonResponse({
                'error': 'Senha atual incorreta',
                'field': 'current_password'
            }, status=400)

        # Validar senhas novas coincidem
        if new_password != confirm_password:
            return JsonResponse({
                'error': 'As senhas não coincidem',
                'field': 'confirm_password'
            }, status=400)

        # Validar que nova senha é diferente da atual
        if current_password == new_password:
            return JsonResponse({
                'error': 'A nova senha deve ser diferente da atual',
                'field': 'new_password'
            }, status=400)

        # Validar força da senha (Django validators)
        try:
            validate_password(new_password, request.user)
        except DjangoValidationError as e:
            return JsonResponse({
                'error': ' '.join(e.messages),
                'field': 'new_password'
            }, status=400)

        # Atualizar senha
        request.user.set_password(new_password)
        request.user.save()

        # Log da ação
        log_user_action(
            request=request,
            action=UserActionLog.ACTION_UPDATE,
            instance=request.user,
            message='Senha alterada com sucesso',
            extra={'ip': get_client_ip(request) or 'N/A'}
        )

        # Re-autenticar usuário para manter sessão
        update_session_auth_hash(request, request.user)

        # Enviar email de notificação (se habilitado)
        try:
            ip_address = get_client_ip(request) or 'Desconhecido'
            send_password_change_notification(request.user, ip_address=ip_address)
        except Exception as e:
            logger.warning(f'[CHANGE_PASSWORD] Email notification failed for user {request.user.id}: {str(e)}')

        return JsonResponse({
            'success': True,
            'message': 'Senha alterada com sucesso!'
        })

    except Exception as e:
        logger.error(f'[CHANGE_PASSWORD] User: {request.user.id}, Error: {str(e)}', exc_info=True)
        return JsonResponse({
            'error': 'Erro ao alterar senha. Tente novamente.'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def change_theme(request):
    """Altera tema do usuário (light/dark/auto)."""
    from .models import UserProfile
    from django.http import JsonResponse

    try:
        theme = request.POST.get('theme', '').strip()

        # Validar tema
        valid_themes = ['light', 'dark', 'auto']
        if theme not in valid_themes:
            return JsonResponse({'error': 'Tema inválido'}, status=400)

        profile = UserProfile.objects.get(user=request.user)
        profile.theme_preference = theme
        profile.save()

        log_user_action(
            request=request,
            action=UserActionLog.ACTION_UPDATE,
            instance=request.user,
            message=f'Tema alterado para {theme}',
            extra={'theme': theme}
        )

        return JsonResponse({
            'success': True,
            'theme': theme,
            'message': 'Tema alterado com sucesso!'
        })

    except UserProfile.DoesNotExist:
        return JsonResponse({'error': 'Perfil não encontrado'}, status=404)
    except Exception as e:
        logger.error(f'[CHANGE_THEME] User: {request.user.id}, Error: {str(e)}', exc_info=True)
        return JsonResponse({'error': 'Erro ao alterar tema'}, status=500)


@login_required
@require_http_methods(["POST"])
def update_notification_preferences(request):
    """Atualiza preferências de notificação por email."""
    from .models import UserProfile
    from django.http import JsonResponse

    try:
        profile = UserProfile.objects.get(user=request.user)

        # Atualizar preferências
        profile.email_on_profile_change = request.POST.get('email_on_profile_change') == 'true'
        profile.email_on_password_change = request.POST.get('email_on_password_change') == 'true'
        profile.email_on_login = request.POST.get('email_on_login') == 'true'
        profile.save()

        log_user_action(
            request=request,
            action=UserActionLog.ACTION_UPDATE,
            instance=request.user,
            message='Preferências de notificação atualizadas'
        )

        return JsonResponse({
            'success': True,
            'message': 'Preferências atualizadas com sucesso!'
        })

    except UserProfile.DoesNotExist:
        return JsonResponse({'error': 'Perfil não encontrado'}, status=404)
    except Exception as e:
        logger.error(f'[UPDATE_NOTIF_PREFS] User: {request.user.id}, Error: {str(e)}', exc_info=True)
        return JsonResponse({'error': 'Erro ao atualizar preferências'}, status=500)


@login_required
@require_http_methods(["POST"])
def update_privacy_settings(request):
    """Atualiza configurações de privacidade."""
    from .models import UserProfile
    from django.http import JsonResponse

    try:
        profile = UserProfile.objects.get(user=request.user)

        # Atualizar configurações
        profile.profile_public = request.POST.get('profile_public') == 'true'
        profile.show_email = request.POST.get('show_email') == 'true'
        profile.show_phone = request.POST.get('show_phone') == 'true'
        profile.show_statistics = request.POST.get('show_statistics') == 'true'
        profile.save()

        log_user_action(
            request=request,
            action=UserActionLog.ACTION_UPDATE,
            instance=request.user,
            message='Configurações de privacidade atualizadas'
        )

        return JsonResponse({
            'success': True,
            'message': 'Configurações atualizadas com sucesso!'
        })

    except UserProfile.DoesNotExist:
        return JsonResponse({'error': 'Perfil não encontrado'}, status=404)
    except Exception as e:
        logger.error(f'[UPDATE_PRIVACY] User: {request.user.id}, Error: {str(e)}', exc_info=True)
        return JsonResponse({'error': 'Erro ao atualizar configurações'}, status=500)


@login_required
@require_http_methods(["POST"])
def setup_2fa(request):
    """Inicia configuração de 2FA gerando secret e retornando URL do QR code."""
    from .models import UserProfile
    from django.http import JsonResponse

    try:
        profile = UserProfile.objects.get(user=request.user)

        # Gerar nova chave secreta
        secret = profile.generate_2fa_secret()
        profile.save()

        # Obter URL para QR code
        qr_uri = profile.get_2fa_qr_code()

        log_user_action(
            request=request,
            action=UserActionLog.ACTION_UPDATE,
            instance=request.user,
            message='Iniciou configuração de 2FA'
        )

        return JsonResponse({
            'success': True,
            'secret': secret,
            'qr_uri': qr_uri,
            'message': 'QR Code gerado com sucesso!'
        })

    except UserProfile.DoesNotExist:
        return JsonResponse({'error': 'Perfil não encontrado'}, status=404)
    except Exception as e:
        logger.error(f'[SETUP_2FA] User: {request.user.id}, Error: {str(e)}', exc_info=True)
        return JsonResponse({'error': 'Erro ao configurar 2FA'}, status=500)


@login_required
@require_http_methods(["POST"])
def enable_2fa(request):
    """Ativa 2FA após validar código de verificação."""
    from .models import UserProfile
    from django.http import JsonResponse

    try:
        code = request.POST.get('code', '').strip()

        if not code:
            return JsonResponse({'error': 'Código não fornecido'}, status=400)

        profile = UserProfile.objects.get(user=request.user)

        if not profile.two_factor_secret:
            return JsonResponse({'error': '2FA não foi configurado'}, status=400)

        # Verificar código
        import pyotp
        totp = pyotp.TOTP(profile.two_factor_secret)
        if not totp.verify(code, valid_window=1):
            return JsonResponse({'error': 'Código inválido'}, status=400)

        # Ativar 2FA
        profile.two_factor_enabled = True

        # Gerar códigos de backup
        backup_codes = profile.generate_backup_codes()
        profile.save()

        log_user_action(
            request=request,
            action=UserActionLog.ACTION_UPDATE,
            instance=request.user,
            message='2FA ativado com sucesso'
        )

        # Enviar email de notificação
        try:
            send_2fa_enabled_notification(request.user)
        except Exception as e:
            logger.warning(f'[ENABLE_2FA] Email notification failed for user {request.user.id}: {str(e)}')

        return JsonResponse({
            'success': True,
            'backup_codes': backup_codes,
            'message': '2FA ativado com sucesso!'
        })

    except UserProfile.DoesNotExist:
        return JsonResponse({'error': 'Perfil não encontrado'}, status=404)
    except Exception as e:
        logger.error(f'[ENABLE_2FA] User: {request.user.id}, Error: {str(e)}', exc_info=True)
        return JsonResponse({'error': 'Erro ao ativar 2FA'}, status=500)


@login_required
@require_http_methods(["POST"])
def disable_2fa(request):
    """Desativa 2FA após validar senha do usuário."""
    from .models import UserProfile
    from django.http import JsonResponse
    from django.contrib.auth import authenticate

    try:
        password = request.POST.get('password', '').strip()

        if not password:
            return JsonResponse({'error': 'Senha não fornecida'}, status=400)

        # Verificar senha
        user = authenticate(username=request.user.username, password=password)
        if not user:
            return JsonResponse({'error': 'Senha incorreta'}, status=400)

        profile = UserProfile.objects.get(user=request.user)

        # Desativar 2FA
        profile.two_factor_enabled = False
        profile.two_factor_secret = None
        profile.two_factor_backup_codes = None
        profile.save()

        log_user_action(
            request=request,
            action=UserActionLog.ACTION_UPDATE,
            instance=request.user,
            message='2FA desativado'
        )

        return JsonResponse({
            'success': True,
            'message': '2FA desativado com sucesso!'
        })

    except UserProfile.DoesNotExist:
        return JsonResponse({'error': 'Perfil não encontrado'}, status=404)
    except Exception as e:
        logger.error(f'[DISABLE_2FA] User: {request.user.id}, Error: {str(e)}', exc_info=True)
        return JsonResponse({'error': 'Erro ao desativar 2FA'}, status=500)


@login_required
@require_http_methods(["POST"])
def regenerate_backup_codes(request):
    """Gera novos códigos de backup para 2FA."""
    from .models import UserProfile
    from django.http import JsonResponse

    try:
        password = request.POST.get('password', '').strip()

        if not password:
            return JsonResponse({'error': 'Senha não fornecida'}, status=400)

        # Verificar senha
        from django.contrib.auth import authenticate
        user = authenticate(username=request.user.username, password=password)
        if not user:
            return JsonResponse({'error': 'Senha incorreta'}, status=400)

        profile = UserProfile.objects.get(user=request.user)

        if not profile.two_factor_enabled:
            return JsonResponse({'error': '2FA não está ativado'}, status=400)

        # Gerar novos códigos
        backup_codes = profile.generate_backup_codes()
        profile.save()

        log_user_action(
            request=request,
            action=UserActionLog.ACTION_UPDATE,
            instance=request.user,
            message='Códigos de backup regenerados'
        )

        return JsonResponse({
            'success': True,
            'backup_codes': backup_codes,
            'message': 'Novos códigos de backup gerados!'
        })

    except UserProfile.DoesNotExist:
        return JsonResponse({'error': 'Perfil não encontrado'}, status=404)
    except Exception as e:
        logger.error(f'[REGEN_BACKUP] User: {request.user.id}, Error: {str(e)}', exc_info=True)
        return JsonResponse({'error': 'Erro ao gerar códigos'}, status=500)


@login_required
@require_http_methods(["GET"])
def get_2fa_qr_code(request):
    """Retorna imagem do QR code para configuração de 2FA."""
    from .models import UserProfile
    from django.http import HttpResponse
    import qrcode
    from io import BytesIO

    try:
        profile = UserProfile.objects.get(user=request.user)

        if not profile.two_factor_secret:
            return HttpResponse('2FA não configurado', status=400)

        # Obter URL do QR code
        qr_uri = profile.get_2fa_qr_code()

        # Gerar QR code
        qr = qrcode.QRCode(version=1, box_size=10, border=5)
        qr.add_data(qr_uri)
        qr.make(fit=True)

        img = qr.make_image(fill_color="black", back_color="white")

        # Salvar em buffer
        buffer = BytesIO()
        img.save(buffer, format='PNG')
        buffer.seek(0)

        return HttpResponse(buffer, content_type='image/png')

    except UserProfile.DoesNotExist:
        return HttpResponse('Perfil não encontrado', status=404)
    except Exception as e:
        logger.error(f'[GET_QR_CODE] User: {request.user.id}, Error: {str(e)}', exc_info=True)
        return HttpResponse('Erro ao gerar QR code', status=500)


@login_required
@require_http_methods(["GET", "POST"])
def edit_horario_envios(request):
    HORARIOS_OBRIGATORIOS = [
        {
            "nome": "mensalidades_a_vencer",
            "tipo_envio": "mensalidades_a_vencer",
            "horario": None,
            "status": False,
            "ativo": True,
        },
        {
            "nome": "obter_mensalidades_vencidas",
            "tipo_envio": "obter_mensalidades_vencidas",
            "horario": None,
            "status": False,
            "ativo": True,
        },
    ]
    usuario = request.user
    sessao = SessaoWpp.objects.filter(usuario=usuario, is_active=True).first()
    sessao_wpp = bool(sessao)

    if request.method == "GET":
        # Criação automática dos horários obrigatórios, se não existirem
        with transaction.atomic():
            for horario_data in HORARIOS_OBRIGATORIOS:
                if not HorarioEnvios.objects.filter(usuario=usuario, tipo_envio=horario_data["tipo_envio"]).exists():
                    novo_horario = HorarioEnvios.objects.create(
                        usuario=usuario,
                        nome=horario_data["nome"],
                        tipo_envio=horario_data["tipo_envio"],
                        horario=horario_data["horario"],
                        status=horario_data["status"],
                        ativo=horario_data["ativo"],
                    )
                    log_user_action(
                        request=request,
                        action=UserActionLog.ACTION_CREATE,
                        instance=novo_horario,
                        message="Horário de envio criado automaticamente.",
                        extra={
                            "tipo_envio": novo_horario.tipo_envio,
                            "status": novo_horario.status,
                            "ativo": novo_horario.ativo,
                        },
                    )
        horarios = HorarioEnvios.objects.filter(usuario=usuario)
        horarios_json = []
        for h in horarios:
            horarios_json.append({
                "id": h.id,
                "nome": h.nome,
                "nome_display": dict(HorarioEnvios.TITULO).get(h.nome, h.nome),
                "tipo_envio": h.tipo_envio,
                "descricao": dict(HorarioEnvios.DESCRICOES).get(h.tipo_envio, ""),
                "exemplo": dict(HorarioEnvios.EXEMPLOS).get(h.tipo_envio, ""),
                "horario": h.horario.strftime("%H:%M") if h.horario else "",
                "status": h.status,
                "ativo": h.ativo,
            })
        return JsonResponse({"horarios": horarios_json, "sessao_wpp": sessao_wpp}, status=200)

    elif request.method == "POST":
        ip_address = get_client_ip(request)
        try:
            data = json.loads(request.body)
        except Exception as e:
            logger.exception(
                "JSON inválido ao atualizar horário de envio (user=%s, ip=%s)",
                request.user,
                ip_address,
            )
            return JsonResponse({"error": f"JSON inválido. Detalhe: {e}"}, status=400)

        horario_id = data.get("id")
        if not horario_id:
            logger.warning(
                "Requisição para editar horário sem ID (user=%s, ip=%s)",
                request.user,
                ip_address,
            )
            return JsonResponse({"error": "ID do horário não informado."}, status=400)

        try:
            horario_envio = HorarioEnvios.objects.get(id=horario_id, usuario=usuario)
        except HorarioEnvios.DoesNotExist:
            logger.warning(
                "Horário de envio %s não encontrado para usuário %s (ip=%s)",
                horario_id,
                request.user,
                ip_address,
            )
            return JsonResponse({"error": "Horário não encontrado."}, status=404)

        original_horario = {
            "horario": horario_envio.horario.strftime("%H:%M") if horario_envio.horario else "",
            "status": horario_envio.status,
        }

        # Atualiza campos permitidos
        if "horario" in data:
            horario_str = data["horario"]
            from datetime import time
            try:
                if horario_str:
                    h, m = [int(x) for x in horario_str.split(":")]
                    horario_envio.horario = time(hour=h, minute=m)
                else:
                    horario_envio.horario = None
            except Exception:
                logger.warning(
                    "Formato de horário inválido recebido: %s (user=%s, horario_id=%s, ip=%s)",
                    horario_str,
                    request.user,
                    horario_id,
                    ip_address,
                )
                return JsonResponse({"error": "Horário inválido (use HH:MM)."}, status=400)
        if "status" in data:
            status = data["status"]
            if isinstance(status, str):
                horario_envio.status = status.lower() in ("1", "true", "on")
            else:
                horario_envio.status = bool(status)

        try:
            horario_envio.save()
        except ValidationError as e:
            logger.warning(
                "Erro de validação ao salvar horário %s do usuário %s (ip=%s): %s",
                horario_id,
                request.user,
                ip_address,
                e,
            )
            return JsonResponse({"error": str(e)}, status=400)
        except Exception as e:
            logger.exception(
                "Erro inesperado ao salvar horário %s do usuário %s (ip=%s)",
                horario_id,
                request.user,
                ip_address,
            )
            return JsonResponse({"error": "Erro interno ao atualizar horário."}, status=500)

        changes = {}
        novo_horario = horario_envio.horario.strftime("%H:%M") if horario_envio.horario else ""
        if original_horario["horario"] != novo_horario:
            changes["horario"] = (original_horario["horario"], novo_horario)
        if original_horario["status"] != horario_envio.status:
            changes["status"] = (original_horario["status"], horario_envio.status)

        log_user_action(
            request=request,
            action=UserActionLog.ACTION_UPDATE,
            instance=horario_envio,
            message="Horário de envio atualizado.",
            extra=changes if changes else None,
        )

        return JsonResponse({
            "success": True,
            "message": "Horário atualizado com sucesso.",
            "horario": {
                "id": horario_envio.id,
                "nome": horario_envio.nome,
                "nome_display": dict(HorarioEnvios.TITULO).get(horario_envio.nome, horario_envio.nome),
                "tipo_envio": horario_envio.tipo_envio,
                "descricao": dict(HorarioEnvios.DESCRICOES).get(horario_envio.tipo_envio, ""),
                "exemplo": dict(HorarioEnvios.EXEMPLOS).get(horario_envio.tipo_envio, ""),
                "horario": horario_envio.horario.strftime("%H:%M") if horario_envio.horario else "",
                "status": horario_envio.status,
                "sessao_wpp": sessao_wpp,
            }
        }, status=200)


@login_required
@require_http_methods(["GET", "POST"])
def edit_reject_call_config(request):
    """Edita configurações de rejeição automática de chamadas."""
    sessao = SessaoWpp.objects.filter(user=request.user, is_active=True).first()

    if request.method == "GET":
        if not sessao:
            return JsonResponse({
                "success": True,
                "sessao_wpp": False,
                "config": {
                    "reject_call_enabled": True,
                    "reject_call_horario_inicio": None,
                    "reject_call_horario_fim": None
                }
            })
        return JsonResponse({
            "success": True,
            "sessao_wpp": True,
            "config": {
                "reject_call_enabled": sessao.reject_call_enabled,
                "reject_call_horario_inicio": sessao.reject_call_horario_inicio.strftime("%H:%M") if sessao.reject_call_horario_inicio else None,
                "reject_call_horario_fim": sessao.reject_call_horario_fim.strftime("%H:%M") if sessao.reject_call_horario_fim else None
            }
        })

    # POST - Atualizar configurações
    if not sessao:
        return JsonResponse({
            "success": False,
            "message": "Você não possui uma sessão WhatsApp ativa."
        }, status=400)

    data = json.loads(request.body)

    sessao.reject_call_enabled = data.get("reject_call_enabled", True)

    horario_inicio = data.get("reject_call_horario_inicio")
    horario_fim = data.get("reject_call_horario_fim")

    if horario_inicio and horario_fim:
        from datetime import datetime
        sessao.reject_call_horario_inicio = datetime.strptime(horario_inicio, "%H:%M").time()
        sessao.reject_call_horario_fim = datetime.strptime(horario_fim, "%H:%M").time()
    else:
        sessao.reject_call_horario_inicio = None
        sessao.reject_call_horario_fim = None

    sessao.save(update_fields=[
        'reject_call_enabled',
        'reject_call_horario_inicio',
        'reject_call_horario_fim'
    ])

    return JsonResponse({
        "success": True,
        "message": "Configurações de chamadas atualizadas."
    })


@login_required
@require_http_methods(["GET", "POST"])
def edit_referral_plan(request):
    PLANOS_OBRIGATORIOS = [
        {"tipo_plano": "desconto", "valor": 0.00, "valor_minimo_mensalidade": 10.00, "limite_indicacoes": 0},
        {"tipo_plano": "dinheiro", "valor": 0.00, "valor_minimo_mensalidade": 10.00, "limite_indicacoes": 0},
        {"tipo_plano": "anuidade", "valor": 0.00, "valor_minimo_mensalidade": 10.00, "limite_indicacoes": 0},
        {"tipo_plano": "desconto_progressivo", "valor": 0.00, "valor_minimo_mensalidade": 10.00, "limite_indicacoes": 0},
    ]
    usuario = request.user
    sessao = SessaoWpp.objects.filter(usuario=usuario, is_active=True).first()
    sessao_wpp = bool(sessao)

    if request.method == "GET":
        # Criação automática dos planos se não existirem
        with transaction.atomic():
            for plano_data in PLANOS_OBRIGATORIOS:
                # Define se o plano é ativo por padrão (anuidade é inativo)
                ativo = True if plano_data["tipo_plano"] != "anuidade" else False

                if not PlanoIndicacao.objects.filter(usuario=usuario, tipo_plano=plano_data["tipo_plano"]).exists():
                    novo_plano = PlanoIndicacao.objects.create(
                        usuario=usuario,
                        nome=plano_data["tipo_plano"],
                        tipo_plano=plano_data["tipo_plano"],
                        valor=Decimal(str(plano_data["valor"])),
                        valor_minimo_mensalidade=Decimal(str(plano_data["valor_minimo_mensalidade"])),
                        limite_indicacoes=plano_data.get("limite_indicacoes", 0),
                        status=False,
                        ativo=ativo,
                    )
                    log_user_action(
                        request=request,
                        action=UserActionLog.ACTION_CREATE,
                        instance=novo_plano,
                        message="Plano de indicação criado automaticamente.",
                        extra={
                            "tipo_plano": novo_plano.tipo_plano,
                            "valor": str(novo_plano.valor),
                            "valor_minimo_mensalidade": str(novo_plano.valor_minimo_mensalidade),
                            "status": novo_plano.status,
                            "ativo": novo_plano.ativo,
                        },
                    )

        planos = PlanoIndicacao.objects.filter(usuario=request.user, ativo=True)
        planos_json = []
        for plano in planos:
            planos_json.append({
                "id": plano.id,
                "nome": plano.nome,
                "nome_display": plano.get_nome_display(),
                "tipo_plano": plano.tipo_plano,
                "descricao": plano.descricao,
                "exemplo": plano.exemplo,
                "valor": float(plano.valor),
                "valor_minimo_mensalidade": float(plano.valor_minimo_mensalidade),
                "limite_indicacoes": plano.limite_indicacoes,
                "status": plano.status,
            })

        return JsonResponse({"planos": planos_json, "sessao_wpp": sessao_wpp}, status=200)

    elif request.method == "POST":
        ip_address = get_client_ip(request)
        try:
            data = json.loads(request.body)
        except Exception as e:
            logger.exception(
                "JSON inválido ao atualizar plano de indicação (user=%s, ip=%s)",
                request.user,
                ip_address,
            )
            return JsonResponse({"error": f"JSON inválido. Detalhe: {e}"}, status=400)

        plano_id = data.get("id")
        if not plano_id:
            logger.warning(
                "Requisição para editar plano sem ID (user=%s, ip=%s)",
                request.user,
                ip_address,
            )
            return JsonResponse({"error": "ID do plano não informado."}, status=400)

        try:
            plano = PlanoIndicacao.objects.get(id=plano_id, usuario=usuario)
        except PlanoIndicacao.DoesNotExist:
            logger.warning(
                "Plano de indicação %s não encontrado para usuário %s (ip=%s)",
                plano_id,
                request.user,
                ip_address,
            )
            return JsonResponse({"error": "Plano não encontrado."}, status=404)

        original_plano = {
            "valor": str(plano.valor),
            "status": plano.status,
            "limite_indicacoes": plano.limite_indicacoes,
        }

        # Atualiza os campos permitidos
        if "valor" in data:
            try:
                plano.valor = Decimal(str(data["valor"]))
            except (InvalidOperation, ValueError):
                logger.warning(
                    "Valor inválido recebido para plano %s do usuário %s (ip=%s): %s",
                    plano_id,
                    request.user,
                    ip_address,
                    data.get("valor"),
                )
                return JsonResponse({"error": "Valor inválido."}, status=400)
        if "status" in data:
            status = data["status"]
            if isinstance(status, str):
                plano.status = status.lower() in ("1", "true", "on")
            else:
                plano.status = bool(status)
        if "limite_indicacoes" in data:
            try:
                plano.limite_indicacoes = int(data["limite_indicacoes"])
                if plano.limite_indicacoes < 0:
                    plano.limite_indicacoes = 0
            except (ValueError, TypeError):
                logger.warning(
                    "Limite de indicações inválido recebido para plano %s do usuário %s (ip=%s): %s",
                    plano_id,
                    request.user,
                    ip_address,
                    data.get("limite_indicacoes"),
                )
                return JsonResponse({"error": "Limite de indicações inválido."}, status=400)

        # Validação de modalidades conflitantes
        if plano.status:  # Se está sendo ativado
            modalidades_conflitantes = []

            # Se está ativando "Desconto Progressivo", verificar se "Desconto" ou "Bônus" estão ativos
            if plano.tipo_plano == "desconto_progressivo":
                conflitos = PlanoIndicacao.objects.filter(
                    usuario=usuario,
                    tipo_plano__in=["desconto", "dinheiro"],
                    ativo=True,
                    status=True
                ).exclude(id=plano.id)

                if conflitos.exists():
                    for conflito in conflitos:
                        nome_display = conflito.get_nome_display()
                        modalidades_conflitantes.append(nome_display)

            # Se está ativando "Desconto" ou "Bônus", verificar se "Desconto Progressivo" está ativo
            elif plano.tipo_plano in ["desconto", "dinheiro"]:
                conflito = PlanoIndicacao.objects.filter(
                    usuario=usuario,
                    tipo_plano="desconto_progressivo",
                    ativo=True,
                    status=True
                ).exclude(id=plano.id).first()

                if conflito:
                    modalidades_conflitantes.append(conflito.get_nome_display())

            # Se há conflitos, retornar erro
            if modalidades_conflitantes:
                modalidades_str = ", ".join(modalidades_conflitantes)
                mensagem_erro = (
                    f"Não é possível ativar '{plano.get_nome_display()}' enquanto "
                    f"'{modalidades_str}' estiver(em) ativo(s). "
                    f"Desative a(s) modalidade(s) conflitante(s) primeiro."
                )
                logger.warning(
                    "Tentativa de ativar planos conflitantes (user=%s, ip=%s): %s",
                    request.user,
                    ip_address,
                    mensagem_erro,
                )
                return JsonResponse({"error": mensagem_erro}, status=400)

        try:
            plano.save()
        except ValidationError as e:
            logger.warning(
                "Erro de validação ao salvar plano %s do usuário %s (ip=%s): %s",
                plano_id,
                request.user,
                ip_address,
                e,
            )
            return JsonResponse({"error": str(e)}, status=400)
        except Exception as e:
            logger.exception(
                "Erro inesperado ao salvar plano %s do usuário %s (ip=%s)",
                plano_id,
                request.user,
                ip_address,
            )
            return JsonResponse({"error": "Erro interno ao atualizar plano."}, status=500)

        changes = {}
        if original_plano["valor"] != str(plano.valor):
            changes["valor"] = (original_plano["valor"], str(plano.valor))
        if original_plano["status"] != plano.status:
            changes["status"] = (original_plano["status"], plano.status)
        if original_plano["limite_indicacoes"] != plano.limite_indicacoes:
            changes["limite_indicacoes"] = (original_plano["limite_indicacoes"], plano.limite_indicacoes)

        log_user_action(
            request=request,
            action=UserActionLog.ACTION_UPDATE,
            instance=plano,
            message="Plano de indicação atualizado.",
            extra=changes if changes else None,
        )

        return JsonResponse({
            "success": True,
            "message": "Plano atualizado com sucesso.",
            "plano": {
                "id": plano.id,
                "nome": plano.nome,
                "nome_display": plano.get_nome_display(),
                "tipo_plano": plano.tipo_plano,
                "descricao": plano.descricao,
                "exemplo": plano.exemplo,
                "valor": float(plano.valor),
                "valor_minimo_mensalidade": float(plano.valor_minimo_mensalidade),
                "limite_indicacoes": plano.limite_indicacoes,
                "status": plano.status,
                "ativo": plano.ativo,
                "sessao_wpp": sessao_wpp,
            }
        }, status=200)


@login_required
def test(request):
    clientes = Cliente.objects.filter(usuario=request.user)

    return render(
        request,
        'teste.html',
        {
            'clientes': clientes,
        },
    )


############################################ CREATE VIEW ############################################

@require_POST
@login_required
def create_app_account(request):
    app_id = request.POST.get('app_id')
    cliente_id = request.POST.get('cliente-id')
    dispositivo_id = request.POST.get('dispositivo_id')
    force_create = request.POST.get('force_create') == 'true'  # Prosseguir mesmo com aviso

    app = get_object_or_404(Aplicativo, id=app_id, usuario=request.user)
    cliente = get_object_or_404(Cliente, id=cliente_id, usuario=request.user)
    dispositivo = get_object_or_404(Dispositivo, id=dispositivo_id, usuario=request.user)

    device_id = request.POST.get('device-id') or None
    device_key = request.POST.get('device-key') or None
    app_email = request.POST.get('app-email') or None

    # ⭐ FASE 1: Validação de limite de dispositivos
    try:
        assinatura = cliente.assinatura

        # Verificar limite ANTES de criar (se não for forced)
        if not force_create:
            validacao = assinatura.validar_limite_dispositivos()

            # DEBUG: Log para verificar o estado atual
            logger.info(
                f"[VALIDACAO_LIMITE] Cliente: {cliente.nome} - "
                f"Plano: {assinatura.plano.nome} - "
                f"Telas: {assinatura.plano.telas} - "
                f"Max_dispositivos: {assinatura.plano.max_dispositivos} - "
                f"Dispositivos_usados: {assinatura.dispositivos_usados} - "
                f"No_limite: {validacao['no_limite']} - "
                f"Excedeu: {validacao['excedeu']}"
            )

            # Só exibe aviso se ao adicionar este dispositivo, vai IGUALAR OU EXCEDER o limite
            # Exemplo: Plano com 3 telas, cliente tem 2 dispositivos
            #   -> 2 < 3, não exibe aviso (ainda pode adicionar)
            # Exemplo: Plano com 3 telas, cliente tem 3 dispositivos
            #   -> 3 >= 3, exibe aviso (vai exceder ao adicionar mais um)
            # Exemplo: Plano com 3 telas, cliente tem 4 dispositivos
            #   -> 4 >= 3, exibe aviso (já excedeu)
            if validacao['no_limite'] or validacao['excedeu']:
                return JsonResponse({
                    'warning': True,
                    'dados_aviso': {
                        'plano': {
                            'nome': assinatura.plano.nome,
                            'max_dispositivos': assinatura.plano.max_dispositivos
                        },
                        'usado': {
                            'dispositivos': assinatura.dispositivos_usados
                        }
                    }
                }, status=200)

    except AttributeError:
        # Cliente não tem assinatura - permitir criação mas logar aviso
        logger.warning(
            f"[ASSINATURA] Cliente {cliente.nome} (ID: {cliente.id}) não possui AssinaturaCliente. "
            f"Criando dispositivo sem validação de limite."
        )

    # Verificar se é a primeira conta deste cliente (deve ser principal)
    total_contas_existentes = ContaDoAplicativo.objects.filter(cliente=cliente).count()
    is_primeira_conta = (total_contas_existentes == 0)

    nova_conta_app = ContaDoAplicativo(
        cliente=cliente,
        dispositivo=dispositivo,
        app=app,
        device_id=device_id,
        device_key=device_key,
        email=app_email,
        usuario=request.user,
        is_principal=is_primeira_conta,  # Primeira conta = principal
    )

    try:
        nova_conta_app.save()

        # ⭐ FASE 1: Incrementar contador após criação bem-sucedida
        try:
            assinatura = cliente.assinatura
            assinatura.dispositivos_usados += 1
            assinatura.save(update_fields=['dispositivos_usados'])

            # Logar se excedeu o limite
            validacao = assinatura.validar_limite_dispositivos()
            if validacao['excedeu']:
                logger.warning(
                    f"[LIMITE_EXCEDIDO] Cliente {cliente.nome} - "
                    f"Dispositivos: {assinatura.dispositivos_usados}/{assinatura.plano.max_dispositivos} "
                    f"(Excesso: +{validacao['excesso']})"
                )
        except AttributeError:
            pass  # Cliente sem assinatura, já logado acima

        log_user_action(
            request=request,
            action=UserActionLog.ACTION_CREATE,
            instance=nova_conta_app,
            message="Conta de aplicativo criada.",
            extra={
                "cliente": cliente.nome,
                "dispositivo": dispositivo.nome,
                "app": app.nome,
                "device_id": device_id or '',
                "email": app_email or '',
            },
        )
        return JsonResponse({'success_message_cancel': 'Conta do aplicativo cadastrada com sucesso!'}, status=200)
    except Exception as erro:
        logger.error('[%s] [USER][%s] [IP][%s] [ERRO][%s]', timezone.localtime(), request.user, get_client_ip(request) or 'N/A', erro, exc_info=True)
        return JsonResponse({'error_message': 'Ocorreu um erro ao tentar realizar o cadastro.'}, status=500)


@login_required
def import_customers(request):
    timestamp = localtime().strftime('%d-%m-%Y %H:%M:%S')
    func_name = inspect.currentframe().f_code.co_name
    usuario = request.user
    token = SessaoWpp.objects.filter(usuario=usuario, is_active=True).first()
    page_group = 'clientes'
    page = 'importar-clientes'
    success, fail, clientes_existentes, clientes_invalidos_whatsapp, erros_importacao, clientes_criados = 0, 0, [], [], [], {}
    error_message = None

    def clean_cell(row, key):
        valor = row.get(key, None)
        return "" if pd.isnull(valor) or valor is None else str(valor).strip()

    if request.method == "POST" and 'importar' in request.POST:
        logger.info("[IMPORT] [%s] Iniciando importação de clientes", usuario)
        arquivo = request.FILES.get('arquivo')
        if not arquivo:
            error_message = "Nenhum arquivo enviado."
        else:
            ext = os.path.splitext(arquivo.name)[1].lower()
            if ext != '.xlsx':
                error_message = "Apenas arquivos .xlsx são suportados."
            else:
                try:
                    dados = pd.read_excel(
                        arquivo,
                        engine='openpyxl',
                        header=2,
                        dtype={
                            "servidor": str,
                            "dispositivo": str,
                            "sistema": str,
                            "device_id": str,
                            "email": str,
                            "device_key": str,
                            "nome": str,
                            "telefone": str,
                            "indicado_por": str,
                            "data_vencimento": str,
                            "forma_pgto": str,
                            "tipo_plano": str,
                            "plano_valor": float,
                            "telas": int,
                            "data_adesao": str
                        }
                    )
                    if "data_adesao" in dados.columns:
                        dados["data_adesao"] = pd.to_datetime(dados["data_adesao"], errors="coerce")
                        dados = dados.sort_values("data_adesao", ascending=True)
                except Exception as e:
                    logger.error("Erro lendo planilha: %s", e, exc_info=True)
                    error_message = "Erro ao ler a planilha. Verifique o arquivo."

        if error_message:
            return render(request, "pages/importar-cliente.html", {
                "error_message": error_message,
                "page_group": page_group, "page": page
            })

        registros = dados.to_dict('records')
        with transaction.atomic():

            # Verifica se o usuário tem uma sessão ativa do WhatsApp
            if not token:
                error_message = (
                    "Você precisa conectar sua conta ao WhatsApp antes de importar clientes. "
                    "Vá até a tela de integração com o WhatsApp e faça a conexão para prosseguir."
                )
                return render(request, "pages/importar-cliente.html", {
                    "error_message": error_message,
                    "page_group": page_group,
                    "page": page,
                })
            
            # 1º loop: salva todos os clientes sem indicado_por
            for idx, row in enumerate(registros, 1):
                try:
                    logger.debug("[%s] [%s] [%s] [IMPORT] Processando linha %d - dados: %s", timestamp, func_name, usuario, idx, row)
                    # Normaliza nomes de servidor, dispositivo e sistema
                    servidor_nome = normalizar_servidor(clean_cell(row, 'servidor'))
                    dispositivo_nome = normalizar_dispositivo(clean_cell(row, 'dispositivo'))
                    sistema_nome = normalizar_aplicativo(clean_cell(row, 'sistema'))
                    device_id = clean_cell(row, 'device_id')
                    email = clean_cell(row, 'email')
                    senha = clean_cell(row, 'device_key')
                    nome = clean_cell(row, 'nome')
                    telefone = clean_cell(row, 'telefone')
                    data_vencimento_str = clean_cell(row, 'data_vencimento')
                    forma_pgto_nome = clean_cell(row, 'forma_pgto')
                    plano_nome = clean_cell(row, 'tipo_plano')
                    plano_valor = clean_cell(row, 'plano_valor')
                    plano_telas = clean_cell(row, 'telas')
                    data_adesao_str = clean_cell(row, 'data_adesao')

                    # Valida obrigatórios
                    if not all([servidor_nome, dispositivo_nome, sistema_nome, nome, telefone, data_vencimento_str, forma_pgto_nome, plano_nome, plano_valor, plano_telas, data_adesao_str]):
                        fail += 1
                        erros_importacao.append(f"Linha {idx}: campos obrigatórios em branco")
                        continue

                    # 1. Validação de telefone
                    resultado_telefone = validar_tel_whatsapp(telefone, token.token, request.user)
                    if not resultado_telefone.get("wpp"):
                        fail += 1
                        clientes_invalidos_whatsapp.append(f"Linha {idx}: {telefone} (número não existe no WhatsApp)")
                        continue
                    if not resultado_telefone.get("telefone_validado_wpp"):
                        fail += 1
                        clientes_invalidos_whatsapp.append(f"Linha {idx}: {telefone} (não foi possível validar o telefone para WhatsApp)")
                        continue
                    if resultado_telefone.get("cliente_existe_telefone"):
                        fail += 1
                        clientes_existentes.append(f"Linha {idx}: {telefone} (já existe cliente com esse telefone)")
                        continue
                    if resultado_telefone.get("telefone_validado_wpp"):
                        telefone = resultado_telefone.get("telefone_validado_wpp")

                    # Valida se o plano informado é um dos permitidos
                    planos_validos = [plano[0] for plano in Plano.CHOICES]
                    # Normaliza o nome do plano (ex: "mensal" -> "Mensal", "MENSAL" -> "Mensal")
                    plano_nome = plano_nome.strip().title()
                    if plano_nome not in planos_validos:
                        fail += 1
                        erros_importacao.append(f"Linha {idx}: plano '{plano_nome}' inválido. Deve ser um dos: {', '.join(planos_validos)}.")
                        continue

                    # Parse plano_valor
                    try:
                        plano_valor = float(str(plano_valor).replace(",", "."))
                    except Exception:
                        fail += 1
                        erros_importacao.append(f"Linha {idx}: valor do plano inválido.")
                        continue

                    # Parse de plano_telas
                    try:
                        plano_telas = int(plano_telas)
                    except Exception:
                        fail += 1
                        erros_importacao.append(f"Linha {idx}: quantidade de telas inválida.")
                        continue

                    # Parse datas
                    data_vencimento = None
                    if data_vencimento_str:
                        try:
                            if len(data_vencimento_str) > 10:
                                data_vencimento = datetime.strptime(data_vencimento_str, "%Y-%m-%d %H:%M:%S").date()
                            else:
                                data_vencimento = datetime.strptime(data_vencimento_str, "%Y-%m-%d").date()
                        except Exception:
                            fail += 1
                            erros_importacao.append(f"Linha {idx}: data de vencimento inválida.")
                            continue
                    try:
                        if len(data_adesao_str) > 10:
                            data_adesao = datetime.strptime(data_adesao_str, "%Y-%m-%d %H:%M:%S").date()
                        else:
                            data_adesao = datetime.strptime(data_adesao_str, "%Y-%m-%d").date()
                    except Exception:
                        fail += 1
                        erros_importacao.append(f"Linha {idx}: data de adesão inválida.")
                        continue

                    # Validar e-mail
                    try:
                        if email:
                            validate_email(email)
                    except ValidationError:
                        fail += 1
                        erros_importacao.append(f"Linha {idx}: E-mail inválido.")
                        continue

                    # Objetos relacionados - usar filter().first() para evitar MultipleObjectsReturned
                    plano = Plano.objects.filter(nome=plano_nome, telas=plano_telas, valor=plano_valor, usuario=usuario).first()
                    if not plano:
                        plano = Plano.objects.create(nome=plano_nome, telas=plano_telas, valor=plano_valor, usuario=usuario)

                    # Usa funções helper com busca case-insensitive e normalização
                    sistema, _ = get_or_create_aplicativo(sistema_nome, usuario)
                    servidor, _ = get_or_create_servidor(servidor_nome, usuario)
                    dispositivo, _ = get_or_create_dispositivo(dispositivo_nome, usuario)

                    forma_pgto = Tipos_pgto.objects.filter(nome__iexact=forma_pgto_nome, usuario=usuario).first()
                    if not forma_pgto:
                        forma_pgto = Tipos_pgto.objects.create(nome=forma_pgto_nome, usuario=usuario)

                    # Salva cliente sem indicado_por
                    cliente = Cliente(
                        nome=nome,
                        telefone=telefone,
                        dispositivo=dispositivo,
                        sistema=sistema,
                        servidor=servidor,
                        forma_pgto=forma_pgto,
                        plano=plano,
                        data_vencimento=data_vencimento,
                        data_adesao=data_adesao,
                        usuario=usuario,
                    )
                    cliente.save()
                    clientes_criados[telefone] = cliente

                    # Conta do App se aplicável
                    conta_criada = False
                    if device_id or email:
                        device_id = re.sub(r'[^A-Fa-f0-9]', '', device_id or '')
                        if device_id:
                            ContaDoAplicativo.objects.create(
                                device_id=device_id,
                                email=email,
                                device_key=senha,
                                app=sistema,
                                cliente=cliente,
                                usuario=usuario,
                            )
                            conta_criada = True

                    # Atualiza contador de dispositivos na AssinaturaCliente
                    if conta_criada:
                        try:
                            assinatura = cliente.assinatura
                            assinatura.dispositivos_usados = 1
                            assinatura.save(update_fields=['dispositivos_usados'])
                        except Exception as e:
                            logger.warning("[IMPORT] [%s] Erro ao atualizar contador de dispositivos para cliente %s: %s", usuario, nome, e)

                    # Cria mensalidade com marcação de dados importados
                    try:
                        Mensalidade.objects.create(
                            cliente=cliente,
                            valor=plano_valor,
                            dt_vencimento=data_vencimento,
                            usuario=usuario,
                            dados_historicos_verificados=False,
                            valor_base_plano=plano_valor,
                        )
                    except Exception as e:
                        logger.error("[IMPORT] [%s] Erro ao criar mensalidade para cliente %s: %s", usuario, nome, e, exc_info=True)

                    # Cria histórico de plano para rastreabilidade
                    try:
                        historico_iniciar(cliente, plano=plano, inicio=data_adesao, motivo='create')
                    except Exception as e:
                        logger.warning("[IMPORT] [%s] Erro ao criar histórico de plano para cliente %s: %s", usuario, nome, e)

                    success += 1
                    logger.info("[IMPORT] [%s] Cliente importado com sucesso: %s (%s)", usuario, nome, telefone)
                except Exception as e:
                    fail += 1
                    logger.error("[IMPORT] [%s] Falha ao importar linha %d: %s", usuario, idx, e, exc_info=True)
                    erros_importacao.append(f"Linha {idx}: {e}")

            # 2º loop: associa indicador
            for idx, row in enumerate(registros, 1):
                telefone_raw = clean_cell(row, 'telefone')
                indicador_raw = clean_cell(row, 'indicado_por')
                try:
                    if indicador_raw:
                        # Formata telefone do indicador para busca (com ou sem +)
                        indicador_formatado = indicador_raw if indicador_raw.startswith('+') else f'+{indicador_raw}'
                        # Busca primeiro nos clientes recém-criados, depois no banco
                        indicador = clientes_criados.get(indicador_formatado) or Cliente.objects.filter(telefone=indicador_formatado, usuario=usuario).first()
                        if indicador:
                            # Busca o cliente pelo telefone já formatado no dicionário
                            telefone_formatado = telefone_raw if telefone_raw.startswith('+') else f'+{telefone_raw}'
                            cliente = clientes_criados.get(telefone_formatado)
                            if cliente:
                                cliente.indicado_por = indicador
                                cliente.save(update_fields=['indicado_por'])
                            else:
                                Cliente.objects.filter(telefone=telefone_formatado, usuario=usuario).update(indicado_por=indicador)
                except Exception as e:
                    fail += 1
                    erros_importacao.append(f"Linha {idx}: Não foi possível associar o Indicador ({indicador_raw}) ao Cliente ({telefone_raw}).")
                    continue

        logger.info(
            "[IMPORT] [%s] Importação concluída - Sucesso: %d | Falhas: %d | Existentes: %d | WhatsApp inválido: %d",
            usuario, success, fail, len(clientes_existentes), len(clientes_invalidos_whatsapp)
        )
        log_user_action(
            request=request,
            action=UserActionLog.ACTION_IMPORT,
            entity="Cliente",
            message="Importação de clientes concluída.",
            extra={
                "sucesso": success,
                "falhas": fail,
                "clientes_existentes": len(clientes_existentes),
                "clientes_invalidos_whatsapp": len(clientes_invalidos_whatsapp),
                "erros_registrados": len(erros_importacao),
            },
        )
        return render(request, "pages/importar-cliente.html", {
            "success_message": "Importação concluída!",
            "num_linhas_importadas": success,
            "num_linhas_nao_importadas": fail,
            "nomes_clientes_existentes": clientes_existentes,
            "nomes_clientes_erro_importacao": erros_importacao,
            "clientes_invalidos_whatsapp": clientes_invalidos_whatsapp,
            "page_group": page_group,
            "page": page,
        })

    return render(request, "pages/importar-cliente.html", {"page_group": page_group, "page": page})


# AÇÃO PARA CRIAR NOVO CLIENTE ATRAVÉS DO FORMULÁRIO
@login_required
def create_customer(request):
    timestamp = localtime().strftime('%d-%m-%Y %H:%M:%S')
    func_name = inspect.currentframe().f_code.co_name
    usuario = request.user
    token = SessaoWpp.objects.filter(usuario=usuario, is_active=True).first()
    page_group = "clientes"
    page = "cadastro-cliente"

    # Querysets para preencher os selects do formulário
    plano_queryset = Plano.objects.filter(usuario=usuario).order_by('nome', 'telas', 'valor')
    forma_pgto_queryset = Tipos_pgto.objects.filter(usuario=usuario).select_related('conta_bancaria__instituicao', 'dados_bancarios').order_by('nome')
    servidor_queryset = Servidor.objects.filter(usuario=usuario).order_by('nome')
    sistema_queryset = Aplicativo.objects.filter(usuario=usuario).order_by('nome')
    indicador_por_queryset = Cliente.objects.filter(usuario=usuario, cancelado=False).order_by('nome')
    dispositivo_queryset = Dispositivo.objects.filter(usuario=usuario).order_by('nome')

    # Processa requisição POST (cadastro)
    if request.method == 'POST' and 'cadastrar' in request.POST:
        
        # Verifica se o usuário tem uma sessão ativa do WhatsApp
        if not token:
            return render(request, "pages/cadastro-cliente.html", {
                'error_message': (
                    "Você precisa conectar sua conta ao WhatsApp antes de cadastrar um cliente.<br>"
                    "Vá até a tela de integração com o WhatsApp e faça a conexão para prosseguir."
                ),
                'servidores': servidor_queryset,
                'dispositivos': dispositivo_queryset,
                'sistemas': sistema_queryset,
                'indicadores': indicador_por_queryset,
                'formas_pgtos': forma_pgto_queryset,
                'planos': plano_queryset,
                'page_group': page_group,
                'page': page,
            })

        # Coleta e sanitiza dados do formulário
        post = request.POST
        nome = post.get('nome', '').strip()
        telefone = post.get('telefone', '').strip()
        notas = post.get('notas', '').strip()
        indicador_nome = post.get('indicador_list', '').strip()
        servidor_nome = normalizar_servidor(post.get('servidor', '').strip())
        forma_pgto_input = post.get('forma_pgto', '').strip()  # Pode ser ID ou nome
        plano_info = post.get('plano', '').replace(' ', '').split('-')

        if not telefone:
            return render(request, "pages/cadastro-cliente.html", {
                "error_message": "O campo telefone não pode estar em branco.",
            })

        print(f"[{timestamp}] [INFO] [{func_name}] [{usuario}] Iniciando cadastro de Novo Cliente.")
        resultado_wpp = validar_tel_whatsapp(telefone, token.token, user=usuario)
        # Verifica se o número possui WhatsApp (antes de qualquer validação de duplicidade)
        if not resultado_wpp.get("wpp"):
            return render(request, "pages/cadastro-cliente.html", {
                "error_message": (
                    "O telefone informado não possui um WhatsApp.<br>"
                    "Cada novo cliente precisa estar cadastrado no WhatsApp."
                ),
            })

        # Se cliente existe com o telefone recebido, retorna os erros conforme cada possibilidade
        if resultado_wpp.get("cliente_existe_telefone"):
            telefone = resultado_wpp.get("cliente_existe_telefone")
            cliente_existente = None

            try:
                cliente_existente = Cliente.objects.filter(telefone=telefone, usuario=usuario).first()
                if cliente_existente:
                    logger.warning(
                        f"[ERRO][CADASTRO CLIENTE] Já existe cliente para telefone {cliente_existente.telefone} (ID: {cliente_existente.id})"
                    )
                    return render(request, "pages/cadastro-cliente.html", {
                        "error_message": (
                            "Há um cliente cadastrado com o telefone informado! <br><br>"
                            f"<strong>Nome:</strong> {cliente_existente.nome} <br>"
                            f"<strong>Telefone:</strong> {cliente_existente.telefone}"
                        ),
                    })
                else:
                    logger.error(
                        f"[ERRO][CADASTRO CLIENTE] resultado_wpp apontou cliente_existe=True, mas nenhum cliente foi localizado no banco para telefone {telefone}"
                    )
                    return render(request, "pages/cadastro-cliente.html", {
                        "error_message": (
                            "Erro inesperado ao validar telefone existente.<br>"
                            "Por favor, tente novamente em poucos instantes ou contate o suporte."
                        ),
                    })
            except Exception as e:
                logger.exception(
                    f"[ERRO][CADASTRO CLIENTE] Erro ao buscar cliente existente para telefone(s) {telefone}: {e}"
                )
                return render(request, "pages/cadastro-cliente.html", {
                    "error_message": (
                        "Ocorreu um erro inesperado ao tentar verificar o número de telefone.<br>"
                        "Por favor, tente novamente ou contate o suporte."
                    ),
                })
            
        # Obtém o telefone validado com WhatsApp
        if resultado_wpp.get("telefone_validado_wpp"):
            telefone = resultado_wpp.get("telefone_validado_wpp")

        # Trata indicador (pode ser nulo)
        indicador = None
        if indicador_nome:
            indicador = Cliente.objects.filter(nome=indicador_nome, usuario=usuario).first()

        # Trata plano
        try:
            plano_nome = plano_info[0]
            plano_valor = float(plano_info[1].replace(',', '.'))
            plano_telas = int(plano_info[2])
        except (IndexError, ValueError):
            return render(request, "pages/cadastro-cliente.html", {
                "error_message": "Plano inválido.",
            })
        plano = Plano.objects.filter(nome=plano_nome, valor=plano_valor, telas=plano_telas, usuario=usuario).first()
        if not plano:
            plano = Plano.objects.create(nome=plano_nome, valor=plano_valor, telas=plano_telas, usuario=usuario)

        # Trata relacionados
        servidor, _ = Servidor.objects.get_or_create(nome=servidor_nome, usuario=usuario)

        # Forma de pagamento: primeiro tenta por ID, depois por nome (compatibilidade)
        forma_pgto = None
        if forma_pgto_input.isdigit():
            forma_pgto = Tipos_pgto.objects.filter(pk=int(forma_pgto_input), usuario=usuario).first()
        if not forma_pgto:
            # Fallback: buscar por nome (compatibilidade com importações e casos antigos)
            forma_pgto, _ = Tipos_pgto.objects.get_or_create(nome=forma_pgto_input, usuario=usuario)

        # Coleta dados de múltiplos dispositivos/apps/contas baseado na quantidade de telas
        telas_data = []
        for i in range(plano_telas):
            dispositivo_nome = normalizar_dispositivo(post.get(f'dispositivo_{i}', '').strip())
            sistema_nome = normalizar_aplicativo(post.get(f'sistema_{i}', '').strip())
            device_id = post.get(f'device_id_{i}', '').strip()
            email = post.get(f'email_{i}', '').strip()
            senha = post.get(f'senha_{i}', '').strip()

            if not dispositivo_nome or not sistema_nome:
                return render(request, "pages/cadastro-cliente.html", {
                    "error_message": f"Dispositivo e Aplicativo são obrigatórios para a Tela {i + 1}.",
                    'servidores': servidor_queryset,
                    'dispositivos': dispositivo_queryset,
                    'sistemas': sistema_queryset,
                    'indicadores': indicador_por_queryset,
                    'formas_pgtos': forma_pgto_queryset,
                    'planos': plano_queryset,
                    'page_group': page_group,
                    'page': page,
                })

            # Get or create dispositivo e sistema (com busca case-insensitive e normalização)
            dispositivo, _ = get_or_create_dispositivo(dispositivo_nome, usuario)
            sistema, _ = get_or_create_aplicativo(sistema_nome, usuario)

            # Valida se conta é obrigatória
            if sistema.device_has_mac and not device_id and not email:
                return render(request, "pages/cadastro-cliente.html", {
                    "error_message": f"Conta do aplicativo é obrigatória para a Tela {i + 1}.",
                    'servidores': servidor_queryset,
                    'dispositivos': dispositivo_queryset,
                    'sistemas': sistema_queryset,
                    'indicadores': indicador_por_queryset,
                    'formas_pgtos': forma_pgto_queryset,
                    'planos': plano_queryset,
                    'page_group': page_group,
                    'page': page,
                })

            telas_data.append({
                'dispositivo': dispositivo,
                'sistema': sistema,
                'device_id': device_id,
                'email': email,
                'senha': senha,
            })

        try:
            with transaction.atomic():
                # SALVAR CLIENTE
                # Define dispositivo e sistema como o primeiro da lista (principal)
                primeiro_dispositivo = telas_data[0]['dispositivo'] if telas_data else None
                primeiro_sistema = telas_data[0]['sistema'] if telas_data else None

                try:
                    cliente = Cliente(
                        nome=nome,
                        telefone=telefone,
                        dispositivo=primeiro_dispositivo,
                        sistema=primeiro_sistema,
                        indicado_por=indicador,
                        servidor=servidor,
                        forma_pgto=forma_pgto,
                        plano=plano,
                        notas=notas,
                        usuario=usuario,
                    )
                    cliente.save()
                except Exception as e:
                    logger.error("Erro ao salvar o cliente: %s", e, exc_info=True)
                    raise Exception("Falha ao salvar os dados do cliente.")

                # CRIAR MÚLTIPLAS CONTAS DO APLICATIVO
                contas_criadas = 0
                try:
                    for idx, tela in enumerate(telas_data):
                        # Primeira conta (idx == 0) deve ser marcada como principal
                        ContaDoAplicativo.objects.create(
                            dispositivo=tela['dispositivo'],
                            app=tela['sistema'],
                            device_id=tela['device_id'] or None,
                            email=tela['email'] or None,
                            device_key=tela['senha'] or None,
                            cliente=cliente,
                            usuario=usuario,
                            is_principal=(idx == 0),  # Primeira conta = principal
                        )
                        contas_criadas += 1

                    # Incrementar contador de dispositivos usados
                    if contas_criadas > 0:
                        try:
                            assinatura = cliente.assinatura
                            assinatura.dispositivos_usados = contas_criadas
                            assinatura.save()
                        except Exception as e:
                            logger.warning(f"Erro ao atualizar contador de dispositivos: {e}")

                except Exception as e:
                    logger.error("Erro ao criar ContaDoAplicativo: %s", e, exc_info=True)
                    raise Exception("Falha ao criar as contas dos aplicativos.<p>Algum dos dados não pôde ser salvo.</p>")

                # ENVIO DE MENSAGEM - Removido: agora a mensagem é enviada apenas após
                # o pagamento da primeira mensalidade (via _enviar_notificacoes_pagamento)
                # try:
                #     envio_apos_novo_cadastro(cliente)
                # except Exception as e:
                #     logger.error("Erro ao enviar mensagem para o cliente: %s", e, exc_info=True)
                #     raise Exception("Erro ao realizar cadastro!<p>Talvez você ainda não tenha conectado a sessão do WhatsApp.</p>")

                # ⭐ FASE 2: Auto-enroll in campaign if eligible (DEVE ser antes da criação da mensalidade)
                try:
                    from nossopainel.utils import enroll_client_in_campaign_if_eligible
                    enroll_client_in_campaign_if_eligible(cliente)
                except Exception as e:
                    logger.error(f"Erro ao inscrever cliente na campanha: {e}", exc_info=True)

                # CRIAÇÃO DE MENSALIDADE (agora com em_campanha=True se elegível)
                try:
                    criar_mensalidade(cliente)
                except Exception as e:
                    logger.error("Erro ao criar a mensalidade: %s", e, exc_info=True)
                    raise Exception("Falha ao criar a mensalidade do cliente.")

                # Histórico de planos (inicial)
                try:
                    inicio_hist = cliente.data_adesao or timezone.localdate()
                    historico_iniciar(cliente, plano=plano, inicio=inicio_hist, motivo='create')
                except Exception:
                    pass

            print(f"[{timestamp}] [SUCCESS] [{func_name}] [{usuario}] Cliente {cliente.nome} ({cliente.telefone}) cadastrado com sucesso!")
            log_user_action(
                request=request,
                action=UserActionLog.ACTION_CREATE,
                instance=cliente,
                message="Cliente criado.",
                extra={
                    "telefone": cliente.telefone,
                    "plano": getattr(cliente.plano, 'nome', ''),
                    "servidor": getattr(cliente.servidor, 'nome', ''),
                    "forma_pgto": getattr(cliente.forma_pgto, 'nome', ''),
                    "data_vencimento": cliente.data_vencimento.strftime('%Y-%m-%d') if cliente.data_vencimento else '',
                },
            )
            return render(request, "pages/cadastro-cliente.html", {
                "success_message": "Novo cliente cadastrado com sucesso!",
            })

        except Exception as e:
            logger.error('[%s] [USER][%s] [IP][%s] [ERRO][%s]',
                        timezone.localtime(), usuario, get_client_ip(request) or 'IP não identificado', e, exc_info=True)
            return render(request, "pages/cadastro-cliente.html", {
                "error_message": str(e),
            })

    # Requisição GET (carrega formulário)

    # Serializar dispositivos e aplicativos para JSON (necessário para JavaScript)
    dispositivos_json = json.dumps([
        {'nome': d.nome}
        for d in dispositivo_queryset
    ])

    sistemas_json = json.dumps([
        {
            'nome': a.nome,
            'device_has_mac': a.device_has_mac
        }
        for a in sistema_queryset
    ])

    return render(request, "pages/cadastro-cliente.html", {
        'servidores': servidor_queryset,
        'dispositivos': dispositivo_queryset,
        'dispositivos_json': dispositivos_json,
        'sistemas': sistema_queryset,
        'sistemas_json': sistemas_json,
        'indicadores': indicador_por_queryset,
        'formas_pgtos': forma_pgto_queryset,
        'planos': plano_queryset,
        'page_group': page_group,
        'page': page,
    })


# =============================================================================
# CADASTRO BÁSICO DE CLIENTE (SEM ASSINATURA)
# =============================================================================

@login_required
def cadastrar_cliente_basico(request):
    """
    Cadastra apenas os dados básicos do cliente, SEM criar assinatura.

    Campos: nome, telefone, email, cpf, notas
    NÃO cria mensalidade nem envia mensagem de confirmação.
    O cliente fica com tem_assinatura=False até criar uma assinatura.
    """
    usuario = request.user
    token = SessaoWpp.objects.filter(usuario=usuario, is_active=True).first()
    page_group = "clientes"
    page = "cadastro-cliente"

    if request.method == 'POST':
        post = request.POST
        nome = post.get('nome', '').strip()
        telefone = post.get('telefone', '').strip()
        email = post.get('email', '').strip() or None
        cpf = post.get('cpf', '').strip() or None
        notas = post.get('notas', '').strip() or None

        # Validações básicas
        if not nome:
            return render(request, "pages/cadastro-cliente.html", {
                "error_message": "O campo nome é obrigatório.",
                'page_group': page_group,
                'page': page,
            })

        if not telefone:
            return render(request, "pages/cadastro-cliente.html", {
                "error_message": "O campo telefone é obrigatório.",
                'page_group': page_group,
                'page': page,
            })

        # Validar WhatsApp (se houver sessão ativa)
        if token:
            resultado_wpp = validar_tel_whatsapp(telefone, token.token, user=usuario)

            if not resultado_wpp.get("wpp"):
                return render(request, "pages/cadastro-cliente.html", {
                    "error_message": (
                        "O telefone informado não possui um WhatsApp.<br>"
                        "Cada cliente precisa estar cadastrado no WhatsApp."
                    ),
                    'page_group': page_group,
                    'page': page,
                })

            # Verificar duplicidade
            if resultado_wpp.get("cliente_existe_telefone"):
                telefone_existente = resultado_wpp.get("cliente_existe_telefone")
                cliente_existente = Cliente.objects.filter(telefone=telefone_existente, usuario=usuario).first()
                if cliente_existente:
                    return render(request, "pages/cadastro-cliente.html", {
                        "error_message": (
                            "Já existe um cliente cadastrado com este telefone!<br><br>"
                            f"<strong>Nome:</strong> {cliente_existente.nome}<br>"
                            f"<strong>Telefone:</strong> {cliente_existente.telefone}"
                        ),
                        'page_group': page_group,
                        'page': page,
                    })

            # Usar telefone normalizado
            telefone = resultado_wpp.get("telefone_validado_wpp")

        try:
            # Criar cliente apenas com dados básicos (SEM assinatura)
            cliente = Cliente.objects.create(
                nome=nome,
                telefone=telefone,
                email=email,
                cpf=cpf,
                notas=notas,
                usuario=usuario,
                tem_assinatura=False,  # Não possui assinatura ainda
                # Campos de assinatura ficam NULL
                servidor=None,
                plano=None,
                forma_pgto=None,
                dispositivo=None,
                sistema=None,
                indicado_por=None,
            )

            # Log de auditoria
            log_user_action(
                request=request,
                action=UserActionLog.ACTION_CREATE,
                instance=cliente,
                message="Cliente cadastrado (sem assinatura).",
                extra={
                    "telefone": cliente.telefone,
                    "tem_assinatura": False,
                },
            )

            return render(request, "pages/cadastro-cliente.html", {
                "success_message": (
                    f"Cliente <strong>{cliente.nome}</strong> cadastrado com sucesso!<br>"
                    "Agora você pode criar uma assinatura para este cliente."
                ),
                'page_group': page_group,
                'page': page,
            })

        except Exception as e:
            logger.exception(f"[ERRO][CADASTRO CLIENTE BÁSICO] {e}")
            return render(request, "pages/cadastro-cliente.html", {
                "error_message": f"Erro ao cadastrar cliente: {str(e)}",
                'page_group': page_group,
                'page': page,
            })

    # GET - Renderiza formulário
    return render(request, "pages/cadastro-cliente.html", {
        'page_group': page_group,
        'page': page,
    })


# =============================================================================
# CADASTRO DE ASSINATURA
# =============================================================================

@login_required
def cadastrar_assinatura(request):
    """
    Cria assinatura para um cliente.

    Dois modos:
    1. Vincular cliente existente (busca por ID)
    2. Criar novo cliente + assinatura (fluxo completo)

    Ao criar assinatura:
    - Preenche campos de assinatura no Cliente
    - Cria primeira Mensalidade
    - Envia mensagem de confirmação de pagamento
    - Processa indicação e descontos progressivos
    """
    usuario = request.user
    token = SessaoWpp.objects.filter(usuario=usuario, is_active=True).first()
    page_group = "clientes"
    page = "cadastro-assinatura"

    # Querysets para preencher os selects do formulário
    plano_queryset = Plano.objects.filter(usuario=usuario).order_by('nome', 'telas', 'valor')
    forma_pgto_queryset = Tipos_pgto.objects.filter(usuario=usuario).select_related('conta_bancaria__instituicao', 'dados_bancarios').order_by('nome')
    servidor_queryset = Servidor.objects.filter(usuario=usuario).order_by('nome')
    sistema_queryset = Aplicativo.objects.filter(usuario=usuario).order_by('nome')
    dispositivo_queryset = Dispositivo.objects.filter(usuario=usuario).order_by('nome')
    indicador_queryset = Cliente.objects.filter(usuario=usuario, cancelado=False).order_by('nome')

    # Clientes sem assinatura (para vincular)
    clientes_sem_assinatura = Cliente.objects.filter(
        usuario=usuario,
        tem_assinatura=False,
        cancelado=False
    ).order_by('nome')

    # Serializar para JavaScript
    dispositivos_json = json.dumps([{'nome': d.nome} for d in dispositivo_queryset])
    sistemas_json = json.dumps([
        {'nome': a.nome, 'device_has_mac': a.device_has_mac}
        for a in sistema_queryset
    ])

    context = {
        'servidores': servidor_queryset,
        'dispositivos': dispositivo_queryset,
        'dispositivos_json': dispositivos_json,
        'sistemas': sistema_queryset,
        'sistemas_json': sistemas_json,
        'indicadores': indicador_queryset,
        'formas_pgtos': forma_pgto_queryset,
        'planos': plano_queryset,
        'clientes_sem_assinatura': clientes_sem_assinatura,
        'page_group': page_group,
        'page': page,
    }

    if request.method == 'POST':
        post = request.POST
        modo = post.get('modo', 'existente')  # 'existente' ou 'novo'
        cliente_id = post.get('cliente_id', '').strip()

        # Verifica sessão WhatsApp
        if not token:
            context['error_message'] = (
                "Você precisa conectar sua conta ao WhatsApp antes de criar uma assinatura.<br>"
                "Vá até a tela de integração com o WhatsApp e faça a conexão."
            )
            return render(request, "pages/cadastro-assinatura.html", context)

        try:
            with transaction.atomic():
                # ===== MODO 1: Cliente existente =====
                if modo == 'existente':
                    if not cliente_id:
                        context['error_message'] = "Selecione um cliente existente."
                        return render(request, "pages/cadastro-assinatura.html", context)

                    cliente = Cliente.objects.get(pk=int(cliente_id), usuario=usuario)

                    if cliente.tem_assinatura:
                        context['error_message'] = f"O cliente {cliente.nome} já possui uma assinatura ativa."
                        return render(request, "pages/cadastro-assinatura.html", context)

                # ===== MODO 2: Novo cliente =====
                else:
                    nome = post.get('nome', '').strip()
                    telefone = post.get('telefone', '').strip()

                    if not nome or not telefone:
                        context['error_message'] = "Nome e telefone são obrigatórios para novo cliente."
                        return render(request, "pages/cadastro-assinatura.html", context)

                    # Validar WhatsApp
                    resultado_wpp = validar_tel_whatsapp(telefone, token.token, user=usuario)

                    if not resultado_wpp.get("wpp"):
                        context['error_message'] = "O telefone informado não possui WhatsApp."
                        return render(request, "pages/cadastro-assinatura.html", context)

                    if resultado_wpp.get("cliente_existe_telefone"):
                        telefone_existente = resultado_wpp.get("cliente_existe_telefone")
                        cliente_existente = Cliente.objects.filter(telefone=telefone_existente, usuario=usuario).first()
                        if cliente_existente:
                            context['error_message'] = f"Já existe cliente com este telefone: {cliente_existente.nome}"
                            return render(request, "pages/cadastro-assinatura.html", context)

                    telefone = resultado_wpp.get("tel")

                    # Criar cliente básico
                    cliente = Cliente.objects.create(
                        nome=nome,
                        telefone=telefone,
                        usuario=usuario,
                        tem_assinatura=False,
                    )

                # ===== Processar dados da assinatura =====
                servidor_nome = normalizar_servidor(post.get('servidor', '').strip())
                forma_pgto_input = post.get('forma_pgto', '').strip()
                plano_id = post.get('plano', '').strip()
                indicador_nome = post.get('indicador_list', '').strip()
                notas = post.get('notas', '').strip()

                # Validar e buscar plano pelo ID
                if not plano_id or not plano_id.isdigit():
                    context['error_message'] = "Selecione um plano válido."
                    return render(request, "pages/cadastro-assinatura.html", context)

                plano = Plano.objects.filter(id=int(plano_id), usuario=usuario).first()
                if not plano:
                    context['error_message'] = "Plano não encontrado."
                    return render(request, "pages/cadastro-assinatura.html", context)

                # Servidor
                servidor, _ = Servidor.objects.get_or_create(
                    nome=servidor_nome,
                    usuario=usuario
                )

                # Forma de pagamento
                forma_pgto = None
                if forma_pgto_input:
                    if forma_pgto_input.isdigit():
                        forma_pgto = Tipos_pgto.objects.filter(pk=int(forma_pgto_input), usuario=usuario).first()
                    else:
                        forma_pgto = Tipos_pgto.objects.filter(nome__iexact=forma_pgto_input, usuario=usuario).first()

                # Valida se a forma de pagamento não está bloqueada por limite
                if forma_pgto and forma_pgto.esta_bloqueada:
                    context['error_message'] = (
                        "A forma de pagamento selecionada atingiu o limite de faturamento "
                        "e não pode receber novos clientes. "
                        "Selecione outra forma de pagamento disponível."
                    )
                    return render(request, "pages/cadastro-assinatura.html", context)

                # Indicador
                indicador = None
                if indicador_nome:
                    indicador = Cliente.objects.filter(nome__iexact=indicador_nome, usuario=usuario).first()

                # Coletar dados de dispositivos/apps
                telas_data = []
                for i in range(plano.telas):
                    disp_nome = post.get(f'dispositivo_{i}', '').strip()
                    app_nome = post.get(f'aplicativo_{i}', '').strip()
                    device_id = post.get(f'device_id_{i}', '').strip()
                    email_app = post.get(f'email_{i}', '').strip()
                    senha_app = post.get(f'senha_{i}', '').strip()

                    if disp_nome and app_nome:
                        dispositivo, _ = Dispositivo.objects.get_or_create(
                            nome=normalizar_dispositivo(disp_nome),
                            usuario=usuario
                        )
                        sistema, _ = Aplicativo.objects.get_or_create(
                            nome=normalizar_aplicativo(app_nome),
                            usuario=usuario
                        )
                        telas_data.append({
                            'dispositivo': dispositivo,
                            'sistema': sistema,
                            'device_id': device_id,
                            'email': email_app,
                            'senha': senha_app,
                        })

                if not telas_data:
                    context['error_message'] = "Preencha pelo menos um dispositivo e aplicativo."
                    return render(request, "pages/cadastro-assinatura.html", context)

                # ===== Atualizar cliente com dados da assinatura =====
                cliente.servidor = servidor
                cliente.plano = plano
                cliente.forma_pgto = forma_pgto
                cliente.dispositivo = telas_data[0]['dispositivo']
                cliente.sistema = telas_data[0]['sistema']
                cliente.indicado_por = indicador
                cliente.data_adesao = date.today()
                cliente.tem_assinatura = True
                cliente.notas = notas if notas else cliente.notas
                cliente.save()

                # Criar contas de aplicativo
                for idx, tela in enumerate(telas_data):
                    ContaDoAplicativo.objects.create(
                        dispositivo=tela['dispositivo'],
                        app=tela['sistema'],
                        device_id=tela['device_id'],
                        email=tela['email'],
                        device_key=tela['senha'],
                        cliente=cliente,
                        usuario=usuario,
                        is_principal=(idx == 0),
                    )

                # Atualizar AssinaturaCliente
                assinatura, _ = AssinaturaCliente.objects.get_or_create(
                    cliente=cliente,
                    defaults={
                        'plano': plano,
                        'data_inicio_assinatura': date.today(),
                        'dispositivos_usados': len(telas_data)
                    }
                )
                assinatura.plano = plano
                assinatura.dispositivos_usados = len(telas_data)
                assinatura.save()

                # ===== Enviar mensagem de confirmação de pagamento =====
                # Removido: agora a mensagem é enviada apenas após o pagamento
                # da primeira mensalidade (via _enviar_notificacoes_pagamento)
                # envio_apos_novo_cadastro(cliente.telefone, cliente.nome, usuario, token.token)

                # Inscrever em campanha se elegível
                enroll_client_in_campaign_if_eligible(cliente)

                # Criar primeira mensalidade
                criar_mensalidade(cliente)

                # Histórico de planos
                historico_iniciar(cliente, cliente.plano)

                # Log de auditoria
                log_user_action(
                    request=request,
                    action=UserActionLog.ACTION_CREATE,
                    instance=cliente,
                    message="Assinatura criada para cliente.",
                    extra={
                        "telefone": cliente.telefone,
                        "plano": cliente.plano.nome,
                        "servidor": cliente.servidor.nome,
                        "forma_pgto": cliente.forma_pgto.nome if cliente.forma_pgto else None,
                    },
                )

                # Verificar se a forma de pagamento é FastDePix ou PIX Manual
                is_fastdepix = False
                link_painel_cliente = None
                pix_manual_dados = None

                if cliente.forma_pgto and cliente.forma_pgto.conta_bancaria:
                    conta = cliente.forma_pgto.conta_bancaria
                    if conta.instituicao and conta.instituicao.tipo_integracao == 'fastdepix':
                        is_fastdepix = True
                        # Buscar subdomínio do painel do cliente
                        from painel_cliente.models import SubdominioPainelCliente
                        subdominio = SubdominioPainelCliente.objects.filter(
                            admin_responsavel=usuario,
                            ativo=True
                        ).first()
                        if subdominio:
                            link_painel_cliente = f"https://{subdominio.dominio_completo}"
                    elif conta.instituicao and conta.instituicao.tipo_integracao == 'manual':
                        # PIX Manual com dados preenchidos
                        if conta.chave_pix:
                            pix_manual_dados = {
                                'tipo_chave': conta.get_tipo_chave_pix_display() if conta.tipo_chave_pix else '',
                                'chave': conta.chave_pix,
                                'banco': conta.instituicao.nome,
                                'beneficiario': conta.beneficiario or '',
                            }

                # Fallback para dados bancários legados
                if not is_fastdepix and not pix_manual_dados and cliente.forma_pgto:
                    dados = cliente.forma_pgto.dados_bancarios
                    if dados and dados.chave:
                        pix_manual_dados = {
                            'tipo_chave': dados.tipo_chave,
                            'chave': dados.chave,
                            'banco': dados.instituicao,
                            'beneficiario': dados.beneficiario,
                        }

                context['assinatura_criada'] = True
                context['cliente_id'] = cliente.id
                context['cliente_nome'] = cliente.nome
                context['is_fastdepix'] = is_fastdepix
                context['link_painel_cliente'] = link_painel_cliente
                context['pix_manual_dados'] = pix_manual_dados

                context['success_message'] = (
                    f"Assinatura criada com sucesso para <strong>{cliente.nome}</strong>!<br>"
                    f"Plano: {plano.nome} - R$ {plano.valor}"
                )
                return render(request, "pages/cadastro-assinatura.html", context)

        except Cliente.DoesNotExist:
            context['error_message'] = "Cliente não encontrado."
            return render(request, "pages/cadastro-assinatura.html", context)
        except Exception as e:
            logger.exception(f"[ERRO][CADASTRO ASSINATURA] {e}")
            context['error_message'] = f"Erro ao criar assinatura: {str(e)}"
            return render(request, "pages/cadastro-assinatura.html", context)

    # GET - Renderiza formulário
    return render(request, "pages/cadastro-assinatura.html", context)


# AÇÃO PARA CRIAR NOVO OBJETO PLANO MENSAL
@login_required
def create_payment_plan(request):
    planos_mensalidades = Plano.objects.filter(usuario=request.user).order_by('nome', 'telas', 'valor')
    usuario = request.user
    page_group = "nossopainel"
    page = "plano_adesao"

    if request.method == "POST":
        nome = request.POST.get("nome")

        if nome:

            try:
                # Obter valores do formulário
                telas = int(request.POST.get('telas'))
                valor = Decimal(request.POST.get('valor').replace(',', '.'))
                campanha_ativa = request.POST.get('campanha_ativa') == 'on'

                # Buscar planos existentes com mesmo nome/valor/telas
                planos_existentes = Plano.objects.filter(
                    nome=nome,
                    valor=valor,
                    telas=telas,
                    usuario=usuario
                )

                # Verificar se já existe plano com mesma configuração de campanha
                plano_com_campanha = planos_existentes.filter(campanha_ativa=True).exists()
                plano_sem_campanha = planos_existentes.filter(campanha_ativa=False).exists()

                # Validar duplicidade
                if campanha_ativa and plano_com_campanha:
                    return render(
                        request,
                        'pages/cadastro-plano-adesao.html',
                        {
                            'planos_mensalidades': planos_mensalidades,
                            "error_message": "Já existe um Plano com campanha para esta configuração!",
                        },
                    )

                if not campanha_ativa and plano_sem_campanha:
                    return render(
                        request,
                        'pages/cadastro-plano-adesao.html',
                        {
                            'planos_mensalidades': planos_mensalidades,
                            "error_message": "Já existe um Plano regular para esta configuração!",
                        },
                    )

                # Criar novo plano
                plano = Plano.objects.create(
                    nome=nome,
                    valor=valor,
                    telas=telas,
                    max_dispositivos=telas,
                    usuario=usuario
                )

                # Configurar dados de campanha
                plano.campanha_ativa = campanha_ativa

                if campanha_ativa:
                    plano.campanha_tipo = request.POST.get('campanha_tipo', 'FIXO')
                    plano.campanha_data_inicio = request.POST.get('campanha_data_inicio') or None
                    plano.campanha_data_fim = request.POST.get('campanha_data_fim') or None
                    plano.campanha_duracao_meses = request.POST.get('campanha_duracao_meses') or None

                    if plano.campanha_tipo == 'FIXO':
                        campanha_valor_fixo = request.POST.get('campanha_valor_fixo')
                        plano.campanha_valor_fixo = Decimal(campanha_valor_fixo.replace(',', '.')) if campanha_valor_fixo else None
                    else:  # PERSONALIZADO
                        for i in range(1, 13):
                            campo = f'campanha_valor_mes_{i}'
                            valor_mes = request.POST.get(campo)
                            setattr(plano, campo, Decimal(valor_mes.replace(',', '.')) if valor_mes else None)

                plano.save()

                log_user_action(
                    request=request,
                    action=UserActionLog.ACTION_CREATE,
                    instance=plano,
                    message="Plano de adesao criado.",
                    extra={
                        "nome": plano.nome,
                        "valor": str(plano.valor),
                        "telas": plano.telas,
                        "max_dispositivos": plano.max_dispositivos,
                        "campanha_ativa": plano.campanha_ativa,
                        "campanha_tipo": plano.campanha_tipo if plano.campanha_ativa else None,
                    },
                )
                return render(
                    request,
                    'pages/cadastro-plano-adesao.html',
                    {
                        'planos_mensalidades': planos_mensalidades,
                        "success_message": "Novo Plano cadastrado com sucesso!",
                    },
                )
                
            except Exception as e:
                logger.error('[%s] [USER][%s] [IP][%s] [ERRO][%s]', timezone.localtime(), request.user, get_client_ip(request) or 'N/A', e, exc_info=True)
                # Capturando outras exceções e renderizando a página novamente com a mensagem de erro
                return render(
                    request,
                    "pages/cadastro-plano-adesao.html",
                    {
                        'planos_mensalidades': planos_mensalidades,
                        "error_message": "Não foi possível cadastrar este novo Plano. Verifique os logs!",
                    },
                )

    return render(
        request, 'pages/cadastro-plano-adesao.html', {'planos_mensalidades': planos_mensalidades, "page_group": page_group, "page": page}
    )

# AÇÃO PARA CRIAR NOVO OBJETO SERVIDOR
@login_required
def create_server(request):
    servidores = Servidor.objects.filter(usuario=request.user).order_by('nome')
    usuario = request.user
    page_group = "nossopainel"
    page = "servidor"

    if request.method == "POST":
        nome = request.POST.get("nome")
        imagem = request.FILES.get("imagem")

        if nome:

            try:
                # Consultando o objeto requisitado. Caso não exista, será criado.
                servidor, created = Servidor.objects.get_or_create(nome=nome, usuario=usuario)

                if created:
                    # Se foi enviada uma imagem, criar ServidorImagem
                    if imagem:
                        from .models import ServidorImagem
                        ServidorImagem.objects.create(
                            servidor=servidor,
                            usuario=usuario,
                            imagem=imagem
                        )

                    log_user_action(
                        request=request,
                        action=UserActionLog.ACTION_CREATE,
                        instance=servidor,
                        message="Servidor criado.",
                        extra={
                            "nome": servidor.nome,
                            "tem_imagem": bool(imagem),
                        },
                    )
                    return render(
                            request,
                        'pages/cadastro-servidor.html',
                        {
                            'servidores': servidores,
                            "success_message": "Novo Servidor cadastrado com sucesso!",
                        },
                    )

                else:
                    return render(
                            request,
                        'pages/cadastro-servidor.html',
                        {
                            'servidores': servidores,
                            "error_message": "Já existe um Servidor com este nome!",
                        },
                    )
                
            except Exception as e:
                logger.error('[%s] [USER][%s] [IP][%s] [ERRO][%s]', timezone.localtime(), request.user, get_client_ip(request) or 'N/A', e, exc_info=True)
                # Capturando outras exceções e renderizando a página novamente com a mensagem de erro
                return render(
                    request,
                    "'pages/cadastro-servidor.html",
                    {
                        'servidores': servidores,
                        "error_message": "Não foi possível cadastrar este novo Servidor. Verifique os logs!",
                    },
                )

    return render(
        request, 'pages/cadastro-servidor.html', {'servidores': servidores, "page_group": page_group, "page": page}
    )


# ACAO PARA CRIAR NOVO OBJETO FORMA DE PAGAMENTO (TIPOS_PGTO) - Versao simples para usuarios comuns
@login_required
def create_payment_method(request):
    formas_pgto = Tipos_pgto.objects.filter(usuario=request.user).annotate(clientes_count=Count('cliente')).order_by('nome')
    usuario = request.user
    page_group = "nossopainel"
    page = "forma_pgto"

    if request.method == "POST":
        nome = request.POST.get("nome")

        if nome:
            try:
                # Consultando o objeto requisitado. Caso nao exista, sera criado.
                formapgto, created = Tipos_pgto.objects.get_or_create(nome=nome, usuario=usuario)

                if created:
                    log_user_action(
                        request=request,
                        action=UserActionLog.ACTION_CREATE,
                        instance=formapgto,
                        message="Forma de pagamento criada.",
                        extra={
                            "nome": formapgto.nome,
                        },
                    )
                    return render(
                            request,
                        'pages/cadastro-forma-pagamento.html',
                        {
                            'formas_pgto': formas_pgto,
                            "success_message": "Nova Forma de Pagamento cadastrada com sucesso!",
                        },
                    )

                else:
                    return render(
                            request,
                        'pages/cadastro-forma-pagamento.html',
                        {
                            'formas_pgto': formas_pgto,
                            "error_message": "Ja existe uma Forma de Pagamento com este nome!",
                        },
                    )

            except Exception as e:
                logger.error('[%s] [USER][%s] [IP][%s] [ERRO][%s]', timezone.localtime(), request.user, get_client_ip(request) or 'N/A', e, exc_info=True)
                return render(
                    request,
                    "pages/cadastro-forma-pagamento.html",
                    {
                        'formas_pgto': formas_pgto,
                        "error_message": "Nao foi possivel cadastrar esta nova Forma de Pagamento. Verifique os logs!",
                    },
                )

    return render(
        request, 'pages/cadastro-forma-pagamento.html', {
            'formas_pgto': formas_pgto,
            "page_group": page_group,
            "page": page
        }
    )


# ACAO PARA CRIAR NOVO OBJETO FORMA DE PAGAMENTO - Versao com integracao bancaria
@login_required
def create_payment_method_admin(request):
    """
    View para cadastro de formas de pagamento com integracao bancaria.
    Nova estrutura: Instituicao como objeto principal.
    Suporta FastDePix, Mercado Pago, Efi Bank e instituicoes manuais.
    """
    usuario = request.user
    page_group = "nossopainel"
    page = "forma_pgto"

    def get_context(extra=None):
        from django.conf import settings as django_settings
        # Recarregar dados atualizados
        formas_pgto = Tipos_pgto.objects.filter(usuario=usuario).annotate(
            clientes_count=Count('cliente', filter=Q(cliente__cancelado=False))
        ).order_by('nome').select_related('conta_bancaria', 'conta_bancaria__instituicao')
        instituicoes = InstituicaoBancaria.objects.filter(ativo=True).order_by('nome')  # Para cards de seleção
        todas_instituicoes = InstituicaoBancaria.objects.all().order_by('nome')  # Para modal de gerenciamento
        contas_bancarias = ContaBancaria.objects.filter(usuario=usuario, ativo=True).select_related('instituicao')

        # Contar clientes ativos sem forma de pagamento
        clientes_sem_forma_pgto = Cliente.objects.filter(
            usuario=usuario,
            cancelado=False,
            forma_pgto__isnull=True
        ).count()

        # ============================================================
        # ALERTAS FASTDEPIX - Planos sem link ou com valor divergente
        # ============================================================
        alertas_fastdepix = {
            'planos_sem_link': [],
            'planos_valor_divergente': [],
            'tem_alertas': False
        }

        # Verificar se usuário tem conta FastDePix com link_fastdepix
        contas_fastdepix_alerta = ContaBancaria.objects.filter(
            usuario=usuario,
            instituicao__tipo_integracao='fastdepix',
            tipo_cobranca_fastdepix='link_fastdepix',
            ativo=True
        ).select_related('instituicao')

        if contas_fastdepix_alerta.exists():
            planos_usuario = Plano.objects.filter(usuario=usuario)

            for conta in contas_fastdepix_alerta:
                # Buscar links existentes para esta conta
                links_conta = PlanoLinkPagamento.objects.filter(
                    conta_bancaria=conta
                ).select_related('plano')
                planos_com_link = {link.plano_id for link in links_conta}

                # Planos sem link
                for plano in planos_usuario:
                    if plano.id not in planos_com_link:
                        alertas_fastdepix['planos_sem_link'].append({
                            'plano_id': plano.id,
                            'plano_nome': f"{plano.nome} ({plano.telas} tela(s))",
                            'plano_valor': float(plano.valor),
                            'conta_id': conta.id,
                            'conta_nome': conta.nome_identificacao
                        })

                # Planos com valor divergente
                for link in links_conta:
                    if link.valor_divergente:
                        alertas_fastdepix['planos_valor_divergente'].append({
                            'plano_id': link.plano.id,
                            'plano_nome': f"{link.plano.nome} ({link.plano.telas} tela(s))",
                            'valor_atual': float(link.plano.valor),
                            'valor_configurado': float(link.valor_configurado),
                            'conta_id': conta.id,
                            'conta_nome': conta.nome_identificacao
                        })

            alertas_fastdepix['tem_alertas'] = (
                len(alertas_fastdepix['planos_sem_link']) > 0 or
                len(alertas_fastdepix['planos_valor_divergente']) > 0
            )

        ctx = {
            'formas_pgto': formas_pgto,
            'instituicoes': instituicoes,  # Apenas ativas (para cards de seleção)
            'todas_instituicoes': todas_instituicoes,  # Todas (para modal de gerenciamento)
            'contas_bancarias': contas_bancarias,
            'clientes_sem_forma_pgto': clientes_sem_forma_pgto,
            'alertas_fastdepix': alertas_fastdepix,
            "page_group": page_group,
            "page": page,
            "debug": django_settings.DEBUG,  # Para controlar visibilidade do sandbox
        }
        if extra:
            ctx.update(extra)
        return ctx

    if request.method == "POST":
        # Dados do formulario
        nome = request.POST.get("nome")  # Tipo de pagamento: PIX, Cartao de Credito, Boleto
        tipo_integracao = request.POST.get("tipo_integracao")  # fastdepix, mercado_pago, efi_bank, manual
        instituicao_id = request.POST.get("instituicao")
        nome_identificacao = request.POST.get("nome_identificacao")
        beneficiario = request.POST.get("beneficiario")
        tipo_conta = request.POST.get("tipo_conta", "pf")
        limite_mensal = request.POST.get("limite_mensal")

        # Campos PIX
        tipo_chave_pix = request.POST.get("tipo_chave_pix")
        chave_pix = request.POST.get("chave_pix")

        # Campo para instituicao manual
        instituicao_nome_manual = request.POST.get("instituicao_nome_manual")

        if nome:
            try:
                conta_bancaria = None
                instituicao = None

                # ============================================================
                # PROCESSAR INSTITUICAO
                # ============================================================

                # Se for manual e informou nome de instituicao, criar nova
                if tipo_integracao == 'manual' and instituicao_nome_manual:
                    instituicao, created = InstituicaoBancaria.objects.get_or_create(
                        nome=instituicao_nome_manual,
                        defaults={
                            'tipo_integracao': 'manual',
                            'ativo': True,
                        }
                    )
                    if created:
                        log_user_action(
                            request=request,
                            action=UserActionLog.ACTION_CREATE,
                            instance=instituicao,
                            message="Instituicao bancaria criada (manual).",
                            extra={"nome": instituicao.nome},
                        )
                elif instituicao_id:
                    try:
                        instituicao = InstituicaoBancaria.objects.get(id=instituicao_id, ativo=True)
                    except InstituicaoBancaria.DoesNotExist:
                        # Tentar buscar pelo tipo de integracao
                        try:
                            instituicao = InstituicaoBancaria.objects.get(tipo_integracao=tipo_integracao, ativo=True)
                        except InstituicaoBancaria.DoesNotExist:
                            return render(request, 'pages/cadastro-forma-pagamento.html',
                                get_context({"error_message": "Instituicao bancaria nao encontrada."}))
                else:
                    # Buscar pelo tipo de integracao
                    try:
                        instituicao = InstituicaoBancaria.objects.get(tipo_integracao=tipo_integracao, ativo=True)
                    except InstituicaoBancaria.DoesNotExist:
                        return render(request, 'pages/cadastro-forma-pagamento.html',
                            get_context({"error_message": f"Instituicao do tipo '{tipo_integracao}' nao encontrada."}))

                # ============================================================
                # VALIDAR CAMPOS OBRIGATORIOS
                # ============================================================

                # Nome de identificacao e sempre obrigatorio
                if not nome_identificacao:
                    return render(request, 'pages/cadastro-forma-pagamento.html',
                        get_context({"error_message": "Nome de identificacao e obrigatorio."}))

                # Instituicoes com API nao precisam de beneficiario/chave PIX manual
                instituicoes_com_api = ['fastdepix', 'mercado_pago', 'efi_bank']

                if tipo_integracao not in instituicoes_com_api:
                    # Instituicao manual: validar beneficiario
                    if not beneficiario:
                        return render(request, 'pages/cadastro-forma-pagamento.html',
                            get_context({"error_message": "Beneficiario e obrigatorio para instituicoes manuais."}))

                    # Para PIX manual, validar chave PIX
                    if nome == "PIX":
                        if not all([tipo_chave_pix, chave_pix]):
                            return render(request, 'pages/cadastro-forma-pagamento.html',
                                get_context({"error_message": "Para PIX, informe o tipo e a chave PIX."}))

                        # Verificar se ja existe conta com essa chave PIX
                        if ContaBancaria.objects.filter(usuario=usuario, chave_pix=chave_pix).exists():
                            return render(request, 'pages/cadastro-forma-pagamento.html',
                                get_context({"error_message": "Ja existe uma conta bancaria com essa chave PIX."}))

                # ============================================================
                # CRIAR CONTA BANCARIA
                # ============================================================

                # Para instituicoes com API, beneficiario e PIX nao sao necessarios
                # (a API gerencia os dados de recebimento)
                # Usar None para campos vazios para evitar conflito de UNIQUE constraint
                conta_bancaria = ContaBancaria(
                    usuario=usuario,
                    instituicao=instituicao,
                    nome_identificacao=nome_identificacao,
                    tipo_conta=tipo_conta,
                    beneficiario=beneficiario if beneficiario else None,
                    tipo_chave_pix=tipo_chave_pix if tipo_chave_pix else None,
                    chave_pix=chave_pix if chave_pix else None,
                )

                # Limite MEI
                if limite_mensal:
                    conta_bancaria.limite_mensal = limite_mensal

                # ============================================================
                # CREDENCIAIS API POR INSTITUICAO
                # ============================================================

                # Inicializar variável para links de planos (FastDePix)
                links_planos_para_salvar = {}

                if tipo_integracao == 'fastdepix':
                    # FastDePix: apenas API Key
                    api_key = request.POST.get("api_key", "")
                    if not api_key:
                        return render(request, 'pages/cadastro-forma-pagamento.html',
                            get_context({"error_message": "API Key e obrigatoria para FastDePix."}))
                    conta_bancaria.api_key = api_key
                    conta_bancaria.ambiente_sandbox = request.POST.get("ambiente_sandbox") == "on"

                    # Configuração de Cobranças FastDePix
                    tipo_cobranca = request.POST.get("tipoCobrancaFastDePix", "painel_cliente")
                    conta_bancaria.tipo_cobranca_fastdepix = tipo_cobranca

                    # Coletar links por plano para salvar após conta_bancaria.save()
                    links_planos_para_salvar = {}
                    if tipo_cobranca == 'link_fastdepix':
                        for key, value in request.POST.items():
                            if key.startswith('link_plano_') and value:
                                plano_id = key.replace('link_plano_', '')
                                links_planos_para_salvar[plano_id] = value

                        # Validar se todos os planos têm link
                        planos_usuario = Plano.objects.filter(usuario=usuario)
                        if planos_usuario.exists():
                            planos_sem_link = []
                            for plano in planos_usuario:
                                if str(plano.id) not in links_planos_para_salvar:
                                    planos_sem_link.append(f"{plano.nome} ({plano.telas} tela(s))")

                            if planos_sem_link:
                                return render(request, 'pages/cadastro-forma-pagamento.html',
                                    get_context({"error_message": f"Informe o link FastDePix para os planos: {', '.join(planos_sem_link)}"}))

                elif tipo_integracao == 'mercado_pago':
                    # Mercado Pago: Client ID, Client Secret, Access Token
                    api_client_id = request.POST.get("api_client_id_mp", "")
                    api_client_secret = request.POST.get("api_client_secret_mp", "")
                    api_access_token = request.POST.get("api_access_token", "")

                    if not all([api_client_id, api_client_secret, api_access_token]):
                        return render(request, 'pages/cadastro-forma-pagamento.html',
                            get_context({"error_message": "Preencha todas as credenciais do Mercado Pago."}))

                    conta_bancaria.api_client_id = api_client_id
                    conta_bancaria.api_client_secret = api_client_secret
                    conta_bancaria.api_access_token = api_access_token
                    conta_bancaria.ambiente_sandbox = request.POST.get("ambiente_sandbox_mp") == "on"

                elif tipo_integracao == 'efi_bank':
                    # Efi Bank: Client ID, Client Secret, Certificado
                    api_client_id = request.POST.get("api_client_id_efi", "")
                    api_client_secret = request.POST.get("api_client_secret_efi", "")

                    if not all([api_client_id, api_client_secret]):
                        return render(request, 'pages/cadastro-forma-pagamento.html',
                            get_context({"error_message": "Preencha todas as credenciais do Efi Bank."}))

                    if 'api_certificado' not in request.FILES:
                        return render(request, 'pages/cadastro-forma-pagamento.html',
                            get_context({"error_message": "Certificado (.p12) e obrigatorio para Efi Bank."}))

                    conta_bancaria.api_client_id = api_client_id
                    conta_bancaria.api_client_secret = api_client_secret
                    conta_bancaria.api_certificado = request.FILES['api_certificado']
                    conta_bancaria.ambiente_sandbox = request.POST.get("ambiente_sandbox_efi") == "on"

                # Salvar conta bancaria
                conta_bancaria.save()

                # Salvar links de planos FastDePix (se aplicável)
                if tipo_integracao == 'fastdepix' and links_planos_para_salvar:
                    for plano_id_str, url in links_planos_para_salvar.items():
                        try:
                            plano = Plano.objects.get(id=int(plano_id_str), usuario=usuario)
                            PlanoLinkPagamento.objects.create(
                                plano=plano,
                                conta_bancaria=conta_bancaria,
                                url=url,
                                valor_configurado=plano.valor
                            )
                        except Plano.DoesNotExist:
                            pass  # Plano não encontrado, ignorar

                log_user_action(
                    request=request,
                    action=UserActionLog.ACTION_CREATE,
                    instance=conta_bancaria,
                    message="Conta bancaria criada.",
                    extra={
                        "nome": conta_bancaria.nome_identificacao,
                        "instituicao": instituicao.nome,
                        "tipo_integracao": tipo_integracao,
                    },
                )

                # ============================================================
                # CRIAR FORMA DE PAGAMENTO
                # ============================================================

                # Verificar se ja existe forma de pagamento com esse nome E mesma conta bancaria
                if Tipos_pgto.objects.filter(usuario=usuario, nome=nome, conta_bancaria=conta_bancaria).exists():
                    return render(request, 'pages/cadastro-forma-pagamento.html',
                        get_context({"error_message": "Ja existe uma Forma de Pagamento com este nome e esta conta bancaria!"}))

                # Criar forma de pagamento
                formapgto = Tipos_pgto.objects.create(
                    nome=nome,
                    usuario=usuario,
                    conta_bancaria=conta_bancaria
                )

                log_user_action(
                    request=request,
                    action=UserActionLog.ACTION_CREATE,
                    instance=formapgto,
                    message="Forma de pagamento criada.",
                    extra={
                        "nome": formapgto.nome,
                        "instituicao": instituicao.nome,
                        "conta_bancaria": conta_bancaria.nome_identificacao,
                    },
                )

                # ============================================================
                # ASSOCIAR CLIENTES A CONTA BANCARIA
                # ============================================================

                clientes_associados_str = request.POST.get('clientes_associados', '')
                clientes_para_transferir_str = request.POST.get('clientes_para_transferir', '')

                clientes_associados_ids = [int(x) for x in clientes_associados_str.split(',') if x.strip()]
                clientes_para_transferir_ids = [int(x) for x in clientes_para_transferir_str.split(',') if x.strip()]

                # Validar mínimo de 1 cliente
                todos_clientes_ids = set(clientes_associados_ids + clientes_para_transferir_ids)
                if not todos_clientes_ids:
                    return render(request, 'pages/cadastro-forma-pagamento.html',
                        get_context({"error_message": "É obrigatório associar ao menos um cliente à forma de pagamento."}))

                # Processar clientes a serem transferidos primeiro
                for cliente_id in clientes_para_transferir_ids:
                    try:
                        cliente = Cliente.objects.get(id=cliente_id, usuario=usuario)
                        ClienteContaBancaria.transferir_cliente(
                            cliente=cliente,
                            nova_conta=conta_bancaria,
                            usuario=usuario
                        )
                        # Atualizar forma de pagamento do cliente
                        cliente.forma_pgto = formapgto
                        cliente.save(update_fields=['forma_pgto'])
                        log_user_action(
                            request=request,
                            action=UserActionLog.ACTION_UPDATE,
                            instance=cliente,
                            message=f"Cliente transferido para nova conta bancaria e forma de pagamento atualizada.",
                            extra={
                                "cliente": cliente.nome,
                                "nova_conta": conta_bancaria.nome_identificacao,
                                "nova_forma_pgto": str(formapgto),
                            },
                        )
                    except Cliente.DoesNotExist:
                        continue
                    except Exception as e:
                        logger.warning(f"Erro ao transferir cliente {cliente_id}: {e}")

                # Criar associacoes para clientes novos (nao transferidos)
                clientes_novos_ids = set(clientes_associados_ids) - set(clientes_para_transferir_ids)
                for cliente_id in clientes_novos_ids:
                    try:
                        cliente = Cliente.objects.get(id=cliente_id, usuario=usuario)
                        # Verificar se cliente nao esta associado a outra conta
                        if not ClienteContaBancaria.objects.filter(cliente=cliente, ativo=True).exists():
                            ClienteContaBancaria.objects.create(
                                cliente=cliente,
                                conta_bancaria=conta_bancaria,
                                ativo=True
                            )
                            # Atualizar forma de pagamento do cliente
                            cliente.forma_pgto = formapgto
                            cliente.save(update_fields=['forma_pgto'])
                    except Cliente.DoesNotExist:
                        continue
                    except Exception as e:
                        logger.warning(f"Erro ao associar cliente {cliente_id}: {e}")

                total_associados = len(clientes_associados_ids)
                total_transferidos = len(clientes_para_transferir_ids)

                msg_sucesso = "Forma de Pagamento cadastrada com sucesso!"
                if total_associados > 0:
                    msg_sucesso += f" {total_associados} cliente(s) associado(s)"
                    if total_transferidos > 0:
                        msg_sucesso += f" ({total_transferidos} transferido(s))."
                    else:
                        msg_sucesso += "."

                return render(request, 'pages/cadastro-forma-pagamento.html',
                    get_context({"success_message": msg_sucesso}))

            except Exception as e:
                logger.error('[%s] [USER][%s] [IP][%s] [ERRO][%s]', timezone.localtime(), request.user, get_client_ip(request) or 'N/A', e, exc_info=True)
                return render(request, "pages/admin/cadastro-forma-pagamento-admin.html",
                    get_context({"error_message": f"Erro ao cadastrar: {str(e)}"}))

    return render(request, 'pages/cadastro-forma-pagamento.html', get_context())


# ============================================================================
# INTEGRAÇÃO BANCÁRIA - VIEWS
# ============================================================================

@login_required
@require_http_methods(["GET"])
def api_instituicoes_bancarias(request):
    """
    API JSON para listar instituições bancárias ativas.
    Retorna lista de instituições disponíveis.
    """
    instituicoes = InstituicaoBancaria.objects.filter(ativo=True).order_by('nome')
    data = [
        {
            'id': inst.id,
            'nome': inst.nome,
            'tipo_integracao': inst.tipo_integracao,
            'tipo_integracao_display': inst.get_tipo_integracao_display(),
            'tem_api': inst.tem_api,
        }
        for inst in instituicoes
    ]
    return JsonResponse({'instituicoes': data})


@login_required
@require_http_methods(["GET"])
def api_contas_bancarias(request):
    """
    API JSON para listar contas bancárias do usuário.
    Retorna lista de contas ativas do usuário logado.
    """
    contas = ContaBancaria.objects.filter(
        usuario=request.user,
        ativo=True
    ).select_related('instituicao').order_by('nome_identificacao')

    data = [
        {
            'id': conta.id,
            'nome_identificacao': conta.nome_identificacao,
            'instituicao_id': conta.instituicao.id,
            'instituicao_nome': conta.instituicao.nome,
            'tipo_conta': conta.tipo_conta,
            'tipo_conta_display': conta.get_tipo_conta_display(),
            'beneficiario': conta.beneficiario,
            'tipo_chave_pix': conta.tipo_chave_pix,
            'chave_pix': conta.chave_pix,
            'tem_api': conta.tem_integracao_api,
            'is_mei': conta.is_mei,
            'limite_mensal': str(conta.limite_mensal) if conta.limite_mensal else None,
            'limite_efetivo': str(conta.limite_efetivo) if conta.limite_efetivo else None,
            'clientes_count': conta.get_clientes_associados_count(),
        }
        for conta in contas
    ]
    return JsonResponse({'contas': data})


@login_required
@require_http_methods(["POST"])
def criar_conta_bancaria(request):
    """
    Cria uma nova conta bancária para o usuário.
    """
    try:
        data = json.loads(request.body)

        instituicao_id = data.get('instituicao_id')
        nome_identificacao = data.get('nome_identificacao', '').strip()
        tipo_conta = data.get('tipo_conta', 'pf')
        beneficiario = data.get('beneficiario', '').strip()
        tipo_chave_pix = data.get('tipo_chave_pix', '')
        chave_pix = data.get('chave_pix', '').strip()
        limite_mensal = data.get('limite_mensal')

        # Credenciais API (opcionais)
        api_client_id = data.get('api_client_id', '').strip()
        api_client_secret = data.get('api_client_secret', '').strip()
        api_access_token = data.get('api_access_token', '').strip()
        ambiente_sandbox = data.get('ambiente_sandbox', True)

        # Validações
        if not instituicao_id:
            return JsonResponse({'success': False, 'error': 'Instituição é obrigatória.'}, status=400)

        if not nome_identificacao:
            return JsonResponse({'success': False, 'error': 'Nome de identificação é obrigatório.'}, status=400)

        if not beneficiario:
            return JsonResponse({'success': False, 'error': 'Beneficiário é obrigatório.'}, status=400)

        if not tipo_chave_pix:
            return JsonResponse({'success': False, 'error': 'Tipo de chave PIX é obrigatório.'}, status=400)

        if not chave_pix:
            return JsonResponse({'success': False, 'error': 'Chave PIX é obrigatória.'}, status=400)

        # MEI deve ter limite
        if tipo_conta == 'mei' and not limite_mensal:
            return JsonResponse({'success': False, 'error': 'Contas MEI devem ter um limite mensal definido.'}, status=400)

        # Buscar instituição
        try:
            instituicao = InstituicaoBancaria.objects.get(id=instituicao_id, ativo=True)
        except InstituicaoBancaria.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Instituição não encontrada.'}, status=404)

        # Verificar duplicidade de chave PIX
        if ContaBancaria.objects.filter(usuario=request.user, chave_pix=chave_pix).exists():
            return JsonResponse({'success': False, 'error': 'Você já possui uma conta com esta chave PIX.'}, status=400)

        # Criar conta
        conta = ContaBancaria.objects.create(
            usuario=request.user,
            instituicao=instituicao,
            nome_identificacao=nome_identificacao,
            tipo_conta=tipo_conta,
            beneficiario=beneficiario,
            tipo_chave_pix=tipo_chave_pix,
            chave_pix=chave_pix,
            limite_mensal=Decimal(limite_mensal) if limite_mensal else None,
            api_client_id=api_client_id,
            api_client_secret=api_client_secret,
            api_access_token=api_access_token,
            ambiente_sandbox=ambiente_sandbox,
        )

        log_user_action(
            request=request,
            action=UserActionLog.ACTION_CREATE,
            instance=conta,
            message="Conta bancária criada.",
            extra={
                "nome_identificacao": conta.nome_identificacao,
                "instituicao": instituicao.nome,
                "tipo_conta": conta.tipo_conta,
            },
        )

        return JsonResponse({
            'success': True,
            'message': 'Conta bancária criada com sucesso!',
            'conta': {
                'id': conta.id,
                'nome_identificacao': conta.nome_identificacao,
                'instituicao_nome': instituicao.nome,
                'tipo_conta': conta.tipo_conta,
            }
        })

    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Dados inválidos.'}, status=400)
    except Exception as e:
        logger.error('[%s] [USER][%s] [IP][%s] [ERRO][%s]', timezone.localtime(), request.user, get_client_ip(request) or 'N/A', e, exc_info=True)
        return JsonResponse({'success': False, 'error': 'Erro ao criar conta bancária.'}, status=500)


@login_required
@require_http_methods(["DELETE"])
def excluir_conta_bancaria(request, pk):
    """
    Exclui uma conta bancária do usuário.
    """
    try:
        conta = ContaBancaria.objects.get(pk=pk, usuario=request.user)

        # Verificar se está em uso por alguma forma de pagamento
        formas_em_uso = Tipos_pgto.objects.filter(conta_bancaria=conta).count()
        if formas_em_uso > 0:
            return JsonResponse({
                'success': False,
                'error': f'Esta conta está em uso por {formas_em_uso} forma(s) de pagamento. Remova as associações primeiro.'
            }, status=400)

        log_extra = {
            "id": conta.id,
            "nome_identificacao": conta.nome_identificacao,
            "instituicao": conta.instituicao.nome,
        }
        conta.delete()

        log_user_action(
            request=request,
            action=UserActionLog.ACTION_DELETE,
            instance=conta,
            message="Conta bancária excluída.",
            extra=log_extra,
        )

        return JsonResponse({'success': True, 'message': 'Conta bancária excluída com sucesso!'})

    except ContaBancaria.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Conta bancária não encontrada.'}, status=404)
    except Exception as e:
        logger.error('[%s] [USER][%s] [IP][%s] [ERRO][%s]', timezone.localtime(), request.user, get_client_ip(request) or 'N/A', e, exc_info=True)
        return JsonResponse({'success': False, 'error': 'Erro ao excluir conta bancária.'}, status=500)


# --- INSTITUIÇÕES BANCÁRIAS (ADMIN) ---

@login_required
@require_http_methods(["POST"])
def criar_instituicao_bancaria(request):
    """
    Cria uma nova instituição bancária.
    Apenas superusuários podem criar.
    """
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Acesso negado.'}, status=403)

    try:
        data = json.loads(request.body)
        nome = data.get('nome', '').strip()
        tipo_integracao = data.get('tipo_integracao', 'manual')

        if not nome:
            return JsonResponse({'success': False, 'error': 'Nome é obrigatório.'}, status=400)

        # Verificar duplicidade
        if InstituicaoBancaria.objects.filter(nome__iexact=nome).exists():
            return JsonResponse({'success': False, 'error': 'Já existe uma instituição com este nome.'}, status=400)

        instituicao = InstituicaoBancaria.objects.create(
            nome=nome,
            tipo_integracao=tipo_integracao,
            ativo=True,
        )

        log_user_action(
            request=request,
            action=UserActionLog.ACTION_CREATE,
            instance=instituicao,
            message="Instituição bancária criada.",
            extra={"nome": nome, "tipo_integracao": tipo_integracao},
        )

        return JsonResponse({
            'success': True,
            'message': 'Instituição criada com sucesso!',
            'instituicao': {
                'id': instituicao.id,
                'nome': instituicao.nome,
                'tipo_integracao': instituicao.tipo_integracao,
                'tipo_integracao_display': instituicao.get_tipo_integracao_display(),
                'tem_api': instituicao.tem_api,
            }
        })

    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Dados inválidos.'}, status=400)
    except Exception as e:
        logger.error('[%s] [USER][%s] [IP][%s] [ERRO][%s]', timezone.localtime(), request.user, get_client_ip(request) or 'N/A', e, exc_info=True)
        return JsonResponse({'success': False, 'error': 'Erro ao criar instituição.'}, status=500)


@login_required
@require_http_methods(["POST"])
def toggle_instituicao_bancaria(request, pk):
    """
    Ativa/desativa uma instituição bancária.
    Apenas superusuários podem alterar.
    """
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Acesso negado.'}, status=403)

    try:
        instituicao = InstituicaoBancaria.objects.get(pk=pk)
        instituicao.ativo = not instituicao.ativo
        instituicao.save()

        status_text = "ativada" if instituicao.ativo else "desativada"
        log_user_action(
            request=request,
            action=UserActionLog.ACTION_UPDATE,
            instance=instituicao,
            message=f"Instituição bancária {status_text}.",
            extra={"nome": instituicao.nome, "ativo": instituicao.ativo},
        )

        return JsonResponse({
            'success': True,
            'message': f'Instituição {status_text} com sucesso!',
            'ativo': instituicao.ativo,
        })

    except InstituicaoBancaria.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Instituição não encontrada.'}, status=404)
    except Exception as e:
        logger.error('[%s] [USER][%s] [IP][%s] [ERRO][%s]', timezone.localtime(), request.user, get_client_ip(request) or 'N/A', e, exc_info=True)
        return JsonResponse({'success': False, 'error': 'Erro ao alterar instituição.'}, status=500)


@login_required
@require_http_methods(["DELETE"])
def excluir_instituicao_bancaria(request, pk):
    """
    Exclui uma instituição bancária.
    Apenas superusuários podem excluir.
    """
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Acesso negado.'}, status=403)

    try:
        instituicao = InstituicaoBancaria.objects.get(pk=pk)

        # Verificar se está em uso
        contas_em_uso = ContaBancaria.objects.filter(instituicao=instituicao).count()
        if contas_em_uso > 0:
            return JsonResponse({
                'success': False,
                'error': f'Esta instituição está em uso por {contas_em_uso} conta(s) bancária(s). Não é possível excluir.'
            }, status=400)

        log_extra = {"id": instituicao.id, "nome": instituicao.nome}
        instituicao.delete()

        log_user_action(
            request=request,
            action=UserActionLog.ACTION_DELETE,
            instance=instituicao,
            message="Instituição bancária excluída.",
            extra=log_extra,
        )

        return JsonResponse({'success': True, 'message': 'Instituição excluída com sucesso!'})

    except InstituicaoBancaria.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Instituição não encontrada.'}, status=404)
    except Exception as e:
        logger.error('[%s] [USER][%s] [IP][%s] [ERRO][%s]', timezone.localtime(), request.user, get_client_ip(request) or 'N/A', e, exc_info=True)
        return JsonResponse({'success': False, 'error': 'Erro ao excluir instituição.'}, status=500)


# ============================================================================
# CONFIGURAÇÃO DE LIMITE MEI - VIEWS
# ============================================================================

@login_required
@require_http_methods(["GET"])
def api_config_limite(request):
    """
    Retorna a configuração atual de limites de faturamento.
    GET /api/config-limite/
    """
    try:
        config = ConfiguracaoLimite.get_config()
        return JsonResponse({
            'success': True,
            'config': {
                'valor_anual': float(config.valor_anual),
                'valor_anual_pf': float(config.valor_anual_pf),
                'margem_seguranca': config.margem_seguranca,
                'valor_alerta': float(config.valor_alerta),
                'valor_bloqueio': float(config.valor_bloqueio),
                'valor_alerta_pf': float(config.valor_alerta_pf),
                'valor_bloqueio_pf': float(config.valor_bloqueio_pf),
            }
        })
    except Exception as e:
        logger.error('[%s] [USER][%s] [IP][%s] [ERRO][%s]', timezone.localtime(), request.user, get_client_ip(request) or 'N/A', e, exc_info=True)
        return JsonResponse({'success': False, 'error': 'Erro ao carregar configuração.'}, status=500)


@login_required
@require_http_methods(["POST"])
def api_config_limite_atualizar(request):
    """
    Atualiza a configuração de limites de faturamento.
    POST /api/config-limite/atualizar/

    Corpo da requisição (JSON):
    {
        "valor_anual": 81000,
        "valor_anual_pf": 60000,
        "margem_seguranca": 10
    }
    """
    if not request.user.is_staff:
        return JsonResponse({'success': False, 'error': 'Acesso negado.'}, status=403)

    try:
        data = json.loads(request.body)
        valor_anual = data.get('valor_anual')
        valor_anual_pf = data.get('valor_anual_pf')
        margem_seguranca = data.get('margem_seguranca')

        if valor_anual is None or valor_anual <= 0:
            return JsonResponse({'success': False, 'error': 'Valor anual MEI inválido.'}, status=400)

        if valor_anual_pf is None or valor_anual_pf <= 0:
            return JsonResponse({'success': False, 'error': 'Valor anual Pessoa Física inválido.'}, status=400)

        if margem_seguranca is None or margem_seguranca < 0 or margem_seguranca > 50:
            return JsonResponse({'success': False, 'error': 'Margem de segurança deve ser entre 0 e 50%.'}, status=400)

        config = ConfiguracaoLimite.get_config()
        config.valor_anual = Decimal(str(valor_anual))
        config.valor_anual_pf = Decimal(str(valor_anual_pf))
        config.margem_seguranca = int(margem_seguranca)
        config.atualizado_por = request.user
        config.save()

        log_user_action(
            request=request,
            action=UserActionLog.ACTION_UPDATE,
            instance=config,
            message=f"Configuração de limite MEI atualizada: R$ {valor_anual:,.2f}, margem {margem_seguranca}%",
            extra={'valor_anual': float(valor_anual), 'margem_seguranca': margem_seguranca},
        )

        return JsonResponse({
            'success': True,
            'message': 'Configuração atualizada com sucesso!',
            'config': {
                'valor_anual': float(config.valor_anual),
                'margem_seguranca': config.margem_seguranca,
                'valor_alerta': float(config.valor_alerta),
                'valor_bloqueio': float(config.valor_bloqueio),
            }
        })

    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Dados inválidos.'}, status=400)
    except Exception as e:
        logger.error('[%s] [USER][%s] [IP][%s] [ERRO][%s]', timezone.localtime(), request.user, get_client_ip(request) or 'N/A', e, exc_info=True)
        return JsonResponse({'success': False, 'error': 'Erro ao atualizar configuração.'}, status=500)


# ============================================================================
# CREDENCIAIS API - VIEWS
# ============================================================================

@login_required
@require_http_methods(["GET"])
def api_credenciais_por_tipo(request, tipo_integracao):
    """
    Lista credenciais API do usuário por tipo de integração.
    GET /api/credenciais/<tipo_integracao>/

    Tipos válidos: fastdepix, mercado_pago, efi_bank

    Para FastDePix, também lista contas bancárias com API Key configurada.
    """
    tipos_validos = ['fastdepix', 'mercado_pago', 'efi_bank']
    if tipo_integracao not in tipos_validos:
        return JsonResponse({
            'success': False,
            'error': f'Tipo de integração inválido. Use: {", ".join(tipos_validos)}'
        }, status=400)

    try:
        lista = []

        # 1. Buscar credenciais do modelo CredencialAPI
        credenciais = CredencialAPI.objects.filter(
            usuario=request.user,
            tipo_integracao=tipo_integracao,
            ativo=True
        ).order_by('-criado_em')

        for cred in credenciais:
            lista.append({
                'id': cred.id,
                'nome_identificacao': cred.nome_identificacao,
                'tipo_integracao': cred.tipo_integracao,
                'is_configured': cred.is_configured,
                'criado_em': cred.criado_em.strftime('%d/%m/%Y %H:%M'),
                'fonte': 'credencial_api',
            })

        # 2. Para FastDePix, também buscar contas bancárias com API Key
        if tipo_integracao == 'fastdepix':
            contas_fastdepix = ContaBancaria.objects.filter(
                usuario=request.user,
                instituicao__tipo_integracao='fastdepix',
                _api_key__isnull=False
            ).exclude(_api_key='').order_by('-criado_em')

            for conta in contas_fastdepix:
                lista.append({
                    'id': f'conta_{conta.id}',  # Prefixo para diferenciar
                    'nome_identificacao': f'{conta.nome_identificacao} (Conta Bancária)',
                    'tipo_integracao': 'fastdepix',
                    'is_configured': True,
                    'criado_em': conta.criado_em.strftime('%d/%m/%Y %H:%M'),
                    'fonte': 'conta_bancaria',
                    'conta_id': conta.id,
                })

        return JsonResponse({
            'success': True,
            'credenciais': lista,
            'total': len(lista)
        })

    except Exception as e:
        logger.error('[%s] [USER][%s] [IP][%s] [ERRO][%s]', timezone.localtime(), request.user, get_client_ip(request) or 'N/A', e, exc_info=True)
        return JsonResponse({'success': False, 'error': 'Erro ao listar credenciais.'}, status=500)


# ============================================================================
# CLIENTES PARA ASSOCIAÇÃO - VIEWS
# ============================================================================

@login_required
@require_http_methods(["GET"])
def api_clientes_ativos_associacao(request):
    """
    Lista clientes ativos do usuário para associação com forma de pagamento.
    Retorna valor anual PROJETADO baseado no tipo de plano (ano cheio).

    Cálculo:
    - Mensal: valor × 12 pagamentos/ano
    - Bimestral: valor × 6 pagamentos/ano
    - Trimestral: valor × 4 pagamentos/ano
    - Semestral: valor × 2 pagamentos/ano
    - Anual: valor × 1 pagamento/ano

    Parâmetros:
    - conta_id (opcional): ID da conta sendo editada. Clientes associados a esta conta
                           serão marcados como 'ja_associado_esta_conta' em vez de bloqueados.
    - forma_pgto_id (opcional): ID da forma de pagamento sendo editada.
                                A conta bancária será buscada automaticamente.

    GET /api/clientes-ativos-associacao/?conta_id=123
    GET /api/clientes-ativos-associacao/?forma_pgto_id=456
    """
    try:
        from django.db.models import Sum

        # Mapeamento de tipo de plano para quantidade de pagamentos por ano
        PAGAMENTOS_POR_ANO = {
            'Mensal': 12,
            'Bimestral': 6,
            'Trimestral': 4,
            'Semestral': 2,
            'Anual': 1,
        }

        ano_atual = timezone.now().year
        conta_id_editando = request.GET.get('conta_id')
        forma_pgto_id = request.GET.get('forma_pgto_id')
        forma_pgto_id_editando = None  # Para formas antigas (sem conta_bancaria)

        # Se forma_pgto_id foi fornecido, buscar a conta bancária associada
        if forma_pgto_id and not conta_id_editando:
            try:
                forma_pgto = Tipos_pgto.objects.get(id=forma_pgto_id, usuario=request.user)
                if forma_pgto.conta_bancaria:
                    conta_id_editando = str(forma_pgto.conta_bancaria.id)
                else:
                    # Forma antiga (sem conta_bancaria) - guardar ID para comparar diretamente
                    forma_pgto_id_editando = forma_pgto.id
            except Tipos_pgto.DoesNotExist:
                pass

        # Buscar todas as associações ativas de clientes com contas bancárias
        associacoes_ativas = ClienteContaBancaria.objects.filter(
            conta_bancaria__usuario=request.user,
            ativo=True
        ).select_related('conta_bancaria').values(
            'cliente_id',
            'conta_bancaria_id',
            'conta_bancaria__nome_identificacao'
        )

        # Criar dicionário de associações: cliente_id -> {conta_id, conta_nome}
        clientes_associados = {}
        for assoc in associacoes_ativas:
            clientes_associados[assoc['cliente_id']] = {
                'conta_id': assoc['conta_bancaria_id'],
                'conta_nome': assoc['conta_bancaria__nome_identificacao']
            }

        # Incluir apenas clientes ATIVOS (cancelados não interferem nos limites)
        clientes = Cliente.objects.filter(
            usuario=request.user,
            cancelado=False
        ).select_related('plano', 'forma_pgto', 'forma_pgto__conta_bancaria').order_by('nome')

        resultado = []
        for cliente in clientes:
            if cliente.plano:
                valor_plano = cliente.plano.valor
                tipo_plano = cliente.plano.nome  # Mensal, Bimestral, etc.
                pagamentos_ano = PAGAMENTOS_POR_ANO.get(tipo_plano, 12)  # Default: mensal
            else:
                valor_plano = Decimal('0')
                tipo_plano = '-'
                pagamentos_ano = 12

            # Valor anual PROJETADO (ano cheio com plano atual)
            valor_anual_projetado = float(valor_plano * pagamentos_ano)

            # Valor já recebido este ano (para informação)
            valor_recebido_ano = Mensalidade.objects.filter(
                cliente=cliente,
                pgto=True,
                dt_pagamento__year=ano_atual
            ).aggregate(total=Sum('valor'))['total'] or Decimal('0')

            # Próximo vencimento em aberto (apenas para clientes ATIVOS)
            prox_vencimento = None
            prox_vencimento_formatado = '-'
            if not cliente.cancelado:
                proxima_mensalidade = Mensalidade.objects.filter(
                    cliente=cliente,
                    pgto=False,
                    cancelado=False,
                ).order_by('dt_vencimento').values('dt_vencimento').first()
                if proxima_mensalidade:
                    prox_vencimento = str(proxima_mensalidade['dt_vencimento'])
                    prox_vencimento_formatado = proxima_mensalidade['dt_vencimento'].strftime('%d/%m/%Y')

            # Verificar se cliente já está associado a outra conta
            associacao_info = clientes_associados.get(cliente.id)
            ja_associado = associacao_info is not None
            bloqueado = False
            ja_associado_esta_conta = False
            conta_associada = None

            # Função auxiliar para formatar "Tipo (Identificação)"
            def formatar_conta_associada(tipo_pgto, identificacao):
                if identificacao:
                    return f"{tipo_pgto} ({identificacao})"
                return tipo_pgto

            if ja_associado:
                # Cliente tem associação via ClienteContaBancaria
                tipo_pgto = cliente.forma_pgto.nome if cliente.forma_pgto else 'PIX'
                conta_associada = formatar_conta_associada(tipo_pgto, associacao_info['conta_nome'])
                # Se estamos editando uma conta específica
                if conta_id_editando:
                    # Cliente está associado à conta que estamos editando = não bloquear, marcar como selecionado
                    if str(associacao_info['conta_id']) == str(conta_id_editando):
                        ja_associado_esta_conta = True
                        bloqueado = False
                    else:
                        # Cliente está associado a OUTRA conta = bloquear
                        bloqueado = True
                else:
                    # Criando nova conta = bloquear clientes já associados
                    bloqueado = True
            else:
                # Cliente NÃO tem associação via ClienteContaBancaria
                # Verificar se tem forma de pagamento com conta bancária
                if cliente.forma_pgto and cliente.forma_pgto.conta_bancaria:
                    forma_pgto_conta_id = cliente.forma_pgto.conta_bancaria.id
                    tipo_pgto = cliente.forma_pgto.nome
                    identificacao = cliente.forma_pgto.conta_bancaria.nome_identificacao
                    conta_associada_formatada = formatar_conta_associada(tipo_pgto, identificacao)

                    if conta_id_editando:
                        # Se estamos editando uma conta específica
                        if str(forma_pgto_conta_id) == str(conta_id_editando):
                            # Cliente usa forma de pagamento da conta que estamos editando
                            ja_associado_esta_conta = True
                            bloqueado = False
                            conta_associada = conta_associada_formatada
                        else:
                            # Cliente usa forma de pagamento de OUTRA conta = bloquear
                            bloqueado = True
                            conta_associada = conta_associada_formatada
                    else:
                        # Criando nova conta = bloquear clientes com forma de pagamento já associada a outra conta
                        bloqueado = True
                        conta_associada = conta_associada_formatada
                elif cliente.forma_pgto:
                    # Cliente tem forma de pagamento ANTIGA (sem conta_bancaria)
                    tipo_pgto = cliente.forma_pgto.nome
                    identificacao = cliente.forma_pgto.nome_identificacao
                    conta_associada_formatada = formatar_conta_associada(tipo_pgto, identificacao)

                    if forma_pgto_id_editando:
                        # Estamos editando uma forma antiga
                        if cliente.forma_pgto_id == forma_pgto_id_editando:
                            # Cliente está associado à forma que estamos editando
                            ja_associado_esta_conta = True
                            conta_associada = conta_associada_formatada
                        else:
                            # Cliente está associado a OUTRA forma antiga = bloquear
                            bloqueado = True
                            conta_associada = conta_associada_formatada
                    else:
                        # Criando nova forma = bloquear cliente com forma antiga
                        bloqueado = True
                        conta_associada = conta_associada_formatada

            # Informações da forma de pagamento do cliente (campo direto)
            forma_pgto_info = None
            if cliente.forma_pgto:
                # Para formas novas: usar nome_identificacao da conta_bancaria
                # Para formas antigas: usar nome_identificacao do próprio Tipos_pgto
                if cliente.forma_pgto.conta_bancaria:
                    identificacao = cliente.forma_pgto.conta_bancaria.nome_identificacao
                else:
                    identificacao = cliente.forma_pgto.nome_identificacao

                forma_pgto_info = {
                    'id': cliente.forma_pgto.id,
                    'nome': cliente.forma_pgto.nome,
                    'identificacao': identificacao,
                    'conta_bancaria_id': cliente.forma_pgto.conta_bancaria.id if cliente.forma_pgto.conta_bancaria else None,
                }

            resultado.append({
                'id': cliente.id,
                'nome': cliente.nome,
                'telefone': cliente.telefone or '',
                'plano': tipo_plano,
                'valor_plano': float(valor_plano),
                'pagamentos_ano': pagamentos_ano,
                'valor_anual': valor_anual_projetado,
                'valor_recebido_ano': float(valor_recebido_ano),
                # Próximo vencimento em aberto
                'prox_vencimento': prox_vencimento,
                'prox_vencimento_formatado': prox_vencimento_formatado,
                # Informações de associação (ClienteContaBancaria)
                'ja_associado': ja_associado,
                'bloqueado': bloqueado,
                'ja_associado_esta_conta': ja_associado_esta_conta,
                'conta_associada': conta_associada,
                # Status do cliente (ativo ou cancelado)
                'cancelado': cliente.cancelado,
                'status': 'Cancelado' if cliente.cancelado else 'Ativo',
                # Informações da forma de pagamento direta do cliente (Cliente.forma_pgto)
                'forma_pgto': forma_pgto_info,
            })

        return JsonResponse({
            'success': True,
            'clientes': resultado,
            'total': len(resultado)
        })

    except Exception as e:
        logger.error('[%s] [USER][%s] [IP][%s] [ERRO][%s]', timezone.localtime(), request.user, get_client_ip(request) or 'N/A', e, exc_info=True)
        return JsonResponse({'success': False, 'error': 'Erro ao listar clientes.'}, status=500)


# ============================================================================
# PLANOS DE ADESÃO - VIEWS
# ============================================================================

@login_required
@require_http_methods(["GET"])
def api_planos(request):
    """
    Lista planos de adesão do usuário para configuração de links FastDePix.
    GET /api/planos/
    GET /api/planos/?conta_bancaria_id=123 (retorna links existentes)
    """
    try:
        conta_bancaria_id = request.GET.get('conta_bancaria_id')
        planos = Plano.objects.filter(usuario=request.user).order_by('nome', 'telas', 'valor')

        # Buscar links existentes se conta_bancaria_id foi fornecido
        links_existentes = {}
        if conta_bancaria_id:
            links = PlanoLinkPagamento.objects.filter(
                conta_bancaria_id=conta_bancaria_id,
                conta_bancaria__usuario=request.user
            ).select_related('plano')
            links_existentes = {link.plano_id: link for link in links}

        resultado = []
        for plano in planos:
            plano_data = {
                'id': plano.id,
                'nome': plano.nome,
                'telas': plano.telas,
                'valor': float(plano.valor)
            }

            # Adicionar link existente se houver
            link = links_existentes.get(plano.id)
            if link:
                plano_data['link_url'] = link.url
                plano_data['link_valor_configurado'] = float(link.valor_configurado)
                plano_data['link_valor_divergente'] = plano.valor != link.valor_configurado

            resultado.append(plano_data)

        return JsonResponse({
            'success': True,
            'planos': resultado,
            'total': len(resultado)
        })

    except Exception as e:
        logger.error('[%s] [USER][%s] [IP][%s] [ERRO][%s]', timezone.localtime(), request.user, get_client_ip(request) or 'N/A', e, exc_info=True)
        return JsonResponse({'success': False, 'error': 'Erro ao listar planos.'}, status=500)


# ============================================================================
# NOTIFICAÇÕES DO SISTEMA - VIEWS
# ============================================================================

@login_required
@require_http_methods(["GET"])
def api_notificacoes_listar(request):
    """
    Lista notificações do usuário (não lidas primeiro).
    GET /api/notificacoes/
    Parâmetros:
      - apenas_nao_lidas: true/false (default: false)
      - limit: número (default: 20)
    """
    try:
        apenas_nao_lidas = request.GET.get('apenas_nao_lidas', 'false').lower() == 'true'
        limit = int(request.GET.get('limit', 20))

        queryset = NotificacaoSistema.objects.filter(usuario=request.user)

        if apenas_nao_lidas:
            queryset = queryset.filter(lida=False)

        notificacoes = queryset.order_by('-criada_em')[:limit]

        total_nao_lidas = NotificacaoSistema.objects.filter(
            usuario=request.user,
            lida=False
        ).count()

        resultado = []
        for notif in notificacoes:
            resultado.append({
                'id': notif.id,
                'tipo': notif.tipo,
                'tipo_display': notif.get_tipo_display(),
                'prioridade': notif.prioridade,
                'titulo': notif.titulo,
                'mensagem': notif.mensagem,
                'dados_extras': notif.dados_extras,
                'lida': notif.lida,
                'criada_em': notif.criada_em.isoformat(),
            })

        return JsonResponse({
            'success': True,
            'notificacoes': resultado,
            'total_nao_lidas': total_nao_lidas,
        })

    except Exception as e:
        logger.error('[%s] [USER][%s] [IP][%s] [ERRO][%s]', timezone.localtime(), request.user, get_client_ip(request) or 'N/A', e, exc_info=True)
        return JsonResponse({'success': False, 'error': 'Erro ao listar notificações.'}, status=500)


@login_required
@require_http_methods(["POST"])
def api_notificacao_marcar_lida(request, notificacao_id):
    """
    Marca uma notificação como lida.
    POST /api/notificacoes/<id>/marcar-lida/
    """
    try:
        notificacao = NotificacaoSistema.objects.get(
            id=notificacao_id,
            usuario=request.user
        )
        notificacao.marcar_como_lida()

        return JsonResponse({'success': True})

    except NotificacaoSistema.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Notificação não encontrada.'}, status=404)
    except Exception as e:
        logger.error('[%s] [USER][%s] [IP][%s] [ERRO][%s]', timezone.localtime(), request.user, get_client_ip(request) or 'N/A', e, exc_info=True)
        return JsonResponse({'success': False, 'error': 'Erro ao marcar notificação.'}, status=500)


@login_required
@require_http_methods(["POST"])
def api_notificacoes_marcar_todas_lidas(request):
    """
    Marca todas as notificações do usuário como lidas.
    POST /api/notificacoes/marcar-todas-lidas/
    """
    try:
        atualizadas = NotificacaoSistema.objects.filter(
            usuario=request.user,
            lida=False
        ).update(lida=True, data_leitura=timezone.now())

        return JsonResponse({
            'success': True,
            'atualizadas': atualizadas
        })

    except Exception as e:
        logger.error('[%s] [USER][%s] [IP][%s] [ERRO][%s]', timezone.localtime(), request.user, get_client_ip(request) or 'N/A', e, exc_info=True)
        return JsonResponse({'success': False, 'error': 'Erro ao marcar notificações.'}, status=500)


# ============================================================================
# PUSH NOTIFICATIONS - SUBSCRIPTION ENDPOINTS
# ============================================================================

@login_required
@require_http_methods(["POST"])
def api_push_subscribe(request):
    """
    Registra uma subscription de push notification para o usuário.

    POST /api/push/subscribe/
    Body: {
        "endpoint": "https://...",
        "keys": {
            "p256dh": "...",
            "auth": "..."
        }
    }
    """
    from nossopainel.models import PushSubscription

    try:
        data = json.loads(request.body)

        endpoint = data.get('endpoint')
        keys = data.get('keys', {})
        p256dh = keys.get('p256dh')
        auth = keys.get('auth')

        if not all([endpoint, p256dh, auth]):
            return JsonResponse({
                'success': False,
                'error': 'Dados de subscription incompletos'
            }, status=400)

        # Criar ou atualizar subscription
        subscription, created = PushSubscription.objects.update_or_create(
            usuario=request.user,
            endpoint=endpoint,
            defaults={
                'p256dh': p256dh,
                'auth': auth,
                'user_agent': request.META.get('HTTP_USER_AGENT', '')[:500],
                'ativo': True,
            }
        )

        logger.info(f'[Push] Subscription {"criada" if created else "atualizada"} para usuário {request.user.id}')

        return JsonResponse({
            'success': True,
            'created': created,
            'subscription_id': subscription.id
        })

    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'JSON inválido'}, status=400)
    except Exception as e:
        logger.error(f'[Push] Erro ao criar subscription: {e}', exc_info=True)
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def api_push_unsubscribe(request):
    """
    Remove uma subscription de push notification.

    POST /api/push/unsubscribe/
    Body: {"endpoint": "https://..."}
    """
    from nossopainel.models import PushSubscription

    try:
        data = json.loads(request.body)
        endpoint = data.get('endpoint')

        if not endpoint:
            return JsonResponse({'success': False, 'error': 'Endpoint não fornecido'}, status=400)

        deleted, _ = PushSubscription.objects.filter(
            usuario=request.user,
            endpoint=endpoint
        ).delete()

        logger.info(f'[Push] Subscription removida para usuário {request.user.id}: {deleted} registro(s)')

        return JsonResponse({
            'success': True,
            'deleted': deleted
        })

    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'JSON inválido'}, status=400)
    except Exception as e:
        logger.error(f'[Push] Erro ao remover subscription: {e}', exc_info=True)
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["GET"])
def api_push_vapid_public_key(request):
    """
    Retorna a chave pública VAPID para o frontend.

    GET /api/push/vapid-key/
    """
    from django.conf import settings

    vapid_public = getattr(settings, 'VAPID_PUBLIC_KEY', '')

    if not vapid_public:
        return JsonResponse({
            'success': False,
            'error': 'Push notifications não configurado'
        }, status=503)

    return JsonResponse({
        'success': True,
        'vapid_public_key': vapid_public
    })


# ============================================================================
# COBRANÇA PIX - VIEWS
# ============================================================================

@login_required
@require_http_methods(["POST"])
def gerar_cobranca_pix(request, mensalidade_id):
    """
    Gera uma cobrança PIX para uma mensalidade.

    POST /api/pix/gerar/<mensalidade_id>/

    Retorna:
        - qr_code: Código para gerar QR Code
        - qr_code_base64: Imagem do QR Code em base64 (se disponível)
        - pix_copia_cola: Código copia e cola
        - valor: Valor da cobrança
        - expira_em: Data/hora de expiração
        - cobranca_id: ID da cobrança no sistema
    """
    from nossopainel.services.payment_integrations import get_payment_integration, PaymentIntegrationError

    try:
        # Buscar mensalidade
        mensalidade = Mensalidade.objects.select_related(
            'cliente', 'cliente__forma_pgto', 'cliente__forma_pgto__conta_bancaria',
            'cliente__forma_pgto__conta_bancaria__instituicao'
        ).get(id=mensalidade_id, usuario=request.user)

        # Verificar se já foi paga
        if mensalidade.data_pagamento:
            return JsonResponse({
                'success': False,
                'error': 'Esta mensalidade já foi paga.'
            }, status=400)

        # Verificar forma de pagamento do cliente
        forma_pgto = mensalidade.cliente.forma_pgto
        if not forma_pgto or forma_pgto.nome != 'PIX':
            return JsonResponse({
                'success': False,
                'error': 'Cliente não possui forma de pagamento PIX configurada.'
            }, status=400)

        # Verificar conta bancária
        conta_bancaria = forma_pgto.conta_bancaria
        if not conta_bancaria:
            return JsonResponse({
                'success': False,
                'error': 'Forma de pagamento PIX não possui conta bancária configurada.'
            }, status=400)

        # Verificar se conta tem integração API
        if not conta_bancaria.tem_integracao_api:
            return JsonResponse({
                'success': False,
                'error': 'Conta bancária não possui integração com API de pagamento.'
            }, status=400)

        # Obter integração de pagamento
        integration = get_payment_integration(conta_bancaria)
        if not integration:
            return JsonResponse({
                'success': False,
                'error': 'Integração de pagamento não disponível para esta conta.'
            }, status=400)

        # Verificar se já existe cobrança pendente para esta mensalidade
        cobranca_existente = CobrancaPix.objects.filter(
            mensalidade=mensalidade,
            status='pending'
        ).first()

        if cobranca_existente and not cobranca_existente.is_expired:
            # Retornar cobrança existente
            return JsonResponse({
                'success': True,
                'cobranca_id': str(cobranca_existente.id),
                'transaction_id': cobranca_existente.transaction_id,
                'qr_code': cobranca_existente.qr_code,
                'qr_code_url': cobranca_existente.qr_code_url,
                'qr_code_base64': cobranca_existente.qr_code_base64,
                'pix_copia_cola': cobranca_existente.pix_copia_cola,
                'valor': float(cobranca_existente.valor),
                'expira_em': cobranca_existente.expira_em.isoformat(),
                'mensagem': 'Cobrança já existente retornada.',
            })

        # Gerar descrição
        descricao = f"Mensalidade {mensalidade.cliente.nome} - {mensalidade.mes_referencia}"

        # Criar cobrança via API
        pix_charge = integration.create_pix_charge(
            amount=mensalidade.valor,
            description=descricao,
            external_id=str(mensalidade.id),
            expiration_minutes=30,
            payer_name=mensalidade.cliente.nome,
            payer_document=mensalidade.cliente.cpf if hasattr(mensalidade.cliente, 'cpf') else None,
        )

        # Salvar cobrança no banco
        cobranca = CobrancaPix.objects.create(
            transaction_id=pix_charge.transaction_id,
            usuario=request.user,
            conta_bancaria=conta_bancaria,
            mensalidade=mensalidade,
            cliente=mensalidade.cliente,
            valor=pix_charge.amount,
            descricao=descricao,
            status='pending',
            qr_code=pix_charge.qr_code,
            qr_code_url=pix_charge.qr_code_url or '',
            qr_code_base64=pix_charge.qr_code_base64 or '',
            pix_copia_cola=pix_charge.pix_copy_paste,
            expira_em=pix_charge.expiration,
            integracao=conta_bancaria.instituicao.tipo_integracao,
            raw_response=pix_charge.raw_response or {},
        )

        log_user_action(
            request=request,
            action=UserActionLog.ACTION_CREATE,
            instance=cobranca,
            message="Cobrança PIX gerada.",
            extra={
                "mensalidade_id": str(mensalidade.id),
                "cliente": mensalidade.cliente.nome,
                "valor": float(cobranca.valor),
                "integracao": cobranca.integracao,
            },
        )

        return JsonResponse({
            'success': True,
            'cobranca_id': str(cobranca.id),
            'transaction_id': cobranca.transaction_id,
            'qr_code': cobranca.qr_code,
            'qr_code_url': cobranca.qr_code_url,  # Link para visualizar QR Code
            'qr_code_base64': cobranca.qr_code_base64,
            'pix_copia_cola': cobranca.pix_copia_cola,
            'valor': float(cobranca.valor),
            'expira_em': cobranca.expira_em.isoformat(),
            'mensagem': 'Cobrança PIX gerada com sucesso!',
        })

    except Mensalidade.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Mensalidade não encontrada.'
        }, status=404)

    except PaymentIntegrationError as e:
        logger.error(f'[PIX] Erro ao gerar cobrança: {e.message}')
        return JsonResponse({
            'success': False,
            'error': f'Erro na integração: {e.message}'
        }, status=400)

    except Exception as e:
        logger.error('[%s] [USER][%s] [IP][%s] [ERRO][%s]', timezone.localtime(), request.user, get_client_ip(request) or 'N/A', e, exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'Erro ao gerar cobrança PIX.'
        }, status=500)


@login_required
@require_http_methods(["GET"])
def consultar_cobranca_pix(request, cobranca_id):
    """
    Consulta o status de uma cobrança PIX.

    GET /api/pix/status/<cobranca_id>/
    """
    from nossopainel.services.payment_integrations import get_payment_integration

    try:
        cobranca = CobrancaPix.objects.select_related(
            'conta_bancaria', 'conta_bancaria__instituicao'
        ).get(id=cobranca_id, usuario=request.user)

        # Se ainda está pendente, consultar na API
        if cobranca.status == 'pending' and not cobranca.is_expired:
            integration = get_payment_integration(cobranca.conta_bancaria)
            if integration:
                try:
                    status = integration.get_charge_status(cobranca.transaction_id)
                    if status.value != cobranca.status:
                        cobranca.status = status.value
                        if status.value == 'paid':
                            cobranca.pago_em = timezone.now()
                        cobranca.save()
                except Exception as e:
                    logger.warning(f'[PIX] Erro ao consultar status: {e}')

        # Verificar expiração
        if cobranca.status == 'pending' and cobranca.is_expired:
            cobranca.mark_as_expired()

        return JsonResponse({
            'success': True,
            'cobranca_id': str(cobranca.id),
            'transaction_id': cobranca.transaction_id,
            'status': cobranca.status,
            'status_display': cobranca.get_status_display(),
            'valor': float(cobranca.valor),
            'is_paid': cobranca.is_paid,
            'is_expired': cobranca.is_expired,
            'pago_em': cobranca.pago_em.isoformat() if cobranca.pago_em else None,
            'expira_em': cobranca.expira_em.isoformat() if cobranca.expira_em else None,
        })

    except CobrancaPix.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Cobrança não encontrada.'
        }, status=404)


@csrf_exempt
@require_http_methods(["POST"])
def webhook_pagamento_pix(request):
    """
    Recebe webhooks de pagamento das integrações PIX.

    POST /api/pix/webhook/

    Headers esperados:
        - X-Webhook-Signature: Assinatura HMAC-SHA256 (FastDePix)
        - X-Integration: Identificador da integração (opcional)
    """
    import logging
    from nossopainel.services.payment_integrations import (
        FastDePixIntegration, PaymentStatus
    )

    # Logger específico para webhook FastDePix
    fdpx_logger = logging.getLogger('fastdepix.webhook')

    # ========== LOG INICIAL: Request recebido ==========
    fdpx_logger.info("=" * 80)
    fdpx_logger.info("WEBHOOK RECEBIDO - INÍCIO DO PROCESSAMENTO")
    fdpx_logger.info("=" * 80)

    # Log de informações da requisição
    fdpx_logger.info(f"Timestamp: {timezone.now().isoformat()}")
    fdpx_logger.info(f"IP Origem: {request.META.get('REMOTE_ADDR', 'N/A')}")
    fdpx_logger.info(f"X-Forwarded-For: {request.META.get('HTTP_X_FORWARDED_FOR', 'N/A')}")
    fdpx_logger.info(f"User-Agent: {request.META.get('HTTP_USER_AGENT', 'N/A')}")
    fdpx_logger.info(f"Content-Type: {request.META.get('CONTENT_TYPE', 'N/A')}")
    fdpx_logger.info(f"Content-Length: {request.META.get('CONTENT_LENGTH', 'N/A')}")

    # Log de todos os headers relevantes
    fdpx_logger.info("-" * 40)
    fdpx_logger.info("HEADERS:")
    for key, value in request.headers.items():
        fdpx_logger.info(f"  {key}: {value}")

    try:
        # ========== PARSE DO BODY ==========
        fdpx_logger.info("-" * 40)
        fdpx_logger.info("BODY (raw):")
        raw_body = request.body.decode('utf-8', errors='replace')
        fdpx_logger.info(raw_body)

        # Tratar requisições com body vazio (possíveis health checks)
        if not raw_body or not raw_body.strip():
            fdpx_logger.info("Requisição vazia recebida (possível health check) - ignorando")
            fdpx_logger.info("=" * 80)
            return JsonResponse({'status': 'ok', 'info': 'empty body ignored'})

        try:
            payload = json.loads(request.body)

            # FastDePix às vezes envia JSON encapsulado como string
            # Se o payload for string, fazer segundo parse
            if isinstance(payload, str):
                fdpx_logger.info("Payload é string, fazendo segundo parse...")
                payload = json.loads(payload)

        except json.JSONDecodeError as e:
            fdpx_logger.error(f"ERRO: JSON inválido - {e}")
            fdpx_logger.info("=" * 80)
            return JsonResponse({'error': 'Invalid JSON'}, status=400)

        fdpx_logger.info("-" * 40)
        fdpx_logger.info("PAYLOAD (parsed JSON):")
        fdpx_logger.info(json.dumps(payload, indent=2, ensure_ascii=False, default=str))

        # ========== EXTRAIR DADOS DO PAYLOAD ==========
        # O payload pode vir com estrutura {data: {...}} ou diretamente como {...}
        data = payload.get('data', payload) if isinstance(payload, dict) else payload
        transaction_id = str(data.get('id', data.get('transaction_id', '')))
        event_type = payload.get('event', '')
        status_str = data.get('status', '').lower()

        fdpx_logger.info("-" * 40)
        fdpx_logger.info("DADOS EXTRAÍDOS:")
        fdpx_logger.info(f"  Transaction ID: {transaction_id}")
        fdpx_logger.info(f"  Event Type: {event_type}")
        fdpx_logger.info(f"  Status: {status_str}")

        if not transaction_id:
            fdpx_logger.warning("ERRO: transaction_id não encontrado no payload")
            fdpx_logger.info("=" * 80)
            return JsonResponse({'error': 'Transaction ID not found'}, status=400)

        # ========== BUSCAR COBRANÇA ==========
        fdpx_logger.info("-" * 40)
        fdpx_logger.info(f"Buscando CobrancaPix com transaction_id: {transaction_id}")

        try:
            cobranca = CobrancaPix.objects.select_related(
                'conta_bancaria', 'conta_bancaria__instituicao', 'mensalidade', 'cliente'
            ).get(transaction_id=transaction_id)

            fdpx_logger.info("COBRANÇA ENCONTRADA:")
            fdpx_logger.info(f"  ID: {cobranca.id}")
            fdpx_logger.info(f"  Status atual: {cobranca.status}")
            fdpx_logger.info(f"  Valor: R$ {cobranca.valor}")
            fdpx_logger.info(f"  Criada em: {cobranca.criado_em}")
            fdpx_logger.info(f"  Expira em: {cobranca.expira_em}")
            fdpx_logger.info(f"  Integração: {cobranca.integracao}")
            if cobranca.mensalidade:
                fdpx_logger.info(f"  Mensalidade ID: {cobranca.mensalidade.id}")
                fdpx_logger.info(f"  Mensalidade Vencimento: {cobranca.mensalidade.dt_vencimento}")
            if cobranca.cliente:
                fdpx_logger.info(f"  Cliente: {cobranca.cliente.nome} (ID: {cobranca.cliente.id})")
            if cobranca.conta_bancaria:
                fdpx_logger.info(f"  Conta Bancária: {cobranca.conta_bancaria.nome_identificacao}")

        except CobrancaPix.DoesNotExist:
            fdpx_logger.warning(f"COBRANÇA NÃO ENCONTRADA: {transaction_id}")
            fdpx_logger.info("Retornando 200 OK (para não reenviar webhook)")
            fdpx_logger.info("=" * 80)
            return JsonResponse({'status': 'ignored', 'reason': 'Charge not found'})

        # ========== VALIDAR ASSINATURA ==========
        signature = request.headers.get('X-Webhook-Signature', '')
        fdpx_logger.info("-" * 40)
        fdpx_logger.info("VALIDAÇÃO DE ASSINATURA:")
        fdpx_logger.info(f"  Assinatura recebida: {signature if signature else '(nenhuma)'}")
        fdpx_logger.info(f"  Webhook Secret configurado: {'Sim' if cobranca.conta_bancaria.webhook_secret else 'Não'}")

        # ========== PROCESSAR EVENTO ==========
        fdpx_logger.info("-" * 40)
        fdpx_logger.info("PROCESSAMENTO DO EVENTO:")

        if event_type == 'transaction.paid' or status_str == 'paid':
            fdpx_logger.info(f"  Tipo: PAGAMENTO CONFIRMADO")

            if cobranca.status != 'paid':
                # Extrair dados do pagador
                payer = data.get('payer', {})
                payer_name = payer.get('name') if isinstance(payer, dict) else None
                payer_doc = payer.get('cpf_cnpj') if isinstance(payer, dict) else None

                fdpx_logger.info(f"  Pagador Nome: {payer_name}")
                fdpx_logger.info(f"  Pagador Documento: {payer_doc}")

                # Extrair data de pagamento
                paid_at = None
                if data.get('paid_at'):
                    try:
                        paid_at = timezone.datetime.fromisoformat(
                            data['paid_at'].replace('Z', '+00:00')
                        )
                        fdpx_logger.info(f"  Data Pagamento (API): {paid_at}")
                    except (ValueError, AttributeError):
                        paid_at = timezone.now()
                        fdpx_logger.info(f"  Data Pagamento (fallback now): {paid_at}")
                else:
                    paid_at = timezone.now()
                    fdpx_logger.info(f"  Data Pagamento (now): {paid_at}")

                # Extrair valores financeiros
                # FastDePix envia:
                # - amount: valor bruto cobrado
                # - commission_amount: valor líquido real (o que você recebe)
                # - net_amount: NÃO usar (é valor arredondado, não é o líquido real)
                amount = data.get('amount')
                # Priorizar commission_amount (valor líquido real do FastDePix)
                amount_received = data.get('commission_amount') or data.get('amount_received') or data.get('net_amount')
                fee = data.get('fee') or data.get('tax')
                fdpx_logger.info(f"  Valor Cobrado: {amount}")
                fdpx_logger.info(f"  Valor Recebido (commission_amount): {amount_received}")
                fdpx_logger.info(f"  Taxa informada: {fee}")

                # Converter valores para Decimal se existirem
                valor_recebido = None
                valor_taxa = None
                if amount_received:
                    try:
                        valor_recebido = Decimal(str(amount_received))
                    except:
                        pass
                if fee:
                    try:
                        valor_taxa = Decimal(str(fee))
                    except:
                        pass

                # Se não veio taxa explícita mas temos amount e valor_recebido, calcular
                if valor_taxa is None and amount and valor_recebido:
                    try:
                        valor_taxa = Decimal(str(amount)) - valor_recebido
                        fdpx_logger.info(f"  Taxa calculada: {valor_taxa}")
                    except:
                        pass

                # Marcar como pago
                fdpx_logger.info("  Executando mark_as_paid()...")
                cobranca.mark_as_paid(
                    paid_at=paid_at,
                    payer_name=payer_name,
                    payer_document=payer_doc,
                    webhook_data=payload,
                    valor_recebido=valor_recebido,
                    valor_taxa=valor_taxa,
                )

                fdpx_logger.info(f"  ✓ Cobrança {transaction_id} marcada como PAGA")
                if cobranca.mensalidade:
                    fdpx_logger.info(f"  ✓ Mensalidade atualizada: pgto={cobranca.mensalidade.pgto}")
            else:
                fdpx_logger.info(f"  Cobrança já estava paga, ignorando")

        elif event_type == 'transaction.expired' or status_str == 'expired':
            fdpx_logger.info(f"  Tipo: EXPIRAÇÃO")
            cobranca.mark_as_expired()
            fdpx_logger.info(f"  ✓ Cobrança {transaction_id} marcada como EXPIRADA")

        elif event_type == 'transaction.cancelled' or status_str in ['cancelled', 'canceled']:
            fdpx_logger.info(f"  Tipo: CANCELAMENTO")
            cobranca.mark_as_cancelled()
            fdpx_logger.info(f"  ✓ Cobrança {transaction_id} marcada como CANCELADA")

        else:
            fdpx_logger.info(f"  Tipo: EVENTO DESCONHECIDO (event={event_type}, status={status_str})")

        # ========== RESPOSTA ==========
        response_data = {
            'status': 'processed',
            'cobranca_id': str(cobranca.id),
            'new_status': cobranca.status,
        }

        fdpx_logger.info("-" * 40)
        fdpx_logger.info("RESPOSTA:")
        fdpx_logger.info(json.dumps(response_data, indent=2))
        fdpx_logger.info("=" * 80)
        fdpx_logger.info("")

        return JsonResponse(response_data)

    except Exception as e:
        fdpx_logger.error("-" * 40)
        fdpx_logger.error(f"ERRO CRÍTICO: {e}")
        fdpx_logger.error("Traceback:", exc_info=True)
        fdpx_logger.info("=" * 80)
        return JsonResponse({'error': 'Internal error'}, status=500)


@login_required
@require_http_methods(["POST"])
def cancelar_cobranca_pix(request, cobranca_id):
    """
    Cancela uma cobrança PIX pendente.

    POST /api/pix/cancelar/<cobranca_id>/
    """
    from nossopainel.services.payment_integrations import get_payment_integration

    try:
        cobranca = CobrancaPix.objects.select_related(
            'conta_bancaria', 'conta_bancaria__instituicao'
        ).get(id=cobranca_id, usuario=request.user)

        if not cobranca.can_cancel:
            return JsonResponse({
                'success': False,
                'error': 'Esta cobrança não pode ser cancelada.'
            }, status=400)

        # Tentar cancelar na API
        integration = get_payment_integration(cobranca.conta_bancaria)
        if integration:
            try:
                integration.cancel_charge(cobranca.transaction_id)
            except Exception as e:
                logger.warning(f'[PIX] Erro ao cancelar na API: {e}')

        # Marcar como cancelada localmente
        cobranca.mark_as_cancelled()

        log_user_action(
            request=request,
            action=UserActionLog.ACTION_UPDATE,
            instance=cobranca,
            message="Cobrança PIX cancelada.",
            extra={"transaction_id": cobranca.transaction_id},
        )

        return JsonResponse({
            'success': True,
            'mensagem': 'Cobrança cancelada com sucesso.'
        })

    except CobrancaPix.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Cobrança não encontrada.'
        }, status=404)


# AÇÃO PARA CRIAR NOVO OBJETO DISPOSITIVO
@login_required
def create_device(request):
    dispositivos = Dispositivo.objects.filter(usuario=request.user).order_by('nome')
    usuario = request.user
    page_group = "nossopainel"
    page = "dispositivo"

    if request.method == "POST":
        nome = request.POST.get("nome")

        if nome:

            try:
                # Consultando o objeto requisitado (case-insensitive). Caso não exista, será criado com nome normalizado.
                dispositivo, created = get_or_create_dispositivo(nome, usuario)

                if created:
                    log_user_action(
                        request=request,
                        action=UserActionLog.ACTION_CREATE,
                        instance=dispositivo,
                        message="Dispositivo criado.",
                        extra={
                            "nome": dispositivo.nome,
                        },
                    )
                    return render(
                            request,
                        "pages/cadastro-dispositivo.html",
                        {
                            'dispositivos': dispositivos,
                            "success_message": "Novo Dispositivo cadastrado com sucesso!",
                        },
                    )
                
                else:
                    return render(
                            request,
                        "pages/cadastro-dispositivo.html",
                        {
                            'dispositivos': dispositivos,
                            "error_message": "Já existe um Dispositivo com este nome!",
                        },
                    )
                
            except Exception as e:
                logger.error('[%s] [USER][%s] [IP][%s] [ERRO][%s]', timezone.localtime(), request.user, get_client_ip(request) or 'N/A', e, exc_info=True)
                # Capturando outras exceções e renderizando a página novamente com a mensagem de erro
                return render(
                    request,
                    "pages/cadastro-dispositivo.html",
                    {
                        'dispositivos': dispositivos,
                        "error_message": "Não foi possível cadastrar este novo Dispositivo. Verifique os logs!",
                    },
                )

    return render(
        request, "pages/cadastro-dispositivo.html", {'dispositivos': dispositivos, "page_group": page_group, "page": page}
    )


# AÇÃO PARA CRIAR NOVO OBJETO APLICATIVO
@login_required
def create_app(request):
    aplicativos = Aplicativo.objects.filter(usuario=request.user).order_by('nome')
    usuario = request.user
    page_group = "nossopainel"
    page = "aplicativo"

    if request.method == "POST":
        nome = request.POST.get("name")
        have_mac = request.POST.get("mac")

        if nome and have_mac:
            # Verifica se o campo "mac" foi preenchido e atribui o valor correspondente
            if have_mac == "true":
                have_mac = True
            else:
                have_mac = False

            try:
                # Consultando o objeto requisitado (case-insensitive). Caso não exista, será criado com nome normalizado.
                aplicativo, created = get_or_create_aplicativo(nome, usuario, device_has_mac=have_mac)

                if created:
                    log_user_action(
                        request=request,
                        action=UserActionLog.ACTION_CREATE,
                        instance=aplicativo,
                        message="Aplicativo criado.",
                        extra={
                            "nome": aplicativo.nome,
                            "device_has_mac": aplicativo.device_has_mac,
                        },
                    )
                    return render(
                            request,
                        "pages/cadastro-aplicativo.html",
                        {
                            'aplicativos': aplicativos,
                            "success_message": "Novo Aplicativo cadastrado com sucesso!",
                        },
                    )
                
                else:
                    return render(
                            request,
                        "pages/cadastro-aplicativo.html",
                        {
                            'aplicativos': aplicativos,
                            "error_message": "Já existe um Aplicativo com este nome!",
                        },
                    )
                
            except Exception as e:
                logger.error('[%s] [USER][%s] [IP][%s] [ERRO][%s]', timezone.localtime(), request.user, get_client_ip(request) or 'N/A', e, exc_info=True)
                # Capturando outras exceções e renderizando a página novamente com a mensagem de erro
                return render(
                    request,
                    "pages/cadastro-aplicativo.html",
                    {
                        'aplicativos': aplicativos,
                        "error_message": "Não foi possível cadastrar este novo Aplicativo. Verifique os logs!",
                    },
                )

    return render(
        request, "pages/cadastro-aplicativo.html", {'aplicativos': aplicativos, "page_group": page_group, "page": page}
    )


############################################ DELETE VIEW ############################################

@login_required
def delete_app_account(request, pk):
    if request.method == "DELETE":
        try:
            conta_app = ContaDoAplicativo.objects.get(pk=pk, usuario=request.user)

            # ========== VALIDAÇÃO 1: Não pode deletar conta principal ==========
            if conta_app.is_principal:
                logger.warning(
                    f"[DELETE_APP_ACCOUNT] Tentativa de excluir conta principal (ID: {pk}) "
                    f"do cliente {conta_app.cliente.nome} (ID: {conta_app.cliente.id})"
                )
                return JsonResponse({
                    'error_message': 'Não é possível excluir a conta principal. Marque outra conta como principal antes de excluir esta.'
                }, status=400)

            # ========== VALIDAÇÃO 2: Cliente deve ter ao menos 1 conta ==========
            total_contas = ContaDoAplicativo.objects.filter(cliente=conta_app.cliente).count()
            if total_contas <= 1:
                logger.warning(
                    f"[DELETE_APP_ACCOUNT] Tentativa de excluir última conta (ID: {pk}) "
                    f"do cliente {conta_app.cliente.nome} (ID: {conta_app.cliente.id})"
                )
                return JsonResponse({
                    'error_message': 'Não é possível excluir a única conta do cliente. Todo cliente deve ter pelo menos 1 conta de aplicativo.'
                }, status=400)

            log_extra = {
                "id": conta_app.id,
                "cliente": conta_app.cliente_id,
                "app": conta_app.app_id,
                "device_id": conta_app.device_id or "",
                "email": conta_app.email or "",
            }
            conta_app.delete()
            log_user_action(
                request=request,
                action=UserActionLog.ACTION_DELETE,
                instance=conta_app,
                message="Conta de aplicativo removida.",
                extra=log_extra,
            )

            return JsonResponse({'success_message': 'deu bom'}, status=200)
        
        except Aplicativo.DoesNotExist as erro1:
            logger.error('[%s] [USER][%s] [IP][%s] [ERRO][%s]', timezone.localtime(), request.user, get_client_ip(request) or 'N/A', erro1, exc_info=True)
            error_msg = 'Você tentou excluir uma conta de aplicativo que não existe.'
            
            return JsonResponse({'error_message': 'erro'}, status=500)
        
        except ProtectedError as erro2:
            logger.error('[%s] [USER][%s] [IP][%s] [ERRO][%s]', timezone.localtime(), request.user, get_client_ip(request) or 'N/A', erro2, exc_info=True)
            error_msg = 'Essa conta de aplicativo não pôde ser excluída.'

            return JsonResponse(status=500)
    else:
        return JsonResponse({'error_message': 'erro'}, status=500)
    

@login_required
def delete_app(request, pk):
    try:
        aplicativo = Aplicativo.objects.get(pk=pk, usuario=request.user)
        log_extra = {"id": aplicativo.id, "nome": aplicativo.nome}
        aplicativo.delete()
        log_user_action(
            request=request,
            action=UserActionLog.ACTION_DELETE,
            instance=aplicativo,
            message="Aplicativo removido.",
            extra=log_extra,
        )
    except Aplicativo.DoesNotExist as erro1:
        logger.error('[%s] [USER][%s] [IP][%s] [ERRO][%s]', timezone.localtime(), request.user, get_client_ip(request) or 'N/A', erro1, exc_info=True)
        return HttpResponseNotFound(
            json.dumps({'error_delete': error_msg}), content_type='application/json'
        )
    except ProtectedError as erro2:
        logger.error('[%s] [USER][%s] [IP][%s] [ERRO][%s]', timezone.localtime(), request.user, get_client_ip(request) or 'N/A', erro2, exc_info=True)
        error_msg = 'Este Aplicativo não pode ser excluído porque está relacionado com algum cliente.'
        return HttpResponseBadRequest(
            json.dumps({'error_delete': error_msg}), content_type='application/json'
        )
    else:
        return redirect('cadastro-aplicativo')
    

@login_required
def delete_device(request, pk):
    try:
        dispositivo = Dispositivo.objects.get(pk=pk, usuario=request.user)
        log_extra = {"id": dispositivo.id, "nome": dispositivo.nome}
        dispositivo.delete()
        log_user_action(
            request=request,
            action=UserActionLog.ACTION_DELETE,
            instance=dispositivo,
            message="Dispositivo removido.",
            extra=log_extra,
        )
    except Dispositivo.DoesNotExist as erro1:
        logger.error('[%s] [USER][%s] [IP][%s] [ERRO][%s]', timezone.localtime(), request.user, get_client_ip(request) or 'N/A', erro1, exc_info=True)
        return HttpResponseNotFound(
            json.dumps({'error_delete': error_msg}), content_type='application/json'
        )
    except ProtectedError as erro2:
        logger.error('[%s] [USER][%s] [IP][%s] [ERRO][%s]', timezone.localtime(), request.user, get_client_ip(request) or 'N/A', erro2, exc_info=True)
        error_msg = 'Este Dispositivo não pode ser excluído porque está relacionado com algum cliente.'
        return HttpResponseBadRequest(
            json.dumps({'error_delete': error_msg}), content_type='application/json'
        )
    else:
        return redirect('cadastro-dispositivo')
    

@login_required
def delete_payment_method(request, pk):
    try:
        formapgto = Tipos_pgto.objects.get(pk=pk, usuario=request.user)

        # Verificar se existem clientes ATIVOS usando esta forma de pagamento
        clientes_count = Cliente.objects.filter(forma_pgto=formapgto, usuario=request.user, cancelado=False).count()
        if clientes_count > 0:
            error_msg = f'Esta Forma de Pagamento não pode ser excluída pois possui {clientes_count} cliente(s) ativo(s) associado(s). Transfira os clientes para outra forma de pagamento antes de excluir.'
            return JsonResponse({'error_delete': error_msg}, status=400)

        # Guardar referência à conta bancária antes de excluir
        conta_bancaria = formapgto.conta_bancaria

        log_extra = {"id": formapgto.id, "nome": formapgto.nome}
        formapgto.delete()
        log_user_action(
            request=request,
            action=UserActionLog.ACTION_DELETE,
            instance=formapgto,
            message="Forma de pagamento removida.",
            extra=log_extra,
        )

        # Excluir a ContaBancaria se não houver outras formas de pagamento usando ela
        if conta_bancaria:
            outras_formas = Tipos_pgto.objects.filter(conta_bancaria=conta_bancaria).count()
            if outras_formas == 0:
                conta_nome = conta_bancaria.nome_identificacao
                conta_bancaria.delete()
                log_user_action(
                    request=request,
                    action=UserActionLog.ACTION_DELETE,
                    instance=conta_bancaria,
                    message="Conta bancária removida (última forma de pagamento excluída).",
                    extra={"nome": conta_nome},
                )

        return JsonResponse({'success': True})

    except Tipos_pgto.DoesNotExist:
        error_msg = 'Forma de Pagamento não encontrada.'
        return JsonResponse({'error_delete': error_msg}, status=404)
    except Exception as e:
        logger.error('[%s] [USER][%s] [IP][%s] [ERRO][%s]',
                     timezone.localtime(), request.user, get_client_ip(request) or 'N/A', e, exc_info=True)
        error_msg = 'Erro ao excluir a Forma de Pagamento.'
        return JsonResponse({'error_delete': error_msg}, status=500)
    

@login_required
def delete_server(request, pk):
    try:
        servidor = Servidor.objects.get(pk=pk, usuario=request.user)
        log_extra = {"id": servidor.id, "nome": servidor.nome}
        servidor.delete()
        log_user_action(
            request=request,
            action=UserActionLog.ACTION_DELETE,
            instance=servidor,
            message="Servidor removido.",
            extra=log_extra,
        )
    except Servidor.DoesNotExist as erro1:
        logger.error('[%s] [USER][%s] [IP][%s] [ERRO][%s]', timezone.localtime(), request.user, get_client_ip(request) or 'N/A', erro1, exc_info=True)
        return HttpResponseNotFound(
            json.dumps({'error_delete': error_msg}), content_type='application/json'
        )
    except ProtectedError as erro2:
        logger.error('[%s] [USER][%s] [IP][%s] [ERRO][%s]', timezone.localtime(), request.user, get_client_ip(request) or 'N/A', erro2, exc_info=True)
        error_msg = 'Este Servidor não pode ser excluído porque está relacionado com algum cliente.'
        return HttpResponseBadRequest(
            json.dumps({'error_delete': error_msg}), content_type='application/json'
        )
    else:
        return redirect('cadastro-servidor')
    

@login_required
def delete_payment_plan(request, pk):
    try:
        plano_mensal = Plano.objects.get(pk=pk, usuario=request.user)
        log_extra = {"id": plano_mensal.id, "nome": plano_mensal.nome}
        plano_mensal.delete()
        log_user_action(
            request=request,
            action=UserActionLog.ACTION_DELETE,
            instance=plano_mensal,
            message="Plano removido.",
            extra=log_extra,
        )
    except Plano.DoesNotExist as erro1:
        logger.error('[%s] [USER][%s] [IP][%s] [ERRO][%s]', timezone.localtime(), request.user, get_client_ip(request) or 'N/A', erro1, exc_info=True)
        return HttpResponseNotFound(
            json.dumps({'error_delete': error_msg}), content_type='application/json'
        )
    except ProtectedError as erro2:
        logger.error('[%s] [USER][%s] [IP][%s] [ERRO][%s]', timezone.localtime(), request.user, get_client_ip(request) or 'N/A', erro2, exc_info=True)
        error_msg = 'Este Plano não pode ser excluído porque está relacionado com algum cliente.'
        return HttpResponseBadRequest(
            json.dumps({'error_delete': error_msg}), content_type='application/json'
        )

    return redirect('cadastro-plano-adesao')


######################################
############## OUTROS ################
######################################

def get_location_from_ip(ip):
    try:
        response = requests.get(f'https://ipapi.co/{ip}/json/', timeout=5)
        data = response.json()
        cidade = data.get('city', 'Desconhecida')
        pais = data.get('country_name', 'Desconhecido')
        return f"{cidade}, {pais}"
    except Exception:
        return "Localização desconhecida"


############################################
# API: Evolução do Patrimônio (JSON)
############################################

def _last_day_of_month(d: date) -> date:
    _, last = calendar.monthrange(d.year, d.month)
    return date(d.year, d.month, last)


@login_required
@require_GET
def evolucao_patrimonio(request):
    """Retorna patrimonio mensal e evolucao para o periodo solicitado (JSON)."""

    months_param_raw = (request.GET.get('months', '12') or '12').strip().lower()
    max_months = 120

    months_value: Optional[int]
    if months_param_raw == 'all':
        months_value = None
    else:
        try:
            months_value = int(months_param_raw)
        except ValueError:
            months_value = 12
        months_value = max(1, min(months_value, max_months))

    today_month = timezone.localdate().replace(day=1)
    months_list = []

    historico_qs = ClientePlanoHistorico.objects.filter(usuario=request.user)

    if months_value is None:
        first_inicio = historico_qs.aggregate(Min('inicio'))['inicio__min']
        if first_inicio:
            if isinstance(first_inicio, datetime):
                first_inicio = first_inicio.date()
            start_month = date(first_inicio.year, first_inicio.month, 1)
            if start_month > today_month:
                start_month = today_month
            cursor = start_month
            end_month = today_month
            while cursor <= end_month:
                months_list.append(cursor)
                if cursor.month == 12:
                    cursor = date(cursor.year + 1, 1, 1)
                else:
                    cursor = date(cursor.year, cursor.month + 1, 1)
        else:
            months_value = 12

    if months_value is not None and not months_list:
        cursor = today_month
        for _ in range(months_value):
            months_list.append(cursor)
            cursor = (cursor - timedelta(days=1)).replace(day=1)
        months_list.reverse()

    if not months_list:
        months_list = [today_month]

    categorias = []
    patrim = []

    for m in months_list:
        ref = _last_day_of_month(m)
        # ativos (sem fim) até a data de referência
        ativos = historico_qs.filter(
            inicio__lte=ref,
            fim__isnull=True,
        ).values_list('valor_plano', flat=True)
        # encerrados mas ainda ativos na data de referência (fim >= ref)
        encerrados = historico_qs.filter(
            inicio__lte=ref,
            fim__gte=ref,
        ).values_list('valor_plano', flat=True)

        total = sum((v or 0) for v in list(ativos) + list(encerrados))
        patrim.append(float(total))
        categorias.append(f"{m.month:02d}/{str(m.year)[2:]}")

    evol = []
    prev = None
    for val in patrim:
        if prev is None:
            evol.append(0.0)
        else:
            evol.append(float(val - prev))
        prev = val

    return JsonResponse({
        'categories': categorias,
        'series': [
            {'name': 'Patrimônio', 'data': patrim},
            {'name': 'Evolução', 'data': evol},
        ],
    })


@login_required
@require_GET
def api_receita_anual(request):
    """
    Retorna a receita anual por forma de pagamento.

    - Ano atual: PROJEÇÃO baseada nos clientes ativos e seus planos
    - Anos anteriores: DADOS REAIS das mensalidades pagas

    GET /api/receita-anual/
    Parâmetros:
        - periodo: 'current' (ano atual, default), 'last5' (últimos 5 anos), 'all' (todos os anos)

    Retorna JSON com:
        - categories: lista de anos
        - series: lista de formas de pagamento com valores por ano
        - summary: resumo total por forma de pagamento
        - meta: informações adicionais
    """
    from collections import defaultdict

    periodo = (request.GET.get('periodo', 'current') or 'current').strip().lower()
    if periodo not in {'current', 'last5', 'all'}:
        periodo = 'current'

    usuario = request.user
    hoje = timezone.localdate()
    ano_atual = hoje.year

    # Mapeamento de tipo de plano para quantidade de pagamentos por ano
    PAGAMENTOS_POR_ANO = {
        'Mensal': 12,
        'Bimestral': 6,
        'Trimestral': 4,
        'Semestral': 2,
        'Anual': 1,
    }

    # Definir range de anos baseado no período
    if periodo == 'current':
        anos = [ano_atual]
    elif periodo == 'last5':
        anos = list(range(ano_atual - 4, ano_atual + 1))
    else:  # 'all'
        # Buscar ano mais antigo de pagamento ou adesão
        primeiro_pagamento = Mensalidade.objects.filter(
            usuario=usuario,
            pgto=True,
            dt_pagamento__isnull=False
        ).aggregate(min_ano=Min('dt_pagamento__year'))['min_ano']

        primeiro_adesao = Cliente.objects.filter(
            usuario=usuario
        ).aggregate(min_ano=Min('data_adesao__year'))['min_ano']

        primeiro_ano = min(filter(None, [primeiro_pagamento, primeiro_adesao, ano_atual]))

        if primeiro_ano:
            anos = list(range(primeiro_ano, ano_atual + 1))
        else:
            anos = [ano_atual]

    # Estrutura para acumular dados
    # {forma_pgto_id: {ano: valor_total, 'nome': '...', 'instituicao': '...'}}
    dados_por_forma = defaultdict(lambda: {'nome': '', 'instituicao': '', 'anos': defaultdict(Decimal)})

    # Cache de formas de pagamento para evitar queries repetidas
    formas_cache = {}

    def get_forma_info(forma):
        """Retorna nome e instituição formatados para uma forma de pagamento."""
        if forma.id in formas_cache:
            return formas_cache[forma.id]

        if forma.conta_bancaria and forma.conta_bancaria.instituicao:
            if forma.conta_bancaria.instituicao.tipo_integracao == 'fastdepix':
                nome = forma.conta_bancaria.nome_identificacao or forma.nome
            else:
                nome = forma.nome
            instituicao = forma.conta_bancaria.instituicao.nome
        elif forma.dados_bancarios and forma.dados_bancarios.instituicao:
            nome = forma.nome
            instituicao = forma.dados_bancarios.instituicao
        else:
            nome = forma.nome
            instituicao = ''

        formas_cache[forma.id] = (nome, instituicao)
        return nome, instituicao

    # ========== ANOS ANTERIORES: DADOS REAIS ==========
    anos_anteriores = [a for a in anos if a < ano_atual]

    if anos_anteriores:
        # Buscar mensalidades pagas em anos anteriores
        mensalidades_pagas = Mensalidade.objects.filter(
            usuario=usuario,
            pgto=True,
            dt_pagamento__year__in=anos_anteriores
        ).select_related(
            'forma_pgto',
            'forma_pgto__conta_bancaria',
            'forma_pgto__conta_bancaria__instituicao',
            'forma_pgto__dados_bancarios',
            'cliente',
            'cliente__forma_pgto',
            'cliente__forma_pgto__conta_bancaria',
            'cliente__forma_pgto__conta_bancaria__instituicao',
            'cliente__forma_pgto__dados_bancarios'
        )

        for mensalidade in mensalidades_pagas:
            # Usar forma_pgto da mensalidade (histórico) ou fallback para cliente
            forma = mensalidade.forma_pgto or (mensalidade.cliente.forma_pgto if mensalidade.cliente else None)

            if not forma:
                continue

            ano = mensalidade.dt_pagamento.year
            nome, instituicao = get_forma_info(forma)

            dados_por_forma[forma.id]['nome'] = nome
            dados_por_forma[forma.id]['instituicao'] = instituicao
            dados_por_forma[forma.id]['anos'][ano] += mensalidade.valor

    # ========== ANO ATUAL: PROJEÇÃO ==========
    if ano_atual in anos:
        # Buscar clientes ativos com forma de pagamento
        clientes = Cliente.objects.filter(
            usuario=usuario,
            cancelado=False
        ).select_related(
            'plano',
            'forma_pgto',
            'forma_pgto__conta_bancaria',
            'forma_pgto__conta_bancaria__instituicao',
            'forma_pgto__dados_bancarios'
        )

        for cliente in clientes:
            if not cliente.plano or not cliente.forma_pgto:
                continue

            forma = cliente.forma_pgto
            nome, instituicao = get_forma_info(forma)

            dados_por_forma[forma.id]['nome'] = nome
            dados_por_forma[forma.id]['instituicao'] = instituicao

            # Calcular receita anual projetada
            valor_plano = cliente.plano.valor
            tipo_plano = cliente.plano.nome
            pagamentos_ano = PAGAMENTOS_POR_ANO.get(tipo_plano, 12)
            receita_anual = valor_plano * pagamentos_ano

            dados_por_forma[forma.id]['anos'][ano_atual] += receita_anual

    # Construir series para o gráfico (ordenado por valor total desc)
    series = []
    summary_list = []

    # Calcular total por forma para ordenação
    totais_forma = []
    for forma_id, data in dados_por_forma.items():
        total = sum(data['anos'].values())
        totais_forma.append((forma_id, data, float(total)))

    # Ordenar por total (maior para menor)
    totais_forma.sort(key=lambda x: x[2], reverse=True)

    # Cores para o gráfico (barras empilhadas)
    CORES = [
        '#624bff', '#10b981', '#f59e0b', '#ef4444', '#8b5cf6',
        '#06b6d4', '#ec4899', '#14b8a6', '#f97316', '#6366f1',
        '#84cc16', '#22d3d8'
    ]

    for idx, (forma_id, data, total) in enumerate(totais_forma):
        nome_display = data['nome']
        if data['instituicao']:
            nome_display = f"{data['nome']} ({data['instituicao']})"

        # Valores por ano
        valores = [float(data['anos'].get(ano, 0)) for ano in anos]

        series.append({
            'label': nome_display,
            'data': valores,
            'backgroundColor': CORES[idx % len(CORES)],
        })

        # Summary: valor conforme o período selecionado
        if periodo == 'current':
            # Apenas ano atual (projeção)
            total_summary = float(data['anos'].get(ano_atual, 0))
        else:
            # Períodos com múltiplos anos: soma de todos os anos
            total_summary = total

        if total_summary > 0:
            summary_list.append({
                'nome': nome_display,
                'valor': total_summary,
            })

    # Formatar categorias (anos como string) com indicador de projeção
    categories = []
    for ano in anos:
        if ano == ano_atual:
            categories.append(f"{ano} (projeção)")
        else:
            categories.append(str(ano))

    # Calcular total geral (ano atual para summary)
    total_geral = sum(item['valor'] for item in summary_list)

    return JsonResponse({
        'categories': categories,
        'series': series,
        'summary': summary_list,
        'meta': {
            'periodo': periodo,
            'ano_atual': ano_atual,
            'total_geral': total_geral,
            'total_formas': len(summary_list),
            'info': 'Anos anteriores: dados reais | Ano atual: projeção'
        }
    })


@csrf_exempt
@require_POST
def internal_send_whatsapp(request):
    """
    Endpoint interno para envio de notificações WhatsApp ao admin.

    Restrito à rede interna (database-network) via InternalAPIMiddleware.
    Usado por scripts cron do servidor MySQL para enviar alertas ao admin.

    O número de destino é fixo (MEU_NUM_TIM do .env).

    Payload JSON esperado:
    {
        "mensagem": "Texto da notificação",
        "tipo": "backup|status|alert|info"  # Opcional, para logging (default: "unknown")
    }

    Segurança:
    - IP-restricted via InternalAPIMiddleware (apenas database-network)
    - CSRF exempt (seguro pois POST-only + IP restrito)
    - Sem autenticação de usuário (não necessária na rede interna)
    - Logging completo de todas as requisições

    Returns:
        JSON: {'success': bool, 'status_code': int, 'response': dict}
    """
    import requests
    from nossopainel.services.logging import append_line

    # Diretório de logs
    log_path = Path("logs/MySQL_Triggers/whatsapp_notifications.log")
    log_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        # Parse do payload JSON
        data = json.loads(request.body.decode('utf-8'))
        mensagem = data.get('mensagem')
        tipo = data.get('tipo', 'unknown')

        # Validação de campo obrigatório
        if not mensagem:
            error_msg = "[ERROR] Campo obrigatório 'mensagem' não fornecido"
            append_line(str(log_path), error_msg)
            return JsonResponse({
                'success': False,
                'error': 'Campo obrigatório: mensagem'
            }, status=400)

        # Busca telefone fixo do admin no .env
        telefone_admin = os.getenv('MEU_NUM_TIM')
        if not telefone_admin:
            error_msg = "[ERROR] Variável MEU_NUM_TIM não configurada no .env"
            append_line(str(log_path), error_msg)
            return JsonResponse({
                'success': False,
                'error': 'Telefone admin não configurado (MEU_NUM_TIM)'
            }, status=500)

        # Busca sessão WhatsApp ativa do usuário admin (id=1)
        try:
            user_admin = User.objects.get(id=1)
            sessao = SessaoWpp.objects.filter(
                usuario=user_admin.username,
                is_active=True
            ).order_by('-dt_inicio').first()

            if not sessao:
                error_msg = f"[ERROR] Nenhuma sessão WhatsApp ativa para usuário {user_admin.username}"
                append_line(str(log_path), error_msg)
                return JsonResponse({
                    'success': False,
                    'error': 'Sessão WhatsApp não encontrada ou inativa'
                }, status=503)

        except User.DoesNotExist:
            error_msg = "[ERROR] Usuário admin (ID=1) não encontrado no banco"
            append_line(str(log_path), error_msg)
            return JsonResponse({
                'success': False,
                'error': 'Usuário admin não encontrado'
            }, status=500)

        # Prepara requisição para WPPConnect API
        url_api_wpp = os.getenv('URL_API_WPP', os.getenv('URL_API', 'http://api.nossopainel.com.br/api'))
        url = f"{url_api_wpp}/{sessao.usuario}/send-message"

        headers = {
            'Content-Type': 'application/json',
            'Accept': 'application/json',
            'Authorization': f'Bearer {sessao.token}'
        }

        body = {
            'phone': telefone_admin,
            'message': mensagem,
            'isGroup': False
        }

        # Envia mensagem via WPPConnect
        response = requests.post(url, headers=headers, json=body, timeout=30)

        # Registra resultado no log
        timestamp = timezone.now().strftime('%Y-%m-%d %H:%M:%S')
        client_ip = request.META.get('REMOTE_ADDR', 'unknown')

        # Sanitiza resposta para evitar registrar HTML de páginas de erro
        sanitized_resp = _sanitize_response(response.text, max_length=200)
        log_entry = (
            f"[{timestamp}] "
            f"Tipo: {tipo} | "
            f"IP: {client_ip} | "
            f"Telefone: {telefone_admin} | "
            f"Status: {response.status_code} | "
            f"Response: {sanitized_resp}"
        )
        append_line(str(log_path), log_entry)

        # Retorna resposta baseada no status code
        if 200 <= response.status_code < 300:
            return JsonResponse({
                'success': True,
                'status_code': response.status_code,
                'response': response.json() if response.text else {}
            }, status=200)
        else:
            return JsonResponse({
                'success': False,
                'status_code': response.status_code,
                'response': response.json() if response.text else {'error': 'Resposta vazia da API'},
                'error': f'API retornou status {response.status_code}'
            }, status=response.status_code)

    except json.JSONDecodeError as e:
        error_msg = f"[ERROR] JSON inválido no body da requisição: {str(e)}"
        append_line(str(log_path), error_msg)
        return JsonResponse({
            'success': False,
            'error': 'Payload JSON inválido'
        }, status=400)

    except requests.Timeout:
        error_msg = f"[ERROR] Timeout ao conectar com API WhatsApp (tipo: {tipo})"
        append_line(str(log_path), error_msg)
        return JsonResponse({
            'success': False,
            'error': 'Timeout ao conectar com API WhatsApp'
        }, status=504)

    except requests.RequestException as e:
        error_msg = f"[ERROR] Erro na requisição HTTP: {str(e)}"
        append_line(str(log_path), error_msg)
        return JsonResponse({
            'success': False,
            'error': f'Erro na requisição: {str(e)}'
        }, status=500)

    except Exception as e:
        error_msg = f"[ERROR] Erro inesperado: {type(e).__name__} - {str(e)}"
        append_line(str(log_path), error_msg)
        return JsonResponse({
            'success': False,
            'error': f'Erro interno: {str(e)}'
        }, status=500)


########################################
# MIGRAÇÃO DE CLIENTES ENTRE USUÁRIOS #
########################################

class SuperuserRequiredMixin(UserPassesTestMixin):
    """Mixin que permite acesso apenas para superusuários"""
    def test_func(self):
        return self.request.user.is_superuser

    def handle_no_permission(self):
        return JsonResponse({
            'error': 'Acesso negado. Apenas superusuários podem acessar esta funcionalidade.'
        }, status=403)


class MigrationClientesListView(LoginRequiredMixin, SuperuserRequiredMixin, View):
    """
    View para listar clientes de um usuário específico para seleção na migração.

    Retorna dados formatados para popular DataTable com informações completas
    dos clientes (nome, servidor, status, plano, telefone, etc.)
    """

    def get(self, request):
        """Retorna lista de clientes do usuário especificado"""
        usuario_origem_id = request.GET.get('usuario_origem_id')

        if not usuario_origem_id:
            return JsonResponse({
                'error': 'ID do usuário de origem não informado.'
            }, status=400)

        try:
            usuario_origem = User.objects.get(id=usuario_origem_id)
        except User.DoesNotExist:
            return JsonResponse({
                'error': 'Usuário de origem não encontrado.'
            }, status=404)

        # Buscar clientes do usuário
        clientes = Cliente.objects.filter(
            usuario=usuario_origem
        ).select_related(
            'servidor', 'dispositivo', 'sistema', 'forma_pgto', 'plano', 'indicado_por'
        ).order_by('-data_adesao')

        # Serializar dados para DataTable
        clientes_data = []
        for cliente in clientes:
            clientes_data.append({
                'id': cliente.id,
                'nome': cliente.nome,
                'telefone': cliente.telefone or '-',
                'servidor': cliente.servidor.nome if cliente.servidor else '-',
                'dispositivo': cliente.dispositivo.nome if cliente.dispositivo else '-',
                'sistema': cliente.sistema.nome if cliente.sistema else '-',
                'plano': cliente.plano.nome if cliente.plano else '-',
                'plano_valor': float(cliente.plano.valor) if cliente.plano else 0,
                'forma_pgto': cliente.forma_pgto.nome if cliente.forma_pgto else '-',
                'status': 'Cancelado' if cliente.cancelado else 'Ativo',
                'status_class': 'danger' if cliente.cancelado else 'success',
                'data_cadastro': cliente.data_adesao.strftime('%d/%m/%Y') if cliente.data_adesao else '-',
                'data_cancelamento': cliente.data_cancelamento.strftime('%d/%m/%Y') if cliente.data_cancelamento else '-',
                'indicado_por': cliente.indicado_por.nome if cliente.indicado_por else '-',
                'uf': cliente.uf or '-',
            })

        return JsonResponse({
            'success': True,
            'clientes': clientes_data,
            'total': len(clientes_data),
            'usuario_origem': {
                'id': usuario_origem.id,
                'username': usuario_origem.username,
                'nome': usuario_origem.get_full_name() or usuario_origem.username,
            }
        })


class MigrationValidationView(LoginRequiredMixin, SuperuserRequiredMixin, View):
    """
    View para validar migração de clientes e retornar resumo detalhado.

    Executa todas as validações (indicações, descontos, entidades) e retorna
    estatísticas completas sobre o que será migrado.
    """

    def post(self, request):
        """Valida a migração e retorna resumo"""
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({
                'error': 'JSON inválido no corpo da requisição.'
            }, status=400)

        usuario_origem_id = data.get('usuario_origem_id')
        usuario_destino_id = data.get('usuario_destino_id')
        clientes_ids = data.get('clientes_ids', [])

        # Validações básicas
        if not usuario_origem_id or not usuario_destino_id:
            return JsonResponse({
                'error': 'IDs de usuário de origem e destino são obrigatórios.'
            }, status=400)

        if usuario_origem_id == usuario_destino_id:
            return JsonResponse({
                'error': 'Usuário de origem e destino não podem ser iguais.'
            }, status=400)

        if not clientes_ids:
            return JsonResponse({
                'error': 'Nenhum cliente selecionado para migração.'
            }, status=400)

        try:
            usuario_origem = User.objects.get(id=usuario_origem_id)
            usuario_destino = User.objects.get(id=usuario_destino_id)
        except User.DoesNotExist:
            return JsonResponse({
                'error': 'Usuário de origem ou destino não encontrado.'
            }, status=404)

        # Importar serviço de migração
        from nossopainel.services.migration_service import (
            ClientMigrationService,
            MigrationValidationError
        )

        # Executar validação
        service = ClientMigrationService(usuario_origem, usuario_destino)

        try:
            validation_result = service.validate_migration(clientes_ids)

            return JsonResponse({
                'success': True,
                'validation': validation_result,
                'usuario_origem': {
                    'id': usuario_origem.id,
                    'username': usuario_origem.username,
                    'nome': usuario_origem.get_full_name() or usuario_origem.username,
                },
                'usuario_destino': {
                    'id': usuario_destino.id,
                    'username': usuario_destino.username,
                    'nome': usuario_destino.get_full_name() or usuario_destino.username,
                }
            })

        except MigrationValidationError as e:
            return JsonResponse({
                'success': False,
                'error': str(e),
                'validation_errors': service.validation_errors,
            }, status=400)

        except Exception as e:
            logger.exception(
                "Erro inesperado ao validar migração (user=%s, origem=%s, destino=%s)",
                request.user.username,
                usuario_origem_id,
                usuario_destino_id,
            )
            return JsonResponse({
                'success': False,
                'error': f'Erro interno ao validar migração: {str(e)}'
            }, status=500)


class MigrationExecuteView(LoginRequiredMixin, SuperuserRequiredMixin, View):
    """
    View para executar a migração de clientes entre usuários.

    Executa a migração de forma transacional (rollback em caso de erro)
    e registra todas as operações em UserActionLog.
    """

    def post(self, request):
        """Executa a migração de clientes"""
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({
                'error': 'JSON inválido no corpo da requisição.'
            }, status=400)

        usuario_origem_id = data.get('usuario_origem_id')
        usuario_destino_id = data.get('usuario_destino_id')
        clientes_ids = data.get('clientes_ids', [])

        # Validações básicas
        if not usuario_origem_id or not usuario_destino_id:
            return JsonResponse({
                'error': 'IDs de usuário de origem e destino são obrigatórios.'
            }, status=400)

        if usuario_origem_id == usuario_destino_id:
            return JsonResponse({
                'error': 'Usuário de origem e destino não podem ser iguais.'
            }, status=400)

        if not clientes_ids:
            return JsonResponse({
                'error': 'Nenhum cliente selecionado para migração.'
            }, status=400)

        try:
            usuario_origem = User.objects.get(id=usuario_origem_id)
            usuario_destino = User.objects.get(id=usuario_destino_id)
        except User.DoesNotExist:
            return JsonResponse({
                'error': 'Usuário de origem ou destino não encontrado.'
            }, status=404)

        # Importar serviço de migração
        from nossopainel.services.migration_service import (
            ClientMigrationService,
            MigrationValidationError
        )

        # Executar migração
        service = ClientMigrationService(usuario_origem, usuario_destino)

        try:
            # Executar migração (transacional)
            result = service.execute_migration(clientes_ids)

            # Registrar log da operação
            UserActionLog.objects.create(
                usuario=request.user,
                acao='migration',
                entidade='Cliente',
                objeto_id='',  # String vazia ao invés de None
                objeto_repr=f'{result["stats"]["clientes_migrados"]} clientes',
                mensagem=f'Migração de {result["stats"]["clientes_migrados"]} clientes de '
                         f'{usuario_origem.username} para {usuario_destino.username}',
                extras={
                    'usuario_origem_id': usuario_origem.id,
                    'usuario_destino_id': usuario_destino.id,
                    'clientes_ids': clientes_ids,
                    'stats': result['stats'],
                    'entities_created': result['entities_created'],
                },
                ip=get_client_ip(request),
                request_path=request.path,
            )

            logger.info(
                "Migração de clientes concluída com sucesso (admin=%s, origem=%s, destino=%s, clientes=%d)",
                request.user.username,
                usuario_origem.username,
                usuario_destino.username,
                result['stats']['clientes_migrados'],
            )

            return JsonResponse({
                'success': True,
                'result': result,
                'message': f'{result["stats"]["clientes_migrados"]} clientes migrados com sucesso!',
            })

        except MigrationValidationError as e:
            logger.warning(
                "Erro de validação ao executar migração (admin=%s, origem=%s, destino=%s): %s",
                request.user.username,
                usuario_origem_id,
                usuario_destino_id,
                str(e),
            )
            return JsonResponse({
                'success': False,
                'error': str(e),
            }, status=400)

        except Exception as e:
            logger.exception(
                "Erro inesperado ao executar migração (admin=%s, origem=%s, destino=%s)",
                request.user.username,
                usuario_origem_id,
                usuario_destino_id,
            )
            return JsonResponse({
                'success': False,
                'error': f'Erro interno ao executar migração: {str(e)}'
            }, status=500)


# ==================== VIEWS: GESTÃO DE DOMÍNIOS DNS (RESELLER AUTOMATION) ====================

@login_required
def gestao_dns_page(request):
    """
    Página principal de Gestão de Domínios DNS para automação reseller.

    Permite ao usuário:
    - Selecionar aplicativo (apenas os que suportam automação)
    - Fazer login manual no painel reseller (se necessário)
    - Configurar e iniciar migração DNS
    - Visualizar progresso em tempo real
    - Consultar histórico de tarefas

    Template: templates/pages/gestao-dns.html
    """
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    from django.db.models import Case, When, Value, IntegerField

    # Busca apenas aplicativos que usam MAC e foram cadastrados pelo superuser
    # Ordenação: aplicativos com automação primeiro (DreamTV = 0), depois alfabético
    aplicativos = Aplicativo.objects.filter(
        device_has_mac=True,
        usuario__is_superuser=True
    ).annotate(
        prioridade=Case(
            When(nome__iexact='dreamtv', then=Value(0)),
            default=Value(1),
            output_field=IntegerField()
        )
    ).order_by('prioridade', 'nome')

    # Contas reseller do usuário
    contas = ContaReseller.objects.filter(usuario=request.user).select_related('aplicativo')

    # Verifica se foi solicitada uma tarefa específica
    tarefa_id = request.GET.get('tarefa_id')
    tarefa_selecionada = None
    dispositivos_tarefa = None

    if tarefa_id:
        try:
            tarefa_selecionada = TarefaMigracaoDNS.objects.get(
                id=tarefa_id,
                usuario=request.user
            )
            # Busca dispositivos da tarefa COM PAGINAÇÃO (10 por página)
            dispositivos_list = tarefa_selecionada.dispositivos.all().order_by('id')

            dispositivos_paginator = Paginator(dispositivos_list, 10)
            dispositivos_page = request.GET.get('dispositivos_page', 1)

            try:
                dispositivos_tarefa = dispositivos_paginator.page(dispositivos_page)
            except PageNotAnInteger:
                dispositivos_tarefa = dispositivos_paginator.page(1)
            except EmptyPage:
                dispositivos_tarefa = dispositivos_paginator.page(dispositivos_paginator.num_pages)
        except TarefaMigracaoDNS.DoesNotExist:
            pass

    # Tarefas recentes com paginação (10 por página)
    tarefas_list = TarefaMigracaoDNS.objects.filter(
        usuario=request.user
    ).select_related('aplicativo').order_by('-criada_em')

    paginator = Paginator(tarefas_list, 10)  # 10 tarefas por página
    page = request.GET.get('page', 1)

    try:
        tarefas_recentes = paginator.page(page)
    except PageNotAnInteger:
        tarefas_recentes = paginator.page(1)
    except EmptyPage:
        tarefas_recentes = paginator.page(paginator.num_pages)

    context = {
        'aplicativos': aplicativos,
        'contas': contas,
        'tarefas_recentes': tarefas_recentes,
        'tarefa_selecionada': tarefa_selecionada,
        'dispositivos_tarefa': dispositivos_tarefa,
        'page_title': 'Gestão de Domínios DNS',
    }

    return render(request, 'pages/gestao-dns.html', context)


@login_required
def obter_dispositivos_paginados_api(request):
    """
    API para obter dispositivos de uma tarefa de migração com paginação AJAX.

    GET Params:
        tarefa_id: ID da tarefa (obrigatório)
        page: Número da página (default: 1)
        status_filter: Filtro de status ('all', 'sucesso', 'erro', 'pulado') (default: 'all')

    Returns:
        JSON:
            - success: Boolean
            - dispositivos: Lista de dispositivos (dict)
            - pagination: Informações de paginação
            - estatisticas: Estatísticas da tarefa
    """
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

    try:
        tarefa_id = request.GET.get('tarefa_id')
        page = request.GET.get('page', 1)
        status_filter = request.GET.get('status_filter', 'all')

        if not tarefa_id:
            return JsonResponse({
                'success': False,
                'error': 'tarefa_id é obrigatório'
            }, status=400)

        # Busca tarefa (apenas do usuário logado)
        try:
            tarefa = TarefaMigracaoDNS.objects.get(
                id=tarefa_id,
                usuario=request.user
            )
        except TarefaMigracaoDNS.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Tarefa não encontrada'
            }, status=404)

        # Filtra dispositivos por status (se não for 'all')
        dispositivos_queryset = tarefa.dispositivos.all().order_by('id')

        if status_filter != 'all':
            dispositivos_queryset = dispositivos_queryset.filter(status=status_filter)

        # Paginação (10 por página)
        paginator = Paginator(dispositivos_queryset, 10)

        try:
            dispositivos_page = paginator.page(page)
        except PageNotAnInteger:
            dispositivos_page = paginator.page(1)
        except EmptyPage:
            dispositivos_page = paginator.page(paginator.num_pages)

        # Serializa dispositivos
        dispositivos_data = []
        for dispositivo in dispositivos_page:
            dispositivos_data.append({
                'device_id': dispositivo.device_id,
                'nome_dispositivo': dispositivo.nome_dispositivo or '-',
                'status': dispositivo.status,
                'dns_encontrado': extrair_dominio_de_url(dispositivo.dns_encontrado) if dispositivo.dns_encontrado else '-',
                'dns_atualizado': extrair_dominio_de_url(dispositivo.dns_atualizado) if dispositivo.dns_atualizado else '-',
                'mensagem_erro': dispositivo.mensagem_erro or ''
            })

        # Informações de paginação
        pagination_data = {
            'current_page': dispositivos_page.number,
            'total_pages': paginator.num_pages,
            'total_items': paginator.count,
            'has_previous': dispositivos_page.has_previous(),
            'has_next': dispositivos_page.has_next(),
            'previous_page': dispositivos_page.previous_page_number() if dispositivos_page.has_previous() else None,
            'next_page': dispositivos_page.next_page_number() if dispositivos_page.has_next() else None,
            'page_range': list(paginator.page_range)
        }

        # Estatísticas da tarefa (sempre baseadas no total, não no filtro)
        estatisticas = {
            'total': tarefa.total_dispositivos,
            'sucessos': tarefa.sucessos,
            'falhas': tarefa.falhas,
            'pulados': tarefa.pulados,
            'processados': tarefa.processados
        }

        return JsonResponse({
            'success': True,
            'dispositivos': dispositivos_data,
            'pagination': pagination_data,
            'estatisticas': estatisticas,
            'status_filter_ativo': status_filter
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Erro interno: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(['GET', 'POST'])
def verificar_conta_reseller_api(request):
    """
    API para verificar se usuário possui conta reseller válida para o aplicativo.

    GET/POST Params:
        aplicativo_id: ID do aplicativo

    Returns:
        JSON:
            - status: 'sem_conta' | 'sessao_expirada' | 'ok'
            - email: Email da conta (se existir)
            - ultimo_login: Data/hora do último login (se existir)
            - sessao_valida: Boolean
            - mensagem: Mensagem descritiva
    """
    try:
        # Aceita tanto GET quanto POST (GET usado no polling de login)
        aplicativo_id = request.POST.get('aplicativo_id') or request.GET.get('aplicativo_id')

        if not aplicativo_id:
            return JsonResponse({
                'success': False,
                'error': 'aplicativo_id é obrigatório'
            }, status=400)

        # Busca aplicativo
        try:
            aplicativo = Aplicativo.objects.get(id=aplicativo_id)
        except Aplicativo.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Aplicativo não encontrado'
            }, status=404)

        # Verifica se aplicativo tem automação implementada
        # NOTA: Por enquanto, apenas DreamTV tem automação
        if aplicativo.nome.lower() != 'dreamtv':
            return JsonResponse({
                'status': 'nao_implementado',
                'mensagem': f'Automação ainda não implementada para {aplicativo.nome}.'
            })

        # Busca conta reseller
        try:
            conta = ContaReseller.objects.get(
                usuario=request.user,
                aplicativo=aplicativo
            )
        except ContaReseller.DoesNotExist:
            return JsonResponse({
                'status': 'sem_conta',
                'mensagem': 'Você ainda não possui credenciais salvas para este aplicativo.'
            })

        # Verifica se sessão ainda é válida
        from nossopainel.services.reseller_automation import DreamTVSeleniumAutomation

        service = DreamTVSeleniumAutomation(user=request.user, aplicativo=aplicativo)
        sessao_valida = service.verificar_sessao_valida()

        if not sessao_valida:
            return JsonResponse({
                'status': 'sessao_expirada',
                'email': conta.email_login,
                'ultimo_login': conta.ultimo_login.isoformat() if conta.ultimo_login else None,
                'sessao_valida': False,
                'login_progresso': conta.login_progresso,  # Progresso do login em andamento
                'mensagem': 'Sua sessão expirou. Faça login novamente.'
            })

        return JsonResponse({
            'status': 'ok',
            'email': conta.email_login,
            'ultimo_login': conta.ultimo_login.isoformat() if conta.ultimo_login else None,
            'sessao_valida': True,
            'login_progresso': conta.login_progresso,  # Progresso do login
            'mensagem': 'Conta autenticada e sessão válida.'
        })

    except Exception as e:
        logger.exception(f"Erro ao verificar conta reseller: {e}")
        return JsonResponse({
            'success': False,
            'error': f'Erro interno: {str(e)}'
        }, status=500)


@login_required
@require_POST
def iniciar_login_manual_api(request):
    """
    API para iniciar processo de login automático no painel reseller.

    Utiliza CapSolver para resolver reCAPTCHA automaticamente.
    O navegador pode ser visível ou headless dependendo da configuração de debug.

    POST Params:
        aplicativo_id: ID do aplicativo
        email: Email/usuário para login
        senha: Senha para login

    Returns:
        JSON:
            - status: 'login_iniciado' | 'erro'
            - mensagem: Mensagem descritiva
    """
    try:
        aplicativo_id = request.POST.get('aplicativo_id')
        email = request.POST.get('email')
        senha = request.POST.get('senha')

        # Validações
        if not all([aplicativo_id, email, senha]):
            return JsonResponse({
                'success': False,
                'error': 'Campos obrigatórios: aplicativo_id, email, senha'
            }, status=400)

        # Busca aplicativo
        try:
            aplicativo = Aplicativo.objects.get(id=aplicativo_id)
        except Aplicativo.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Aplicativo não encontrado'
            }, status=404)

        # Verifica se aplicativo tem automação
        if aplicativo.nome.lower() != 'dreamtv':
            return JsonResponse({
                'success': False,
                'error': f'Automação não implementada para {aplicativo.nome}'
            }, status=400)

        # Salva ou atualiza credenciais (senha será criptografada)
        from nossopainel.utils import encrypt_password

        conta, created = ContaReseller.objects.update_or_create(
            usuario=request.user,
            aplicativo=aplicativo,
            defaults={
                'email_login': email,
                'senha_login': encrypt_password(senha),
                'sessao_valida': False,  # Ainda não logou
            }
        )

        logger.info(
            f"[USER:{request.user.username}] Credenciais salvas para {aplicativo.nome} "
            f"({'criadas' if created else 'atualizadas'})"
        )

        # Inicia login automático em thread separada
        def fazer_login_thread():
            try:
                from nossopainel.services.reseller_automation import DreamTVSeleniumAutomation

                service = DreamTVSeleniumAutomation(user=request.user, aplicativo=aplicativo)

                logger.info(
                    f"[USER:{request.user.username}] Iniciando login AUTOMÁTICO com CapSolver"
                )
                sucesso = service.fazer_login_automatico()

                if sucesso:
                    logger.info(
                        f"[USER:{request.user.username}] Login automático concluído com sucesso"
                    )
                else:
                    logger.warning(
                        f"[USER:{request.user.username}] Login automático falhou (timeout ou erro)"
                    )

            except Exception as e:
                logger.exception(
                    f"[USER:{request.user.username}] Erro na thread de login automático: {e}"
                )

        thread = threading.Thread(target=fazer_login_thread, daemon=True)
        thread.start()

        return JsonResponse({
            'status': 'login_iniciado',
            'mensagem': 'Automação iniciada. O sistema resolverá o reCAPTCHA automaticamente.',
        })

    except Exception as e:
        logger.exception(f"Erro ao iniciar login automático: {e}")
        return JsonResponse({
            'success': False,
            'error': f'Erro interno: {str(e)}'
        }, status=500)


@login_required
@require_POST
def iniciar_migracao_dns_api(request):
    """
    API para iniciar tarefa de migração DNS em background.

    Valida inputs, cria tarefa no banco e inicia execução em thread separada.

    POST Params:
        aplicativo_id: ID do aplicativo
        tipo_migracao: 'todos' | 'especifico'
        mac_alvo: MAC Address (obrigatório se tipo='especifico')
        dominio_origem: Domínio DNS atual (protocolo+host+porta, ex: http://dominio.com:8080)
        dominio_destino: Novo domínio DNS (protocolo+host+porta)

    Returns:
        JSON:
            - status: 'iniciado' | 'erro'
            - tarefa_id: ID da tarefa criada (para polling)
            - mensagem: Mensagem descritiva
    """
    from nossopainel.utils import validar_formato_dominio, extrair_dominio_de_url

    try:
        # Parse dados
        aplicativo_id = request.POST.get('aplicativo_id')
        tipo_migracao = request.POST.get('tipo_migracao')
        mac_alvo = request.POST.get('mac_alvo', '').strip()
        dominio_origem = request.POST.get('dominio_origem', '').strip()
        dominio_destino = request.POST.get('dominio_destino', '').strip()

        # Validações básicas
        if not all([aplicativo_id, tipo_migracao, dominio_origem, dominio_destino]):
            return JsonResponse({
                'success': False,
                'error': 'Campos obrigatórios: aplicativo_id, tipo_migracao, dominio_origem, dominio_destino'
            }, status=400)

        # Validação de formato de domínio
        if not validar_formato_dominio(dominio_origem):
            return JsonResponse({
                'success': False,
                'error': 'Domínio origem inválido. Formato esperado: http://dominio.com ou http://dominio.com:8080'
            }, status=400)

        if not validar_formato_dominio(dominio_destino):
            return JsonResponse({
                'success': False,
                'error': 'Domínio destino inválido. Formato esperado: http://dominio.com ou http://dominio.com:8080'
            }, status=400)

        if tipo_migracao not in [TarefaMigracaoDNS.TIPO_TODOS, TarefaMigracaoDNS.TIPO_ESPECIFICO]:
            return JsonResponse({
                'success': False,
                'error': 'tipo_migracao deve ser "todos" ou "especifico"'
            }, status=400)

        if tipo_migracao == TarefaMigracaoDNS.TIPO_ESPECIFICO and not mac_alvo:
            return JsonResponse({
                'success': False,
                'error': 'mac_alvo é obrigatório quando tipo_migracao="especifico"'
            }, status=400)

        # Busca aplicativo
        try:
            aplicativo = Aplicativo.objects.get(id=aplicativo_id)
        except Aplicativo.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Aplicativo não encontrado'
            }, status=404)

        # Busca conta reseller
        try:
            conta = ContaReseller.objects.get(
                usuario=request.user,
                aplicativo=aplicativo
            )
        except ContaReseller.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Conta reseller não encontrada. Faça login primeiro.'
            }, status=400)

        # Verifica se sessão é válida
        if not conta.sessao_valida:
            return JsonResponse({
                'success': False,
                'error': 'Sessão expirada. Faça login novamente.'
            }, status=400)

        # Validação específica: se tipo='especifico', validar se dispositivo existe
        if tipo_migracao == TarefaMigracaoDNS.TIPO_ESPECIFICO:
            # NOTA: Esta validação assume que você tem ContaDoAplicativo vinculado ao usuário
            # Ajuste conforme seu modelo de dados
            try:
                device = ContaDoAplicativo.objects.get(
                    usuario=request.user,
                    device_id=mac_alvo
                )

                # VALIDAÇÃO CRÍTICA: Extrai domínio da URL do dispositivo e compara
                if device.url_lista:
                    dominio_device = extrair_dominio_de_url(device.url_lista)

                    if dominio_device and dominio_device != dominio_origem:
                        return JsonResponse({
                            'success': False,
                            'error': f'Domínio atual do dispositivo ({dominio_device}) não corresponde ao domínio origem informado ({dominio_origem}).'
                        }, status=400)

            except ContaDoAplicativo.DoesNotExist:
                # Dispositivo não encontrado no banco local
                # Isso não é necessariamente um erro, pois pode existir apenas no painel reseller
                # Vamos permitir e deixar a validação para o momento da execução
                logger.warning(
                    f"[USER:{request.user.username}] Dispositivo {mac_alvo} não encontrado no banco local. "
                    "Migração será tentada mesmo assim."
                )

        # ===== OTIMIZAÇÃO v2.0: Receber cache de devices do frontend =====
        cached_devices_json = request.POST.get('cached_devices')

        # Cria tarefa
        tarefa = TarefaMigracaoDNS.objects.create(
            usuario=request.user,
            aplicativo=aplicativo,
            conta_reseller=conta,
            tipo_migracao=tipo_migracao,
            mac_alvo=mac_alvo if tipo_migracao == TarefaMigracaoDNS.TIPO_ESPECIFICO else '',
            dominio_origem=dominio_origem,
            dominio_destino=dominio_destino,
            status=TarefaMigracaoDNS.STATUS_INICIANDO,
            cached_devices=cached_devices_json if cached_devices_json else None,  # NOVO: cache temporário
        )

        logger.info(
            f"[USER:{request.user.username}] Tarefa de migração DNS criada: #{tarefa.id} "
            f"(tipo={tipo_migracao}, origem={dominio_origem}, destino={dominio_destino})"
        )

        # Inicia execução em thread
        def executar_migracao_thread():
            """Thread que executa a migração DNS."""
            try:
                from nossopainel.services.reseller_automation import DreamTVSeleniumAutomation

                service = DreamTVSeleniumAutomation(user=request.user, aplicativo=aplicativo)
                service.executar_migracao(tarefa_id=tarefa.id)

            except Exception as e:
                logger.exception(
                    f"[TAREFA:{tarefa.id}] Erro na thread de migração: {e}"
                )

        thread = threading.Thread(target=executar_migracao_thread, daemon=True)
        thread.start()

        return JsonResponse({
            'status': 'iniciado',
            'tarefa_id': tarefa.id,
            'mensagem': 'Migração DNS iniciada. Aguarde o progresso.'
        })

    except Exception as e:
        logger.exception(f"Erro ao iniciar migração DNS: {e}")
        return JsonResponse({
            'success': False,
            'error': f'Erro interno: {str(e)}'
        }, status=500)


@login_required
@require_GET
def consultar_progresso_migracao_api(request, tarefa_id):
    """
    API para consultar progresso de uma tarefa de migração DNS (polling).

    URL: /api/gestao-dns/progresso/<tarefa_id>/
    Method: GET

    Returns:
        JSON:
            - status: Status da tarefa
            - total_dispositivos: Total a serem migrados
            - processados: Quantidade já processada
            - sucessos: Quantidade com sucesso
            - falhas: Quantidade com erro
            - erro_geral: Mensagem de erro geral (se houver)
            - concluida: Boolean indicando se tarefa finalizou
            - dispositivos: Lista com status de cada dispositivo
    """
    try:
        # Busca tarefa (apenas do usuário logado - segurança)
        try:
            tarefa = TarefaMigracaoDNS.objects.get(
                id=tarefa_id,
                usuario=request.user
            )
        except TarefaMigracaoDNS.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Tarefa não encontrada ou você não tem permissão para acessá-la.'
            }, status=404)

        # Busca dispositivos processados
        dispositivos = tarefa.dispositivos.all().values(
            'device_id',
            'nome_dispositivo',
            'status',
            'dns_encontrado',
            'dns_atualizado',
            'mensagem_erro',
            'processado_em'
        )

        # Serializa processado_em para ISO format e extrai domínios
        from nossopainel.utils import extrair_dominio_de_url

        dispositivos_list = []
        for disp in dispositivos:
            disp_dict = dict(disp)
            if disp_dict['processado_em']:
                disp_dict['processado_em'] = disp_dict['processado_em'].isoformat()

            # Extrair apenas domínio (não URL completa) para DNS encontrado e atualizado
            if disp_dict['dns_encontrado']:
                disp_dict['dns_encontrado'] = extrair_dominio_de_url(disp_dict['dns_encontrado'])
            if disp_dict['dns_atualizado']:
                disp_dict['dns_atualizado'] = extrair_dominio_de_url(disp_dict['dns_atualizado'])

            dispositivos_list.append(disp_dict)

        return JsonResponse({
            'status': tarefa.status,
            'etapa_atual': tarefa.etapa_atual,  # NEW: Etapa atual (iniciando, analisando, processando, concluida, cancelada)
            'mensagem_progresso': tarefa.mensagem_progresso,  # NEW: Mensagem dinâmica de progresso
            'progresso_percentual': tarefa.progresso_percentual,  # UPDATED: Usa valor direto do banco (0-100)
            'total_dispositivos': tarefa.total_dispositivos,
            'processados': tarefa.processados,
            'sucessos': tarefa.sucessos,
            'falhas': tarefa.falhas,
            'pulados': tarefa.pulados,
            'erro_geral': tarefa.erro_geral,
            'dominio_origem': tarefa.dominio_origem,  # Domínio DNS de origem (para cache update)
            'dominio_destino': tarefa.dominio_destino,  # Domínio DNS de destino (para cache update)
            'concluida': tarefa.esta_concluida(),
            'criada_em': tarefa.criada_em.isoformat(),
            'iniciada_em': tarefa.iniciada_em.isoformat() if tarefa.iniciada_em else None,
            'concluida_em': tarefa.concluida_em.isoformat() if tarefa.concluida_em else None,
            'dispositivos': dispositivos_list
        })

    except Exception as e:
        logger.exception(f"Erro ao consultar progresso da tarefa {tarefa_id}: {e}")
        return JsonResponse({
            'success': False,
            'error': f'Erro interno: {str(e)}'
        }, status=500)


@login_required
def listar_dominios_api(request):
    """
    API para listar domínios únicos de todos os dispositivos do reseller account.

    GET Params:
        aplicativo_id: ID do aplicativo

    Returns:
        JSON:
            - success: bool
            - dominios: List[{'dominio': str, 'count': int}]
            - error: str (se houver erro)
    """
    try:
        aplicativo_id = request.GET.get('aplicativo_id')

        if not aplicativo_id:
            return JsonResponse({'success': False, 'error': 'aplicativo_id é obrigatório'}, status=400)

        try:
            aplicativo = Aplicativo.objects.get(id=aplicativo_id)
        except Aplicativo.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Aplicativo não encontrado'}, status=404)

        # Obter conta reseller
        conta_reseller = ContaReseller.objects.filter(
            usuario=request.user,
            aplicativo=aplicativo
        ).first()

        if not conta_reseller or not conta_reseller.session_data:
            return JsonResponse({
                'success': False,
                'error': 'Conta reseller não encontrada ou sem sessão válida'
            }, status=400)

        # Extrair JWT do session_data
        try:
            session_data = json.loads(conta_reseller.session_data)
            jwt = session_data.get('jwt') or session_data.get('token')
        except:
            jwt = None

        if not jwt:
            return JsonResponse({
                'success': False,
                'error': 'JWT não encontrado na sessão. Faça login novamente.'
            }, status=400)

        # Inicializar API client
        from nossopainel.services.lib.dream_tv_api import DreamTVAPI
        from nossopainel.utils import extrair_dominio_de_url
        from nossopainel.services.api_raw_logger import get_api_raw_logger
        from collections import Counter
        import time

        api = DreamTVAPI(jwt=jwt)

        # Inicializar logger RAW para registrar dados completos da API
        api_raw_logger = get_api_raw_logger()
        collection_start_time = time.time()

        # Log RAW: Início da coleta
        api_raw_logger.log_collection_start(
            user=request.user.username,
            app_name=aplicativo.nome,
            total_devices=0  # Será atualizado conforme dispositivos são processados
        )

        # v2.0: Coletar devices completos + domínios
        dominios_counter = Counter()
        devices_completos = []  # NOVO: armazena devices com playlists
        page = 1
        limit = 100

        logger.info(f"[listar_dominios_api] Iniciando coleta de devices completos para user={request.user.username}, app={aplicativo.nome}")

        while True:
            try:
                # Listar dispositivos
                devices_data = api.list_devices(page=page, limit=limit)
                devices = devices_data.get('rows', [])

                if not devices:
                    break  # Sem mais dispositivos

                logger.debug(f"[listar_dominios_api] Página {page}: {len(devices)} dispositivos")

                # Para cada dispositivo, listar playlists
                for device in devices:
                    device_id = device.get('id')  # ID numérico (usado para API)
                    device_mac = device.get('mac')  # MAC address real (usado no cache)

                    # Extrair nome do dispositivo (campo pode variar: comment, name, id)
                    nome_dispositivo = (
                        device.get('reseller_activation', {}).get('comment') or
                        device.get('comment') or
                        device.get('name') or
                        device_mac or
                        str(device_id)
                    )

                    # Log RAW: Device completo
                    api_raw_logger.log_device_raw(
                        device_id=device_id,
                        device_mac=device_mac or str(device_id),
                        device_name=nome_dispositivo,
                        raw_data=device  # JSON completo do device
                    )

                    try:
                        playlists = api.list_playlists(device_id=device_id)

                        # NOVO: Estruturar playlists com domínio extraído
                        playlists_estruturadas = []
                        for playlist in playlists:
                            url = playlist.get('url', '')
                            dominio = extrair_dominio_de_url(url) if url else None

                            # Log RAW: Playlist completa
                            api_raw_logger.log_playlist_raw(
                                device_id=device_id,
                                playlist_id=playlist.get('id'),
                                playlist_name=playlist.get('name', 'Sem nome'),
                                raw_data=playlist  # JSON completo da playlist
                            )

                            playlists_estruturadas.append({
                                'id': playlist.get('id'),
                                'name': playlist.get('name', 'Sem nome'),
                                'url': url,
                                'dominio': dominio,
                                'is_selected': playlist.get('is_selected', False),
                                'deviceId': playlist.get('deviceId')  # Preservar deviceId numérico para updates
                            })

                            # Contar domínios para lista de domínios únicos
                            if dominio:
                                dominios_counter[dominio] += 1

                        # NOVO: Adicionar device completo ao array
                        devices_completos.append({
                            'device_id': device_mac,  # MAC real (ex: 00:1A:79:XX:XX:XX)
                            'nome_dispositivo': nome_dispositivo,
                            'playlists': playlists_estruturadas
                        })

                    except Exception as e:
                        logger.warning(f"[listar_dominios_api] Erro ao buscar playlists do device {device_mac}: {e}")
                        # NOVO: Mesmo com erro, adicionar device sem playlists
                        devices_completos.append({
                            'device_id': device_mac,  # MAC real (ex: 00:1A:79:XX:XX:XX)
                            'nome_dispositivo': nome_dispositivo,
                            'playlists': []
                        })
                        continue

                    # Delay para evitar rate limiting da API (429) - reduzido de 0.2 para 0.05
                    import time
                    time.sleep(0.05)  # 50ms entre cada requisição (economia de ~26s para 175 devices)

                # Verificar se há mais páginas
                total = devices_data.get('count', 0)
                if page * limit >= total:
                    break

                page += 1

            except Exception as e:
                logger.error(f"[listar_dominios_api] Erro ao listar devices na página {page}: {e}")
                break

        # Formatar resultado (ordenar por count DESC)
        dominios_list = [
            {'dominio': dominio, 'count': count}
            for dominio, count in dominios_counter.most_common()
        ]

        logger.info(f"[listar_dominios_api] Coleta finalizada: {len(devices_completos)} dispositivos, {len(dominios_list)} domínios únicos")

        # Log RAW: Fim da coleta (sucesso)
        collection_duration = time.time() - collection_start_time
        total_playlists = sum(len(d['playlists']) for d in devices_completos)
        api_raw_logger.log_collection_end(
            user=request.user.username,
            app_name=aplicativo.nome,
            status="success",
            duration_seconds=collection_duration,
            devices_processed=len(devices_completos),
            playlists_total=total_playlists
        )

        return JsonResponse({
            'success': True,
            'devices': devices_completos,  # NOVO: devices completos
            'dominios': dominios_list,      # mantém para compatibilidade
            'total_dispositivos': len(devices_completos),
            'total_dominios': len(dominios_list)
        })

    except Exception as e:
        # Log RAW: Fim da coleta (erro)
        try:
            collection_duration = time.time() - collection_start_time
            total_playlists = sum(len(d['playlists']) for d in devices_completos) if devices_completos else 0
            api_raw_logger.log_collection_end(
                user=request.user.username,
                app_name=aplicativo.nome,
                status="error",
                duration_seconds=collection_duration,
                devices_processed=len(devices_completos),
                playlists_total=total_playlists,
                error=str(e)
            )
        except:
            pass  # Se log falhar, não impedir o erro original

        logger.exception(f"[listar_dominios_api] Erro geral: {e}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_POST
def buscar_dispositivo_api(request):
    """
    API para buscar dispositivo por MAC e listar suas playlists.

    POST Body:
        aplicativo_id: ID do aplicativo
        mac_address: MAC address do dispositivo

    Returns:
        JSON:
            - success: bool
            - device: Dict com dados do dispositivo
            - playlists: List[Dict] com playlists e domínios extraídos
            - error: str (se houver erro)
    """
    try:
        data = json.loads(request.body)
        aplicativo_id = data.get('aplicativo_id')
        mac_address = data.get('mac_address', '').strip().upper()

        if not aplicativo_id or not mac_address:
            return JsonResponse({
                'success': False,
                'error': 'aplicativo_id e mac_address são obrigatórios'
            }, status=400)

        try:
            aplicativo = Aplicativo.objects.get(id=aplicativo_id)
        except Aplicativo.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Aplicativo não encontrado'}, status=404)

        # Obter conta reseller
        conta_reseller = ContaReseller.objects.filter(
            usuario=request.user,
            aplicativo=aplicativo
        ).first()

        if not conta_reseller or not conta_reseller.session_data:
            return JsonResponse({
                'success': False,
                'error': 'Conta reseller não encontrada ou sem sessão válida'
            }, status=400)

        # Extrair JWT
        try:
            session_data = json.loads(conta_reseller.session_data)
            jwt = session_data.get('jwt') or session_data.get('token')
        except:
            jwt = None

        if not jwt:
            return JsonResponse({
                'success': False,
                'error': 'JWT não encontrado na sessão. Faça login novamente.'
            }, status=400)

        # Inicializar API client
        from nossopainel.services.lib.dream_tv_api import DreamTVAPI
        from nossopainel.utils import extrair_dominio_de_url

        api = DreamTVAPI(jwt=jwt)

        # Buscar dispositivo por MAC
        logger.info(f"[buscar_dispositivo_api] Buscando dispositivo MAC={mac_address} para user={request.user.username}")

        devices_data = api.list_devices(page=1, limit=10, search={'mac': mac_address})
        devices = devices_data.get('rows', [])

        if not devices:
            return JsonResponse({
                'success': False,
                'error': f'Dispositivo com MAC {mac_address} não encontrado'
            }, status=404)

        device = devices[0]  # Primeiro resultado (MAC é único)
        device_id = device.get('id')

        # Extrair comment do reseller_activation
        reseller_activation = device.get('reseller_activation', {})
        comment = reseller_activation.get('comment', '')

        logger.debug(f"[buscar_dispositivo_api] Dispositivo encontrado: id={device_id}, comment={comment}")

        # Buscar playlists do dispositivo
        playlists_raw = api.list_playlists(device_id=device_id)

        # Extrair domínio de cada playlist
        playlists_formatted = []
        for playlist in playlists_raw:
            url = playlist.get('url', '')
            dominio = extrair_dominio_de_url(url) if url else ''

            playlists_formatted.append({
                'id': playlist.get('id'),
                'name': playlist.get('name', 'Sem nome'),
                'url': url,
                'dominio': dominio,
                'is_selected': playlist.get('is_selected', False)
            })

        logger.info(f"[buscar_dispositivo_api] Dispositivo {mac_address} possui {len(playlists_formatted)} playlist(s)")

        return JsonResponse({
            'success': True,
            'device': {
                'id': device.get('id'),
                'mac': device.get('mac'),
                'comment': comment,
                'activation_expired': device.get('activation_expired')
            },
            'playlists': playlists_formatted
        })

    except Exception as e:
        logger.exception(f"[buscar_dispositivo_api] Erro geral: {e}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# =====================================================================
# API - CONFIGURAÇÃO DEBUG HEADLESS (ADMIN)
# =====================================================================

@login_required
@require_http_methods(["POST"])
def toggle_debug_headless(request):
    """
    Toggle do modo debug headless (apenas admin).
    Quando ativado, o navegador Playwright fica visível durante automação.
    """
    from nossopainel.models import ConfiguracaoAutomacao

    # Apenas admins podem ativar debug mode
    if not request.user.is_staff:
        return JsonResponse({
            'success': False,
            'erro': 'Permissão negada. Apenas administradores podem alterar o modo debug.'
        }, status=403)

    try:
        # Pega ou cria config para o usuário
        config, created = ConfiguracaoAutomacao.objects.get_or_create(user=request.user)

        # Inverte o estado
        config.debug_headless_mode = not config.debug_headless_mode
        config.save()

        status_msg = "ATIVADO" if config.debug_headless_mode else "DESATIVADO"
        logger.info(
            f"[ADMIN:{request.user.username}] Modo debug headless {status_msg}"
        )

        return JsonResponse({
            'success': True,
            'debug_mode': config.debug_headless_mode,
            'mensagem': f"Modo debug {status_msg}. "
                       f"{'Navegador ficará VISÍVEL durante próximas automações.' if config.debug_headless_mode else 'Navegador voltará a ser OCULTO (headless).'}"
        })

    except Exception as e:
        logger.exception(f"Erro ao toggle debug headless: {e}")
        return JsonResponse({
            'success': False,
            'erro': f'Erro ao alterar configuração: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["GET"])
def get_debug_status(request):
    """
    Retorna o status atual do modo debug headless.
    Usado para inicializar o estado do botão na interface.
    """
    from nossopainel.models import ConfiguracaoAutomacao

    try:
        config = ConfiguracaoAutomacao.objects.filter(user=request.user).first()

        return JsonResponse({
            'success': True,
            'debug_mode': config.debug_headless_mode if config else False,
            'is_admin': request.user.is_staff
        })

    except Exception as e:
        logger.exception(f"Erro ao consultar status de debug: {e}")
        return JsonResponse({
            'success': False,
            'erro': f'Erro ao consultar status: {str(e)}'
        }, status=500)


# ============================================================
# VIEWS - TAREFAS DE ENVIO WHATSAPP
# ============================================================

from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from .models import TarefaEnvio, HistoricoExecucaoTarefa, TemplateMensagem
from .forms import TarefaEnvioForm


class TarefaEnvioListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    """Lista todas as tarefas de envio do usuário."""
    model = TarefaEnvio
    template_name = 'tarefas_envio/lista.html'
    context_object_name = 'tarefas'
    paginate_by = 10

    def test_func(self):
        return self.request.user.is_superuser

    def get_queryset(self):
        return TarefaEnvio.objects.filter(
            usuario=self.request.user
        ).order_by('-criado_em')

    def get_context_data(self, **kwargs):
        from .models import HistoricoExecucaoTarefa
        from django.utils import timezone
        from django.db.models import Avg

        context = super().get_context_data(**kwargs)
        qs = TarefaEnvio.objects.filter(usuario=self.request.user)
        context['total_ativas'] = qs.filter(ativo=True).count()
        context['total_inativas'] = qs.filter(ativo=False).count()
        context['total_envios'] = qs.aggregate(total=Sum('total_envios'))['total'] or 0

        # Novas métricas
        agora = timezone.localtime()
        context['total_pausadas'] = qs.filter(pausado_ate__gt=agora).count()

        # Taxa de sucesso dos últimos 30 dias
        data_30_dias = agora - timezone.timedelta(days=30)
        historico_recente = HistoricoExecucaoTarefa.objects.filter(
            tarefa__usuario=self.request.user,
            data_execucao__gte=data_30_dias
        )
        total_execucoes = historico_recente.count()
        if total_execucoes > 0:
            sucesso_count = historico_recente.filter(status='sucesso').count()
            context['taxa_sucesso'] = round((sucesso_count / total_execucoes) * 100, 1)
        else:
            context['taxa_sucesso'] = None

        # Média de envios por execução
        media_envios = historico_recente.aggregate(media=Avg('quantidade_enviada'))['media']
        context['media_envios_execucao'] = round(media_envios, 1) if media_envios else 0

        # Próxima tarefa a executar (hoje, baseado no horário)
        tarefas_ativas = qs.filter(ativo=True, pausado_ate__isnull=True) | qs.filter(ativo=True, pausado_ate__lte=agora)
        proxima_tarefa = None
        for tarefa in tarefas_ativas.order_by('horario'):
            if tarefa.deve_executar_hoje() and tarefa.horario > agora.time():
                proxima_tarefa = tarefa
                break
        context['proxima_tarefa'] = proxima_tarefa

        context['page'] = 'tarefas-envio'
        context['page_group'] = 'admin'
        return context


class TarefaEnvioCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    """Cria uma nova tarefa de envio."""
    model = TarefaEnvio
    form_class = TarefaEnvioForm
    template_name = 'tarefas_envio/form.html'
    success_url = reverse_lazy('tarefas-envio-lista')

    def test_func(self):
        return self.request.user.is_superuser

    def post(self, request, *args, **kwargs):
        """Log do POST antes de processar o form."""
        import logging
        logger = logging.getLogger(__name__)

        # IMPORTANTE: CreateView precisa definir self.object = None antes de processar o form
        self.object = None

        logger.info(f"TarefaEnvioCreateView POST - FILES: {request.FILES}")
        logger.info(f"TarefaEnvioCreateView POST - FILES keys: {list(request.FILES.keys())}")
        logger.info(f"TarefaEnvioCreateView POST - POST keys: {list(request.POST.keys())}")
        logger.info(f"TarefaEnvioCreateView POST - content_type: {request.content_type}")
        return super().post(request, *args, **kwargs)

    def form_valid(self, form):
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"TarefaEnvioCreateView form_valid - FILES: {self.request.FILES}")
        logger.info(f"TarefaEnvioCreateView form_valid - imagem in cleaned_data: {form.cleaned_data.get('imagem')}")
        form.instance.usuario = self.request.user

        try:
            response = super().form_valid(form)
            logger.info(f"TarefaEnvioCreateView - tarefa salva com sucesso, ID: {self.object.id}")
            messages.success(self.request, 'Tarefa de envio criada com sucesso!')
            return response
        except Exception as e:
            logger.error(f"TarefaEnvioCreateView - erro ao salvar: {e}", exc_info=True)
            messages.error(self.request, f'Erro ao salvar tarefa: {e}')
            return self.form_invalid(form)

    def form_invalid(self, form):
        import logging
        logger = logging.getLogger(__name__)
        logger.warning(f"TarefaEnvioCreateView form_invalid - errors: {form.errors}")
        logger.warning(f"TarefaEnvioCreateView form_invalid - FILES: {self.request.FILES}")
        logger.warning(f"TarefaEnvioCreateView form_invalid - POST: {self.request.POST}")
        messages.error(self.request, f'Erro ao criar tarefa: {form.errors}')
        return super().form_invalid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Nova Tarefa de Envio'
        context['botao_submit'] = 'Criar Tarefa'
        context['page'] = 'tarefas-envio'
        context['page_group'] = 'admin'
        return context


class TarefaEnvioUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    """Edita uma tarefa de envio existente."""
    model = TarefaEnvio
    form_class = TarefaEnvioForm
    template_name = 'tarefas_envio/form.html'
    success_url = reverse_lazy('tarefas-envio-lista')

    def test_func(self):
        return self.request.user.is_superuser

    def get_queryset(self):
        return TarefaEnvio.objects.filter(usuario=self.request.user)

    def post(self, request, *args, **kwargs):
        """Log do POST antes de processar o form."""
        import logging
        logger = logging.getLogger(__name__)

        # IMPORTANTE: UpdateView precisa definir self.object antes de processar o form
        self.object = self.get_object()

        logger.info(f"TarefaEnvioUpdateView POST - FILES: {request.FILES}")
        logger.info(f"TarefaEnvioUpdateView POST - FILES keys: {list(request.FILES.keys())}")
        logger.info(f"TarefaEnvioUpdateView POST - content_type: {request.content_type}")
        logger.info(f"TarefaEnvioUpdateView POST - object.imagem: {self.object.imagem}")
        return super().post(request, *args, **kwargs)

    def form_valid(self, form):
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"TarefaEnvioUpdateView form_valid - FILES: {self.request.FILES}")
        logger.info(f"TarefaEnvioUpdateView form_valid - imagem in cleaned_data: {form.cleaned_data.get('imagem')}")
        logger.info(f"TarefaEnvioUpdateView form_valid - form.instance.imagem: {form.instance.imagem}")

        # Trata remoção de imagem (checkbox imagem-clear)
        if self.request.POST.get('imagem-clear') == 'true':
            logger.info("TarefaEnvioUpdateView - removendo imagem atual")
            form.instance.imagem = None

        try:
            response = super().form_valid(form)
            logger.info(f"TarefaEnvioUpdateView - tarefa atualizada com sucesso, ID: {self.object.id}")
            messages.success(self.request, 'Tarefa de envio atualizada com sucesso!')
            return response
        except Exception as e:
            logger.error(f"TarefaEnvioUpdateView - erro ao salvar: {e}", exc_info=True)
            messages.error(self.request, f'Erro ao salvar tarefa: {e}')
            return self.form_invalid(form)

    def form_invalid(self, form):
        import logging
        logger = logging.getLogger(__name__)
        logger.warning(f"TarefaEnvioUpdateView form_invalid - errors: {form.errors}")
        logger.warning(f"TarefaEnvioUpdateView form_invalid - FILES: {self.request.FILES}")
        logger.warning(f"TarefaEnvioUpdateView form_invalid - POST keys: {list(self.request.POST.keys())}")
        messages.error(self.request, f'Erro ao atualizar tarefa: {form.errors}')
        return super().form_invalid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Editar Tarefa de Envio'
        context['botao_submit'] = 'Salvar Alterações'
        context['editando'] = True
        context['page'] = 'tarefas-envio'
        context['page_group'] = 'admin'
        return context


class TarefaEnvioDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    """Deleta uma tarefa de envio."""
    model = TarefaEnvio
    success_url = reverse_lazy('tarefas-envio-lista')

    def test_func(self):
        return self.request.user.is_superuser

    def get_queryset(self):
        return TarefaEnvio.objects.filter(usuario=self.request.user)

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        nome = self.object.nome
        self.object.delete()

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'message': f'Tarefa "{nome}" excluída com sucesso'
            })

        messages.success(request, f'Tarefa "{nome}" excluída com sucesso!')
        return redirect(self.success_url)


@login_required
@require_POST
def tarefa_envio_excluir_ajax(request, pk):
    """Exclui uma tarefa de envio via AJAX."""
    if not request.user.is_superuser:
        return JsonResponse({
            'success': False,
            'error': 'Permissão negada'
        }, status=403)

    tarefa = get_object_or_404(TarefaEnvio, pk=pk, usuario=request.user)
    nome = tarefa.nome
    tarefa.delete()

    return JsonResponse({
        'success': True,
        'message': f'Tarefa "{nome}" excluída com sucesso'
    })


@login_required
@require_POST
def tarefa_envio_toggle(request, pk):
    """Alterna o status ativo/inativo de uma tarefa via AJAX."""
    if not request.user.is_superuser:
        return JsonResponse({
            'success': False,
            'error': 'Permissão negada'
        }, status=403)

    tarefa = get_object_or_404(TarefaEnvio, pk=pk, usuario=request.user)
    tarefa.ativo = not tarefa.ativo
    tarefa.save(update_fields=['ativo'])

    return JsonResponse({
        'success': True,
        'ativo': tarefa.ativo,
        'message': f'Tarefa {"ativada" if tarefa.ativo else "desativada"} com sucesso'
    })


@login_required
@require_POST
def tarefa_envio_duplicar(request, pk):
    """Duplica uma tarefa de envio existente."""
    if not request.user.is_superuser:
        return JsonResponse({
            'success': False,
            'error': 'Permissão negada'
        }, status=403)

    tarefa = get_object_or_404(TarefaEnvio, pk=pk, usuario=request.user)

    # Cria cópia da tarefa
    nova_tarefa = TarefaEnvio.objects.create(
        nome=f"[Cópia] {tarefa.nome}",
        tipo_envio=tarefa.tipo_envio,
        dias_semana=tarefa.dias_semana,
        periodo_mes=tarefa.periodo_mes,
        horario=tarefa.horario,
        imagem=tarefa.imagem,
        mensagem=tarefa.mensagem,
        mensagem_plaintext=tarefa.mensagem_plaintext,
        filtro_estados=tarefa.filtro_estados,
        filtro_cidades=tarefa.filtro_cidades,
        ativo=False,  # Começa inativa
        usuario=request.user,
    )

    return JsonResponse({
        'success': True,
        'message': f'Tarefa duplicada: "{nova_tarefa.nome}"',
        'id': nova_tarefa.id
    })


@login_required
@require_POST
def tarefa_envio_preview(request):
    """Retorna preview da mensagem convertida para WhatsApp."""
    if not request.user.is_superuser:
        return JsonResponse({
            'success': False,
            'error': 'Permissão negada'
        }, status=403)

    html_content = request.POST.get('mensagem', '')

    # Cria instância temporária para usar o método de conversão
    tarefa_temp = TarefaEnvio(mensagem=html_content)
    preview = tarefa_temp.converter_html_para_whatsapp()

    return JsonResponse({
        'success': True,
        'preview': preview
    })


@login_required
def tarefa_envio_sugestao_horarios(request):
    """
    Analisa o histórico de execuções e sugere os melhores horários.
    Baseado na taxa de sucesso por faixa horária.
    """
    from django.db.models import Count, Case, When, FloatField
    from django.db.models.functions import ExtractHour

    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Permissão negada'}, status=403)

    # Busca histórico dos últimos 90 dias
    data_limite = timezone.localtime() - timezone.timedelta(days=90)
    historico = HistoricoExecucaoTarefa.objects.filter(
        tarefa__usuario=request.user,
        data_execucao__gte=data_limite
    ).annotate(
        hora=ExtractHour('data_execucao')
    )

    # Agrupa por hora e calcula taxa de sucesso
    stats_por_hora = {}
    for h in range(6, 22):  # 6h às 21h
        registros = historico.filter(hora=h)
        total = registros.count()
        if total > 0:
            sucesso = registros.filter(status='sucesso').count()
            media_envios = registros.aggregate(
                media=Avg('quantidade_enviada')
            )['media'] or 0
            stats_por_hora[h] = {
                'hora': f"{h:02d}:00",
                'total_execucoes': total,
                'taxa_sucesso': round((sucesso / total) * 100, 1),
                'media_envios': round(media_envios, 1)
            }

    # Ordena por taxa de sucesso e média de envios
    if stats_por_hora:
        melhores_horarios = sorted(
            stats_por_hora.items(),
            key=lambda x: (x[1]['taxa_sucesso'], x[1]['media_envios']),
            reverse=True
        )[:5]
        sugestoes = [
            {
                'horario': item[1]['hora'],
                'motivo': f"Taxa de sucesso: {item[1]['taxa_sucesso']}% | Média: {item[1]['media_envios']} envios"
            }
            for item in melhores_horarios
        ]
    else:
        # Horários sugeridos padrão (sem histórico)
        sugestoes = [
            {'horario': '09:00', 'motivo': 'Horário comercial - início do expediente'},
            {'horario': '14:00', 'motivo': 'Horário comercial - após almoço'},
            {'horario': '17:00', 'motivo': 'Horário comercial - final do dia'},
        ]

    return JsonResponse({
        'success': True,
        'sugestoes': sugestoes,
        'baseado_em_dados': len(stats_por_hora) > 0
    })


class TarefaEnvioHistoricoView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    """Exibe o histórico de execuções de uma tarefa."""
    model = HistoricoExecucaoTarefa
    template_name = 'tarefas_envio/historico.html'
    context_object_name = 'historicos'
    paginate_by = 20

    def test_func(self):
        return self.request.user.is_superuser

    def get_queryset(self):
        self.tarefa = get_object_or_404(
            TarefaEnvio,
            pk=self.kwargs['pk'],
            usuario=self.request.user
        )
        return HistoricoExecucaoTarefa.objects.filter(
            tarefa=self.tarefa
        ).order_by('-data_execucao')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['tarefa'] = self.tarefa

        # Estatísticas
        historicos = HistoricoExecucaoTarefa.objects.filter(tarefa=self.tarefa)
        context['total_execucoes'] = historicos.count()
        context['total_sucesso'] = historicos.filter(status='sucesso').count()
        context['total_parcial'] = historicos.filter(status='parcial').count()
        context['total_erro'] = historicos.filter(status='erro').count()

        # Taxa de sucesso
        if context['total_execucoes'] > 0:
            context['taxa_sucesso'] = round(
                (context['total_sucesso'] / context['total_execucoes']) * 100, 1
            )
        else:
            context['taxa_sucesso'] = 0

        context['page'] = 'tarefas-envio'
        context['page_group'] = 'admin'
        return context


@login_required
@require_GET
def tarefas_envio_stats_api(request):
    """Retorna estatísticas das tarefas de envio em JSON para atualização em tempo real."""
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Não autorizado'}, status=403)

    from django.db.models import Avg
    from django.utils import timezone

    qs = TarefaEnvio.objects.filter(usuario=request.user)
    agora = timezone.localtime()

    # Estatísticas básicas
    total_ativas = qs.filter(ativo=True).count()
    total_inativas = qs.filter(ativo=False).count()
    total_pausadas = qs.filter(pausado_ate__gt=agora).count()
    total_envios = qs.aggregate(total=Sum('total_envios'))['total'] or 0

    # Taxa de sucesso dos últimos 30 dias
    data_30_dias = agora - timezone.timedelta(days=30)
    historico_recente = HistoricoExecucaoTarefa.objects.filter(
        tarefa__usuario=request.user,
        data_execucao__gte=data_30_dias
    )
    total_execucoes = historico_recente.count()
    if total_execucoes > 0:
        sucesso_count = historico_recente.filter(status='sucesso').count()
        taxa_sucesso = round((sucesso_count / total_execucoes) * 100, 1)
    else:
        taxa_sucesso = None

    # Média de envios por execução
    media_envios = historico_recente.aggregate(media=Avg('quantidade_enviada'))['media']
    media_envios_execucao = round(media_envios, 1) if media_envios else 0

    # Próxima tarefa a executar
    tarefas_ativas = qs.filter(ativo=True, pausado_ate__isnull=True) | qs.filter(ativo=True, pausado_ate__lte=agora)
    proxima_tarefa = None
    proxima_nome = None
    proxima_horario = None
    for tarefa in tarefas_ativas.order_by('horario'):
        if tarefa.deve_executar_hoje() and tarefa.horario > agora.time():
            proxima_tarefa = tarefa
            proxima_nome = tarefa.nome[:25] + '...' if len(tarefa.nome) > 25 else tarefa.nome
            proxima_horario = tarefa.horario.strftime('%H:%M')
            break

    # Tarefas em execução
    tarefas_em_execucao = list(
        qs.filter(em_execucao=True).values_list('id', flat=True)
    )

    return JsonResponse({
        'success': True,
        'stats': {
            'total_ativas': total_ativas,
            'total_inativas': total_inativas,
            'total_pausadas': total_pausadas,
            'total_envios': total_envios,
            'taxa_sucesso': taxa_sucesso,
            'media_envios_execucao': media_envios_execucao,
            'proxima_tarefa': {
                'nome': proxima_nome,
                'horario': proxima_horario
            } if proxima_tarefa else None
        },
        'tarefas_em_execucao': tarefas_em_execucao,
        'timestamp': agora.isoformat()
    })


# ============================================================================
# TAREFAS ENVIO - VIEWS AJAX PARA FORMULÁRIO
# ============================================================================

@login_required
@require_GET
def tarefa_envio_preview_alcance(request):
    """
    Retorna preview da quantidade de clientes que serão atingidos.

    Parâmetros GET:
    - tipo_envio: 'ativos' ou 'cancelados'
    - filtro_estados[]: lista de UFs (opcional)

    Retorna JSON com contagem de clientes.
    """
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Permissão negada'}, status=403)

    tipo_envio = request.GET.get('tipo_envio', '')
    filtro_estados = request.GET.getlist('filtro_estados[]', [])

    # Base query
    qs = Cliente.objects.filter(usuario=request.user, nao_enviar_msgs=False)

    if tipo_envio == 'ativos':
        qs = qs.filter(cancelado=False)
    elif tipo_envio == 'cancelados':
        data_limite = timezone.now() - timedelta(days=7)
        qs = qs.filter(cancelado=True, data_cancelamento__lte=data_limite)
    else:
        return JsonResponse({
            'success': False,
            'error': 'Tipo de envio inválido'
        }, status=400)

    # Aplicar filtro de estados se fornecido
    if filtro_estados:
        qs = qs.filter(uf__in=filtro_estados)

    total = qs.count()

    # Estatísticas extras por estado (top 5)
    from django.db.models import Count
    stats_estados = list(
        qs.values('uf').annotate(total=Count('id')).order_by('-total')[:5]
    )

    return JsonResponse({
        'success': True,
        'total': total,
        'tipo_envio': tipo_envio,
        'filtro_estados': filtro_estados,
        'stats_estados': stats_estados
    })


@login_required
@require_GET
def tarefa_envio_verificar_conflito(request):
    """
    Verifica se há tarefas com mesmo horário nos mesmos dias.

    Parâmetros GET:
    - horario: HH:MM
    - dias_semana[]: lista de dias (0-6)
    - tarefa_id: ID da tarefa atual (para excluir da verificação em edição)

    Retorna lista de tarefas conflitantes.
    """
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Permissão negada'}, status=403)

    horario_str = request.GET.get('horario', '')
    dias_semana = request.GET.getlist('dias_semana[]', [])
    tarefa_id = request.GET.get('tarefa_id', '')

    if not horario_str or not dias_semana:
        return JsonResponse({
            'success': True,
            'conflitos': [],
            'tem_conflito': False
        })

    try:
        from datetime import datetime as dt
        horario = dt.strptime(horario_str, '%H:%M').time()
    except ValueError:
        return JsonResponse({
            'success': False,
            'error': 'Formato de horário inválido'
        }, status=400)

    # Converte dias para inteiros
    try:
        dias_semana_int = [int(d) for d in dias_semana]
    except ValueError:
        return JsonResponse({
            'success': False,
            'error': 'Dias da semana inválidos'
        }, status=400)

    # Busca tarefas ativas com mesmo horário
    qs = TarefaEnvio.objects.filter(
        usuario=request.user,
        ativo=True,
        horario=horario
    )

    # Exclui a tarefa atual se em modo de edição
    if tarefa_id:
        try:
            qs = qs.exclude(pk=int(tarefa_id))
        except ValueError:
            pass

    conflitos = []
    for tarefa in qs:
        # Verifica se há interseção de dias
        dias_tarefa = set(tarefa.dias_semana or [])
        dias_novos = set(dias_semana_int)
        dias_conflito = dias_tarefa & dias_novos

        if dias_conflito:
            DIAS_NOMES = ['Seg', 'Ter', 'Qua', 'Qui', 'Sex', 'Sáb', 'Dom']
            dias_conflito_nomes = [DIAS_NOMES[d] for d in sorted(dias_conflito)]
            conflitos.append({
                'id': tarefa.id,
                'nome': tarefa.nome,
                'tipo_envio': tarefa.get_tipo_envio_display(),
                'dias_conflito': dias_conflito_nomes
            })

    return JsonResponse({
        'success': True,
        'conflitos': conflitos,
        'tem_conflito': len(conflitos) > 0
    })


@login_required
@require_GET
def tarefa_envio_listar_templates(request):
    """
    Lista templates de mensagem do usuário.

    Parâmetros GET:
    - categoria: filtro opcional por categoria

    Retorna lista de templates agrupados por categoria.
    """
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Permissão negada'}, status=403)

    categoria = request.GET.get('categoria', '')

    qs = TemplateMensagem.objects.filter(usuario=request.user, ativo=True)

    if categoria:
        qs = qs.filter(categoria=categoria)

    templates = []
    for t in qs.order_by('categoria', 'nome'):
        templates.append({
            'id': t.id,
            'nome': t.nome,
            'descricao': t.descricao,
            'categoria': t.categoria,
            'categoria_display': t.get_categoria_display(),
            'mensagem_html': t.mensagem_html,
            'tem_imagem': bool(t.imagem),
            'imagem_url': t.imagem.url if t.imagem else None
        })

    # Agrupa por categoria
    categorias = {}
    for t in templates:
        cat = t['categoria']
        if cat not in categorias:
            categorias[cat] = {
                'nome': t['categoria_display'],
                'templates': []
            }
        categorias[cat]['templates'].append(t)

    return JsonResponse({
        'success': True,
        'templates': templates,
        'categorias': categorias,
        'total': len(templates)
    })


@login_required
@require_POST
def tarefa_envio_salvar_template(request):
    """
    Salva uma nova mensagem como template.

    Parâmetros POST (JSON):
    - nome: nome do template
    - categoria: categoria (promocao, lembrete, boas_vindas, cobranca, geral)
    - mensagem: conteúdo HTML da mensagem
    - descricao: descrição opcional

    Retorna o template criado.
    """
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Permissão negada'}, status=403)

    try:
        import json
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'JSON inválido'
        }, status=400)

    nome = data.get('nome', '').strip()
    categoria = data.get('categoria', 'geral')
    mensagem = data.get('mensagem', '').strip()
    descricao = data.get('descricao', '').strip()

    # Validações
    if not nome:
        return JsonResponse({
            'success': False,
            'error': 'Nome do template é obrigatório'
        }, status=400)

    if len(nome) > 100:
        return JsonResponse({
            'success': False,
            'error': 'Nome muito longo (máximo 100 caracteres)'
        }, status=400)

    if not mensagem:
        return JsonResponse({
            'success': False,
            'error': 'Mensagem é obrigatória'
        }, status=400)

    # Verifica se categoria é válida
    categorias_validas = ['promocao', 'lembrete', 'boas_vindas', 'cobranca', 'geral']
    if categoria not in categorias_validas:
        categoria = 'geral'

    # Verifica duplicidade de nome
    if TemplateMensagem.objects.filter(usuario=request.user, nome=nome).exists():
        return JsonResponse({
            'success': False,
            'error': f'Já existe um template com o nome "{nome}"'
        }, status=400)

    # Cria o template
    template = TemplateMensagem.objects.create(
        usuario=request.user,
        nome=nome,
        categoria=categoria,
        mensagem_html=mensagem,
        descricao=descricao
    )

    return JsonResponse({
        'success': True,
        'template': {
            'id': template.id,
            'nome': template.nome,
            'categoria': template.categoria,
            'categoria_display': template.get_categoria_display(),
            'mensagem_html': template.mensagem_html,
            'descricao': template.descricao
        },
        'message': f'Template "{nome}" salvo com sucesso!'
    })


@login_required
@require_GET
def tarefa_envio_historico_api(request, pk):
    """
    Retorna historico filtrado via AJAX para a pagina de historico.

    Parametros GET:
    - status: 'sucesso', 'parcial', 'erro' ou '' (todos)
    - data_inicio: YYYY-MM-DD
    - data_fim: YYYY-MM-DD
    - page: numero da pagina (default: 1)
    """
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Permissao negada'}, status=403)

    # Valida que a tarefa pertence ao usuario
    tarefa = get_object_or_404(TarefaEnvio, pk=pk, usuario=request.user)

    # Parametros de filtro
    status_filtro = request.GET.get('status', '')
    data_inicio = request.GET.get('data_inicio', '')
    data_fim = request.GET.get('data_fim', '')
    page = request.GET.get('page', 1)

    # Query base
    qs = HistoricoExecucaoTarefa.objects.filter(tarefa=tarefa).order_by('-data_execucao')

    # Aplica filtros
    if status_filtro in ['sucesso', 'parcial', 'erro']:
        qs = qs.filter(status=status_filtro)

    if data_inicio:
        try:
            from datetime import datetime as dt
            data_inicio_parsed = dt.strptime(data_inicio, '%Y-%m-%d')
            qs = qs.filter(data_execucao__date__gte=data_inicio_parsed.date())
        except ValueError:
            pass

    if data_fim:
        try:
            from datetime import datetime as dt
            data_fim_parsed = dt.strptime(data_fim, '%Y-%m-%d')
            qs = qs.filter(data_execucao__date__lte=data_fim_parsed.date())
        except ValueError:
            pass

    # Paginacao
    from django.core.paginator import Paginator, EmptyPage
    paginator = Paginator(qs, 20)

    try:
        page_obj = paginator.page(page)
    except EmptyPage:
        page_obj = paginator.page(1)

    # Serializa historicos
    historicos = []
    for h in page_obj:
        historicos.append({
            'id': h.id,
            'data_execucao': h.data_execucao.strftime('%d/%m/%Y'),
            'hora_execucao': h.data_execucao.strftime('%H:%M:%S'),
            'data_iso': h.data_execucao.isoformat(),
            'status': h.status,
            'quantidade_enviada': h.quantidade_enviada,
            'quantidade_erros': h.quantidade_erros,
            'duracao': h.get_duracao_formatada(),
            'detalhes': h.detalhes or ''
        })

    # Estatisticas filtradas
    total_filtrado = qs.count()
    sucesso_filtrado = qs.filter(status='sucesso').count()
    parcial_filtrado = qs.filter(status='parcial').count()
    erro_filtrado = qs.filter(status='erro').count()

    taxa_sucesso = 0
    if total_filtrado > 0:
        taxa_sucesso = round((sucesso_filtrado / total_filtrado) * 100, 1)

    # Dados para grafico (ultimas 10 execucoes da query original sem paginacao)
    ultimas_10 = list(qs[:10].values('data_execucao', 'quantidade_enviada', 'status'))
    grafico_data = []
    for item in reversed(ultimas_10):
        grafico_data.append({
            'data': item['data_execucao'].strftime('%d/%m'),
            'enviados': item['quantidade_enviada'],
            'status': item['status']
        })

    return JsonResponse({
        'success': True,
        'historicos': historicos,
        'pagination': {
            'current_page': page_obj.number,
            'total_pages': paginator.num_pages,
            'total_items': paginator.count,
            'has_previous': page_obj.has_previous(),
            'has_next': page_obj.has_next(),
            'previous_page': page_obj.previous_page_number() if page_obj.has_previous() else None,
            'next_page': page_obj.next_page_number() if page_obj.has_next() else None
        },
        'stats': {
            'total': total_filtrado,
            'sucesso': sucesso_filtrado,
            'parcial': parcial_filtrado,
            'erro': erro_filtrado,
            'taxa_sucesso': taxa_sucesso
        },
        'grafico': grafico_data
    })


############################################ API FORMA DE PAGAMENTO ############################################

@login_required
@require_http_methods(["GET"])
def api_formas_pagamento_disponiveis(request):
    """
    Lista formas de pagamento disponíveis com informações de limite.
    Usado no modal de reativação de clientes cancelados.

    GET /api/formas-pagamento-disponiveis/

    Retorna:
    - Lista de formas de pagamento com status de limite detalhado
    - FastDePix: nunca bloqueado (sem limite de faturamento)
    - MEI/PF: bloqueado se percentual >= 98%
    """
    try:
        from decimal import Decimal

        # Mapeamento de tipo de plano para quantidade de pagamentos por ano
        PAGAMENTOS_POR_ANO = {
            'Mensal': 12,
            'Bimestral': 6,
            'Trimestral': 4,
            'Semestral': 2,
            'Anual': 1,
        }

        # Obter configuração de limites
        config = ConfiguracaoLimite.get_config()
        limite_mei = float(config.valor_anual)
        limite_pf = float(config.valor_anual_pf)

        # Buscar formas de pagamento do usuário
        formas_pgto = Tipos_pgto.objects.filter(
            usuario=request.user
        ).select_related(
            'conta_bancaria',
            'conta_bancaria__instituicao',
            'dados_bancarios'
        ).order_by('nome_identificacao', 'nome')

        resultado = []

        for forma in formas_pgto:
            conta = forma.conta_bancaria
            instituicao = conta.instituicao if conta else None
            tipo_integracao = instituicao.tipo_integracao if instituicao else None

            # Nome da instituição (prioriza conta_bancaria, fallback para dados_bancarios legado)
            instituicao_nome = None
            if instituicao:
                instituicao_nome = instituicao.nome
            elif forma.dados_bancarios and forma.dados_bancarios.instituicao:
                instituicao_nome = forma.dados_bancarios.instituicao

            # Dados básicos
            item = {
                'id': forma.id,
                'nome': forma.nome,
                'nome_identificacao': forma.nome_identificacao or forma.nome,
                'tipo_conta': conta.tipo_conta if conta else forma.tipo_conta,
                'tipo_conta_label': '',
                'tipo_integracao': tipo_integracao,
                'instituicao_nome': instituicao_nome,
                'tem_conta_bancaria': conta is not None,
                'limite_anual': 0,
                'total_projetado': 0,
                'percentual_utilizado': 0,
                'margem_disponivel': 0,
                'clientes_ativos': 0,
                'bloqueado': False,
                'motivo_bloqueio': None,
                'status': 'disponivel',  # disponivel, proximo_limite, bloqueado
            }

            # Se não tem conta bancária, calcular baseado nos clientes associados via forma_pgto
            if not conta:
                # Usar limite PF como padrão para formas legadas
                tipo_conta_forma = forma.tipo_conta if hasattr(forma, 'tipo_conta') and forma.tipo_conta else 'pf'
                if tipo_conta_forma == 'mei':
                    limite_aplicavel = limite_mei
                    tipo_label = 'MEI'
                else:
                    limite_aplicavel = limite_pf
                    tipo_label = 'Pessoa Física'

                item['limite_anual'] = limite_aplicavel
                item['tipo_conta_label'] = tipo_label

                # Calcular total projetado dos clientes associados diretamente à forma de pagamento
                from .models import Cliente
                clientes_forma = Cliente.objects.filter(
                    forma_pgto=forma,
                    cancelado=False
                ).select_related('plano')

                total_projetado = Decimal('0')
                clientes_count = 0
                for cliente in clientes_forma:
                    clientes_count += 1
                    if cliente.plano:
                        pagamentos = PAGAMENTOS_POR_ANO.get(cliente.plano.nome, 12)
                        total_projetado += cliente.plano.valor * pagamentos

                item['total_projetado'] = float(total_projetado)
                item['clientes_ativos'] = clientes_count

                # Calcular percentual e margem
                if limite_aplicavel > 0:
                    percentual = (float(total_projetado) / limite_aplicavel) * 100
                    margem = limite_aplicavel - float(total_projetado)
                else:
                    percentual = 0
                    margem = 0

                item['percentual_utilizado'] = round(percentual, 1)
                item['margem_disponivel'] = max(0, margem)

                # Determinar status
                if percentual >= 98:
                    item['bloqueado'] = True
                    item['motivo_bloqueio'] = f'Limite {tipo_label} atingido ({percentual:.1f}%)'
                    item['status'] = 'bloqueado'
                elif percentual >= 80:
                    item['status'] = 'proximo_limite'
                else:
                    item['status'] = 'disponivel'

                resultado.append(item)
                continue

            # FastDePix: sem limite de faturamento, mas calcular total projetado
            if tipo_integracao == 'fastdepix':
                item['limite_anual'] = 0
                item['tipo_conta_label'] = 'FastDePix (Sem limite)'
                item['bloqueado'] = False
                item['motivo_bloqueio'] = None
                item['status'] = 'disponivel'

                # Calcular total projetado dos clientes associados
                clientes_conta = ClienteContaBancaria.objects.filter(
                    conta_bancaria=conta,
                    ativo=True,
                    cliente__cancelado=False
                ).select_related('cliente__plano')

                total_projetado = Decimal('0')
                clientes_ids = set()
                for cc in clientes_conta:
                    clientes_ids.add(cc.cliente_id)
                    if cc.cliente.plano:
                        pagamentos = PAGAMENTOS_POR_ANO.get(cc.cliente.plano.nome, 12)
                        total_projetado += cc.cliente.plano.valor * pagamentos

                # Também buscar clientes associados via forma_pgto
                from .models import Cliente
                clientes_forma = Cliente.objects.filter(
                    forma_pgto=forma,
                    cancelado=False
                ).exclude(id__in=clientes_ids).select_related('plano')

                for cliente in clientes_forma:
                    clientes_ids.add(cliente.id)
                    if cliente.plano:
                        pagamentos = PAGAMENTOS_POR_ANO.get(cliente.plano.nome, 12)
                        total_projetado += cliente.plano.valor * pagamentos

                item['total_projetado'] = float(total_projetado)
                item['clientes_ativos'] = len(clientes_ids)

                resultado.append(item)
                continue

            # Calcular limite aplicável (MEI ou PF)
            tipo_conta = conta.tipo_conta
            if tipo_conta == 'mei':
                limite_aplicavel = limite_mei
                tipo_label = 'MEI'
            else:
                limite_aplicavel = limite_pf
                tipo_label = 'Pessoa Física'

            item['limite_anual'] = limite_aplicavel
            item['tipo_conta_label'] = tipo_label

            # Calcular total projetado dos clientes ATIVOS associados
            # Primeiro, buscar via ClienteContaBancaria
            clientes_conta = ClienteContaBancaria.objects.filter(
                conta_bancaria=conta,
                ativo=True,
                cliente__cancelado=False
            ).select_related('cliente__plano')

            total_projetado = Decimal('0')
            clientes_ids = set()
            for cc in clientes_conta:
                clientes_ids.add(cc.cliente_id)
                if cc.cliente.plano:
                    pagamentos = PAGAMENTOS_POR_ANO.get(cc.cliente.plano.nome, 12)
                    total_projetado += cc.cliente.plano.valor * pagamentos

            # Também buscar clientes associados diretamente via forma_pgto (que não estão em ClienteContaBancaria)
            from .models import Cliente
            clientes_forma = Cliente.objects.filter(
                forma_pgto=forma,
                cancelado=False
            ).exclude(id__in=clientes_ids).select_related('plano')

            for cliente in clientes_forma:
                clientes_ids.add(cliente.id)
                if cliente.plano:
                    pagamentos = PAGAMENTOS_POR_ANO.get(cliente.plano.nome, 12)
                    total_projetado += cliente.plano.valor * pagamentos

            item['total_projetado'] = float(total_projetado)
            item['clientes_ativos'] = len(clientes_ids)

            # Calcular percentual utilizado e margem
            if limite_aplicavel > 0:
                percentual = (float(total_projetado) / limite_aplicavel) * 100
                margem = limite_aplicavel - float(total_projetado)
            else:
                percentual = 0
                margem = 0

            item['percentual_utilizado'] = round(percentual, 1)
            item['margem_disponivel'] = max(0, margem)

            # Determinar status baseado no percentual
            if percentual >= 98:
                item['bloqueado'] = True
                item['motivo_bloqueio'] = f'Limite {tipo_label} atingido ({percentual:.1f}%)'
                item['status'] = 'bloqueado'
            elif percentual >= 80:
                item['status'] = 'proximo_limite'
            else:
                item['status'] = 'disponivel'

            resultado.append(item)

        return JsonResponse({
            'success': True,
            'formas': resultado,
            'limite_mei': limite_mei,
            'limite_pf': limite_pf,
        })

    except Exception as e:
        logger.error(f"[API] Erro ao listar formas de pagamento disponíveis: {e}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["GET"])
def api_forma_pagamento_detalhes(request, pk):
    """
    Retorna os detalhes de uma forma de pagamento.
    Inclui dados da conta bancaria, instituicao e credenciais (mascaradas).
    """
    try:
        forma_pgto = Tipos_pgto.objects.select_related(
            'conta_bancaria',
            'conta_bancaria__instituicao',
            'dados_bancarios'
        ).get(pk=pk, usuario=request.user)

        conta = forma_pgto.conta_bancaria
        instituicao = conta.instituicao if conta else None
        dados_bancarios = forma_pgto.dados_bancarios

        data = {
            'success': True,
            'forma_pgto': {
                'id': forma_pgto.id,
                'nome': forma_pgto.nome,
                'nome_identificacao': forma_pgto.nome_identificacao,
                'tipo_conta': forma_pgto.tipo_conta,
            },
            'conta_bancaria': None,
            'instituicao': None,
            'dados_bancarios': None,
        }

        if conta:
            data['conta_bancaria'] = {
                'id': conta.id,
                'nome_identificacao': conta.nome_identificacao,
                'beneficiario': conta.beneficiario,
                'tipo_conta': conta.tipo_conta,
                'tipo_conta_display': 'Pessoa Física' if conta.tipo_conta == 'pf' else 'MEI',
                'limite_mensal': float(conta.limite_mensal) if conta.limite_mensal else None,
                'tipo_chave_pix': conta.tipo_chave_pix,
                'tipo_chave_pix_display': dict(ContaBancaria.TIPO_CHAVE_PIX).get(conta.tipo_chave_pix, conta.tipo_chave_pix),
                'chave_pix': conta.chave_pix,
                'ambiente_sandbox': conta.ambiente_sandbox,
                'has_api_key': bool(conta.api_key),
                'has_client_id': bool(conta.api_client_id),
                'has_client_secret': bool(conta.api_client_secret),
                'has_access_token': bool(conta.api_access_token),
            }

        if instituicao:
            data['instituicao'] = {
                'id': instituicao.id,
                'nome': instituicao.nome,
                'tipo_integracao': instituicao.tipo_integracao,
            }

        # Dados bancarios legados (para formas antigas sem conta bancaria)
        if dados_bancarios:
            data['dados_bancarios'] = {
                'id': dados_bancarios.id,
                'beneficiario': dados_bancarios.beneficiario,
                'instituicao': dados_bancarios.instituicao,
                'tipo_chave': dados_bancarios.tipo_chave,
                'chave': dados_bancarios.chave,
            }
        elif not conta:
            # Se nao tem dados_bancarios vinculado E é forma antiga (sem conta_bancaria),
            # buscar DadosBancarios do usuario para pre-preencher (PIX, Cartao ou Boleto antigos)
            dados_usuario = DadosBancarios.objects.filter(usuario=request.user).first()
            if dados_usuario:
                data['dados_bancarios'] = {
                    'id': dados_usuario.id,
                    'beneficiario': dados_usuario.beneficiario,
                    'instituicao': dados_usuario.instituicao,
                    'tipo_chave': dados_usuario.tipo_chave,
                    'chave': dados_usuario.chave,
                }

        return JsonResponse(data)

    except Tipos_pgto.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Forma de pagamento nao encontrada.'
        }, status=404)
    except Exception as e:
        logger.error(f"Erro ao buscar detalhes da forma de pagamento: {e}")
        return JsonResponse({
            'success': False,
            'error': 'Erro ao buscar detalhes.'
        }, status=500)


@login_required
@require_http_methods(["GET"])
def api_forma_pagamento_clientes_count(request, pk):
    """Retorna quantidade de clientes ATIVOS associados a uma forma de pagamento."""
    try:
        forma_pgto = Tipos_pgto.objects.get(pk=pk, usuario=request.user)
        count = Cliente.objects.filter(
            forma_pgto=forma_pgto,
            usuario=request.user,
            cancelado=False
        ).count()
        return JsonResponse({'success': True, 'count': count})
    except Tipos_pgto.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Forma de pagamento nao encontrada.'
        }, status=404)


@login_required
@require_http_methods(["POST", "PUT"])
def api_forma_pagamento_atualizar(request, pk):
    """
    Atualiza uma forma de pagamento existente.
    Permite atualizar dados da conta bancaria e credenciais.
    """
    try:
        forma_pgto = Tipos_pgto.objects.select_related(
            'conta_bancaria',
            'conta_bancaria__instituicao'
        ).get(pk=pk, usuario=request.user)

        conta = forma_pgto.conta_bancaria
        if not conta:
            return JsonResponse({
                'success': False,
                'error': 'Esta forma de pagamento nao possui conta bancaria associada.'
            }, status=400)

        # Parse dos dados
        if request.content_type == 'application/json':
            data = json.loads(request.body)
        else:
            data = request.POST

        # Atualizar campos basicos
        if data.get('nome_identificacao'):
            conta.nome_identificacao = data['nome_identificacao']

        if data.get('beneficiario'):
            conta.beneficiario = data['beneficiario']

        if data.get('tipo_conta'):
            conta.tipo_conta = data['tipo_conta']

        if data.get('limite_mensal'):
            try:
                conta.limite_mensal = Decimal(str(data['limite_mensal']))
            except (InvalidOperation, ValueError):
                pass

        # Dados PIX
        if data.get('tipo_chave_pix'):
            conta.tipo_chave_pix = data['tipo_chave_pix']

        if data.get('chave_pix'):
            # Verificar se ja existe outra conta com essa chave
            if ContaBancaria.objects.filter(
                usuario=request.user,
                chave_pix=data['chave_pix']
            ).exclude(id=conta.id).exists():
                return JsonResponse({
                    'success': False,
                    'error': 'Ja existe outra conta bancaria com essa chave PIX.'
                }, status=400)
            conta.chave_pix = data['chave_pix']

        # Ambiente sandbox
        if 'ambiente_sandbox' in data:
            conta.ambiente_sandbox = data['ambiente_sandbox'] in [True, 'true', 'on', '1', 1]

        # Credenciais API (apenas atualizar se fornecidas)
        if data.get('api_key'):
            conta.api_key = data['api_key']

        if data.get('api_client_id'):
            conta.api_client_id = data['api_client_id']

        if data.get('api_client_secret'):
            conta.api_client_secret = data['api_client_secret']

        if data.get('api_access_token'):
            conta.api_access_token = data['api_access_token']

        conta.save()

        # Processar clientes associados (para instituições com API)
        clientes_associados_str = data.get('clientes_associados', '')
        clientes_transferir_str = data.get('clientes_para_transferir', '')

        # Combinar clientes selecionados + clientes a transferir
        todos_clientes_ids = set()
        if clientes_associados_str:
            try:
                ids = [int(x) for x in clientes_associados_str.split(',') if x.strip()]
                todos_clientes_ids.update(ids)
            except (ValueError, TypeError):
                pass

        if clientes_transferir_str:
            try:
                ids = [int(x) for x in clientes_transferir_str.split(',') if x.strip()]
                todos_clientes_ids.update(ids)
            except (ValueError, TypeError):
                pass

        # Validar mínimo de 1 cliente
        if not todos_clientes_ids:
            return JsonResponse({
                'success': False,
                'error': 'É obrigatório manter ao menos um cliente associado à forma de pagamento.'
            }, status=400)

        # Identificar clientes que estavam associados ANTES (para limpar forma_pgto dos desmarcados)
        # Via ClienteContaBancaria
        clientes_associados_antes = set(
            ClienteContaBancaria.objects.filter(
                conta_bancaria=conta,
                ativo=True
            ).values_list('cliente_id', flat=True)
        )
        # Via Cliente.forma_pgto direto (formas antigas)
        clientes_forma_pgto_antiga = set(
            Cliente.objects.filter(
                usuario=request.user,
                forma_pgto=forma_pgto,
                cancelado=False
            ).values_list('id', flat=True)
        )
        # Todos os clientes que estavam associados de alguma forma
        clientes_antes = clientes_associados_antes | clientes_forma_pgto_antiga

        try:
            clientes_ids = list(todos_clientes_ids)

            # Desativar associações antigas desta conta
            ClienteContaBancaria.objects.filter(
                conta_bancaria=conta,
                ativo=True
            ).update(ativo=False)

            # Criar/reativar novas associações para clientes selecionados
            for cliente_id in clientes_ids:
                try:
                    cliente = Cliente.objects.get(id=cliente_id, usuario=request.user)

                    # Desativar associações deste cliente com outras contas
                    ClienteContaBancaria.objects.filter(
                        cliente=cliente,
                        ativo=True
                    ).exclude(conta_bancaria=conta).update(ativo=False)

                    # Criar ou atualizar associação
                    assoc, created = ClienteContaBancaria.objects.update_or_create(
                        cliente=cliente,
                        conta_bancaria=conta,
                        defaults={'ativo': True}
                    )

                    # Atualizar forma de pagamento do cliente
                    cliente.forma_pgto = forma_pgto
                    cliente.save(update_fields=['forma_pgto'])

                except Cliente.DoesNotExist:
                    continue

            # Limpar forma_pgto dos clientes DESMARCADOS (estavam antes mas não estão mais)
            clientes_desmarcados = clientes_antes - todos_clientes_ids
            if clientes_desmarcados:
                Cliente.objects.filter(
                    id__in=clientes_desmarcados,
                    usuario=request.user,
                    forma_pgto=forma_pgto  # Só limpa se ainda aponta para esta forma
                ).update(forma_pgto=None)

        except Exception:
            pass

        log_user_action(
            request=request,
            action=UserActionLog.ACTION_UPDATE,
            instance=conta,
            message="Conta bancaria atualizada via API.",
            extra={
                "forma_pgto_id": forma_pgto.id,
                "nome_identificacao": conta.nome_identificacao,
            },
        )

        return JsonResponse({
            'success': True,
            'message': 'Forma de pagamento atualizada com sucesso!'
        })

    except Tipos_pgto.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Forma de pagamento nao encontrada.'
        }, status=404)
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Dados invalidos.'
        }, status=400)
    except Exception as e:
        logger.error(f"Erro ao atualizar forma de pagamento: {e}")
        return JsonResponse({
            'success': False,
            'error': 'Erro ao atualizar forma de pagamento.'
        }, status=500)


@login_required
@require_http_methods(["POST", "PUT"])
def api_forma_pagamento_antiga_atualizar(request, pk):
    """
    Atualiza clientes associados a uma forma de pagamento antiga (sem conta bancaria).
    Usado para formas PIX, Cartao, Boleto criadas antes do novo sistema.
    Tambem permite editar dados bancarios (modelo DadosBancarios).
    """
    try:
        forma_pgto = Tipos_pgto.objects.select_related('dados_bancarios').get(pk=pk, usuario=request.user)

        # Parse dos dados
        if request.content_type == 'application/json':
            data = json.loads(request.body)
        else:
            data = request.POST

        # ===== PROCESSAR DADOS BASICOS =====
        nome_identificacao = (data.get('nome_identificacao') or '').strip()
        tipo_conta = (data.get('tipo_conta') or '').strip()
        tipo_pagamento = (data.get('tipo_pagamento') or '').strip()
        beneficiario = (data.get('beneficiario') or '').strip()
        instituicao = (data.get('instituicao') or '').strip()
        tipo_chave = (data.get('tipo_chave') or '').strip()
        chave = (data.get('chave') or '').strip()

        # Atualizar campos básicos da forma de pagamento
        campos_atualizados = []

        if nome_identificacao:
            forma_pgto.nome_identificacao = nome_identificacao
            campos_atualizados.append('nome_identificacao')

        if tipo_conta in ['pf', 'mei']:
            forma_pgto.tipo_conta = tipo_conta
            campos_atualizados.append('tipo_conta')

        # Atualizar nome da forma de pagamento se fornecido
        if tipo_pagamento and tipo_pagamento in ['PIX', 'Cartao de Credito', 'Boleto']:
            forma_pgto.nome = tipo_pagamento
            campos_atualizados.append('nome')

        if campos_atualizados:
            forma_pgto.save(update_fields=campos_atualizados)

        # Se algum campo de dados bancarios foi preenchido, criar/atualizar DadosBancarios
        if any([beneficiario, instituicao, tipo_chave, chave]):
            if forma_pgto.dados_bancarios:
                # Atualizar existente (já vinculado à forma de pagamento)
                dados_bancarios = forma_pgto.dados_bancarios
                dados_bancarios.beneficiario = beneficiario or dados_bancarios.beneficiario
                dados_bancarios.instituicao = instituicao or dados_bancarios.instituicao
                dados_bancarios.tipo_chave = tipo_chave or dados_bancarios.tipo_chave
                dados_bancarios.chave = chave or dados_bancarios.chave
                dados_bancarios.save()
            else:
                # Verificar se já existe DadosBancarios para o usuário (dados antigos do Perfil)
                dados_bancarios_existente = DadosBancarios.objects.filter(
                    usuario=request.user
                ).first()

                if dados_bancarios_existente:
                    # Reutilizar registro existente e vincular à forma de pagamento
                    dados_bancarios = dados_bancarios_existente
                    dados_bancarios.beneficiario = beneficiario or dados_bancarios.beneficiario
                    dados_bancarios.instituicao = instituicao or dados_bancarios.instituicao
                    dados_bancarios.tipo_chave = tipo_chave or dados_bancarios.tipo_chave
                    dados_bancarios.chave = chave or dados_bancarios.chave
                    dados_bancarios.save()
                else:
                    # Criar novo apenas se não existir nenhum para o usuário
                    dados_bancarios = DadosBancarios.objects.create(
                        usuario=request.user,
                        beneficiario=beneficiario,
                        instituicao=instituicao,
                        tipo_chave=tipo_chave,
                        chave=chave,
                    )

                forma_pgto.dados_bancarios = dados_bancarios
                forma_pgto.save(update_fields=['dados_bancarios'])

        # ===== PROCESSAR CLIENTES ASSOCIADOS =====
        # Processar clientes associados
        clientes_associados_str = data.get('clientes_associados') or ''
        clientes_transferir_str = data.get('clientes_para_transferir') or ''

        # Combinar clientes selecionados + clientes a transferir
        todos_clientes_ids = set()
        if clientes_associados_str:
            try:
                ids = [int(x) for x in clientes_associados_str.split(',') if x.strip()]
                todos_clientes_ids.update(ids)
            except (ValueError, TypeError):
                pass

        if clientes_transferir_str:
            try:
                ids = [int(x) for x in clientes_transferir_str.split(',') if x.strip()]
                todos_clientes_ids.update(ids)
            except (ValueError, TypeError):
                pass

        # Validar mínimo de 1 cliente
        if not todos_clientes_ids:
            return JsonResponse({
                'success': False,
                'error': 'É obrigatório manter ao menos um cliente associado à forma de pagamento.'
            }, status=400)

        # Identificar clientes que estavam associados ANTES
        clientes_antes = set(
            Cliente.objects.filter(
                usuario=request.user,
                forma_pgto=forma_pgto,
                cancelado=False
            ).values_list('id', flat=True)
        )

        # Atualizar clientes selecionados para usar esta forma de pagamento
        for cliente_id in todos_clientes_ids:
            try:
                cliente = Cliente.objects.get(id=cliente_id, usuario=request.user)
                cliente.forma_pgto = forma_pgto
                cliente.save(update_fields=['forma_pgto'])
            except Cliente.DoesNotExist:
                continue

        # Limpar forma_pgto dos clientes DESMARCADOS
        clientes_desmarcados = clientes_antes - todos_clientes_ids
        if clientes_desmarcados:
            Cliente.objects.filter(
                id__in=clientes_desmarcados,
                usuario=request.user,
                forma_pgto=forma_pgto
            ).update(forma_pgto=None)

        log_user_action(
            request=request,
            action=UserActionLog.ACTION_UPDATE,
            instance=forma_pgto,
            message="Forma de pagamento antiga atualizada.",
            extra={
                "forma_pgto_id": forma_pgto.id,
                "forma_pgto_nome": forma_pgto.nome,
                "clientes_associados": len(todos_clientes_ids),
                "clientes_removidos": len(clientes_desmarcados),
                "dados_bancarios_atualizados": any([beneficiario, instituicao, tipo_chave, chave]),
            },
        )

        return JsonResponse({
            'success': True,
            'message': 'Forma de pagamento atualizada com sucesso!'
        })

    except Tipos_pgto.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Forma de pagamento nao encontrada.'
        }, status=404)
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Dados invalidos.'
        }, status=400)
    except Exception as e:
        logger.error(f"Erro ao atualizar clientes da forma de pagamento antiga: {e}")
        return JsonResponse({
            'success': False,
            'error': 'Erro ao atualizar clientes.'
        }, status=500)


# =============================================================================
# VIEWS DE INTEGRACOES API
# =============================================================================

@login_required
@user_passes_test(lambda u: u.is_superuser)
def integracoes_api_index(request):
    """
    Pagina inicial das integracoes API.
    Lista todas as integracoes disponiveis com status.
    """
    from django.utils import timezone
    from datetime import timedelta

    # Calcular estatisticas FastDePix
    contas_fastdepix = ContaBancaria.objects.filter(
        instituicao__tipo_integracao='fastdepix'
    )

    fastdepix_contas = contas_fastdepix.count()
    fastdepix_status = 'not_configured'

    if fastdepix_contas > 0:
        # Verifica se alguma conta tem api_key configurada (campo encriptado é _api_key)
        contas_com_api = contas_fastdepix.filter(_api_key__isnull=False).exclude(_api_key='')
        if contas_com_api.exists():
            # Verifica se tem webhook configurado
            contas_com_webhook = contas_com_api.filter(webhook_id__isnull=False).exclude(webhook_id='')
            if contas_com_webhook.exists():
                fastdepix_status = 'connected'
            else:
                fastdepix_status = 'partial'

    # Cobrancas do mês atual (primeiro ao último dia)
    hoje = timezone.now()
    inicio_mes = hoje.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    fastdepix_cobrancas = CobrancaPix.objects.filter(
        criado_em__gte=inicio_mes
    ).count()

    context = {
        'page': 'integracoes-api',
        'page_group': 'admin',
        'fastdepix_status': fastdepix_status,
        'fastdepix_contas': fastdepix_contas,
        'fastdepix_cobrancas': fastdepix_cobrancas,
    }

    return render(request, 'pages/admin/integracoes/index.html', context)


@login_required
@user_passes_test(lambda u: u.is_superuser)
def integracoes_fastdepix(request):
    """
    Pagina de configuracao do FastDePix.
    Suporta seleção de conta específica via query param ?conta=ID
    """
    from django.utils import timezone
    from django.conf import settings
    from datetime import timedelta
    from django.db.models import Count
    from nossopainel.services.payment_integrations import get_payment_integration

    # Buscar contas FastDePix com contagem de cobranças
    contas_fastdepix = ContaBancaria.objects.filter(
        instituicao__tipo_integracao='fastdepix'
    ).select_related('usuario', 'instituicao').annotate(
        total_cobrancas=Count('cobrancas_pix')
    )

    # Não há conta selecionada por padrão - seleção é feita via JavaScript
    conta_selecionada = None
    conta_ativa = None

    # Status da API
    api_status = 'not_configured'
    api_key_masked = None
    webhook_status = 'not_configured'

    if conta_ativa:
        # Mascarar API Key
        if conta_ativa.api_key:
            key = conta_ativa.api_key
            api_key_masked = f"{key[:8]}...{key[-4:]}" if len(key) > 12 else "****"
            api_status = 'connected'  # API Key configurada = conectado
        else:
            api_status = 'no_api_key'

        # Status do webhook
        if conta_ativa.webhook_id:
            webhook_status = 'configured'

    # URL do webhook do sistema
    # Em produção, usa o domínio principal do ALLOWED_HOSTS
    # Em desenvolvimento (DEBUG=True), permite URL customizada
    is_development = getattr(settings, 'DEBUG', False)

    # Determinar URL base para webhook
    if is_development:
        # Em desenvolvimento, usar HTTP local
        webhook_url_default = request.build_absolute_uri('/api/pix/webhook/')
    else:
        # Em produção, usar domínio configurado
        allowed_hosts = getattr(settings, 'ALLOWED_HOSTS', [])
        # Pegar primeiro domínio que não seja localhost ou *
        production_domain = None
        for host in allowed_hosts:
            if host not in ['localhost', '127.0.0.1', '*', '']:
                production_domain = host
                break

        if production_domain:
            # Usar HTTPS em produção
            webhook_url_default = f"https://{production_domain}/api/pix/webhook/"
        else:
            # Fallback: forçar HTTPS para evitar redirect que perde o body
            webhook_url_default = f"https://{request.get_host()}/api/pix/webhook/"

    # Webhook URL customizada salva na sessao (para desenvolvimento)
    webhook_url_custom = ''
    if is_development:
        webhook_url_custom = request.session.get('fastdepix_webhook_url_custom', '')

    # Estatisticas do mês atual (primeiro ao último dia)
    hoje = timezone.now()
    inicio_mes = hoje.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    # Base queryset: cobranças via FastDePix
    # Se há conta selecionada, filtrar por ela; senão, mostrar todas do usuário
    cobrancas_base = CobrancaPix.objects.filter(
        integracao='fastdepix',
        criado_em__gte=inicio_mes
    )

    if conta_selecionada:
        # Filtrar por conta selecionada
        cobrancas_base = cobrancas_base.filter(conta_bancaria=conta_selecionada)
    else:
        # Mostrar todas as contas do usuário logado
        cobrancas_base = cobrancas_base.filter(usuario=request.user)

    cobrancas_pagas = cobrancas_base.filter(status='paid')

    # Calcular valores agregados
    valores_agregados = cobrancas_pagas.aggregate(
        total_valor=Sum('valor'),
        total_recebido=Sum('valor_recebido'),
        total_taxa=Sum('valor_taxa')
    )

    valor_pago = valores_agregados['total_valor'] or 0
    valor_recebido = valores_agregados['total_recebido'] or 0
    valor_taxa = valores_agregados['total_taxa'] or 0

    # Se não temos valor_taxa mas temos valor_pago e valor_recebido, calcular
    if valor_taxa == 0 and valor_pago > 0 and valor_recebido > 0:
        valor_taxa = valor_pago - valor_recebido

    # Calcular indicadores de taxa
    total_pagas = cobrancas_pagas.count()
    taxa_media = valor_taxa / total_pagas if total_pagas > 0 else 0
    taxa_percentual = (valor_taxa / valor_pago * 100) if valor_pago > 0 else 0

    stats = {
        'total_cobrancas': cobrancas_base.count(),
        'total_pagas': total_pagas,
        'total_pendentes': cobrancas_base.filter(status='pending').count(),
        'total_expiradas': cobrancas_base.filter(status='expired').count(),
        'valor_pago': valor_pago,
        'valor_recebido': valor_recebido,
        'taxa_media': taxa_media,
        'taxa_percentual': taxa_percentual,
        'valor_taxa': valor_taxa,
    }

    context = {
        'page': 'integracoes-fastdepix',
        'page_group': 'admin',
        'contas_fastdepix': contas_fastdepix,
        'conta_ativa': conta_ativa,
        'conta_selecionada': conta_selecionada,
        'api_status': api_status,
        'api_key_masked': api_key_masked,
        'webhook_status': webhook_status,
        'webhook_url': webhook_url_default,
        'webhook_url_custom': webhook_url_custom,
        'is_development': is_development,
        'stats': stats,
    }

    return render(request, 'pages/admin/integracoes/fastdepix.html', context)


@login_required
@user_passes_test(lambda u: u.is_superuser)
@require_http_methods(["GET"])
def integracoes_fastdepix_conta_dados(request, conta_id=None):
    """
    Retorna dados de uma conta específica ou estatísticas gerais (JSON).
    GET /admin/integracoes-api/fastdepix/conta/<conta_id>/
    GET /admin/integracoes-api/fastdepix/conta/ (sem conta = todas)
    """
    from django.utils import timezone
    from datetime import timedelta

    desde = timezone.now() - timedelta(days=30)

    # Base queryset para estatísticas
    cobrancas_base = CobrancaPix.objects.filter(
        integracao='fastdepix',
        criado_em__gte=desde
    )

    conta_data = None

    if conta_id:
        # Buscar conta específica
        try:
            conta = ContaBancaria.objects.select_related('usuario', 'instituicao').get(
                id=conta_id,
                instituicao__tipo_integracao='fastdepix'
            )
            cobrancas_base = cobrancas_base.filter(conta_bancaria=conta)

            # Mascarar API Key
            api_key_masked = None
            if conta.api_key:
                key = conta.api_key
                api_key_masked = f"{key[:8]}...{key[-4:]}" if len(key) > 12 else "****"

            conta_data = {
                'id': conta.id,
                'nome': conta.nome_identificacao or conta.beneficiario,
                'usuario': conta.usuario.username,
                'api_key_masked': api_key_masked,
                'has_api_key': bool(conta.api_key),
                'webhook_configured': bool(conta.webhook_id),
                'api_status': 'connected' if conta.api_key else 'no_api_key',
            }
        except ContaBancaria.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Conta não encontrada'}, status=404)
    else:
        # Sem conta selecionada = todas do usuário
        cobrancas_base = cobrancas_base.filter(usuario=request.user)

    # Calcular estatísticas
    cobrancas_pagas = cobrancas_base.filter(status='paid')
    valores_agregados = cobrancas_pagas.aggregate(
        total_valor=Sum('valor'),
        total_recebido=Sum('valor_recebido'),
        total_taxa=Sum('valor_taxa')
    )

    valor_pago = valores_agregados['total_valor'] or 0
    valor_recebido = valores_agregados['total_recebido'] or 0
    valor_taxa = valores_agregados['total_taxa'] or 0

    if valor_taxa == 0 and valor_pago > 0 and valor_recebido > 0:
        valor_taxa = valor_pago - valor_recebido

    # Calcular indicadores de taxa
    total_pagas = cobrancas_pagas.count()
    taxa_media = float(valor_taxa / total_pagas) if total_pagas > 0 else 0
    taxa_percentual = float(valor_taxa / valor_pago * 100) if valor_pago > 0 else 0

    stats = {
        'total_cobrancas': cobrancas_base.count(),
        'total_pagas': total_pagas,
        'total_pendentes': cobrancas_base.filter(status='pending').count(),
        'total_expiradas': cobrancas_base.filter(status='expired').count(),
        'valor_pago': float(valor_pago),
        'valor_recebido': float(valor_recebido),
        'valor_taxa': float(valor_taxa),
        'taxa_media': taxa_media,
        'taxa_percentual': taxa_percentual,
    }

    return JsonResponse({
        'success': True,
        'conta': conta_data,
        'stats': stats,
    })


@login_required
@user_passes_test(lambda u: u.is_superuser)
@require_http_methods(["POST"])
def integracoes_fastdepix_testar(request):
    """
    Testa a conexao com a API FastDePix.
    """
    from nossopainel.services.payment_integrations import get_payment_integration

    # Buscar conta ativa (campo encriptado é _api_key)
    conta_ativa = ContaBancaria.objects.filter(
        instituicao__tipo_integracao='fastdepix',
        _api_key__isnull=False
    ).exclude(_api_key='').first()

    if not conta_ativa:
        return JsonResponse({
            'success': False,
            'message': 'Nenhuma conta FastDePix configurada com API Key.'
        })

    # Testar conexao
    integration = get_payment_integration(conta_ativa)
    if not integration:
        return JsonResponse({
            'success': False,
            'message': 'Erro ao inicializar integracao.'
        })

    success, message = integration.test_connection()

    return JsonResponse({
        'success': success,
        'message': message
    })


@login_required
@user_passes_test(lambda u: u.is_superuser)
@require_http_methods(["POST"])
def integracoes_fastdepix_webhook_registrar(request):
    """
    Registra o webhook na API FastDePix.
    Aceita URL customizada via POST JSON body (para desenvolvimento).
    """
    from django.conf import settings
    from nossopainel.services.payment_integrations import get_payment_integration, PaymentIntegrationError

    # Buscar conta ativa (campo encriptado é _api_key)
    conta_ativa = ContaBancaria.objects.filter(
        instituicao__tipo_integracao='fastdepix',
        _api_key__isnull=False
    ).exclude(_api_key='').first()

    if not conta_ativa:
        return JsonResponse({
            'success': False,
            'message': 'Nenhuma conta FastDePix configurada com API Key.'
        })

    # URL do webhook - pode vir do body (desenvolvimento) ou usar padrao
    webhook_path = '/api/pix/webhook/'
    # Forçar HTTPS para evitar redirect que perde o body do POST
    webhook_url = f"https://{request.get_host()}{webhook_path}"

    # Em desenvolvimento, permitir URL base customizada
    if getattr(settings, 'DEBUG', False):
        try:
            body = json.loads(request.body) if request.body else {}
            custom_base_url = body.get('url', '').strip()
            if custom_base_url:
                # Remove barra final se existir e adiciona o path correto
                custom_base_url = custom_base_url.rstrip('/')
                webhook_url = f"{custom_base_url}{webhook_path}"
        except json.JSONDecodeError:
            pass

    # Registrar webhook
    integration = get_payment_integration(conta_ativa)
    if not integration:
        return JsonResponse({
            'success': False,
            'message': 'Erro ao inicializar integracao.'
        })

    try:
        result = integration.register_webhook(webhook_url)

        # Salvar dados do webhook na conta
        conta_ativa.webhook_id = str(result.get('id', ''))
        conta_ativa.webhook_secret = result.get('secret', '')
        conta_ativa.save(update_fields=['webhook_id', 'webhook_secret'])

        logger.info(f'[Integracoes] Webhook registrado para conta {conta_ativa.id} - URL: {webhook_url}')

        return JsonResponse({
            'success': True,
            'message': 'Webhook registrado com sucesso!',
            'webhook_id': conta_ativa.webhook_id,
            'webhook_url': webhook_url
        })

    except PaymentIntegrationError as e:
        logger.error(f'[Integracoes] Erro ao registrar webhook: {e}')
        return JsonResponse({
            'success': False,
            'message': f'Erro ao registrar webhook: {e.message}'
        })
    except Exception as e:
        logger.error(f'[Integracoes] Erro inesperado ao registrar webhook: {e}')
        return JsonResponse({
            'success': False,
            'message': f'Erro inesperado: {str(e)}'
        })


@login_required
@user_passes_test(lambda u: u.is_superuser)
@require_http_methods(["POST"])
def integracoes_fastdepix_webhook_salvar_url(request):
    """
    Salva URL customizada de webhook (apenas em desenvolvimento).
    Armazena em cache/sessao para uso durante desenvolvimento.
    """
    from django.conf import settings

    if not getattr(settings, 'DEBUG', False):
        return JsonResponse({
            'success': False,
            'message': 'Funcionalidade disponivel apenas em desenvolvimento.'
        })

    try:
        body = json.loads(request.body) if request.body else {}
        url = body.get('url', '').strip()

        # Salvar na sessao do usuario
        request.session['fastdepix_webhook_url_custom'] = url

        return JsonResponse({
            'success': True,
            'message': 'URL salva com sucesso.',
            'url': url
        })

    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Dados invalidos.'
        }, status=400)


@login_required
@user_passes_test(lambda u: u.is_superuser)
@require_http_methods(["POST"])
def integracoes_fastdepix_webhook_atualizar(request):
    """
    Atualiza o webhook na API FastDePix.
    """
    from nossopainel.services.payment_integrations import get_payment_integration, PaymentIntegrationError

    # Buscar conta ativa com webhook (campo encriptado é _api_key)
    conta_ativa = ContaBancaria.objects.filter(
        instituicao__tipo_integracao='fastdepix',
        _api_key__isnull=False,
        webhook_id__isnull=False
    ).exclude(_api_key='').exclude(webhook_id='').first()

    if not conta_ativa:
        return JsonResponse({
            'success': False,
            'message': 'Nenhum webhook configurado para atualizar.'
        })

    # URL do webhook - forçar HTTPS para evitar redirect que perde o body
    webhook_url = f"https://{request.get_host()}/api/pix/webhook/"

    # Atualizar webhook
    integration = get_payment_integration(conta_ativa)
    if not integration:
        return JsonResponse({
            'success': False,
            'message': 'Erro ao inicializar integracao.'
        })

    try:
        result = integration.update_webhook(conta_ativa.webhook_id, url=webhook_url)

        logger.info(f'[Integracoes] Webhook atualizado para conta {conta_ativa.id}')

        return JsonResponse({
            'success': True,
            'message': 'Webhook atualizado com sucesso!'
        })

    except PaymentIntegrationError as e:
        logger.error(f'[Integracoes] Erro ao atualizar webhook: {e}')
        return JsonResponse({
            'success': False,
            'message': f'Erro ao atualizar webhook: {e.message}'
        })
    except Exception as e:
        logger.error(f'[Integracoes] Erro inesperado ao atualizar webhook: {e}')
        return JsonResponse({
            'success': False,
            'message': f'Erro inesperado: {str(e)}'
        })


@login_required
@user_passes_test(lambda u: u.is_superuser)
@require_http_methods(["POST"])
def integracoes_fastdepix_webhook_remover(request):
    """
    Remove o webhook da API FastDePix.
    """
    from nossopainel.services.payment_integrations import get_payment_integration, PaymentIntegrationError

    # Buscar conta ativa com webhook (campo encriptado é _api_key)
    conta_ativa = ContaBancaria.objects.filter(
        instituicao__tipo_integracao='fastdepix',
        _api_key__isnull=False,
        webhook_id__isnull=False
    ).exclude(_api_key='').exclude(webhook_id='').first()

    if not conta_ativa:
        return JsonResponse({
            'success': False,
            'message': 'Nenhum webhook configurado para remover.'
        })

    # Remover webhook
    integration = get_payment_integration(conta_ativa)
    if not integration:
        return JsonResponse({
            'success': False,
            'message': 'Erro ao inicializar integracao.'
        })

    try:
        integration.delete_webhook(conta_ativa.webhook_id)

        # Limpar dados do webhook na conta
        conta_ativa.webhook_id = ''
        conta_ativa.webhook_secret = ''
        conta_ativa.save(update_fields=['webhook_id', 'webhook_secret'])

        logger.info(f'[Integracoes] Webhook removido da conta {conta_ativa.id}')

        return JsonResponse({
            'success': True,
            'message': 'Webhook removido com sucesso.'
        })

    except PaymentIntegrationError as e:
        logger.error(f'[Integracoes] Erro ao remover webhook: {e}')
        return JsonResponse({
            'success': False,
            'message': f'Erro ao remover webhook: {e.message}'
        })
    except Exception as e:
        logger.error(f'[Integracoes] Erro inesperado ao remover webhook: {e}')
        return JsonResponse({
            'success': False,
            'message': f'Erro inesperado: {str(e)}'
        })


# =============================================================================
# SINCRONIZAÇÃO DE COBRANÇAS PIX
# =============================================================================

@login_required
@user_passes_test(lambda u: u.is_superuser)
@require_http_methods(["POST"])
def integracoes_fastdepix_sincronizar(request):
    """
    Sincroniza cobranças do FastDePix com o sistema local.

    Esta função realiza duas operações principais:
    1. Busca na API todas as transações do primeiro dia do mês até hoje
    2. Para cada transação:
       - Se não existe no sistema: cria registro (se estiver paga)
       - Se existe com status diferente: atualiza o status

    Funciona com TODAS as contas FastDePix configuradas.
    """
    from nossopainel.services.payment_integrations import get_payment_integration, PaymentStatus
    from django.utils import timezone
    from decimal import Decimal

    # Buscar TODAS as contas FastDePix ativas
    contas_fastdepix = ContaBancaria.objects.filter(
        instituicao__tipo_integracao='fastdepix',
        _api_key__isnull=False
    ).exclude(_api_key='').select_related('instituicao', 'usuario')

    if not contas_fastdepix.exists():
        return JsonResponse({
            'success': False,
            'message': 'Nenhuma conta FastDePix configurada.'
        })

    # Definir período: primeiro dia do mês até hoje
    hoje = timezone.now().date()
    primeiro_dia_mes = hoje.replace(day=1)

    # Contadores gerais
    total_verificadas_api = 0
    total_pendentes_atualizadas = 0
    total_novas_encontradas = 0
    total_erros = 0
    detalhes = []
    contas_processadas = 0

    for conta in contas_fastdepix:
        integration = get_payment_integration(conta)
        if not integration:
            detalhes.append({
                'conta': conta.nome_identificacao or f'Conta {conta.id}',
                'erro': 'Erro ao inicializar integração'
            })
            total_erros += 1
            continue

        contas_processadas += 1
        conta_nome = conta.nome_identificacao or f'Conta {conta.id}'

        # ===== PARTE 1: Buscar transações da API =====
        try:
            # Buscar todas as páginas de transações do período
            page = 1
            per_page = 100
            transacoes_api = []

            while True:
                response = integration.list_transactions(
                    start_date=datetime.combine(primeiro_dia_mes, datetime.min.time()),
                    end_date=datetime.combine(hoje, datetime.max.time()),
                    page=page,
                    per_page=per_page
                )

                # Extrair transações da resposta
                data = response.get('data', response)
                if isinstance(data, list):
                    transacoes = data
                else:
                    transacoes = data.get('transactions', data.get('items', []))

                if not transacoes:
                    break

                transacoes_api.extend(transacoes)

                # Verificar se há mais páginas
                meta = response.get('meta', {})
                total_pages = meta.get('total_pages', meta.get('last_page', 1))
                if page >= total_pages:
                    break
                page += 1

            total_verificadas_api += len(transacoes_api)

            # Processar cada transação da API
            for tx in transacoes_api:
                tx_id = str(tx.get('id', tx.get('transaction_id', '')))
                if not tx_id:
                    continue

                tx_status = tx.get('status', 'pending').lower()
                tx_amount = Decimal(str(tx.get('amount', 0)))

                # Verificar se já existe no sistema
                cobranca_existente = CobrancaPix.objects.filter(
                    transaction_id=tx_id
                ).first()

                if cobranca_existente:
                    # ===== Cobrança existe: verificar se precisa atualizar status =====
                    status_map = {
                        'pending': 'pending',
                        'paid': 'paid',
                        'expired': 'expired',
                        'cancelled': 'cancelled',
                        'canceled': 'cancelled',
                        'refunded': 'refunded',
                    }
                    novo_status = status_map.get(tx_status, 'pending')

                    if cobranca_existente.status != novo_status:
                        if novo_status == 'paid':
                            # Buscar detalhes para pagamento
                            try:
                                details = integration.get_charge_details(tx_id)
                                paid_at = None
                                if details.get('paid_at'):
                                    try:
                                        paid_at = timezone.datetime.fromisoformat(
                                            details['paid_at'].replace('Z', '+00:00')
                                        )
                                    except (ValueError, AttributeError):
                                        paid_at = timezone.now()

                                payer = details.get('payer', {})
                                valor_recebido = None
                                valor_taxa = None

                                for key in ['commission_amount', 'net_amount', 'amount_received']:
                                    if key in details and details[key] is not None:
                                        try:
                                            valor_recebido = Decimal(str(details[key]))
                                            break
                                        except (ValueError, TypeError):
                                            pass

                                for key in ['fee', 'tax', 'taxa', 'fee_amount']:
                                    if key in details and details[key] is not None:
                                        try:
                                            valor_taxa = Decimal(str(details[key]))
                                            break
                                        except (ValueError, TypeError):
                                            pass

                                cobranca_existente.mark_as_paid(
                                    paid_at=paid_at,
                                    payer_name=payer.get('name') if isinstance(payer, dict) else None,
                                    payer_document=payer.get('cpf_cnpj') if isinstance(payer, dict) else None,
                                    webhook_data={'data': details},
                                    valor_recebido=valor_recebido,
                                    valor_taxa=valor_taxa,
                                )
                            except Exception:
                                cobranca_existente.mark_as_paid(paid_at=timezone.now())

                            total_pendentes_atualizadas += 1
                            detalhes.append({
                                'conta': conta_nome,
                                'transaction_id': tx_id,
                                'acao': 'ATUALIZADA',
                                'status': 'paid',
                                'mensagem': 'Cobrança existente marcada como PAGA'
                            })

                        elif novo_status == 'expired':
                            cobranca_existente.mark_as_expired()
                            total_pendentes_atualizadas += 1
                            detalhes.append({
                                'conta': conta_nome,
                                'transaction_id': tx_id,
                                'acao': 'ATUALIZADA',
                                'status': 'expired',
                                'mensagem': 'Cobrança existente marcada como EXPIRADA'
                            })

                        elif novo_status == 'cancelled':
                            cobranca_existente.mark_as_cancelled()
                            total_pendentes_atualizadas += 1
                            detalhes.append({
                                'conta': conta_nome,
                                'transaction_id': tx_id,
                                'acao': 'ATUALIZADA',
                                'status': 'cancelled',
                                'mensagem': 'Cobrança existente marcada como CANCELADA'
                            })

                else:
                    # ===== Cobrança NÃO existe: criar se estiver paga =====
                    if tx_status == 'paid':
                        try:
                            details = integration.get_charge_details(tx_id)

                            # Extrair dados do pagamento
                            paid_at = timezone.now()
                            if details.get('paid_at'):
                                try:
                                    paid_at = timezone.datetime.fromisoformat(
                                        details['paid_at'].replace('Z', '+00:00')
                                    )
                                except (ValueError, AttributeError):
                                    pass

                            # Extrair data de criação
                            created_at = timezone.now()
                            if details.get('created_at'):
                                try:
                                    created_at = timezone.datetime.fromisoformat(
                                        details['created_at'].replace('Z', '+00:00')
                                    )
                                except (ValueError, AttributeError):
                                    pass

                            # Extrair expiração
                            expires_at = timezone.now() + timedelta(hours=24)
                            if details.get('qr_code_expires_at') or details.get('expires_at'):
                                exp_str = details.get('qr_code_expires_at') or details.get('expires_at')
                                try:
                                    expires_at = timezone.datetime.fromisoformat(
                                        exp_str.replace('Z', '+00:00')
                                    )
                                except (ValueError, AttributeError):
                                    pass

                            payer = details.get('payer', {})
                            valor_recebido = None
                            valor_taxa = None

                            for key in ['commission_amount', 'net_amount', 'amount_received']:
                                if key in details and details[key] is not None:
                                    try:
                                        valor_recebido = Decimal(str(details[key]))
                                        break
                                    except (ValueError, TypeError):
                                        pass

                            for key in ['fee', 'tax', 'taxa', 'fee_amount']:
                                if key in details and details[key] is not None:
                                    try:
                                        valor_taxa = Decimal(str(details[key]))
                                        break
                                    except (ValueError, TypeError):
                                        pass

                            # Criar nova cobrança
                            nova_cobranca = CobrancaPix.objects.create(
                                transaction_id=tx_id,
                                usuario=conta.usuario,
                                conta_bancaria=conta,
                                valor=tx_amount,
                                descricao=details.get('description', f'Cobrança sincronizada - {tx_id}'),
                                status='paid',
                                integracao='fastdepix',
                                qr_code=details.get('qr_code_text', ''),
                                qr_code_url=details.get('qr_code', ''),
                                pix_copia_cola=details.get('qr_code_text', ''),
                                expira_em=expires_at,
                                pago_em=paid_at,
                                pagador_nome=payer.get('name', '') if isinstance(payer, dict) else '',
                                pagador_documento=payer.get('cpf_cnpj', '') if isinstance(payer, dict) else '',
                                valor_recebido=valor_recebido,
                                valor_taxa=valor_taxa,
                                raw_response=details,
                                webhook_data={'synced': True, 'data': details},
                            )

                            total_novas_encontradas += 1
                            detalhes.append({
                                'conta': conta_nome,
                                'transaction_id': tx_id,
                                'acao': 'CRIADA',
                                'status': 'paid',
                                'valor': str(tx_amount),
                                'mensagem': f'Nova cobrança PAGA encontrada na API (R$ {tx_amount})'
                            })
                            logger.info(f'[Sync PIX] Nova cobrança criada: {tx_id} - R$ {tx_amount}')

                        except Exception as e:
                            total_erros += 1
                            detalhes.append({
                                'conta': conta_nome,
                                'transaction_id': tx_id,
                                'erro': f'Erro ao criar cobrança: {str(e)}'
                            })
                            logger.warning(f'[Sync PIX] Erro ao criar cobrança {tx_id}: {e}')

        except Exception as e:
            total_erros += 1
            detalhes.append({
                'conta': conta_nome,
                'erro': f'Erro ao buscar transações: {str(e)}'
            })
            logger.exception(f'[Sync PIX] Erro ao processar conta {conta_nome}: {e}')

    # ===== PARTE 2: Sincronizar cobranças pendentes locais =====
    # (Para cobranças que existem no sistema mas podem ter mudado de status)
    cobrancas_pendentes_locais = CobrancaPix.objects.filter(
        integracao='fastdepix',
        status='pending',
        criado_em__gte=timezone.make_aware(datetime.combine(primeiro_dia_mes, datetime.min.time()))
    ).select_related('conta_bancaria')

    for cobranca in cobrancas_pendentes_locais:
        if not cobranca.conta_bancaria or not cobranca.conta_bancaria.api_key:
            continue

        integration = get_payment_integration(cobranca.conta_bancaria)
        if not integration:
            continue

        try:
            status_api = integration.get_charge_status(cobranca.transaction_id)

            if status_api == PaymentStatus.PAID:
                details = integration.get_charge_details(cobranca.transaction_id)
                paid_at = timezone.now()
                if details.get('paid_at'):
                    try:
                        paid_at = timezone.datetime.fromisoformat(
                            details['paid_at'].replace('Z', '+00:00')
                        )
                    except (ValueError, AttributeError):
                        pass

                payer = details.get('payer', {})
                cobranca.mark_as_paid(
                    paid_at=paid_at,
                    payer_name=payer.get('name') if isinstance(payer, dict) else None,
                    payer_document=payer.get('cpf_cnpj') if isinstance(payer, dict) else None,
                    webhook_data={'data': details},
                )
                total_pendentes_atualizadas += 1

            elif status_api == PaymentStatus.EXPIRED:
                cobranca.mark_as_expired()
                total_pendentes_atualizadas += 1

            elif status_api == PaymentStatus.CANCELLED:
                cobranca.mark_as_cancelled()
                total_pendentes_atualizadas += 1

        except Exception as e:
            logger.warning(f'[Sync PIX] Erro ao verificar pendente {cobranca.transaction_id}: {e}')

    # Resumo final
    mensagem = f'Sincronização concluída. '
    if total_novas_encontradas > 0:
        mensagem += f'{total_novas_encontradas} cobrança(s) NOVA(S) encontrada(s) na API. '
    if total_pendentes_atualizadas > 0:
        mensagem += f'{total_pendentes_atualizadas} atualizada(s). '
    if total_erros > 0:
        mensagem += f'{total_erros} erro(s).'

    return JsonResponse({
        'success': True,
        'message': mensagem,
        'periodo': f'{primeiro_dia_mes.strftime("%d/%m/%Y")} até {hoje.strftime("%d/%m/%Y")}',
        'contas_processadas': contas_processadas,
        'total_verificadas_api': total_verificadas_api,
        'novas_encontradas': total_novas_encontradas,
        'atualizadas': total_pendentes_atualizadas,
        'erros': total_erros,
        'detalhes': detalhes
    })


@login_required
@user_passes_test(lambda u: u.is_superuser)
@require_http_methods(["POST"])
def integracoes_fastdepix_revisar_valores(request):
    """
    Revisa TODAS as cobranças PAGAS do FastDePix e atualiza valores financeiros.

    Busca todas as cobranças com status='paid', consulta a API FastDePix
    e atualiza sempre que os valores forem diferentes dos retornados.
    """
    from nossopainel.services.payment_integrations import get_payment_integration
    from decimal import Decimal

    # Buscar TODAS as cobranças pagas do FastDePix
    cobrancas = CobrancaPix.objects.filter(
        usuario=request.user,
        integracao='fastdepix',
        status='paid'
    ).select_related('conta_bancaria')

    total = cobrancas.count()
    atualizadas = 0
    erros = 0
    detalhes = []

    for cobranca in cobrancas:
        if not cobranca.conta_bancaria or not cobranca.conta_bancaria.api_key:
            continue

        integration = get_payment_integration(cobranca.conta_bancaria)
        if not integration:
            continue

        try:
            details = integration.get_charge_details(cobranca.transaction_id)

            # Extrair valores - priorizar commission_amount
            amount_received = (
                details.get('commission_amount') or
                details.get('amount_received') or
                details.get('net_amount')
            )
            fee = details.get('fee') or details.get('tax')

            valor_recebido = None
            valor_taxa = None

            if amount_received:
                valor_recebido = Decimal(str(amount_received))
            if fee:
                valor_taxa = Decimal(str(fee))

            # Calcular taxa se não veio
            if valor_taxa is None and details.get('amount') and valor_recebido:
                valor_taxa = Decimal(str(details['amount'])) - valor_recebido

            # Atualizar cobrança se valores forem diferentes
            updated = False
            if valor_recebido and cobranca.valor_recebido != valor_recebido:
                cobranca.valor_recebido = valor_recebido
                updated = True
            if valor_taxa and cobranca.valor_taxa != valor_taxa:
                cobranca.valor_taxa = valor_taxa
                updated = True

            if updated:
                cobranca.save(update_fields=['valor_recebido', 'valor_taxa'])
                atualizadas += 1
                detalhes.append({
                    'transaction_id': cobranca.transaction_id[:8] + '...',
                    'valor_recebido': str(valor_recebido),
                    'valor_taxa': str(valor_taxa),
                    'mensagem': 'Valores atualizados'
                })

        except Exception as e:
            erros += 1
            detalhes.append({
                'transaction_id': cobranca.transaction_id[:8] + '...',
                'erro': str(e)
            })

    return JsonResponse({
        'success': True,
        'message': f'{atualizadas} de {total} cobranças atualizadas.',
        'total': total,
        'atualizadas': atualizadas,
        'erros': erros,
        'detalhes': detalhes
    })


# =============================================================================
# CONFIGURAÇÃO DE AGENDAMENTOS
# =============================================================================

@login_required
@user_passes_test(lambda u: u.is_superuser)
def config_agendamentos(request):
    """
    Página de configuração dos agendamentos do sistema.
    Permite visualizar, ativar/desativar e configurar templates de mensagem.
    """
    from .models import ConfiguracaoAgendamento

    if request.method == 'POST':
        # Atualizar configuração de um job
        job_nome = request.POST.get('job_nome')
        acao = request.POST.get('acao')

        if not job_nome:
            return JsonResponse({'success': False, 'message': 'Job não especificado.'})

        try:
            job = ConfiguracaoAgendamento.objects.get(nome=job_nome)

            # Verificar se o job é editável
            if job.bloqueado and acao != 'visualizar':
                return JsonResponse({'success': False, 'message': 'Este job não pode ser editado.'})

            if acao == 'toggle':
                job.ativo = not job.ativo
                job.save(update_fields=['ativo'])
                return JsonResponse({
                    'success': True,
                    'message': f'Job {"ativado" if job.ativo else "desativado"} com sucesso.',
                    'ativo': job.ativo
                })

            elif acao == 'atualizar_template':
                template_key = request.POST.get('template_key')
                template_value = request.POST.get('template_value', '')

                if not template_key:
                    return JsonResponse({'success': False, 'message': 'Chave do template não especificada.'})

                templates = job.templates_mensagem or {}
                templates[template_key] = template_value
                job.templates_mensagem = templates
                job.save(update_fields=['templates_mensagem'])

                return JsonResponse({
                    'success': True,
                    'message': 'Template atualizado com sucesso.'
                })

            elif acao == 'atualizar_horario':
                import re
                horario = request.POST.get('horario', '')

                # Validar formato HH:MM
                if not re.match(r'^([01]?[0-9]|2[0-3]):[0-5][0-9]$', horario):
                    return JsonResponse({'success': False, 'message': 'Formato de horário inválido. Use HH:MM.'})

                job.horario = horario
                job.save(update_fields=['horario'])

                return JsonResponse({
                    'success': True,
                    'message': f'Horário atualizado para {horario}. Reinicie o scheduler para aplicar.'
                })

            else:
                return JsonResponse({'success': False, 'message': 'Ação inválida.'})

        except ConfiguracaoAgendamento.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Job não encontrado.'})
        except Exception as e:
            logger.error(f'[Agendamentos] Erro ao processar ação: {e}')
            return JsonResponse({'success': False, 'message': f'Erro: {str(e)}'})

    # GET - Exibir página (ordem alfabética pelo nome de exibição)
    jobs = ConfiguracaoAgendamento.objects.all().order_by('nome_exibicao')

    context = {
        'page': 'config-agendamentos',
        'page_group': 'admin',
        'jobs': jobs,
    }

    return render(request, 'pages/admin/agendamentos.html', context)


# ============================================================================
# RELATÓRIO DE PAGAMENTOS
# ============================================================================

@login_required
def relatorio_pagamentos(request):
    """
    Relatório de pagamentos e transações PIX.

    Exibe pagamentos via PIX (API FastDePix), transações pendentes/expiradas,
    e pagamentos registrados manualmente.

    GET /admin/relatorios/pagamentos/

    Query params:
        - data_inicio: Data inicial (YYYY-MM-DD)
        - data_fim: Data final (YYYY-MM-DD)
        - tipo: 'pix', 'manual' ou 'todos' (default)
        - status: 'paid', 'pending', 'expired' ou 'todos' (default)
    """
    from django.db.models import Sum, Count
    from datetime import datetime, timedelta

    # Filtros
    data_inicio_str = request.GET.get('data_inicio', '')
    data_fim_str = request.GET.get('data_fim', '')
    tipo_filtro = request.GET.get('tipo', 'todos')
    status_filtro = request.GET.get('status', 'todos')

    # Parse de datas (default: mês atual - primeiro dia até hoje)
    hoje = timezone.now().date()
    primeiro_dia_mes = hoje.replace(day=1)
    try:
        data_inicio = datetime.strptime(data_inicio_str, '%Y-%m-%d').date() if data_inicio_str else primeiro_dia_mes
    except ValueError:
        data_inicio = primeiro_dia_mes

    try:
        data_fim = datetime.strptime(data_fim_str, '%Y-%m-%d').date() if data_fim_str else hoje
    except ValueError:
        data_fim = hoje

    # ========== TRANSAÇÕES PIX ==========
    # Base query para todas as cobranças PIX do período
    cobrancas_pix_base = CobrancaPix.objects.filter(
        usuario=request.user,
        criado_em__date__gte=data_inicio,
        criado_em__date__lte=data_fim
    ).select_related('cliente', 'mensalidade', 'conta_bancaria')

    # Pagamentos PIX (status='paid')
    pagamentos_pix = cobrancas_pix_base.filter(status='paid').order_by('-pago_em')

    # Transações pendentes (status='pending')
    transacoes_pendentes = cobrancas_pix_base.filter(status='pending').order_by('-criado_em')

    # Transações expiradas (status='expired')
    transacoes_expiradas = cobrancas_pix_base.filter(status='expired').order_by('-criado_em')

    # ========== PAGAMENTOS MANUAIS (Mensalidades pagas sem CobrancaPix) ==========
    from django.db.models import Exists, OuterRef

    cobranca_paga_subquery = CobrancaPix.objects.filter(
        mensalidade=OuterRef('pk'),
        status='paid'
    )

    pagamentos_manuais = Mensalidade.objects.filter(
        cliente__usuario=request.user,
        pgto=True,
        dt_pagamento__isnull=False,
        dt_pagamento__gte=data_inicio,
        dt_pagamento__lte=data_fim
    ).exclude(
        Exists(cobranca_paga_subquery)
    ).select_related('cliente', 'cliente__plano').order_by('-dt_pagamento')

    # Aplicar filtro de tipo
    if tipo_filtro == 'pix':
        pagamentos_manuais = Mensalidade.objects.none()
    elif tipo_filtro == 'manual':
        pagamentos_pix = CobrancaPix.objects.none()
        transacoes_pendentes = CobrancaPix.objects.none()
        transacoes_expiradas = CobrancaPix.objects.none()

    # Aplicar filtro de status (apenas para PIX)
    if status_filtro == 'paid':
        transacoes_pendentes = CobrancaPix.objects.none()
        transacoes_expiradas = CobrancaPix.objects.none()
    elif status_filtro == 'pending':
        pagamentos_pix = CobrancaPix.objects.none()
        transacoes_expiradas = CobrancaPix.objects.none()
        pagamentos_manuais = Mensalidade.objects.none()
    elif status_filtro == 'expired':
        pagamentos_pix = CobrancaPix.objects.none()
        transacoes_pendentes = CobrancaPix.objects.none()
        pagamentos_manuais = Mensalidade.objects.none()

    # ========== ESTATÍSTICAS ==========
    stats_pix = pagamentos_pix.aggregate(
        total=Sum('valor'),
        total_recebido=Sum('valor_recebido'),
        total_taxa=Sum('valor_taxa'),
        qtd=Count('id')
    )

    stats_manual = pagamentos_manuais.aggregate(
        total=Sum('valor'),
        qtd=Count('id')
    )

    stats_pendentes = transacoes_pendentes.aggregate(
        total=Sum('valor'),
        qtd=Count('id')
    )

    stats_expiradas = transacoes_expiradas.aggregate(
        total=Sum('valor'),
        qtd=Count('id')
    )

    # Calcular taxa média e percentual
    total_pix_valor = stats_pix['total'] or 0
    total_pix_recebido = stats_pix['total_recebido'] or 0
    total_pix_taxa = stats_pix['total_taxa'] or 0
    qtd_pix = stats_pix['qtd'] or 0

    # Se não temos valor_taxa mas temos valor_pago e valor_recebido, calcular
    if total_pix_taxa == 0 and total_pix_valor > 0 and total_pix_recebido > 0:
        total_pix_taxa = total_pix_valor - total_pix_recebido

    taxa_media = total_pix_taxa / qtd_pix if qtd_pix > 0 else 0
    taxa_percentual = (total_pix_taxa / total_pix_valor * 100) if total_pix_valor > 0 else 0

    # Preparar lista unificada de transações para exibição
    pagamentos_lista = []

    # Função auxiliar para adicionar transação PIX à lista
    def adicionar_transacao_pix(pix, status):
        webhook_data_json = None
        if pix.webhook_data:
            try:
                webhook_data_json = json.dumps(pix.webhook_data, ensure_ascii=False, default=str)
            except (TypeError, ValueError):
                webhook_data_json = None

        # Usar pago_em para pagos, criado_em para outros status
        data_ref = pix.pago_em if status == 'paid' and pix.pago_em else pix.criado_em

        pagamentos_lista.append({
            'tipo': 'pix',
            'status': status,
            'data': data_ref,
            'cliente_nome': pix.cliente.nome if pix.cliente else 'N/A',
            'cliente_id': pix.cliente.id if pix.cliente else None,
            'valor': pix.valor,
            'valor_recebido': pix.valor_recebido,
            'valor_taxa': pix.valor_taxa,
            'conta': pix.conta_bancaria.nome_identificacao if pix.conta_bancaria else 'N/A',
            'transaction_id': pix.transaction_id,
            'cobranca_id': str(pix.id),
            'mensalidade_id': pix.mensalidade_id,
            'webhook_data': webhook_data_json,
            'pagador_nome': pix.pagador_nome,
            'pagador_documento': pix.pagador_documento,
            'criado_em': pix.criado_em,
        })

    # Adicionar pagamentos PIX (status='paid')
    for pix in pagamentos_pix:
        adicionar_transacao_pix(pix, 'paid')

    # Adicionar transações pendentes (status='pending')
    for pix in transacoes_pendentes:
        adicionar_transacao_pix(pix, 'pending')

    # Adicionar transações expiradas (status='expired')
    for pix in transacoes_expiradas:
        adicionar_transacao_pix(pix, 'expired')

    # Adicionar pagamentos manuais
    for mens in pagamentos_manuais:
        pagamentos_lista.append({
            'tipo': 'manual',
            'status': 'paid',
            'data': timezone.make_aware(datetime.combine(mens.dt_pagamento, datetime.min.time())),
            'cliente_nome': mens.cliente.nome if mens.cliente else 'N/A',
            'cliente_id': mens.cliente.id if mens.cliente else None,
            'valor': mens.valor,
            'valor_recebido': None,
            'valor_taxa': None,
            'conta': 'Manual',
            'transaction_id': None,
            'cobranca_id': None,
            'mensalidade_id': mens.id,
        })

    # Ordenar por data (mais recente primeiro)
    pagamentos_lista.sort(key=lambda x: x['data'] if x['data'] else timezone.now(), reverse=True)

    context = {
        'page': 'relatorio_pagamentos',
        'page_group': 'relatorios',
        'pagamentos': pagamentos_lista,

        # Estatísticas - Pagamentos confirmados
        'total_pix': total_pix_valor,
        'total_pix_recebido': total_pix_recebido,
        'total_pix_taxa': total_pix_taxa,
        'qtd_pix': qtd_pix,
        'total_manual': stats_manual['total'] or 0,
        'qtd_manual': stats_manual['qtd'] or 0,
        'total_geral': total_pix_valor + (stats_manual['total'] or 0),
        'qtd_total': qtd_pix + (stats_manual['qtd'] or 0),

        # Estatísticas - Pendentes e Expiradas
        'total_pendentes': stats_pendentes['total'] or 0,
        'qtd_pendentes': stats_pendentes['qtd'] or 0,
        'total_expiradas': stats_expiradas['total'] or 0,
        'qtd_expiradas': stats_expiradas['qtd'] or 0,

        # Estatísticas - Taxas
        'taxa_media': taxa_media,
        'taxa_percentual': taxa_percentual,

        # Filtros (para manter no form)
        'data_inicio': data_inicio.strftime('%Y-%m-%d'),
        'data_fim': data_fim.strftime('%Y-%m-%d'),
        'tipo_filtro': tipo_filtro,
        'status_filtro': status_filtro,
    }

    return render(request, 'pages/relatorios/pagamentos.html', context)


@login_required
@require_http_methods(["GET"])
def api_cliente_mensalidades(request, cliente_id):
    """
    Retorna a lista de mensalidades de um cliente.

    GET /api/clientes/<cliente_id>/mensalidades/

    Returns:
        JSON com lista de mensalidades
    """
    try:
        cliente = Cliente.objects.get(id=cliente_id, usuario=request.user)
    except Cliente.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Cliente não encontrado'}, status=404)

    mensalidades = Mensalidade.objects.filter(
        cliente=cliente
    ).select_related('cliente').order_by('-dt_vencimento')[:24]  # Últimas 24 mensalidades

    # Buscar cobranças PIX relacionadas
    from django.db.models import Prefetch
    cobrancas_por_mensalidade = {}
    cobrancas = CobrancaPix.objects.filter(
        mensalidade__in=mensalidades
    ).order_by('-criado_em')

    for cob in cobrancas:
        if cob.mensalidade_id not in cobrancas_por_mensalidade:
            cobrancas_por_mensalidade[cob.mensalidade_id] = cob

    lista = []
    for mens in mensalidades:
        cobranca = cobrancas_por_mensalidade.get(mens.id)
        lista.append({
            'id': mens.id,
            'dt_vencimento': mens.dt_vencimento.strftime('%d/%m/%Y') if mens.dt_vencimento else None,
            'dt_pagamento': mens.dt_pagamento.strftime('%d/%m/%Y') if mens.dt_pagamento else None,
            'dt_cancelamento': mens.dt_cancelamento.strftime('%d/%m/%Y') if mens.dt_cancelamento else None,
            'valor': float(mens.valor) if mens.valor else 0,
            'pgto': mens.pgto,
            'cancelado': mens.cancelado,
            'status': 'Pago' if mens.pgto else ('Cancelada' if mens.cancelado or (mens.dt_vencimento and mens.dt_vencimento < timezone.now().date()) else 'Em aberto'),
            'tipo_pagamento': 'PIX' if cobranca and cobranca.status == 'paid' else ('Manual' if mens.pgto else None),
            'transaction_id': cobranca.transaction_id if cobranca else None,
        })

    return JsonResponse({
        'success': True,
        'cliente': {
            'id': cliente.id,
            'nome': cliente.nome,
            'telefone': cliente.telefone,
        },
        'mensalidades': lista,
        'total': len(lista),
    })
